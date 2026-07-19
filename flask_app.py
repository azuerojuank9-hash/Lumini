import os, sys
from dotenv import load_dotenv
_basedir = os.path.abspath(os.path.dirname(__file__))
load_dotenv(os.path.join(_basedir, '.env'))
from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, send_file, abort, jsonify, g, Response
import sqlite3, hashlib, time, secrets, logging, json, uuid, bcrypt
from datetime import timedelta, datetime
from io import BytesIO
import html

app = Flask(__name__)
ENV = os.environ.get('FLASK_ENV', 'production')

# ── LOGGING (antes que la config para que logger esté disponible) ──────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(_basedir, 'lumini.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ── Environment-aware config ──────────────────────────────────────────────
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 86400 * 7 if ENV == 'production' else 86400
_raw_secret = (os.environ.get('SECRET_KEY') or '').strip()
if not _raw_secret:
    raise RuntimeError(
        "SECRET_KEY no está definido en .env. "
        "Genera una con: python -c \"import secrets; print(secrets.token_hex(32))\""
    )
app.secret_key = _raw_secret
app.permanent_session_lifetime = timedelta(hours=4)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# SESSION_COOKIE_SECURE: environment-appropriate default, explicit override OK
_secure_override = os.environ.get('SESSION_COOKIE_SECURE', '')
if _secure_override:
    app.config['SESSION_COOKIE_SECURE'] = _secure_override.lower() in ('true', '1', 'yes')
else:
    app.config['SESSION_COOKIE_SECURE'] = (ENV == 'production')

if ENV == 'production' and app.config['SESSION_COOKIE_SECURE']:
    logger.info("Producción — SESSION_COOKIE_SECURE=True, asegúrate de tener HTTPS.")
elif ENV == 'production' and not app.config['SESSION_COOKIE_SECURE']:
    logger.warning("Producción con SESSION_COOKIE_SECURE=False. Usa HTTPS y establece SESSION_COOKIE_SECURE=true.")

app.config['JSON_AS_ASCII'] = False
app.config['TEMPLATES_AUTO_RELOAD'] = ENV != 'production'

if ENV == 'production':
    try:
        from flask_compress import Compress
        Compress(app)
        app.config['COMPRESS_ALGORITHM'] = 'gzip'
        app.config['COMPRESS_LEVEL'] = 6
        app.config['COMPRESS_MIN_SIZE'] = 500
    except ImportError:
        pass

@app.template_filter('parse_json')
def parse_json_filter(val):
    try: return json.loads(val) if val else {}
    except Exception: return {}

# ── CONFIGURACIÓN ─────────────────────────────────────────────────────────────
DB_FOLDER   = os.path.join(os.path.dirname(__file__), 'colegios_db')
MASTER_DB   = os.path.join(os.path.dirname(__file__), 'master.db')
LOGO_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'logos')
os.makedirs(DB_FOLDER, exist_ok=True)
os.makedirs(LOGO_FOLDER, exist_ok=True)

# ── CREDENCIALES DESDE .env ───────────────────────────────────────────────────
ADMIN_PASSWORD   = (os.environ.get('ADMIN_PASSWORD') or '').strip()
SENDGRID_API_KEY = (os.environ.get('SENDGRID_API_KEY') or '').strip()
EMAIL_ORIGEN     = (os.environ.get('EMAIL_ORIGEN') or 'lumini.appag@gmail.com').strip()

if not ADMIN_PASSWORD:
    raise RuntimeError("ADMIN_PASSWORD no está definido en .env. Crea el archivo .env con: ADMIN_PASSWORD=tu_clave")
if not SENDGRID_API_KEY:
    logger.warning("SENDGRID_API_KEY no definido — el envío de correos estará deshabilitado.")

app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024

def extension_permitida(filename):
    ext = ('.' + filename.rsplit('.', 1)[-1]).lower() if '.' in filename else ''
    return ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp')

def validar_imagen(ruta):
    try:
        from PIL import Image
        img = Image.open(ruta)
        img.verify()
        return True
    except Exception:
        return False

JORNADAS = ['Mañana', 'Tarde', 'Nocturna']

MATERIAS = [
    'Artes', 'Matemáticas', 'Cipol y Econ', 'Física', 'Química',
    'Español', 'Inglés', 'Biología', 'Sociales',
    'Tecnología e Informática', 'Filosofía', 'Educación Física'
]

PREGUNTAS_SECRETAS = [
    '¿Cuál es el nombre de tu mascota?',
    '¿En qué ciudad naciste?',
    '¿Cuál es el nombre de tu colegio favorito?',
    '¿Cuál es tu comida favorita?',
    '¿Cuál es el nombre de tu mejor amigo(a)?',
    '¿Cuál es tu color favorito?',
    '¿Cuál es el nombre de tu madre?',
    '¿Cuál es tu deporte favorito?',
]

# ── CSRF ──────────────────────────────────────────────────────────────────────
def generar_csrf():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)
    return session['_csrf_token']

def validar_csrf():
    token = request.form.get('_csrf_token') or request.headers.get('X-CSRF-Token')
    return bool(token and token == session.get('_csrf_token'))

app.jinja_env.globals['csrf_token'] = generar_csrf

@app.context_processor
def inject_theme():
    def accent_css(colegio):
        primary = (colegio and colegio['primary_color']) or '#7C3AED'
        secondary = (colegio and colegio['secondary_color']) or '#6D28D9'
        h = primary.lstrip('#')
        rgb = f'{int(h[0:2], 16)},{int(h[2:4], 16)},{int(h[4:6], 16)}' if len(h) == 6 else '124,58,237'
        return f'--accent:{primary};--accent2:{secondary};--accent-rgb:{rgb};'
    return dict(accent_css=accent_css)

@app.context_processor
def inject_rector_defaults():
    return dict(
        total_estudiantes=0, total_profesores=0, total_cursos=0,
        total_materias=0, total_directoras=0, asistencia_hoy=0, asis_pct_r=0,
    )

# ── API v1 BLUEPRINT ────────────────────────────────────────────────────────────
try:
    from api.v1.auth import bp as api_v1_bp
    app.register_blueprint(api_v1_bp)
    logger.info("API v1 blueprint registrado.")
except ImportError as e:
    logger.warning(f"No se pudo registrar API v1: {e}")

# ── NUEVOS BLUEPRINTS (app/) ──────────────────────────────────────────────
try:
    from app.routes import rector_bp, admin_bp, parent_bp, student_bp
    from app.routes.main_routes import main_bp
    from app.routes.auth import auth_bp
    app.register_blueprint(rector_bp)
    app.register_blueprint(admin_bp)
    app.register_blueprint(parent_bp)
    app.register_blueprint(student_bp)
    app.register_blueprint(main_bp)
    app.register_blueprint(auth_bp)
    logger.info("Blueprints modulares (app/routes/) registrados.")
except ImportError as e:
    logger.warning(f"No se pudieron registrar blueprints modulares: {e}")

# ── FUERZA BRUTA ──────────────────────────────────────────────────────────────
login_intentos = {}

def ip_bloqueada(ip, prefijo=''):
    clave = f'{prefijo}_{ip}'
    d = login_intentos.get(clave)
    if not d: return False
    if d['bloqueado_hasta'] and time.time() < d['bloqueado_hasta']:
        return int(d['bloqueado_hasta'] - time.time())
    return False

def registrar_fallo(ip, prefijo=''):
    _purgar_intentos_antiguos()
    clave = f'{prefijo}_{ip}'
    d = login_intentos.setdefault(clave, {'intentos': 0, 'bloqueado_hasta': None})
    d['intentos'] += 1
    if d['intentos'] >= 5:
        d['bloqueado_hasta'] = time.time() + 300
        logger.warning(f"IP bloqueada por fuerza bruta: {ip} (ctx={prefijo})")
    return d['intentos']

def _purgar_intentos_antiguos():
    ahora = time.time()
    viejas = [k for k, v in login_intentos.items()
              if v['bloqueado_hasta'] and ahora > v['bloqueado_hasta'] + 3600]
    for k in viejas:
        del login_intentos[k]

def limpiar_intentos(ip, prefijo=''):
    login_intentos.pop(f'{prefijo}_{ip}', None)

# ── HASH ──────────────────────────────────────────────────────────────────────
def hash_pw(pw, _sal=None):
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verificar_pw(plano, guardada):
    if not guardada: return False
    if guardada.startswith('$2b$') or guardada.startswith('$2a$'):
        return bcrypt.checkpw(plano.encode(), guardada.encode())
    if '$' in guardada:
        partes = guardada.split('$', 1)
        if len(partes) == 2:
            sal, h = partes
            return hashlib.sha256((sal + plano).encode()).hexdigest() == h
        return False
    return hashlib.sha256(plano.encode()).hexdigest() == guardada

def necesita_rehash(guardada):
    return not (guardada.startswith('$2b$') or guardada.startswith('$2a$'))

# ── SCHEMA VERSIONING ──────────────────────────────────────────────────────────
SCHEMA_VERSION = 20

def _ejecutar_migraciones(slug, conn):
    conn.execute('''CREATE TABLE IF NOT EXISTS schema_meta (
        version INTEGER PRIMARY KEY,
        applied_at TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.commit()
    row = conn.execute('SELECT COALESCE(MAX(version), 0) as v FROM schema_meta').fetchone()
    current = row['v'] if row else 0
    for v in range(current + 1, SCHEMA_VERSION + 1):
        mig_fn = MIGRACIONES.get(v)
        if mig_fn:
            logger.info(f"[{slug}] Migrando a versión {v}...")
            mig_fn(conn, slug)
            conn.execute('INSERT OR IGNORE INTO schema_meta (version) VALUES (?)', (v,))
            conn.commit()
            logger.info(f"[{slug}] Versión {v} aplicada.")

def _migrar_v6(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        email TEXT NOT NULL,
        password_hash TEXT,
        nombre TEXT NOT NULL,
        apellido TEXT DEFAULT '',
        tipo_documento TEXT DEFAULT '',
        documento TEXT DEFAULT '',
        telefono TEXT DEFAULT '',
        avatar TEXT DEFAULT '',
        activo INTEGER DEFAULT 1,
        creado TEXT DEFAULT (datetime('now','localtime')),
        actualizado TEXT DEFAULT (datetime('now','localtime')),
        ultimo_acceso TEXT,
        UNIQUE(slug, email)
    )''')

def _migrar_v7(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS roles_base (
        codigo TEXT PRIMARY KEY,
        nombre_default TEXT NOT NULL,
        nivel INTEGER NOT NULL,
        descripcion TEXT
    )''')
    roles_default = [
        ('admin',     'Administrador',     0, 'Acceso global al sistema'),
        ('rector',    'Rector',            1, 'Máxima autoridad institucional'),
        ('authority', 'Autoridad Académica',2, 'Coordinadores, decanos, directores'),
        ('teacher',   'Docente',           3, 'Profesores e instructores'),
        ('student',   'Estudiante',         4, 'Alumnos y participantes'),
        ('guardian',  'Acudiente',          5, 'Padres y representantes'),
    ]
    for cod, nom, niv, desc in roles_default:
        conn.execute('INSERT OR IGNORE INTO roles_base (codigo, nombre_default, nivel, descripcion) VALUES (?,?,?,?)',
                    (cod, nom, niv, desc))
    conn.execute('''CREATE TABLE IF NOT EXISTS roles_instancia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        codigo TEXT NOT NULL,
        nombre TEXT NOT NULL,
        jerarquia INTEGER DEFAULT 1,
        activo INTEGER DEFAULT 1,
        UNIQUE(slug, codigo)
    )''')

def _migrar_v8(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS usuarios_roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER NOT NULL,
        rol_id INTEGER NOT NULL,
        entidad_tipo TEXT,
        entidad_id INTEGER,
        asignado_por INTEGER,
        creado TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(usuario_id, rol_id, entidad_tipo, entidad_id)
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS password_resets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER NOT NULL,
        token TEXT UNIQUE NOT NULL,
        expira TEXT NOT NULL,
        usado INTEGER DEFAULT 0,
        creado TEXT DEFAULT (datetime('now','localtime'))
    )''')

def _migrar_v9(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS config_institucion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL UNIQUE,
        tipo_evaluacion TEXT DEFAULT 'numerica',
        escala_min REAL DEFAULT 1.0,
        escala_max REAL DEFAULT 10.0,
        nota_minima_aprobar REAL DEFAULT 6.0,
        decimales_notas INTEGER DEFAULT 1,
        creditos_activo INTEGER DEFAULT 0,
        escala_conceptual TEXT DEFAULT '["A","B","C","D","E","F"]',
        num_periodos INTEGER DEFAULT 4,
        periodos_json TEXT,
        jornadas_json TEXT,
        jerarquia_activa INTEGER DEFAULT 0,
        niveles_json TEXT,
        roles_json TEXT,
        acuse_recibo INTEGER DEFAULT 1,
        firmas_activas INTEGER DEFAULT 0,
        notas_publicas_entre_pares INTEGER DEFAULT 0,
        idioma TEXT DEFAULT 'es',
        huso_horario TEXT DEFAULT 'America/Bogota',
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    )''')
    if slug:
        conn.execute('''INSERT OR IGNORE INTO config_institucion
            (slug, num_periodos, jornadas_json, roles_json)
            VALUES (?, 4, '["Mañana","Tarde","Nocturna"]',
            '{"rector":"Rector","authority":"Coordinador","teacher":"Docente","student":"Estudiante","guardian":"Acudiente"}')''',
            (slug,))

def _migrar_v10(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER,
        accion TEXT NOT NULL,
        tabla TEXT NOT NULL,
        registro_id INTEGER,
        valor_anterior TEXT,
        valor_nuevo TEXT,
        ip TEXT,
        user_agent TEXT,
        creado TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_audit_tabla ON audit_log(tabla, registro_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_audit_usuario ON audit_log(usuario_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_audit_fecha ON audit_log(creado)')

def _migrar_v11(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS asistencia_v2 (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        aid INTEGER, fecha TEXT, estado TEXT,
        UNIQUE(aid, fecha)
    )''')
    conn.execute('''INSERT OR IGNORE INTO asistencia_v2 (id, aid, fecha, estado)
        SELECT id, aid, fecha, estado FROM asistencia''')
    conn.execute('DROP TABLE asistencia')
    conn.execute('ALTER TABLE asistencia_v2 RENAME TO asistencia')

    conn.execute('''CREATE TABLE IF NOT EXISTS estructura_academica (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        nivel INTEGER NOT NULL,
        nombre TEXT NOT NULL,
        nombre_tipo TEXT DEFAULT '',
        padre_id INTEGER,
        activo INTEGER DEFAULT 1,
        UNIQUE(slug, nivel, nombre)
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS curso_nuevo (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        estructura_id INTEGER,
        nombre TEXT NOT NULL,
        jornada TEXT DEFAULT 'Mañana',
        activo INTEGER DEFAULT 1
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS materias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        nombre TEXT NOT NULL,
        activo INTEGER DEFAULT 1,
        UNIQUE(slug, nombre)
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS curso_materias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        curso_id INTEGER NOT NULL,
        materia_id INTEGER NOT NULL,
        docente_id INTEGER,
        UNIQUE(curso_id, materia_id)
    )''')

def _recrear_si_unique_incorrecto(conn, slug, tabla, unique_deseado, sql_insert, sql_select):
    import re as _re
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name=?",
        (tabla,)
    ).fetchone()
    if not row:
        return False
    sql_actual = row['sql']
    m = _re.search(r'UNIQUE\s*\(([^)]+)\)', sql_actual, _re.IGNORECASE)
    if m:
        cols_actuales = [c.strip().lower() for c in m.group(1).split(',')]
        cols_deseadas = [c.strip().lower() for c in unique_deseado.strip('()').split(',')]
        if cols_actuales == cols_deseadas:
            return False
    logger.warning(f'[{slug}] Recreando tabla {tabla} (UNIQUE incorrecto)')
    conn.execute(f'ALTER TABLE {tabla} RENAME TO {tabla}_old')
    conn.execute(sql_insert)
    conn.execute(f'INSERT OR IGNORE INTO {tabla} {sql_select}')
    conn.execute(f'DROP TABLE {tabla}_old')
    conn.commit()
    return True

def _migrar_v12(conn, slug=None):
    _recrear_si_unique_incorrecto(conn, slug, 'evaluaciones',
        '(aid,profesor_id,materia,jornada,periodo)',
        '''CREATE TABLE evaluaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
            materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
            evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
            UNIQUE(aid,profesor_id,materia,jornada,periodo))''',
        '''(id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
           SELECT id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,
                  COALESCE(periodo,1) FROM evaluaciones_old''')
    _recrear_si_unique_incorrecto(conn, slug, 'horarios_curso',
        '(curso,jornada,dia,franja)',
        '''CREATE TABLE horarios_curso (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            curso TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
            dia TEXT NOT NULL, franja TEXT NOT NULL,
            num TEXT DEFAULT "", materia TEXT DEFAULT "", profesor TEXT DEFAULT "",
            UNIQUE(curso, jornada, dia, franja))''',
        'SELECT * FROM horarios_curso_old')

def _migrar_v13(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS auditoria_notas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER NOT NULL,
        rol TEXT NOT NULL,
        creado TEXT DEFAULT (datetime('now','localtime')),
        ip TEXT,
        curso TEXT,
        materia TEXT,
        periodo INTEGER,
        tipo_accion TEXT NOT NULL,
        tabla TEXT NOT NULL,
        registro_id INTEGER,
        aid INTEGER NOT NULL,
        actividad_id INTEGER,
        campo TEXT,
        valor_anterior TEXT,
        valor_nuevo TEXT,
        motivo TEXT
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_auditoria_notas_aid ON auditoria_notas(aid)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_auditoria_notas_curso ON auditoria_notas(curso, materia, periodo)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_auditoria_notas_curso_prof ON auditoria_notas(curso, materia, periodo, profesor_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_auditoria_notas_fecha ON auditoria_notas(creado)')

def _migrar_v14(conn, slug=None):
    # Handle legacy solicitudes_modificacion table (old schema from pre-v14)
    cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='solicitudes_modificacion'")
    if cur.fetchone():
        cols = {r[1] for r in conn.execute("PRAGMA table_info(solicitudes_modificacion)").fetchall()}
        if 'slug' not in cols:
            conn.execute("DROP TABLE IF EXISTS solicitudes_modificacion")
    conn.execute('''CREATE TABLE IF NOT EXISTS solicitudes_modificacion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        aid INTEGER NOT NULL,
        profesor_id INTEGER NOT NULL,
        materia TEXT NOT NULL,
        curso TEXT NOT NULL,
        jornada TEXT NOT NULL,
        periodo INTEGER NOT NULL DEFAULT 1,
        tipo TEXT NOT NULL CHECK(tipo IN ('actividad', 'evaluacion', 'autoevaluacion')),
        actividad_id INTEGER,
        valor_actual TEXT,
        valor_solicitado TEXT NOT NULL,
        motivo TEXT NOT NULL,
        estado TEXT NOT NULL DEFAULT 'pendiente' CHECK(estado IN ('pendiente', 'aprobada', 'rechazada')),
        aprobado_por INTEGER,
        fecha_solicitud TEXT DEFAULT (datetime('now','localtime')),
        fecha_respuesta TEXT
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_solicitudes_slug_estado ON solicitudes_modificacion(slug, estado)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_solicitudes_profesor ON solicitudes_modificacion(profesor_id, slug)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_solicitudes_aid ON solicitudes_modificacion(aid)')
    # Ensure slug default for existing rows
    conn.execute("UPDATE solicitudes_modificacion SET slug=? WHERE slug IS NULL OR slug=''", (slug or '',))

def _migrar_v15(conn, slug=None):
    # Expand asistencia: new states + metadata columns
    cur = conn.execute("PRAGMA table_info(asistencia)")
    cols = {r[1] for r in cur.fetchall()}
    for col, ddl in [
        ('observacion',   'observacion TEXT DEFAULT ""'),
        ('hora',          'hora TEXT DEFAULT ""'),
        ('usuario_tipo',  'usuario_tipo TEXT DEFAULT "profesor"'),
        ('usuario_id',    'usuario_id INTEGER DEFAULT 0'),
    ]:
        if col not in cols:
            conn.execute(f'ALTER TABLE asistencia ADD COLUMN {ddl}')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_asistencia_fecha_estado ON asistencia(fecha, estado)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_asistencia_aid_estado ON asistencia(aid, estado)')

def _migrar_v16(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS firmas_digitales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        usuario_tipo TEXT NOT NULL,
        usuario_id INTEGER NOT NULL,
        nombre TEXT NOT NULL,
        documento_tipo TEXT NOT NULL,
        documento_id INTEGER NOT NULL,
        hash_documento TEXT NOT NULL,
        firma_hash TEXT NOT NULL,
        metodo TEXT DEFAULT 'hmac-sha256',
        ip TEXT DEFAULT '',
        user_agent TEXT DEFAULT '',
        creado TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_firmas_doc ON firmas_digitales(documento_tipo, documento_id)')

    conn.execute('''CREATE TABLE IF NOT EXISTS enterprise_audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        usuario_id INTEGER,
        usuario_tipo TEXT DEFAULT '',
        accion TEXT NOT NULL,
        categoria TEXT DEFAULT '',
        descripcion TEXT DEFAULT '',
        tabla TEXT DEFAULT '',
        registro_id INTEGER,
        valor_anterior TEXT,
        valor_nuevo TEXT,
        ip TEXT DEFAULT '',
        user_agent TEXT DEFAULT '',
        dispositivo TEXT DEFAULT '',
        navegador TEXT DEFAULT '',
        sesion_id TEXT DEFAULT '',
        nivel TEXT DEFAULT 'info',
        creado TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_enterprise_audit_slug ON enterprise_audit_log(slug, creado)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_enterprise_audit_accion ON enterprise_audit_log(accion)')

    conn.execute('''CREATE TABLE IF NOT EXISTS observador_registros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        aid INTEGER NOT NULL,
        tipo TEXT NOT NULL CHECK(tipo IN ('positivo','llamado','compromiso','seguimiento')),
        texto TEXT NOT NULL,
        docente TEXT DEFAULT '',
        materia TEXT DEFAULT '',
        estado TEXT DEFAULT 'pendiente' CHECK(estado IN ('pendiente','aprobado','rechazado')),
        aprobado_por_tipo TEXT,
        aprobado_por_id INTEGER,
        fecha TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_observador_aid ON observador_registros(aid)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_observador_tipo ON observador_registros(tipo)')

    conn.execute('''CREATE TABLE IF NOT EXISTS expediente_documentos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        aid INTEGER NOT NULL,
        tipo TEXT NOT NULL,
        nombre TEXT NOT NULL,
        archivo TEXT DEFAULT '',
        descripcion TEXT DEFAULT '',
        subido_por_tipo TEXT DEFAULT 'rector',
        subido_por_id INTEGER DEFAULT 0,
        fecha TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_expediente_aid ON expediente_documentos(aid)')

    conn.execute('''CREATE TABLE IF NOT EXISTS eventos_calendario (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        tipo TEXT NOT NULL DEFAULT 'evento',
        titulo TEXT NOT NULL,
        descripcion TEXT DEFAULT '',
        fecha_inicio TEXT NOT NULL,
        fecha_fin TEXT,
        todo_el_dia INTEGER DEFAULT 1,
        curso TEXT DEFAULT '',
        creado_por_tipo TEXT DEFAULT 'rector',
        creado_por_id INTEGER DEFAULT 0,
        color TEXT DEFAULT '#6c63ff',
        fecha_creacion TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_eventos_fecha ON eventos_calendario(slug, fecha_inicio)')

    conn.execute('''CREATE TABLE IF NOT EXISTS pagos_estructura (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT NOT NULL,
        alumno_id INTEGER NOT NULL,
        concepto TEXT NOT NULL,
        monto REAL NOT NULL,
        descuento REAL DEFAULT 0,
        pagado REAL DEFAULT 0,
        estado TEXT DEFAULT 'pendiente' CHECK(estado IN ('pendiente','pagado','parcial','anulado')),
        fecha_vencimiento TEXT,
        fecha_pago TEXT,
        metodo_pago TEXT DEFAULT '',
        referencia TEXT DEFAULT '',
        notas TEXT DEFAULT '',
        creado TEXT DEFAULT (datetime('now','localtime')),
        actualizado TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_pagos_alumno ON pagos_estructura(alumno_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_pagos_estado ON pagos_estructura(estado)')

def _migrar_v17(conn, slug=None):
    cols = {r[1] for r in conn.execute("PRAGMA table_info(actividades)").fetchall()}
    for col, ddl in {
        'tipo': "ALTER TABLE actividades ADD COLUMN tipo TEXT DEFAULT 'taller'",
        'peso': "ALTER TABLE actividades ADD COLUMN peso REAL",
        'categoria': "ALTER TABLE actividades ADD COLUMN categoria TEXT DEFAULT 'evaluacion'",
        'fecha_limite': "ALTER TABLE actividades ADD COLUMN fecha_limite TEXT",
        'hora_limite': "ALTER TABLE actividades ADD COLUMN hora_limite TEXT",
        'descripcion': "ALTER TABLE actividades ADD COLUMN descripcion TEXT DEFAULT ''",
        'observaciones': "ALTER TABLE actividades ADD COLUMN observaciones TEXT DEFAULT ''",
        'estado_act': "ALTER TABLE actividades ADD COLUMN estado_act TEXT DEFAULT 'publicada'",
        'competencia': "ALTER TABLE actividades ADD COLUMN competencia TEXT DEFAULT ''",
        'entrega_digital': "ALTER TABLE actividades ADD COLUMN entrega_digital INTEGER DEFAULT 0",
        'adjuntos': "ALTER TABLE actividades ADD COLUMN adjuntos TEXT DEFAULT '[]'",
        'integration': "ALTER TABLE actividades ADD COLUMN integration TEXT DEFAULT ''",
    }.items():
        if col not in cols:
            conn.execute(ddl)
    conn.execute('''CREATE TABLE IF NOT EXISTS entregas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        actividad_id INTEGER NOT NULL,
        alumno_id INTEGER NOT NULL,
        fecha_entrega TEXT DEFAULT (datetime('now','localtime')),
        archivos TEXT DEFAULT '[]',
        comentario TEXT DEFAULT '',
        estado TEXT DEFAULT 'pendiente',
        calificacion REAL,
        retroalimentacion TEXT DEFAULT '',
        calificado_por INTEGER,
        fecha_calificacion TEXT,
        UNIQUE(actividad_id, alumno_id)
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_entregas_actividad ON entregas(actividad_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_entregas_alumno ON entregas(alumno_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_entregas_estado ON entregas(estado)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_actividades_tipo ON actividades(tipo)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_actividades_estado ON actividades(estado_act)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_actividades_fecha_limite ON actividades(fecha_limite)')

def _migrar_v18(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS plantillas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        profesor_id INTEGER NOT NULL,
        nombre TEXT NOT NULL,
        tipo TEXT DEFAULT 'tarea',
        peso REAL DEFAULT 10,
        descripcion TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

def _migrar_v19(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS historial_academico (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alumno_id INTEGER NOT NULL,
        curso TEXT NOT NULL,
        jornada TEXT DEFAULT '',
        periodo INTEGER DEFAULT 1,
        promedio_final REAL DEFAULT 0,
        estado TEXT DEFAULT 'cursando',
        observaciones TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS padres (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        pin TEXT NOT NULL,
        telefono TEXT DEFAULT '',
        activo INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS alumno_padre (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alumno_id INTEGER NOT NULL,
        padre_id INTEGER NOT NULL,
        parentesco TEXT DEFAULT ''
    )''')

def _migrar_v20(conn, slug=None):
    conn.execute('''CREATE TABLE IF NOT EXISTS matriculas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alumno_id INTEGER DEFAULT 0,
        nombre TEXT NOT NULL,
        documento TEXT DEFAULT '',
        email TEXT DEFAULT '',
        telefono TEXT DEFAULT '',
        curso_solicitado TEXT DEFAULT '',
        jornada TEXT DEFAULT 'mañana',
        sede TEXT DEFAULT '',
        estado TEXT DEFAULT 'pendiente',
        documentos TEXT DEFAULT '',
        observaciones TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS tesoreria_facturas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alumno_id INTEGER DEFAULT 0,
        concepto TEXT NOT NULL,
        monto REAL DEFAULT 0,
        descuento REAL DEFAULT 0,
        estado TEXT DEFAULT 'pendiente',
        fecha_emision DATE DEFAULT (date('now')),
        fecha_vencimiento DATE DEFAULT (date('now','+30 days')),
        fecha_pago DATE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS tesoreria_pagos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        factura_id INTEGER NOT NULL,
        monto REAL NOT NULL,
        metodo TEXT DEFAULT 'efectivo',
        referencia TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

MIGRACIONES = {
    6:  _migrar_v6,
    7:  _migrar_v7,
    8:  _migrar_v8,
    9:  _migrar_v9,
    10: _migrar_v10,
    11: _migrar_v11,
    12: _migrar_v12,
    13: _migrar_v13,
    14: _migrar_v14,
    15: _migrar_v15,
    16: _migrar_v16,
    17: _migrar_v17,
    18: _migrar_v18,
    19: _migrar_v19,
    20: _migrar_v20,
}

# ── MASTER DB ─────────────────────────────────────────────────────────────────
def conectar_master():
    c = sqlite3.connect(MASTER_DB, timeout=30)
    c.row_factory = sqlite3.Row
    c.execute('PRAGMA journal_mode=WAL')
    c.execute('PRAGMA foreign_keys=ON')
    return c

def init_master_db():
    conn = conectar_master()
    conn.execute('''CREATE TABLE IF NOT EXISTS colegios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT UNIQUE NOT NULL, nombre TEXT NOT NULL,
        logo TEXT DEFAULT '', activo INTEGER DEFAULT 1,
        creado TEXT DEFAULT (date('now')),
        vencimiento TEXT DEFAULT NULL,
        num_periodos INTEGER DEFAULT 4,
        codigo_registro TEXT DEFAULT '',
        primary_color TEXT DEFAULT '#6c63ff',
        secondary_color TEXT DEFAULT '#3498db'
    )''')
    # Migraciones de columnas nuevas
    for col in [
        'logo TEXT DEFAULT ""',
        'vencimiento TEXT DEFAULT NULL',
        'num_periodos INTEGER DEFAULT 4',
        'codigo_registro TEXT DEFAULT ""',
        'primary_color TEXT DEFAULT "#6c63ff"',
        'secondary_color TEXT DEFAULT "#3498db"',
        'codigo_profesores TEXT DEFAULT ""',
        'codigo_directoras TEXT DEFAULT ""',
        'codigo_rectores TEXT DEFAULT ""',
        'schema_version INTEGER DEFAULT 0',
    ]:
        try: conn.execute(f'ALTER TABLE colegios ADD COLUMN {col}')
        except sqlite3.OperationalError:
            logger.debug(f'Columna ya existe en colegios: {col.split()[0]}')
    # Migrar codigo_registro a las columnas específicas si están vacías
    for c in conn.execute('SELECT slug, codigo_registro, codigo_profesores, codigo_directoras, codigo_rectores FROM colegios').fetchall():
        updates = []
        if c['codigo_registro'] and not c['codigo_profesores']:
            updates.append(('codigo_profesores', c['codigo_registro']))
        if c['codigo_registro'] and not c['codigo_directoras']:
            updates.append(('codigo_directoras', c['codigo_registro']))
        if c['codigo_registro'] and not c['codigo_rectores']:
            updates.append(('codigo_rectores', c['codigo_registro']))
        for col_name, val in updates:
            conn.execute(f'UPDATE colegios SET {col_name}=? WHERE slug=?', (val, c['slug']))
    conn.commit()
    conn.close()




def get_codigo_registro(slug, rol=None):
    """Devuelve el código de invitación del colegio para un rol específico.
    rol: 'profesores', 'directoras', 'rectores' o None (usa codigo_registro genérico)."""
    c = get_colegio(slug)
    if not c: return ''
    if rol == 'profesores':
        val = c['codigo_profesores'] or c['codigo_registro'] or ''
    elif rol == 'directoras':
        val = c['codigo_directoras'] or c['codigo_registro'] or ''
    elif rol == 'rectores':
        val = c['codigo_rectores'] or c['codigo_registro'] or ''
    else:
        val = c['codigo_registro'] or ''
    return val

# ── DB POR COLEGIO ────────────────────────────────────────────────────────────
def db_path(slug): return os.path.join(DB_FOLDER, f'{slug}.db')

def conectar(slug):
    c = sqlite3.connect(db_path(slug), timeout=30)
    c.row_factory = sqlite3.Row
    c.execute('PRAGMA journal_mode=WAL')
    c.execute('PRAGMA foreign_keys=ON')
    return c

def migrar_db(slug):
    conn = conectar(slug)
    try:
        cols_prof = [r[1] for r in conn.execute('PRAGMA table_info(profesores)').fetchall()]
        if 'materia' in cols_prof:
            conn.execute('''CREATE TABLE IF NOT EXISTS profesores_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL, usuario TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL, activo INTEGER DEFAULT 1,
                email TEXT DEFAULT '', telefono TEXT DEFAULT '',
                pregunta_secreta TEXT DEFAULT '', respuesta_secreta TEXT DEFAULT '')''')
            conn.execute('''INSERT OR IGNORE INTO profesores_new
                (id,nombre,usuario,password,activo,email,telefono,pregunta_secreta,respuesta_secreta)
                SELECT id,nombre,usuario,password,activo,
                       COALESCE(email,''), COALESCE(telefono,''),
                       COALESCE(pregunta_secreta,''), COALESCE(respuesta_secreta,'')
                FROM profesores''')
            conn.execute('DROP TABLE profesores')
            conn.execute('ALTER TABLE profesores_new RENAME TO profesores')
            conn.commit()

        cols_dir = [r[1] for r in conn.execute('PRAGMA table_info(directoras)').fetchall()]
        for col, defval in [
            ('jornada',           'TEXT NOT NULL DEFAULT "Mañana"'),
            ('activo',            'INTEGER DEFAULT 1'),
            ('email',             'TEXT DEFAULT ""'),
            ('pregunta_secreta',  'TEXT DEFAULT ""'),
            ('respuesta_secreta', 'TEXT DEFAULT ""'),
        ]:
            if col not in cols_dir:
                conn.execute(f'ALTER TABLE directoras ADD COLUMN {col} {defval}')
                conn.commit()

        cols_alum = [r[1] for r in conn.execute('PRAGMA table_info(alumnos)').fetchall()]
        if 'jornada' not in cols_alum:
            conn.execute('ALTER TABLE alumnos ADD COLUMN jornada TEXT NOT NULL DEFAULT "Mañana"')
            conn.commit()
        if 'email_acudiente' not in cols_alum:
            conn.execute('ALTER TABLE alumnos ADD COLUMN email_acudiente TEXT DEFAULT ""')
            conn.commit()
        if 'pin' not in cols_alum:
            conn.execute('ALTER TABLE alumnos ADD COLUMN pin TEXT DEFAULT ""')
            conn.commit()

        cols_act = [r[1] for r in conn.execute('PRAGMA table_info(actividades)').fetchall()]
        if 'periodo' not in cols_act:
            conn.execute('ALTER TABLE actividades ADD COLUMN periodo INTEGER DEFAULT 1')
            conn.commit()
        if 'jornada' not in cols_act:
            conn.execute('ALTER TABLE actividades ADD COLUMN jornada TEXT DEFAULT "Mañana"')
            conn.execute('UPDATE actividades SET jornada="Mañana" WHERE jornada IS NULL OR jornada=""')
            conn.commit()

        cols_ev = [r[1] for r in conn.execute('PRAGMA table_info(evaluaciones)').fetchall()]
        if 'periodo' not in cols_ev:
            conn.execute('ALTER TABLE evaluaciones ADD COLUMN periodo INTEGER DEFAULT 1')
            conn.commit()
        if 'jornada' not in cols_ev:
            conn.execute('ALTER TABLE evaluaciones ADD COLUMN jornada TEXT DEFAULT "Mañana"')
            conn.execute('UPDATE evaluaciones SET jornada="Mañana" WHERE jornada IS NULL OR jornada=""')
            conn.commit()
        # Verificar que UNIQUE incluya periodo — si no, recrear la tabla
        _recrear_si_unique_incorrecto(conn, slug, 'evaluaciones',
            '(aid,profesor_id,materia,jornada,periodo)',
            '''CREATE TABLE evaluaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
                materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
                evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
                UNIQUE(aid,profesor_id,materia,jornada,periodo))''',
            '''(id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
               SELECT id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,
                      COALESCE(periodo,1) FROM evaluaciones_old''')

        cols_comp = [r[1] for r in conn.execute('PRAGMA table_info(compromisos)').fetchall()]
        if 'jornada' not in cols_comp:
            conn.execute('ALTER TABLE compromisos ADD COLUMN jornada TEXT DEFAULT "Mañana"')
            conn.commit()

        cols_hor = [r[1] for r in conn.execute('PRAGMA table_info(horarios_curso)').fetchall()]
        if 'jornada' not in cols_hor:
            conn.execute('ALTER TABLE horarios_curso ADD COLUMN jornada TEXT DEFAULT "Mañana"')
            conn.commit()

        tablas = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if 'asignaciones_materia' not in tablas:
            conn.execute('''CREATE TABLE asignaciones_materia (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                profesor_id INTEGER NOT NULL, materia TEXT NOT NULL, jornada TEXT NOT NULL,
                UNIQUE(profesor_id, materia, jornada))''')
            conn.commit()
        if 'asignaciones_curso' not in tablas:
            conn.execute('''CREATE TABLE asignaciones_curso (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                profesor_id INTEGER NOT NULL, materia TEXT NOT NULL,
                jornada TEXT NOT NULL, curso TEXT NOT NULL,
                UNIQUE(profesor_id, materia, jornada, curso))''')
            conn.commit()
        if 'horarios_curso' not in tablas:
            conn.execute('''CREATE TABLE horarios_curso (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                curso TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
                dia TEXT NOT NULL, franja TEXT NOT NULL,
                num TEXT DEFAULT "", materia TEXT DEFAULT "", profesor TEXT DEFAULT "",
                UNIQUE(curso, jornada, dia, franja))''')
            conn.commit()
        if 'directoras' not in tablas:
            conn.execute('''CREATE TABLE directoras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL, usuario TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL, curso TEXT NOT NULL,
                jornada TEXT NOT NULL DEFAULT "Mañana",
                email TEXT DEFAULT "", activo INTEGER DEFAULT 1,
                pregunta_secreta TEXT DEFAULT "",
                respuesta_secreta TEXT DEFAULT "")''')
            conn.commit()

        _recrear_si_unique_incorrecto(conn, slug, 'horarios_curso',
            '(curso,jornada,dia,franja)',
            '''CREATE TABLE horarios_curso (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                curso TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
                dia TEXT NOT NULL, franja TEXT NOT NULL,
                num TEXT DEFAULT "", materia TEXT DEFAULT "", profesor TEXT DEFAULT "",
                UNIQUE(curso, jornada, dia, franja))''',
            'SELECT * FROM horarios_curso_old')

        cols_rec = [r[1] for r in conn.execute('PRAGMA table_info(rectores)').fetchall()]
        if 'es_principal' not in cols_rec:
            conn.execute('ALTER TABLE rectores ADD COLUMN es_principal INTEGER DEFAULT 0')
            conn.commit()
        if 'jornada' not in cols_rec:
            conn.execute('ALTER TABLE rectores ADD COLUMN jornada TEXT DEFAULT ""')
            conn.commit()

        cols_cl = [r[1] for r in conn.execute('PRAGMA table_info(comunicaciones_leidas)').fetchall()]
        if 'leido' not in cols_cl:
            conn.execute('ALTER TABLE comunicaciones_leidas ADD COLUMN leido INTEGER DEFAULT 0')
            conn.commit()

        tablas_actuales = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]
        if 'profesores' not in tablas_actuales:
            return
        profs = conn.execute('SELECT id FROM profesores').fetchall()
        for p in profs:
            combos = conn.execute(
                'SELECT DISTINCT materia, jornada, curso FROM actividades WHERE profesor_id=?',
                (p['id'],)
            ).fetchall()
            for c in combos:
                conn.execute('INSERT OR IGNORE INTO asignaciones_materia (profesor_id,materia,jornada) VALUES (?,?,?)',
                             (p['id'], c['materia'], c['jornada']))
                conn.execute('INSERT OR IGNORE INTO asignaciones_curso (profesor_id,materia,jornada,curso) VALUES (?,?,?,?)',
                             (p['id'], c['materia'], c['jornada'], c['curso']))
        conn.commit()

    except Exception as e:
        logger.error(f'[{slug}] Error en migración legacy: {e}', exc_info=True)
        raise
    finally:
        conn.close()

def init_db(slug):
    conn = conectar(slug)
    stmts = [
        '''CREATE TABLE IF NOT EXISTS profesores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL, usuario TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, activo INTEGER DEFAULT 1,
            email TEXT DEFAULT '', telefono TEXT DEFAULT '',
            pregunta_secreta TEXT DEFAULT '', respuesta_secreta TEXT DEFAULT '')''',
        '''CREATE TABLE IF NOT EXISTS asignaciones_materia (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profesor_id INTEGER NOT NULL, materia TEXT NOT NULL, jornada TEXT NOT NULL,
            UNIQUE(profesor_id, materia, jornada))''',
        '''CREATE TABLE IF NOT EXISTS asignaciones_curso (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profesor_id INTEGER NOT NULL, materia TEXT NOT NULL,
            jornada TEXT NOT NULL, curso TEXT NOT NULL,
            UNIQUE(profesor_id, materia, jornada, curso))''',
        '''CREATE TABLE IF NOT EXISTS alumnos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL, curso TEXT NOT NULL,
            jornada TEXT NOT NULL DEFAULT "Mañana",
            num_curso INTEGER DEFAULT 0, activo INTEGER DEFAULT 1,
            email_acudiente TEXT DEFAULT '',
            pin TEXT DEFAULT '')''',
        '''CREATE TABLE IF NOT EXISTS asistencia (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aid INTEGER, fecha TEXT, estado TEXT,
            UNIQUE(aid, fecha))''',
        '''CREATE TABLE IF NOT EXISTS compromisos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT, fecha TEXT, materia TEXT,
            curso TEXT, jornada TEXT DEFAULT "Mañana")''',
        '''CREATE TABLE IF NOT EXISTS observaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aid INTEGER, materia TEXT, texto TEXT, fecha TEXT)''',
        '''CREATE TABLE IF NOT EXISTS actividades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profesor_id INTEGER NOT NULL, materia TEXT NOT NULL,
            jornada TEXT NOT NULL DEFAULT "Mañana",
            curso TEXT NOT NULL, nombre TEXT NOT NULL,
            orden INTEGER DEFAULT 0, periodo INTEGER DEFAULT 1)''',
        '''CREATE TABLE IF NOT EXISTS notas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aid INTEGER NOT NULL, actividad_id INTEGER NOT NULL, val REAL NOT NULL,
            UNIQUE(aid,actividad_id))''',
        '''CREATE TABLE IF NOT EXISTS evaluaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
            materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
            evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
            UNIQUE(aid,profesor_id,materia,jornada,periodo))''',
        '''CREATE TABLE IF NOT EXISTS horarios_curso (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            curso TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
            dia TEXT NOT NULL, franja TEXT NOT NULL,
            num TEXT DEFAULT '', materia TEXT DEFAULT '', profesor TEXT DEFAULT '',
            UNIQUE(curso, jornada, dia, franja))''',
        '''CREATE TABLE IF NOT EXISTS directoras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL, usuario TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, curso TEXT NOT NULL,
            jornada TEXT NOT NULL DEFAULT "Mañana",
            email TEXT DEFAULT '', activo INTEGER DEFAULT 1,
            pregunta_secreta TEXT DEFAULT '',
            respuesta_secreta TEXT DEFAULT '')''',
        '''CREATE TABLE IF NOT EXISTS rectores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL, usuario TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, email TEXT DEFAULT '',
            activo INTEGER DEFAULT 1,
            es_principal INTEGER DEFAULT 0,
            pregunta_secreta TEXT DEFAULT '',
            respuesta_secreta TEXT DEFAULT '')''',
        '''CREATE TABLE IF NOT EXISTS comunicaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rector_id INTEGER NOT NULL,
            titulo TEXT NOT NULL,
            contenido TEXT NOT NULL,
            destinatario_tipo TEXT NOT NULL,
            destinatario_valor TEXT DEFAULT '',
            prioridad TEXT NOT NULL DEFAULT 'normal',
            estado TEXT NOT NULL DEFAULT 'borrador',
            fecha_creacion TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            fecha_programada TEXT DEFAULT NULL,
            fecha_publicacion TEXT DEFAULT NULL,
            activo INTEGER DEFAULT 1)''',
        '''CREATE TABLE IF NOT EXISTS comunicaciones_leidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            comunicacion_id INTEGER NOT NULL,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            leido INTEGER DEFAULT 0,
            fecha_lectura TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            UNIQUE(comunicacion_id, usuario_tipo, usuario_id))''',
        '''CREATE TABLE IF NOT EXISTS notificaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            titulo TEXT NOT NULL,
            mensaje TEXT DEFAULT '',
            tipo TEXT NOT NULL DEFAULT 'info',
            link TEXT DEFAULT '',
            leida INTEGER DEFAULT 0,
            fecha_creacion TEXT NOT NULL DEFAULT (datetime('now','localtime')))''',
        '''CREATE TABLE IF NOT EXISTS canales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT NOT NULL,
            rector_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            nombre TEXT NOT NULL,
            descripcion TEXT DEFAULT '',
            curso TEXT DEFAULT '',
            materia TEXT DEFAULT '',
            activo INTEGER DEFAULT 1,
            fecha_creacion TEXT DEFAULT (datetime('now','localtime')))''',
        '''CREATE TABLE IF NOT EXISTS canal_miembros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            canal_id INTEGER NOT NULL,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            fecha_ingreso TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(canal_id, usuario_tipo, usuario_id))''',
        '''CREATE TABLE IF NOT EXISTS mensajes_canal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            canal_id INTEGER NOT NULL,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            mensaje TEXT NOT NULL,
            fecha TEXT DEFAULT (datetime('now','localtime')),
            editado INTEGER DEFAULT 0)''',
        '''CREATE TABLE IF NOT EXISTS mensajes_leidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mensaje_id INTEGER NOT NULL,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            fecha_lectura TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(mensaje_id, usuario_tipo, usuario_id))''',
        '''CREATE TABLE IF NOT EXISTS periodos_estado (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo INTEGER NOT NULL UNIQUE,
            estado TEXT NOT NULL DEFAULT 'abierto',
            fecha_apertura TEXT,
            fecha_cierre TEXT,
            abierto_por INTEGER,
            cerrado_por INTEGER)''',
        # ── solicitudes_modificacion is created by migration v14 ──
        # ── Fase 5 – Comunicación v2 ─────────────────────────────────
        '''CREATE TABLE IF NOT EXISTS mensajes_archivos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mensaje_id INTEGER,
            canal_id INTEGER NOT NULL,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            nombre_original TEXT NOT NULL,
            nombre_archivo TEXT NOT NULL,
            tipo_mime TEXT NOT NULL,
            tamano INTEGER NOT NULL,
            es_imagen INTEGER DEFAULT 0,
            ancho INTEGER,
            alto INTEGER,
            fecha TEXT DEFAULT (datetime('now','localtime')))''',
        '''CREATE TABLE IF NOT EXISTS mensajes_reacciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mensaje_id INTEGER NOT NULL,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            reaccion TEXT NOT NULL,
            fecha TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(mensaje_id, usuario_tipo, usuario_id, reaccion))''',
        '''CREATE TABLE IF NOT EXISTS mensajes_fijados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            canal_id INTEGER NOT NULL,
            mensaje_id INTEGER NOT NULL,
            fijado_por_tipo TEXT NOT NULL,
            fijado_por_id INTEGER NOT NULL,
            fecha TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(canal_id, mensaje_id))''',
        '''CREATE TABLE IF NOT EXISTS canal_enlaces (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            canal_id INTEGER NOT NULL,
            titulo TEXT,
            url TEXT NOT NULL,
            agregado_por_tipo TEXT NOT NULL,
            agregado_por_id INTEGER NOT NULL,
            fecha TEXT DEFAULT (datetime('now','localtime')))''',
        '''CREATE TABLE IF NOT EXISTS canal_actividad (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            canal_id INTEGER NOT NULL,
            usuario_tipo TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            estado TEXT DEFAULT 'online',
            ultima_vista TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(canal_id, usuario_tipo, usuario_id))''',
    ]
    for s in stmts:
        try: conn.execute(s)
        except sqlite3.OperationalError as e:
            logger.warning(f'init_db table: {e}')
    alter_stmts = [
        "ALTER TABLE mensajes_canal ADD COLUMN responde_a INTEGER REFERENCES mensajes_canal(id)",
        "ALTER TABLE mensajes_canal ADD COLUMN editado_en TEXT",
        "ALTER TABLE mensajes_canal ADD COLUMN eliminado INTEGER DEFAULT 0",
        "ALTER TABLE mensajes_canal ADD COLUMN tiene_archivos INTEGER DEFAULT 0",
        "ALTER TABLE config_institucion ADD COLUMN max_tamano_archivo INTEGER DEFAULT 10485760",
    ]
    for stmt in alter_stmts:
        try: conn.execute(stmt)
        except sqlite3.OperationalError:
            pass
    indexes = [
        'CREATE INDEX IF NOT EXISTS idx_notas_aid ON notas(aid)',
        'CREATE INDEX IF NOT EXISTS idx_notas_actividad ON notas(actividad_id)',
        'CREATE INDEX IF NOT EXISTS idx_asistencia_aid ON asistencia(aid)',
        'CREATE INDEX IF NOT EXISTS idx_observaciones_aid ON observaciones(aid)',
        'CREATE INDEX IF NOT EXISTS idx_evaluaciones_aid ON evaluaciones(aid)',
        'CREATE INDEX IF NOT EXISTS idx_actividades_prof ON actividades(profesor_id,materia,curso,jornada,periodo)',
        'CREATE INDEX IF NOT EXISTS idx_alumnos_nombre ON alumnos(nombre,jornada)',
        'CREATE INDEX IF NOT EXISTS idx_mensajes_canal ON mensajes_canal(canal_id, id)',
        'CREATE INDEX IF NOT EXISTS idx_archivos_canal ON mensajes_archivos(canal_id, mensaje_id)',
        'CREATE INDEX IF NOT EXISTS idx_archivos_mensaje ON mensajes_archivos(mensaje_id)',
        'CREATE INDEX IF NOT EXISTS idx_reacciones_mensaje ON mensajes_reacciones(mensaje_id)',
        'CREATE INDEX IF NOT EXISTS idx_fijados_canal ON mensajes_fijados(canal_id)',
        'CREATE INDEX IF NOT EXISTS idx_enlaces_canal ON canal_enlaces(canal_id)',
        'CREATE INDEX IF NOT EXISTS idx_actividad_canal ON canal_actividad(canal_id)',
        # ── Performance indexes ─────────────────────────────────────────
        'CREATE INDEX IF NOT EXISTS idx_alumnos_curso_jornada ON alumnos(curso, jornada, activo)',
        'CREATE INDEX IF NOT EXISTS idx_asistencia_aid_fecha ON asistencia(aid, fecha)',
        'CREATE INDEX IF NOT EXISTS idx_compromisos_materia ON compromisos(materia, curso, jornada)',
        'CREATE INDEX IF NOT EXISTS idx_horarios_materia ON horarios_curso(materia, jornada, dia)',
        'CREATE INDEX IF NOT EXISTS idx_notificaciones_usuario ON notificaciones(usuario_tipo, usuario_id, leida)',
        'CREATE INDEX IF NOT EXISTS idx_comunicaciones_rector ON comunicaciones(rector_id, activo)',
        'CREATE INDEX IF NOT EXISTS idx_asignaciones_curso_prof ON asignaciones_curso(profesor_id, materia, jornada)',
        'CREATE INDEX IF NOT EXISTS idx_canal_miembros_usuario ON canal_miembros(usuario_tipo, usuario_id)',
        'CREATE INDEX IF NOT EXISTS idx_canales_slug ON canales(slug)',
        'CREATE INDEX IF NOT EXISTS idx_evaluaciones_aid_periodo ON evaluaciones(aid, periodo)',
        'CREATE INDEX IF NOT EXISTS idx_actividades_curso ON actividades(curso, jornada, periodo)',
        'CREATE INDEX IF NOT EXISTS idx_asignaciones_materia_prof ON asignaciones_materia(profesor_id)',
        'CREATE INDEX IF NOT EXISTS idx_profesores_usuario ON profesores(usuario, activo)',
        'CREATE INDEX IF NOT EXISTS idx_directoras_usuario ON directoras(usuario, activo)',
        'CREATE INDEX IF NOT EXISTS idx_rectores_usuario ON rectores(usuario, activo)',
        'CREATE INDEX IF NOT EXISTS idx_comunicaciones_leidas_user ON comunicaciones_leidas(usuario_tipo, usuario_id, leido)',
        'CREATE INDEX IF NOT EXISTS idx_periodos_estado_periodo ON periodos_estado(periodo)',
        'CREATE INDEX IF NOT EXISTS idx_config_institucion_slug ON config_institucion(slug)',
        'CREATE INDEX IF NOT EXISTS idx_alumnos_id_curso ON alumnos(id, curso)',
        'CREATE INDEX IF NOT EXISTS idx_comunicaciones_estado ON comunicaciones(rector_id, activo, estado)',
        # ── Phase 11: Performance indexes ──────────────────────────────
        'CREATE INDEX IF NOT EXISTS idx_ml_mensaje_tipo ON mensajes_leidos(mensaje_id, usuario_tipo, usuario_id)',
        'CREATE INDEX IF NOT EXISTS idx_obs_aid_materia ON observaciones(aid, materia)',
        'CREATE INDEX IF NOT EXISTS idx_comunicaciones_rector_fecha ON comunicaciones(rector_id, activo, fecha_creacion)',
        'CREATE INDEX IF NOT EXISTS idx_audit_log_tabla ON audit_log(tabla)',
        'CREATE INDEX IF NOT EXISTS idx_asistencia_fecha ON asistencia(fecha)',
        'CREATE INDEX IF NOT EXISTS idx_actividades_prof_periodo ON actividades(profesor_id, materia, jornada, curso, periodo)',
        'CREATE INDEX IF NOT EXISTS idx_solicitudes_fecha ON solicitudes_modificacion(fecha_solicitud)',
    ]
    for idx in indexes:
        try: conn.execute(idx)
        except sqlite3.OperationalError as e:
            logger.warning(f'init_db index: {e}')
    conn.commit()
    _ejecutar_migraciones(slug, conn)
    conn.close()
    migrar_db(slug)

# ── INTELLIGENT CACHE (TTL-based, for safe data only) ─────────────────────────
import threading
_cache = {}
_cache_lock = threading.Lock()
_CACHE_TTL = {'config': 60, 'cursos': 60, 'materias': 60, 'jornadas': 60, 'colegio': 300}

def _cache_get(key):
    with _cache_lock:
        entry = _cache.get(key)
        if entry and entry['expires'] > time.time():
            return entry['value']
        if entry:
            del _cache[key]
    return None

def _cache_set(key, value, ttl=60):
    with _cache_lock:
        _cache[key] = {'value': value, 'expires': time.time() + ttl}

def _cache_invalidate(slug=None, prefix=None):
    with _cache_lock:
        to_del = [k for k in _cache if (slug and slug in k) or (prefix and k.startswith(prefix))]
        for k in to_del:
            del _cache[k]

def config_get(slug):
    """Cached config lookup. Invalidated on config save."""
    key = f'cfg_{slug}'
    cached = _cache_get(key)
    if cached: return cached
    conn = conectar(slug)
    cfg = conn.execute('SELECT * FROM config_institucion WHERE slug=?', (slug,)).fetchone()
    conn.close()
    if cfg:
        _cache_set(key, dict(cfg), ttl=_CACHE_TTL['config'])
        return dict(cfg)
    return {}

def get_colegio(slug):
    """Cached colegio lookup. Invalidated when colegio is updated."""
    key = f'col_{slug}'
    cached = _cache_get(key)
    if cached: return cached
    conn = conectar_master()
    row = conn.execute('SELECT * FROM colegios WHERE slug=?', (slug,)).fetchone()
    conn.close()
    if row:
        val = dict(row)
        _cache_set(key, val, ttl=_CACHE_TTL['colegio'])
        return val
    return None

def get_cursos_cache(slug, jornada=None):
    """Cached distinct course list."""
    key = f'cursos_{slug}_{jornada or "all"}'
    cached = _cache_get(key)
    if cached: return cached
    conn = conectar(slug)
    if jornada:
        rows = conn.execute('SELECT DISTINCT curso FROM alumnos WHERE activo=1 AND jornada=? ORDER BY curso', (jornada,)).fetchall()
    else:
        rows = conn.execute('SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()
    conn.close()
    val = [r['curso'] for r in rows]
    _cache_set(key, val, ttl=_CACHE_TTL['cursos'])
    return val

def get_materias_cache(slug, jornada=None):
    """Cached distinct materia list."""
    key = f'mats_{slug}_{jornada or "all"}'
    cached = _cache_get(key)
    if cached: return cached
    conn = conectar(slug)
    if jornada:
        rows = conn.execute('SELECT DISTINCT materia FROM asignaciones_materia WHERE jornada=? ORDER BY materia', (jornada,)).fetchall()
    else:
        rows = conn.execute('SELECT DISTINCT materia FROM asignaciones_materia ORDER BY materia').fetchall()
    conn.close()
    val = [r['materia'] for r in rows]
    _cache_set(key, val, ttl=_CACHE_TTL['materias'])
    return val

def get_jornadas_cache(slug):
    """Cached distinct jornada list."""
    key = f'jorn_{slug}'
    cached = _cache_get(key)
    if cached: return cached
    conn = conectar(slug)
    rows = conn.execute('SELECT DISTINCT jornada FROM alumnos WHERE activo=1 ORDER BY jornada').fetchall()
    conn.close()
    val = [r['jornada'] for r in rows]
    _cache_set(key, val, ttl=_CACHE_TTL['jornadas'])
    return val

# ── HELPERS ───────────────────────────────────────────────────────────────────
def get_profesor(slug):
    cache_key = f'_prof_{slug}'
    if hasattr(g, cache_key):
        return getattr(g, cache_key)
    pid = session.get(f'profesor_id_{slug}')
    if not pid: setattr(g, cache_key, None); return None
    conn = conectar(slug)
    p = conn.execute('SELECT * FROM profesores WHERE id=? AND activo=1', (pid,)).fetchone()
    conn.close()
    if not p:
        session.pop(f'profesor_id_{slug}', None)
        session.pop(f'rol_{slug}', None)
    setattr(g, cache_key, p)
    return p

def get_sesion_jornada_materia(slug):
    return (session.get(f'jornada_{slug}'), session.get(f'materia_{slug}'))

def get_materias_profesor(slug, pid):
    conn = conectar(slug)
    rows = conn.execute(
        'SELECT materia, jornada FROM asignaciones_materia WHERE profesor_id=? ORDER BY jornada, materia',
        (pid,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_cursos_profesor(slug, pid, materia, jornada):
    conn = conectar(slug)
    rows = conn.execute(
        'SELECT curso FROM asignaciones_curso WHERE profesor_id=? AND materia=? AND jornada=? ORDER BY curso',
        (pid, materia, jornada)
    ).fetchall()
    conn.close()
    return [r['curso'] for r in rows]

def require_colegio(slug):
    colegio = get_colegio(slug)
    if not colegio: abort(404)
    if not colegio['activo']: abort(403)
    return colegio

# ── CANALES HELPERS ─────────────────────────────────────────────────────────────
def get_usuario_actual(slug):
    prof = get_profesor(slug)
    if prof: return ('profesor', prof['id'])
    aid = session.get(f'alumno_id_{slug}')
    if aid: return ('estudiante', aid)
    direc = get_directora(slug)
    if direc: return ('directora', direc['id'])
    rector = get_rector(slug)
    if rector: return ('rector', rector['id'])
    return (None, None)

# ── PERMISSION SYSTEM ────────────────────────────────────────────────────────
def obtener_roles_usuario(slug, usuario_id):
    conn = conectar(slug)
    rows = conn.execute('''
        SELECT r.codigo, ri.nombre as rol_nombre, ri.jerarquia,
               ur.entidad_tipo, ur.entidad_id
        FROM usuarios_roles ur
        JOIN roles_instancia ri ON ri.id = ur.rol_id
        JOIN roles_base r ON r.codigo = ri.codigo
        WHERE ur.usuario_id = ? AND ri.activo = 1
    ''', (usuario_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

PERMISOS_POR_CODIGO = {
    'admin':     ['*'],
    'rector':    ['*'],
    'authority': ['people.teachers.view', 'people.students.view',
        'structure.courses.manage', 'structure.subjects.manage',
        'academic.grades.view', 'academic.grades.write', 'academic.grades.approve',
        'academic.grades.history', 'academico.periodos.cerrar',
        'academico.periodos.abrir', 'academico.notas.aprobar',
        'academico.notas.modificar_cerrado',
        'academic.attendance.view',
        'academic.observations.view', 'academic.observations.write',
        'academic.evaluations.create', 'academic.evaluations.edit',
        'communication.communicados.view', 'communication.communicados.create',
        'communication.channels.read', 'communication.channels.send',
        'reports.grades', 'reports.attendance', 'reports.export',
        'audit.log.view'],
    'teacher': ['people.students.view',
        'academic.grades.view', 'academic.grades.write',
        'academic.attendance.view', 'academic.attendance.write',
        'academic.observations.view', 'academic.observations.write',
        'academic.evaluations.create', 'academic.evaluations.edit',
        'academic.activities.create', 'academic.activities.edit',
        'communication.communicados.view',
        'communication.channels.read', 'communication.channels.send'],
    'student': ['academic.grades.view', 'academic.attendance.view',
        'communication.communicados.view',
        'communication.channels.read', 'communication.channels.send'],
    'guardian': ['academic.grades.view', 'academic.attendance.view',
        'communication.communicados.view',
        'communication.channels.read'],
}

NIVELES_ROL = {'admin': 0, 'rector': 1, 'authority': 2, 'teacher': 3, 'student': 4, 'guardian': 5}

def _permisos_para_rol(codigo):
    permisos = set()
    nivel = NIVELES_ROL.get(codigo, 99)
    for rc, rn in NIVELES_ROL.items():
        if rn >= nivel:
            permisos.update(PERMISOS_POR_CODIGO.get(rc, []))
    return list(permisos)

def tiene_permiso(slug, usuario_id, permiso, entidad_tipo=None, entidad_id=None):
    roles = obtener_roles_usuario(slug, usuario_id)
    for rol in roles:
        if rol['codigo'] in ('admin', 'rector'):
            return True
        if permiso not in _permisos_para_rol(rol['codigo']) and '*' not in _permisos_para_rol(rol['codigo']):
            continue
        if entidad_tipo and rol['entidad_tipo']:
            if rol['entidad_tipo'] != entidad_tipo or rol['entidad_id'] != entidad_id:
                continue
        return True
    return False

from functools import wraps

def requiere_permiso(permiso, obtener_entidad=None):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            slug = kwargs.get('slug')
            if not slug: abort(400)
            usuario_tipo, usuario_id = get_usuario_actual(slug)
            if not usuario_id:
                return redirect(url_for('auth.login', slug=slug))
            if obtener_entidad:
                e_tipo, e_id = obtener_entidad(kwargs)
            else:
                e_tipo, e_id = None, None
            if not tiene_permiso(slug, usuario_id, permiso, e_tipo, e_id):
                abort(403)
            return f(*args, **kwargs)
        return wrapper
    return decorator

# ── PERIOD HELPERS ────────────────────────────────────────────────────────────
def periodo_cerrado(slug, periodo):
    conn = conectar(slug)
    row = conn.execute(
        'SELECT estado FROM periodos_estado WHERE periodo=?',
        (periodo,)).fetchone()
    conn.close()
    return row is not None and row['estado'] == 'cerrado'

# ── AUDIT HELPER ──────────────────────────────────────────────────────────────
def audit_log(slug, usuario_id, accion, tabla, registro_id=None, valor_anterior=None, valor_nuevo=None):
    from flask import request as flask_request
    conn = None
    try:
        conn = conectar(slug)
        conn.execute(
            '''INSERT INTO audit_log (usuario_id, accion, tabla, registro_id, valor_anterior, valor_nuevo, ip, user_agent)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (usuario_id, accion, tabla, registro_id,
             json.dumps(valor_anterior) if valor_anterior else None,
             json.dumps(valor_nuevo) if valor_nuevo else None,
             flask_request.remote_addr,
             flask_request.user_agent.string if flask_request.user_agent else None)
        )
        conn.commit()
    except Exception as e:
        logger.warning(f"[audit] {e}")
    finally:
        if conn: conn.close()

def auditar_nota(slug, usuario_id, rol, tipo_accion, tabla, aid, curso, materia, periodo, campo=None, actividad_id=None, registro_id=None, valor_anterior=None, valor_nuevo=None, motivo=None):
    from flask import request as flask_request
    conn = None
    try:
        conn = conectar(slug)
        conn.execute(
            '''INSERT INTO auditoria_notas
               (usuario_id, rol, ip, curso, materia, periodo, tipo_accion, tabla, registro_id, aid, actividad_id, campo, valor_anterior, valor_nuevo, motivo)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (usuario_id, rol, flask_request.remote_addr, curso, materia, periodo,
             tipo_accion, tabla, registro_id, aid, actividad_id, campo,
             json.dumps(valor_anterior) if valor_anterior is not None else None,
             json.dumps(valor_nuevo) if valor_nuevo is not None else None,
             motivo)
        )
        conn.commit()
    except Exception as e:
        logger.warning(f"[auditar_nota] {e}")
    finally:
        if conn: conn.close()

# ── CANALES HELPERS ─────────────────────────────────────────────────────────────
def canales_usuario(slug, usuario_tipo, usuario_id):
    conn = conectar(slug)
    rows = conn.execute('''
        SELECT c.*,
            (SELECT mensaje FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultimo_mensaje,
            (SELECT usuario_tipo FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultimo_autor_tipo,
            (SELECT usuario_id FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultimo_autor_id,
            (SELECT fecha FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultima_fecha,
            (SELECT COUNT(*) FROM mensajes_canal mc
             LEFT JOIN mensajes_leidos ml ON ml.mensaje_id=mc.id AND ml.usuario_tipo=? AND ml.usuario_id=?
             WHERE mc.canal_id=c.id AND ml.id IS NULL) as no_leidos
        FROM canales c
        JOIN canal_miembros cm ON cm.canal_id=c.id
        WHERE cm.usuario_tipo=? AND cm.usuario_id=? AND c.activo=1
        ORDER BY ultima_fecha DESC''', (usuario_tipo, usuario_id, usuario_tipo, usuario_id)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def agregar_miembro_canal(conn, canal_id, usuario_tipo, usuario_id):
    conn.execute('INSERT OR IGNORE INTO canal_miembros (canal_id, usuario_tipo, usuario_id) VALUES (?,?,?)',
                 (canal_id, usuario_tipo, usuario_id))

def asignar_miembros_auto(conn, slug, canal_id, tipo, curso='', materia=''):
    if tipo in ('institucional','rectoria','profesores'):
        for p in conn.execute('SELECT id FROM profesores WHERE activo=1').fetchall():
            agregar_miembro_canal(conn, canal_id, 'profesor', p['id'])
    if tipo == 'institucional':
        for a in conn.execute('SELECT id FROM alumnos WHERE activo=1').fetchall():
            agregar_miembro_canal(conn, canal_id, 'estudiante', a['id'])
    elif tipo == 'director_curso' and curso:
        for d in conn.execute('SELECT id FROM directoras WHERE curso=? AND activo=1', (curso,)).fetchall():
            agregar_miembro_canal(conn, canal_id, 'directora', d['id'])
        for p in conn.execute('SELECT DISTINCT profesor_id FROM asignaciones_curso WHERE curso=?', (curso,)).fetchall():
            agregar_miembro_canal(conn, canal_id, 'profesor', p['profesor_id'])
    elif tipo == 'curso' and curso:
        for p in conn.execute('SELECT DISTINCT profesor_id FROM asignaciones_curso WHERE curso=?', (curso,)).fetchall():
            agregar_miembro_canal(conn, canal_id, 'profesor', p['profesor_id'])
        for a in conn.execute('SELECT id FROM alumnos WHERE curso=? AND activo=1', (curso,)).fetchall():
            agregar_miembro_canal(conn, canal_id, 'estudiante', a['id'])
    elif tipo == 'materia' and materia:
        for p in conn.execute('SELECT DISTINCT profesor_id FROM asignaciones_materia WHERE materia=?', (materia,)).fetchall():
            agregar_miembro_canal(conn, canal_id, 'profesor', p['profesor_id'])
        for cr in conn.execute('SELECT DISTINCT curso FROM actividades WHERE materia=?', (materia,)).fetchall():
            for a in conn.execute('SELECT id FROM alumnos WHERE curso=? AND activo=1', (cr['curso'],)).fetchall():
                agregar_miembro_canal(conn, canal_id, 'estudiante', a['id'])
    rector = get_rector(slug)
    if rector:
        agregar_miembro_canal(conn, canal_id, 'rector', rector['id'])

def nombre_usuario_canal(conn, tipo, uid):
    if tipo == 'profesor':
        r = conn.execute('SELECT nombre FROM profesores WHERE id=?', (uid,)).fetchone()
    elif tipo == 'estudiante':
        r = conn.execute('SELECT nombre FROM alumnos WHERE id=?', (uid,)).fetchone()
    elif tipo == 'rector':
        r = conn.execute('SELECT nombre FROM rectores WHERE id=?', (uid,)).fetchone()
    elif tipo == 'directora':
        r = conn.execute('SELECT nombre FROM directoras WHERE id=?', (uid,)).fetchone()
    else:
        return 'Desconocido'
    return r['nombre'] if r else 'Desconocido'

# ── FILE STORAGE (Fase 5) ─────────────────────────────────────────────────────
def max_tamano_archivo(slug):
    conn = conectar(slug)
    cfg = conn.execute('SELECT max_tamano_archivo FROM config_institucion WHERE slug=?', (slug,)).fetchone()
    conn.close()
    return cfg['max_tamano_archivo'] if cfg else 10485760

EXTENSIONES_PERMITIDAS = {
    '.pdf': 'application/pdf',
    '.doc': 'application/msword',
    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    '.xls': 'application/vnd.ms-excel',
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.ppt': 'application/vnd.ms-powerpoint',
    '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
    '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
    '.png': 'image/png', '.gif': 'image/gif', '.webp': 'image/webp',
    '.svg': 'image/svg+xml',
    '.txt': 'text/plain',
    '.csv': 'text/csv',
    '.zip': 'application/zip',
}

def guardar_archivo_mensaje(slug, canal_id, f, usuario_tipo, usuario_id):
    import os
    nombre_original = f.filename
    ext = os.path.splitext(nombre_original)[1].lower()
    if ext not in EXTENSIONES_PERMITIDAS:
        return None, 'Extensión no permitida'
    tamano = len(f.read())
    f.seek(0)
    max_sz = max_tamano_archivo(slug)
    if tamano > max_sz:
        return None, f'Archivo muy grande (máx {max_sz//1048576} MB)'
    es_img = ext in ('.jpg','.jpeg','.png','.gif','.webp')
    if es_img:
        try:
            from PIL import Image
            import io
            img = Image.open(io.BytesIO(f.read()))
            img.verify()
            f.seek(0)
        except Exception:
            return None, 'Archivo de imagen inválido o corrupto'
    nombre_archivo = f'{uuid.uuid4().hex}{ext}'
    upload_dir = os.path.join(app.root_path, 'static', 'uploads', slug)
    os.makedirs(upload_dir, exist_ok=True)
    ruta = os.path.join(upload_dir, nombre_archivo)
    f.save(ruta)
    ancho = alto = None
    if es_img:
        try:
            from PIL import Image
            img = Image.open(ruta)
            ancho, alto = img.size
        except Exception: pass
    conn = conectar(slug)
    try:
        fid = conn.execute(
            '''INSERT INTO mensajes_archivos
               (canal_id, usuario_tipo, usuario_id, nombre_original, nombre_archivo, tipo_mime, tamano, es_imagen, ancho, alto)
               VALUES (?,?,?,?,?,?,?,?,?,?)''',
            (canal_id, usuario_tipo, usuario_id, nombre_original, nombre_archivo,
             EXTENSIONES_PERMITIDAS[ext], tamano, 1 if es_img else 0, ancho, alto)).lastrowid
        conn.commit()
    finally:
        conn.close()
    return fid, None

def archivos_por_mensaje(conn, mensaje_id):
    return [dict(r) for r in conn.execute(
        'SELECT * FROM mensajes_archivos WHERE mensaje_id=? ORDER BY id', (mensaje_id,)).fetchall()]

def reacciones_por_mensaje(conn, mensaje_id):
    rows = conn.execute(
        'SELECT reaccion, usuario_tipo, usuario_id FROM mensajes_reacciones WHERE mensaje_id=?',
        (mensaje_id,)).fetchall()
    result = {}
    for r in rows:
        result.setdefault(r['reaccion'], []).append({'tipo': r['usuario_tipo'], 'id': r['usuario_id']})
    return result

# ── PDF REUTILIZABLE ──────────────────────────────────────────────────────────
def generar_pdf_alumno(alumno, slug, colegio, curso, jornada, periodo, conn):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        raise ImportError(
            'reportlab no está instalado. '
            'PDF no disponible. Instálelo con: pip install reportlab'
        )

    lista_materias = [r['materia'] for r in conn.execute(
        'SELECT DISTINCT materia FROM actividades WHERE curso=? AND jornada=? AND COALESCE(periodo,1)=? ORDER BY materia',
        (curso, jornada, periodo)
    ).fetchall()]

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    try:
        pri_color = colors.HexColor(colegio['primary_color']) if colegio and colegio['primary_color'] else colors.HexColor('#6c63ff')
    except (KeyError, AttributeError, TypeError):
        pri_color = colors.HexColor('#6c63ff')
    try:
        sec_color = colors.HexColor(colegio['secondary_color']) if colegio and colegio['secondary_color'] else colors.HexColor('#3498db')
    except (KeyError, AttributeError, TypeError):
        sec_color = colors.HexColor('#3498db')
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('t', fontSize=16, fontName='Helvetica-Bold',
                                  textColor=pri_color, spaceAfter=4)
    sub_style    = ParagraphStyle('s', fontSize=10, fontName='Helvetica',
                                  textColor=colors.grey, spaceAfter=10)
    mat_style    = ParagraphStyle('m', fontSize=11, fontName='Helvetica-Bold',
                                  textColor=pri_color, spaceBefore=10, spaceAfter=4)
    story = []
    story.append(Paragraph('LUMINI', titulo_style))
    story.append(Paragraph(
        f'Boletín — {colegio["nombre"] if colegio else slug} · {jornada} · Periodo {periodo}', sub_style))
    story.append(Paragraph(f'Estudiante: {alumno["nombre"]}   |   Curso: {curso}', styles['Normal']))
    story.append(Spacer(1, 0.4*cm))

    todos_finales = []
    ph = ','.join('?' * len(lista_materias))
    notas_all = conn.execute(
        f'''SELECT ac.materia, n.val FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
            WHERE n.aid=? AND ac.materia IN ({ph}) AND ac.curso=? AND ac.jornada=?
            AND COALESCE(ac.periodo,1)=?''',
        (alumno['id'],) + tuple(lista_materias) + (curso, jornada, periodo)
    ).fetchall()
    ev_all = conn.execute(
        f'''SELECT materia, evaluacion, autoevaluacion FROM evaluaciones
            WHERE aid=? AND materia IN ({ph}) AND jornada=? AND COALESCE(periodo,1)=?''',
        (alumno['id'],) + tuple(lista_materias) + (jornada, periodo)
    ).fetchall()
    notas_por_mat = {}
    for r in notas_all:
        notas_por_mat.setdefault(r['materia'], []).append(r['val'])
    ev_por_mat = {}
    for r in ev_all:
        ev_por_mat[r['materia']] = r
    for mat in lista_materias:
        notas_vals = notas_por_mat.get(mat, [])
        ev = ev_por_mat.get(mat)
        eval_v   = ev['evaluacion']     if ev and ev['evaluacion']     is not None else None
        auto_v   = ev['autoevaluacion'] if ev and ev['autoevaluacion'] is not None else None
        final = _promedio_ponderado(notas_vals, eval_v, auto_v)
        act_prom = round(sum(notas_vals) / len(notas_vals), 2) if notas_vals else None

        story.append(Paragraph(mat, mat_style))
        data = [['Actividades', 'Evaluación', 'Autoevaluación', 'Nota Final'],
                [str(act_prom) if act_prom is not None else '—',
                 str(eval_v)   if eval_v   is not None else '—',
                 str(auto_v)   if auto_v   is not None else '—',
                 str(final)    if final    is not None else '—']]
        t = Table(data, colWidths=[4*cm, 3.5*cm, 3.5*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), pri_color),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ]))
        story.append(t)
        if final is not None: todos_finales.append(final)

    prom_general = round(sum(todos_finales) / len(todos_finales), 2) if todos_finales else None
    story.append(Spacer(1, 0.5*cm))
    estado = 'Pendiente' if prom_general is None else ('Aprobado' if prom_general >= 3.0 else 'Reprobado')
    bg_color = pri_color if prom_general is not None and prom_general >= 3.0 else colors.HexColor('#e74c3c') if prom_general is not None else colors.HexColor('#64748B')
    resumen = Table(
        [['PROMEDIO GENERAL', str(prom_general) if prom_general is not None else '—', 'ESTADO', estado]],
        colWidths=[5*cm, 3*cm, 3*cm, 3*cm]
    )
    resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), bg_color),
        ('TEXTCOLOR',  (0, 0), (-1, -1), colors.white),
        ('FONTNAME',   (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 11),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(resumen)
    story.append(Spacer(1, 1*cm))
    doc.build(story)
    buf.seek(0)
    return buf.read(), prom_general

# ── ENVIAR CORREO ─────────────────────────────────────────────────────────────
def enviar_correo(destino, asunto, cuerpo_html, adjunto_bytes=None, adjunto_nombre=None, adjunto_tipo=None):
    if not SENDGRID_API_KEY:
        logger.error(f'Intento de envío a {destino} sin SENDGRID_API_KEY configurado.')
        return False
    try:
        import sendgrid
        from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
        import base64
        sg = sendgrid.SendGridAPIClient(api_key=SENDGRID_API_KEY)
        mensaje = Mail(from_email=EMAIL_ORIGEN, to_emails=destino,
                       subject=asunto, html_content=cuerpo_html)
        if adjunto_bytes and adjunto_nombre and adjunto_tipo:
            adjunto = Attachment(
                FileContent(base64.b64encode(adjunto_bytes).decode()),
                FileName(adjunto_nombre),
                FileType(adjunto_tipo),
                Disposition('attachment'))
            mensaje.attachment = adjunto
        sg.client.mail.send.post(request_body=mensaje.get())
        return True
    except Exception as e:
        logger.error(f'Error al enviar correo a {destino}: {e}')
        return False

# ── Auth routes migrated to app/routes/auth.py ──────────────────────────────

# ── SELECTOR DE JORNADA/MATERIA ───────────────────────────────────────────────
@app.route('/<slug>/seleccionar', methods=['GET', 'POST'])
def seleccionar_jornada(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    materias_jornadas = get_materias_profesor(slug, prof['id'])

    if not materias_jornadas:
        return render_template('seleccionar_jornada.html',
                               slug=slug, colegio=colegio, profesor=prof,
                               materias_jornadas=[],
                               error='No tienes materias asignadas. Contacta al administrador.')

    if request.method == 'POST':
        if not validar_csrf(): return ('Error CSRF', 403)
        materia = request.form.get('materia', '').strip()
        jornada = request.form.get('jornada', '').strip()
        if materia and jornada:
            session[f'materia_{slug}'] = materia
            session[f'jornada_{slug}'] = jornada
            return redirect(url_for('home', slug=slug))

    if len(materias_jornadas) == 1:
        session[f'materia_{slug}'] = materias_jornadas[0]['materia']
        session[f'jornada_{slug}'] = materias_jornadas[0]['jornada']
        return redirect(url_for('home', slug=slug))

    return render_template('seleccionar_jornada.html',
                           slug=slug, colegio=colegio, profesor=prof,
                           materias_jornadas=materias_jornadas)

# ── logout migrated to app/routes/auth.py ──────────────────────────────────

# ── HOME ──────────────────────────────────────────────────────────────────────
@app.route('/<slug>/')
@app.route('/<slug>')
def home(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))

    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('seleccionar_jornada', slug=slug))

    colegio     = get_colegio(slug)
    mis_cursos  = get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel   = request.args.get('curso', mis_cursos[0] if mis_cursos else None)
    periodo_sel = request.args.get('periodo', 1, type=int)

    conn = conectar(slug)
    try:
        alumnos = actividades = agenda = []

        if curso_sel and curso_sel in mis_cursos:
            alumnos = conn.execute(
                'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
                (curso_sel, jornada)).fetchall()
            # Batch update num_curso in a single query instead of N individual UPDATES
            if alumnos:
                case_parts = [f'WHEN ? THEN ?' for _ in alumnos]
                id_list = [a['id'] for a in alumnos]
                params = []
                for i, a in enumerate(alumnos, 1):
                    params.extend([a['id'], i])
                params.extend(id_list)
                conn.execute(
                    f'UPDATE alumnos SET num_curso = CASE id {" ".join(case_parts)} END WHERE id IN ({",".join("?" * len(alumnos))})',
                    params
                )
                conn.commit()
            alumnos = conn.execute(
                'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
                (curso_sel, jornada)).fetchall()
            actividades = conn.execute(
                '''SELECT * FROM actividades
                   WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
                   AND COALESCE(periodo,1)=? ORDER BY orden''',
                (prof['id'], materia, jornada, curso_sel, periodo_sel)).fetchall()
            agenda = conn.execute(
                'SELECT * FROM compromisos WHERE materia=? AND curso=? AND jornada=? ORDER BY fecha',
                (materia, curso_sel, jornada)).fetchall()

        MESES = {'01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr', '05': 'May', '06': 'Jun',
                 '07': 'Jul', '08': 'Ago', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'}
        datos = []
        # Pre-fetch grade data for all students (avoids N+1 queries)
        aid_list = [a['id'] for a in alumnos]
        if aid_list:
            placeholders = ','.join('?' * len(aid_list))
            notas_all = conn.execute(
                f'''SELECT n.aid, n.actividad_id, n.val, n.id FROM notas n
                    JOIN actividades ac ON ac.id=n.actividad_id
                    WHERE n.aid IN ({placeholders}) AND ac.materia=? AND ac.jornada=? AND ac.curso=?
                    AND COALESCE(ac.periodo,1)=? AND ac.profesor_id=? ORDER BY n.aid''',
                (*aid_list, materia, jornada, curso_sel, periodo_sel, prof['id'])).fetchall()
            notas_by_aid = {}
            for r in notas_all:
                notas_by_aid.setdefault(r['aid'], []).append(r)
            evals_all = conn.execute(
                f'''SELECT aid, evaluacion, autoevaluacion FROM evaluaciones
                    WHERE aid IN ({placeholders}) AND profesor_id=? AND materia=? AND jornada=?
                    AND COALESCE(periodo,1)=?''',
                (*aid_list, prof['id'], materia, jornada, periodo_sel)).fetchall()
            evals_by_aid = {r['aid']: r for r in evals_all}
        else:
            notas_by_aid = {}
            evals_by_aid = {}
        # Batch-fetch asistencia and observaciones to avoid N+1 queries
        asis_all = {}
        asis_ultimo = {}
        obs_all = {}
        if aid_list:
            rows_asistencia = conn.execute(
                f'SELECT aid, fecha, estado FROM asistencia WHERE aid IN ({placeholders}) ORDER BY aid, fecha',
                aid_list).fetchall()
            for r in rows_asistencia:
                asis_all.setdefault(r['aid'], []).append(r)
            rows_ultimo = conn.execute(
                f'SELECT aid, estado FROM asistencia WHERE aid IN ({placeholders}) AND fecha=date("now")',
                aid_list).fetchall()
            asis_ultimo = {r['aid']: r['estado'] for r in rows_ultimo}
            rows_obs = conn.execute(
                f'SELECT id, aid, materia, texto, fecha FROM observaciones WHERE aid IN ({placeholders}) AND materia=? ORDER BY aid, fecha DESC',
                (*aid_list, materia)).fetchall()
            for r in rows_obs:
                obs_all.setdefault(r['aid'], []).append(r)
        for a in alumnos:
            notas_raw = notas_by_aid.get(a['id'], [])
            notas_map = {nr['actividad_id']: {'val': nr['val'], 'id': nr['id']} for nr in notas_raw}
            ev = evals_by_aid.get(a['id'])
            eval_v = ev['evaluacion']     if ev and ev['evaluacion']     is not None else None
            auto_v = ev['autoevaluacion'] if ev and ev['autoevaluacion'] is not None else None
            vals = [nr['val'] for nr in notas_raw]
            prom = _promedio_simple(vals)
            nota_final = _promedio_ponderado(vals, eval_v, auto_v)
            historial_raw = asis_all.get(a['id'], [])
            hist_meses = {}
            for h in historial_raw:
                if h['fecha']:
                    p2 = h['fecha'].split('-')
                    if len(p2) >= 2:
                        label = f"{MESES.get(p2[1], p2[1])} {p2[0]}"
                        hist_meses.setdefault(label, []).append({'fecha': h['fecha'], 'estado': h['estado']})
            ult_estado = asis_ultimo.get(a['id'])
            obs = obs_all.get(a['id'], [])
            datos.append({
                'id': a['id'], 'num_curso': a['num_curso'],
                'nombre': a['nombre'], 'curso': a['curso'],
                'promedio': prom, 'nota_final': nota_final, 'notas_map': notas_map,
                'evaluacion':     eval_v if eval_v is not None else '',
                'autoevaluacion': auto_v if auto_v is not None else '',
                'asistencia': ult_estado or '-',
                'historial_meses': hist_meses,
                'observaciones': [dict(o) for o in obs],
            })

        promedios = [d['promedio'] for d in datos if d['promedio'] is not None]
        prom_gral = round(sum(promedios) / len(promedios), 2) if promedios else None
        mejor     = max(datos, key=lambda x: x['promedio'] or 0, default={'nombre': 'N/A', 'promedio': None})

        # ── Dashboard data ──────────────────────────────────────────────────
        DIAS = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
        hoy_idx = datetime.today().weekday()
        hoy_nombre = DIAS[hoy_idx] if hoy_idx < 7 else ''
        hoy_fecha = datetime.today().strftime('%Y-%m-%d')
        total_alumnos = conn.execute(
            f'SELECT COUNT(*) as c FROM alumnos WHERE curso IN ({",".join("?" * len(mis_cursos))}) AND jornada=? AND activo=1',
            (*mis_cursos, jornada)
        ).fetchone()['c'] if mis_cursos else 0
        horario_hoy = conn.execute(
            'SELECT * FROM horarios_curso WHERE materia=? AND jornada=? AND dia=? ORDER BY franja',
            (materia, jornada, hoy_nombre)
        ).fetchall() if curso_sel else []
        asis_hoy = conn.execute(
            "SELECT COUNT(DISTINCT aid) as total FROM asistencia WHERE fecha=?",
            (hoy_fecha,)
        ).fetchone()
        asistencia_hoy = asis_hoy['total'] if asis_hoy else 0
        notas_pend = 0
        if curso_sel and actividades:
            act_ids = [a['id'] for a in actividades]
            placeholders = ','.join('?' * len(act_ids))
            rows_present = conn.execute(
                f'SELECT actividad_id, COUNT(DISTINCT aid) as cnt FROM notas n JOIN alumnos al ON al.id=n.aid AND al.activo=1 WHERE n.actividad_id IN ({placeholders}) GROUP BY n.actividad_id',
                act_ids
            ).fetchall()
            present_sum = sum(r['cnt'] for r in rows_present)
            total_alumnos_curso = conn.execute(
                'SELECT COUNT(*) as c FROM alumnos WHERE curso=? AND jornada=? AND activo=1',
                (curso_sel, jornada)
            ).fetchone()['c']
            notas_pend = total_alumnos_curso * len(act_ids) - present_sum
        alertas = []
        if curso_sel:
            rows = conn.execute(
                "SELECT a.nombre, a.id, COUNT(*) as faltas FROM asistencia asis JOIN alumnos a ON a.id=asis.aid WHERE asis.estado='A' AND a.curso=? AND a.jornada=? AND a.activo=1 GROUP BY asis.aid HAVING faltas > 1 ORDER BY faltas DESC LIMIT 5",
                (curso_sel, jornada)
            ).fetchall()
            for r in rows: alertas.append({'nombre': r['nombre'], 'faltas': r['faltas']})
            for e in datos:
                if e['promedio'] is not None and e['promedio'] < 3.0:
                    alertas.append({'nombre': e['nombre'], 'promedio': e['promedio']})
            alertas = alertas[:5]
        pendientes = comunicaciones_pendientes(slug, 'profesor', prof['id'])
        num_periodos = int(colegio['num_periodos']) if colegio and colegio['num_periodos'] else 4
        pc = periodo_cerrado(slug, periodo_sel) if curso_sel and materia else False
        error_msg = request.args.get('error', '')
        if error_msg == 'periodo_cerrado':
            error_msg = 'El per\u00edodo est\u00e1 cerrado. No se pueden crear actividades.'
        solicitudes_pend = conn.execute(
            'SELECT COUNT(*) as c FROM solicitudes_modificacion WHERE profesor_id=? AND estado=? AND slug=?',
            (prof['id'], 'pendiente', slug)).fetchone()['c'] if curso_sel else 0
    finally:
        conn.close()
    return render_template('index.html',
                           profesor=prof, mis_cursos=mis_cursos, curso_sel=curso_sel,
                           estudiantes=datos, actividades=actividades, compromisos=agenda,
                           prom_general=prom_gral, mejor=mejor, slug=slug, colegio=colegio,
                           num_periodos=num_periodos, periodo_sel=periodo_sel,
                           materia=materia, jornada=jornada,
                           materias_jornadas=get_materias_profesor(slug, prof['id']),
                           hoy_nombre=hoy_nombre, hoy_fecha=hoy_fecha,
                           total_alumnos=total_alumnos, horario_hoy=horario_hoy,
                           asistencia_hoy=asistencia_hoy, notas_pend=notas_pend,
                             alertas=alertas,
                             periodo_cerrado=pc,
                             error_msg=error_msg,
                             solicitudes_pendientes_mod=solicitudes_pend,
                            comunicaciones_pendientes=pendientes)

# ── ACTIVIDADES ───────────────────────────────────────────────────────────────
@app.route('/<slug>/nueva_actividad', methods=['POST'])
def nueva_actividad(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    nombre    = request.form.get('nombre', '').strip()
    curso_sel = request.form.get('curso_sel', '')
    periodo   = request.form.get('periodo_sel', 1, type=int)
    if nombre and curso_sel and materia and jornada:
        if periodo_cerrado(slug, periodo):
            return redirect(url_for('home', slug=slug, curso=curso_sel, periodo=periodo, error='periodo_cerrado'))
        conn = conectar(slug)
        max_ord = conn.execute(
            '''SELECT COALESCE(MAX(orden),0) FROM actividades
               WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
               AND COALESCE(periodo,1)=?''',
            (prof['id'], materia, jornada, curso_sel, periodo)).fetchone()[0]
        conn.execute(
            'INSERT INTO actividades (profesor_id,materia,jornada,curso,nombre,orden,periodo) VALUES (?,?,?,?,?,?,?)',
            (prof['id'], materia, jornada, curso_sel, nombre, max_ord + 1, periodo))
        conn.commit(); conn.close()
    return redirect(url_for('home', slug=slug, curso=curso_sel, periodo=periodo))

@app.route('/<slug>/borrar_actividad/<int:act_id>', methods=['POST'])
def borrar_actividad(slug, act_id):
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    conn = conectar(slug)
    act = conn.execute('SELECT profesor_id, curso, materia, COALESCE(periodo,1) as p FROM actividades WHERE id=?', (act_id,)).fetchone()
    if not act or act['profesor_id'] != prof['id']:
        conn.close()
        return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
    if periodo_cerrado(slug, act['p']):
        conn.close()
        return jsonify({'status':'error','codigo':'PERIODO_CERRADO','mensaje':'El per\u00edodo est\u00e1 cerrado.'}), 403
    notas_borradas = conn.execute('SELECT aid, val FROM notas WHERE actividad_id=?', (act_id,)).fetchall()
    conn.execute('DELETE FROM notas WHERE actividad_id=?', (act_id,))
    conn.execute('DELETE FROM actividades WHERE id=?', (act_id,))
    conn.commit()
    curso = act['curso']
    jornada_ctx, materia_ctx = get_sesion_jornada_materia(slug)
    for n in notas_borradas:
        auditar_nota(slug, prof['id'], 'profesor', 'eliminacion', 'notas', n['aid'],
                     act['curso'], materia_ctx or act['materia'], act['p'],
                     campo='nota', actividad_id=act_id,
                     valor_anterior=n['val'], valor_nuevo=None)
    auditar_nota(slug, prof['id'], 'profesor', 'eliminacion', 'actividades', None,
                 act['curso'], materia_ctx or act['materia'], act['p'],
                 actividad_id=act_id, valor_anterior=act_id, valor_nuevo=None,
                 motivo='Actividad eliminada')
    conn.close()
    return jsonify({'status':'ok','actividad_id':act_id,'redirect':url_for('home', slug=slug, curso=curso)})

# ── ACTIVIDADES JSON API (Fase 17) ────────────────────────────────────────────
@app.route('/<slug>/actividades/crear', methods=['POST'])
def actividades_crear(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    data = request.get_json(silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    curso_sel = (data.get('curso') or '').strip()
    periodo = data.get('periodo', 1)
    tipo = data.get('tipo', 'taller')
    peso = data.get('peso')
    fecha_limite = data.get('fecha_limite') or ''
    hora_limite = data.get('hora_limite') or ''
    descripcion = (data.get('descripcion') or '').strip()
    observaciones = (data.get('observaciones') or '').strip()
    estado_act = data.get('estado', 'borrador')
    competencia = (data.get('competencia') or '').strip()
    entrega_digital = 1 if data.get('entrega_digital') else 0
    if not nombre or not curso_sel:
        return jsonify({'status':'error','mensaje':'Nombre y curso requeridos'}), 400
    if periodo_cerrado(slug, periodo):
        return jsonify({'status':'error','codigo':'PERIODO_CERRADO','mensaje':'Periodo cerrado'}), 403
    conn = conectar(slug)
    try:
        max_ord = conn.execute(
            '''SELECT COALESCE(MAX(orden),0) FROM actividades
               WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
               AND COALESCE(periodo,1)=?''',
            (prof['id'], materia, jornada, curso_sel, periodo)).fetchone()[0]
        c = conn.execute(
            '''INSERT INTO actividades
               (profesor_id,materia,jornada,curso,nombre,orden,periodo,tipo,peso,fecha_limite,
                hora_limite,descripcion,observaciones,estado_act,competencia,entrega_digital)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (prof['id'], materia, jornada, curso_sel, nombre, max_ord + 1, periodo,
             tipo, peso, fecha_limite or None, hora_limite or None, descripcion,
             observaciones, estado_act, competencia, entrega_digital))
        act_id = c.lastrowid
        conn.commit()
        audit_log(slug, prof['id'], 'actividad_creada', 'actividades',
                  registro_id=act_id,
                  valor_nuevo={'nombre':nombre,'tipo':tipo,'curso':curso_sel,'materia':materia})
        auditar_nota(slug, prof['id'], 'profesor', 'creacion', 'actividades', None,
                     curso_sel, materia, periodo, actividad_id=act_id,
                     valor_nuevo=nombre, motivo='Actividad creada via API')
        # Auto-create calendar event if fecha_limite is set
        if fecha_limite:
            conn.execute(
                '''INSERT INTO eventos_calendario
                   (slug,tipo,titulo,descripcion,fecha_inicio,fecha_fin,todo_el_dia,curso,
                    creado_por_tipo,creado_por_id,color)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (slug, 'actividad', nombre + ' (' + tipo + ')',
                 descripcion[:200] if descripcion else '',
                 fecha_limite + ('T' + hora_limite if hora_limite else ''),
                 fecha_limite + ('T' + hora_limite if hora_limite else ''),
                 0 if hora_limite else 1,
                 curso_sel, 'profesor', prof['id'], '#6c63ff'))
        return jsonify({
            'status': 'ok',
            'actividad': {
                'id': act_id, 'nombre': nombre, 'tipo': tipo, 'orden': max_ord + 1,
                'peso': peso, 'fecha_limite': fecha_limite, 'hora_limite': hora_limite,
                'descripcion': descripcion, 'observaciones': observaciones,
                'estado_act': estado_act, 'competencia': competencia,
                'entrega_digital': entrega_digital, 'periodo': periodo
            }
        })
    except Exception as e:
        conn.close()
        logger.error(f'Error creando actividad: {e}')
        return jsonify({'status':'error','mensaje':'Error al guardar'}), 500
    finally:
        conn.close()

@app.route('/<slug>/actividades/<int:act_id>', methods=['PUT'])
def actividades_editar(slug, act_id):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
    changes = []
    for field in ('nombre','tipo','peso','fecha_limite','hora_limite','descripcion','observaciones','estado_act','competencia','entrega_digital'):
        if field in data:
            old_val = act[field]
            new_val = data[field]
            if str(old_val) != str(new_val):
                changes.append(field + '=?')
                conn.execute('UPDATE actividades SET ' + field + '=? WHERE id=?', (new_val, act_id))
                audit_log(slug, prof['id'], 'actividad_' + field, 'actividades', registro_id=act_id,
                          valor_anterior={field: old_val}, valor_nuevo={field: new_val})
    if not changes:
        conn.close()
        act_dict = dict(act)
        act_dict.pop('profesor_id', None)
        return jsonify({'status':'ok','actividad':act_dict})
    conn.commit()
    updated = conn.execute('SELECT * FROM actividades WHERE id=?', (act_id,)).fetchone()
    conn.close()
    act_dict = dict(updated)
    act_dict.pop('profesor_id', None)
    return jsonify({'status':'ok','actividad':act_dict})

@app.route('/<slug>/actividades/<int:act_id>/estado', methods=['POST'])
def actividades_cambiar_estado(slug, act_id):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    if request.is_json:
        estado = (request.get_json(silent=True) or {}).get('estado', '').strip()
    else:
        estado = (request.form.get('estado') or '').strip()
    if estado not in ('borrador','publicada','cerrada','archivada'):
        return jsonify({'status':'error','mensaje':'Estado invalido'}), 400
    conn = conectar(slug)
    act = conn.execute('SELECT id, estado_act FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
    if estado == 'archivada' and act['estado_act'] not in ('cerrada', 'publicada'):
        conn.close()
        return jsonify({'status':'error','mensaje':'Solo actividades cerradas o publicadas pueden archivarse'}), 400
    conn.execute('UPDATE actividades SET estado_act=? WHERE id=?', (estado, act_id))
    conn.commit()
    audit_log(slug, prof['id'], 'actividad_estado', 'actividades', registro_id=act_id,
              valor_anterior={'estado':act['estado_act']}, valor_nuevo={'estado':estado})
    conn.close()
    return jsonify({'status':'ok','estado':estado})

@app.route('/<slug>/actividades/<int:act_id>/detalle', methods=['GET'])
def actividades_detalle(slug, act_id):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
    act_dict = dict(act)
    act_dict.pop('profesor_id', None)
    conn.close()
    return jsonify({'status':'ok','actividad':act_dict})

@app.route('/<slug>/actividades/<int:act_id>/duplicar', methods=['POST'])
def actividades_duplicar(slug, act_id):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
    max_ord = conn.execute(
        'SELECT COALESCE(MAX(orden),0) FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=?',
        (prof['id'], act['materia'], act['jornada'], act['curso'], act['periodo'] or 1)).fetchone()[0]
    nuevo_nombre = (act['nombre'] or '').strip()
    import re
    m = re.search(r'(\d+)$', nuevo_nombre)
    if m:
        nuevo_nombre = nuevo_nombre[:m.start()] + str(int(m.group(1)) + 1)
    else:
        nuevo_nombre = nuevo_nombre + ' 2'
    c = conn.execute(
        '''INSERT INTO actividades
           (profesor_id,materia,jornada,curso,nombre,orden,periodo,tipo,peso,fecha_limite,
            hora_limite,descripcion,observaciones,estado_act,competencia,entrega_digital)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (prof['id'], act['materia'], act['jornada'], act['curso'], nuevo_nombre, max_ord + 1,
         act['periodo'] or 1, act['tipo'], act['peso'], act['fecha_limite'], act['hora_limite'],
         act['descripcion'], act['observaciones'], 'borrador', act['competencia'], act['entrega_digital']))
    new_id = c.lastrowid
    conn.commit()
    audit_log(slug, prof['id'], 'actividad_duplicada', 'actividades', registro_id=new_id,
              valor_nuevo={'desde':act_id,'nombre':nuevo_nombre})
    nueva = dict(conn.execute('SELECT * FROM actividades WHERE id=?', (new_id,)).fetchone())
    conn.close()
    nueva.pop('profesor_id', None)
    return jsonify({'status':'ok','actividad':nueva})

@app.route('/<slug>/actividades/<int:act_id>/historial', methods=['GET'])
def actividades_historial(slug, act_id):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    conn = conectar(slug)
    act = conn.execute('SELECT id FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
    rows = conn.execute(
        '''SELECT usuario_id, accion, valor_anterior, valor_nuevo, creado, ip
           FROM audit_log WHERE tabla='actividades' AND registro_id=?
           ORDER BY id DESC LIMIT 100''', (act_id,)).fetchall()
    conn.close()
    return jsonify({'status':'ok','historial':[dict(r) for r in rows]})

@app.route('/<slug>/actividades/<int:act_id>/estadisticas', methods=['GET'])
def actividades_estadisticas(slug, act_id):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
    notas = conn.execute(
        'SELECT n.val, a.nombre, a.id as aid FROM notas n JOIN alumnos a ON a.id=n.aid WHERE n.actividad_id=?',
        (act_id,)).fetchall()
    vals = [float(r['val']) for r in notas if r['val'] is not None]
    total = len(vals)
    prom = round(sum(vals)/total, 2) if total else None
    mx = max(vals) if total else None
    mn = min(vals) if total else None
    aprobados = sum(1 for v in vals if v >= 3.0)
    reprobados = sum(1 for v in vals if v < 3.0 and v > 0)
    distribucion = {'1':0,'2':0,'3':0,'4':0,'5':0}
    for v in vals:
        k = str(int(v))
        if k in distribucion: distribucion[k] += 1
    conn.close()
    return jsonify({'status':'ok','estadisticas':{
        'total_notas':total,'promedio':prom,'max':mx,'min':mn,
        'aprobados':aprobados,'reprobados':reprobados,'distribucion':distribucion
    }})

# ── CENTRAL WEIGHTED AVERAGE (65/25/10) ──────────────────────────────────────
def _promedio_simple(notas_actividades):
    """Average of graded activities only (empty = not counted)."""
    if not notas_actividades:
        return None
    vals = [v for v in notas_actividades if v is not None]
    if not vals:
        return None
    return round(sum(vals) / len(vals), 2)

def _promedio_ponderado(notas_actividades, evaluacion, autoevaluacion):
    act_prom = _promedio_simple(notas_actividades)
    logger.debug('_promedio_ponderado: act_prom=%s evaluacion=%s autoevaluacion=%s', act_prom, evaluacion, autoevaluacion)
    nota_final = 0
    tiene_datos = False
    if act_prom is not None:
        nota_final += act_prom * 0.65
        tiene_datos = True
    if evaluacion is not None:
        nota_final += evaluacion * 0.25
        tiene_datos = True
    if autoevaluacion is not None:
        nota_final += autoevaluacion * 0.10
        tiene_datos = True
    resultado = round(nota_final, 2) if tiene_datos else None
    logger.debug('_promedio_ponderado: resultado=%s', resultado)
    return resultado

def calcular_stats_estudiante(conn, slug, aid, curso_sel, materia, jornada, periodo, profesor_id):
    notas_raw = conn.execute(
        '''SELECT n.val FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
           WHERE n.aid=? AND ac.materia=? AND ac.jornada=? AND ac.curso=?
           AND COALESCE(ac.periodo,1)=? AND ac.profesor_id=?''',
        (aid, materia, jornada, curso_sel, periodo, profesor_id)).fetchall()
    vals = [r['val'] for r in notas_raw] if notas_raw else []
    logger.debug('calcular_stats_estudiante slug=%s aid=%d curso=%s materia=%s jornada=%s periodo=%d prof=%d notas_raw=%d vals=%s',
                 slug, aid, curso_sel, materia, jornada, periodo, profesor_id,
                 len(notas_raw), vals)
    return _promedio_simple(vals)

def calcular_nota_final_estudiante(conn, slug, aid, curso_sel, materia, jornada, periodo, profesor_id):
    notas_raw = conn.execute(
        '''SELECT n.val FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
           WHERE n.aid=? AND ac.materia=? AND ac.jornada=? AND ac.curso=?
           AND COALESCE(ac.periodo,1)=? AND ac.profesor_id=?''',
        (aid, materia, jornada, curso_sel, periodo, profesor_id)).fetchall()
    ev = conn.execute(
        '''SELECT evaluacion, autoevaluacion FROM evaluaciones
           WHERE aid=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?''',
        (aid, materia, jornada, periodo)).fetchone()
    vals = [r['val'] for r in notas_raw] if notas_raw else []
    eval_v   = ev['evaluacion']     if ev and ev['evaluacion']     is not None else None
    auto_v   = ev['autoevaluacion'] if ev and ev['autoevaluacion'] is not None else None
    return _promedio_ponderado(vals, eval_v, auto_v)

def calcular_stats_curso(conn, slug, curso_sel, materia, jornada, periodo, profesor_id):
    alumnos = conn.execute(
        'SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1',
        (curso_sel, jornada)).fetchall()
    promedios = []
    for a in alumnos:
        p = calcular_stats_estudiante(conn, slug, a['id'], curso_sel, materia, jornada, periodo, profesor_id)
        if p is not None: promedios.append(p)
    prom_curso = round(sum(promedios) / len(promedios), 2) if promedios else None
    # Pending grades
    total_est = len(alumnos)
    act_ids = conn.execute(
        '''SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? AND profesor_id=?''',
        (materia, jornada, curso_sel, periodo, profesor_id)).fetchall()
    act_count = len(act_ids)
    notas_count = conn.execute(
        '''SELECT COUNT(*) as c FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
           WHERE ac.materia=? AND ac.jornada=? AND ac.curso=? AND COALESCE(ac.periodo,1)=?
           AND ac.profesor_id=?''',
        (materia, jornada, curso_sel, periodo, profesor_id)).fetchone()['c']
    pend = total_est * act_count - notas_count if total_est and act_count else 0
    return {'promedio_curso': prom_curso, 'notas_pendientes': max(pend, 0)}

# ── NOTAS (FASE 3: BATCH + REORDER) ──────────────────────────────────────────
@app.route('/<slug>/reordenar_actividades', methods=['POST'])
def reordenar_actividades(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    ordenes = data.get('orden', [])
    if not ordenes: return jsonify({'status':'error','mensaje':'Sin datos'}), 400
    conn = conectar(slug)
    try:
        for item in ordenes:
            act_id = item.get('id')
            orden = item.get('orden', 0)
            conn.execute('UPDATE actividades SET orden=? WHERE id=? AND profesor_id=?',
                         (orden, act_id, prof['id']))
        conn.commit()
        return jsonify({'status':'ok'})
    except Exception as e:
        conn.close()
        logger.error(f'Error reordenando: {e}')
        return jsonify({'status':'error','mensaje':'Error al reordenar'}), 500
    finally:
        conn.close()

@app.route('/<slug>/notas/batch', methods=['POST'])
def notas_batch(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    notas = data.get('notas', [])
    if not notas: return jsonify({'status':'error','mensaje':'Sin datos'}), 400
    conn = conectar(slug)
    errors = []
    saved = []
    try:
        for item in notas:
            aid = item.get('aid')
            actividad_id = item.get('actividad_id')
            val = item.get('val')
            if None in (aid, actividad_id, val):
                errors.append({'aid':aid,'actividad_id':actividad_id,'error':'Datos invalidos'})
                continue
            act = conn.execute(
                'SELECT a.id, a.profesor_id, a.curso, COALESCE(a.periodo,1) as p FROM actividades a WHERE a.id=?',
                (actividad_id,)).fetchone()
            if not act or act['profesor_id'] != prof['id']:
                errors.append({'aid':aid,'actividad_id':actividad_id,'error':'No autorizado'})
                continue
            if periodo_cerrado(slug, act['p']):
                errors.append({'aid':aid,'actividad_id':actividad_id,'error':'Periodo cerrado'})
                continue
            old = conn.execute(
                'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
                (aid, actividad_id)).fetchone()
            old_val = old['val'] if old else None
            if old:
                conn.execute('UPDATE notas SET val=? WHERE aid=? AND actividad_id=?',
                             (val, aid, actividad_id))
            else:
                conn.execute('INSERT INTO notas (aid, actividad_id, val) VALUES (?,?,?)',
                             (aid, actividad_id, val))
            auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'notas', aid,
                         act['curso'], act['materia'], act['p'],
                         campo='nota', actividad_id=actividad_id,
                         valor_anterior=old_val, valor_nuevo=val)
            saved.append({'aid':aid,'actividad_id':actividad_id,'val_anterior':old_val,'val_nuevo':val})
        conn.commit()
        return jsonify({'status':'ok','saved':len(saved),'errors':errors,
                        'snapshot':saved[-5:] if saved else []})
    except Exception as e:
        conn.close()
        logger.error(f'Error batch notas: {e}')
        return jsonify({'status':'error','mensaje':'Error al guardar'}), 500
    finally:
        conn.close()

@app.route('/<slug>/notas/deshacer', methods=['POST'])
def notas_deshacer(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    aid = data.get('aid')
    actividad_id = data.get('actividad_id')
    val_anterior = data.get('val')
    if None in (aid, actividad_id):
        return jsonify({'status':'error','mensaje':'Datos invalidos'}), 400
    conn = conectar(slug)
    try:
        act = conn.execute(
            'SELECT profesor_id, curso, materia, COALESCE(periodo,1) as p FROM actividades WHERE id=?',
            (actividad_id,)).fetchone()
        if not act or act['profesor_id'] != prof['id']:
            conn.close()
            return jsonify({'status':'error','mensaje':'No autorizado'}), 403
        if val_anterior is not None:
            conn.execute('UPDATE notas SET val=? WHERE aid=? AND actividad_id=?',
                         (val_anterior, aid, actividad_id))
        else:
            conn.execute('DELETE FROM notas WHERE aid=? AND actividad_id=?',
                         (aid, actividad_id))
        conn.commit()
        auditar_nota(slug, prof['id'], 'profesor', 'deshacer', 'notas', aid,
                     act['curso'], act['materia'], act['p'],
                     campo='nota', actividad_id=actividad_id,
                     valor_nuevo=val_anterior, motivo='Deshacer')
        conn.close()
        return jsonify({'status':'ok','val':val_anterior})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','mensaje':str(e)}), 500

@app.route('/<slug>/observaciones_json', methods=['POST'])
def observaciones_json(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'observaciones':[]})
    data = request.get_json(silent=True) or {}
    aid = data.get('aid')
    if not aid: return jsonify({'observaciones':[]})
    conn = conectar(slug)
    rows = conn.execute(
        'SELECT id, materia, texto, fecha FROM observaciones WHERE aid=? ORDER BY fecha DESC LIMIT 20',
        (aid,)).fetchall()
    conn.close()
    return jsonify({'observaciones':[dict(r) for r in rows]})

@app.route('/<slug>/recalcular/<int:aid>')
def recalcular(slug, aid):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso_sel = request.args.get('curso') or ''
        periodo = request.args.get('periodo', 1, type=int)
        if not curso_sel:
            al = conn.execute('SELECT curso FROM alumnos WHERE id=?', (aid,)).fetchone()
            curso_sel = al['curso'] if al else ''
        prom = calcular_stats_estudiante(conn, slug, aid, curso_sel, materia, jornada, periodo, prof['id'])
        nf = calcular_nota_final_estudiante(conn, slug, aid, curso_sel, materia, jornada, periodo, prof['id'])
        return jsonify({'promedio':prom,'nota_final':nf})
    except Exception as e:
        return jsonify({'error':str(e)}), 500
    finally:
        conn.close()

# ── ANALYTICS ─────────────────────────────────────────────────────────────────
@app.route('/<slug>/curso/analitica')
def curso_analitica(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso:
            conn.close(); return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute(
            'SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
        act_ids = [r['id'] for r in conn.execute(
            'SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()]
        notas_por_act = {}
        for act_id in act_ids:
            vals = [float(r['val']) for r in conn.execute(
                'SELECT val FROM notas WHERE actividad_id=?', (act_id,)).fetchall() if r['val'] is not None]
            notas_por_act[act_id] = vals
        todos_vals = [v for vals in notas_por_act.values() for v in vals]
        promedios = []
        estudiantes_data = []
        for al in alumnos:
            vals_al = []
            for act_id in act_ids:
                r = conn.execute(
                    'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
                    (al['id'], act_id)).fetchone()
                if r and r['val'] is not None:
                    vals_al.append(float(r['val']))
            prom = round(sum(vals_al)/len(vals_al),2) if vals_al else None
            if prom is not None:
                promedios.append(prom)
            estudiantes_data.append({'id':al['id'],'nombre':al['nombre'],'promedio':prom,'notas_count':len(vals_al)})
        max_val = max(todos_vals) if todos_vals else None
        min_val = min(todos_vals) if todos_vals else None
        prom_curso = round(sum(promedios)/len(promedios),2) if promedios else None
        aprobando = sum(1 for p in promedios if p is not None and p >= 3.0)
        perdiendo = sum(1 for p in promedios if p is not None and p < 3.0)
        sin_notas = sum(1 for e in estudiantes_data if e['promedio'] is None)
        riesgo_alto = sum(1 for p in promedios if p is not None and p < 2.8)
        riesgo_medio = sum(1 for p in promedios if p is not None and 2.8 <= p <= 3.5)
        # Actividades stats
        acts_info = []
        for act_id in act_ids:
            vals = notas_por_act.get(act_id, [])
            a = conn.execute('SELECT nombre, tipo, peso, competencia FROM actividades WHERE id=?', (act_id,)).fetchone()
            if a:
                acts_info.append({
                    'id':act_id,'nombre':a['nombre'],'tipo':a['tipo'],'peso':a['peso'],
                    'promedio':round(sum(vals)/len(vals),2) if vals else None,
                    'cantidad':len(vals),
                    'aprobados':sum(1 for v in vals if v>=3.0),
                    'reprobados':sum(1 for v in vals if v<3.0 and v>0),
                    'total_estudiantes':len(alumnos)
                })
        # Distribution
        dist = {'1':0,'2':0,'3':0,'4':0,'5':0}
        for v in todos_vals:
            k = str(int(v))
            if k in dist: dist[k] += 1
        return jsonify({
            'promedio_curso':prom_curso,'max':max_val,'min':min_val,
            'aprobando':aprobando,'perdiendo':perdiendo,'sin_notas':sin_notas,
            'riesgo_alto':riesgo_alto,'riesgo_medio':riesgo_medio,
            'total_estudiantes':len(alumnos),'actividades':acts_info,
            'distribucion':dist,'total_notas':len(todos_vals)
        })
    finally:
        conn.close()

@app.route('/<slug>/curso/ranking')
def curso_ranking(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute(
            'SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
        act_ids = [r['id'] for r in conn.execute(
            'SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()]
        data = []
        for al in alumnos:
            vals = []
            for act_id in act_ids:
                r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                if r and r['val'] is not None: vals.append(float(r['val']))
            prom = round(sum(vals)/len(vals),2) if vals else None
            ev = conn.execute(
                'SELECT evaluacion, autoevaluacion FROM evaluaciones WHERE aid=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                (al['id'], materia, jornada, periodo)).fetchone()
            nf = None
            if prom is not None:
                e = ev['evaluacion'] if ev else None
                a = ev['autoevaluacion'] if ev else None
                nf = round(prom*0.65 + (e or 0)*0.25 + (a or 0)*0.10, 2)
            data.append({'id':al['id'],'nombre':al['nombre'],'promedio':prom,'nota_final':nf,'notas_count':len(vals)})
        data.sort(key=lambda x: (x['promedio'] or 0), reverse=True)
        for i, d in enumerate(data):
            d['posicion'] = i + 1
        return jsonify({'ranking':data[:30],'total':len(data)})
    finally:
        conn.close()

@app.route('/<slug>/estudiante/<int:aid>/tendencia')
def estudiante_tendencia(slug, aid):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        rows = conn.execute(
            '''SELECT n.val, n.actividad_id, ac.nombre as act_nombre, ac.orden, ac.periodo
               FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
               WHERE n.aid=? AND ac.materia=? AND ac.jornada=? AND ac.profesor_id=?
               ORDER BY ac.periodo, ac.orden''',
            (aid, materia, jornada, prof['id'])).fetchall()
        puntos = [{'orden':r['orden'],'valor':float(r['val']),'nombre':r['act_nombre'],
                    'periodo':r['periodo']} for r in rows if r['val'] is not None]
        # Calculate running average
        acum = []; running = []
        for p in puntos:
            acum.append(p['valor'])
            running.append(round(sum(acum)/len(acum),2))
        for i, p in enumerate(puntos):
            p['promedio_acumulado'] = running[i] if i < len(running) else p['valor']
        # Prediction (simple linear extrapolation)
        if len(puntos) >= 2:
            vals = [p['valor'] for p in puntos]
            n = len(vals)
            x_avg = (n - 1) / 2
            y_avg = sum(vals) / n
            num = sum((i - x_avg) * (v - y_avg) for i, v in enumerate(vals))
            den = sum((i - x_avg) ** 2 for i in range(n))
            slope = num / den if den != 0 else 0
            pred = y_avg + slope * (n + 2 - x_avg)  # predict 2 steps ahead
            pred = max(0, min(5, round(pred, 2)))
            confianza = min(95, max(30, int(100 - abs(slope) * 20)))
        else:
            pred = puntos[-1]['valor'] if puntos else None
            confianza = 30 if pred else 0
        # Course comparison
        all_vals = []
        for p in puntos:
            r = conn.execute('SELECT val FROM notas WHERE actividad_id=?', (p['actividad_id'],)).fetchall()
            all_vals.extend([float(x['val']) for x in r if x['val'] is not None])
        prom_curso = round(sum(all_vals)/len(all_vals),2) if all_vals else None
        prom_est = puntos[-1]['promedio_acumulado'] if puntos else None
        diff = None
        if prom_est is not None and prom_curso is not None and prom_curso > 0:
            diff = round((prom_est - prom_curso) / prom_curso * 100, 1)
        return jsonify({
            'puntos':puntos,
            'prediccion':pred,
            'confianza':confianza,
            'promedio_estudiante':prom_est,
            'promedio_curso':prom_curso,
            'diferencia_porcentual':diff
        })
    finally:
        conn.close()

@app.route('/<slug>/observaciones/sugerir', methods=['POST'])
def observaciones_sugerir(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    aid = data.get('aid')
    cambio = data.get('cambio', '')  # 'subio' or 'bajo'
    if not aid: return jsonify({'sugerencia':''})
    conn = conectar(slug)
    try:
        al = conn.execute('SELECT nombre FROM alumnos WHERE id=?', (aid,)).fetchone()
        nombre = al['nombre'] if al else 'El estudiante'
        if cambio == 'bajo':
            sugerencia = f'{nombre} disminuyó su rendimiento académico.'
        elif cambio == 'subio':
            sugerencia = f'{nombre} presenta una mejora constante en su rendimiento.'
        else:
            sugerencia = ''
        return jsonify({'sugerencia':sugerencia})
    finally:
        conn.close()

# ── ALERTS ────────────────────────────────────────────────────────────────────
@app.route('/<slug>/alertas')
def curso_alertas(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute(
            'SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
        act_ids = [r['id'] for r in conn.execute(
            'SELECT id, nombre, fecha_limite FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()]
        alertas = []
        # 1) Students with low promedio (< 2.8)
        for al in alumnos:
            vals = []
            for act_id in act_ids:
                r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                if r and r['val'] is not None: vals.append(float(r['val']))
            prom = round(sum(vals)/len(vals),2) if vals else None
            if prom is not None and prom < 2.8:
                ausencias = conn.execute('SELECT COUNT(*) as c FROM asistencia WHERE alumno_id=? AND estado=?', (al['id'], 'A')).fetchone()
                aus = ausencias['c'] if ausencias else 0
                sin_entregar = sum(1 for act_id in act_ids if not conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone())
                alertas.append({'tipo':'riesgo','alumno_id':al['id'],'nombre':al['nombre'],'promedio':prom,'ausencias':aus,'sin_entregar':sin_entregar})
            elif prom is None:
                alertas.append({'tipo':'sin_notas','alumno_id':al['id'],'nombre':al['nombre']})
        # 2) Activities approaching deadline
        for act_id in act_ids:
            a = conn.execute('SELECT nombre, fecha_limite FROM actividades WHERE id=?', (act_id,)).fetchone()
            if a and a['fecha_limite']:
                try:
                    from datetime import datetime, timedelta
                    fl = datetime.strptime(a['fecha_limite'][:10], '%Y-%m-%d') if isinstance(a['fecha_limite'], str) else a['fecha_limite']
                    if fl < datetime.now() + timedelta(days=3) and fl >= datetime.now():
                        alertas.append({'tipo':'proximo_vencer','actividad_id':act_id,'nombre':a['nombre'],'fecha':a['fecha_limite'][:10]})
                except: pass
        # 3) Students with many absences
        for al in alumnos:
            aus = conn.execute('SELECT COUNT(*) as c FROM asistencia WHERE alumno_id=? AND estado=?', (al['id'], 'A')).fetchone()
            if aus and aus['c'] >= 4:
                if not any(a.get('alumno_id')==al['id'] for a in alertas):
                    alertas.append({'tipo':'muchas_ausencias','alumno_id':al['id'],'nombre':al['nombre'],'ausencias':aus['c']})
        return jsonify({'alertas':alertas})
    finally:
        conn.close()

@app.route('/<slug>/institucional/dashboard')
def institucional_dashboard(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        # Institutional stats across all courses taught by this teacher
        jornada, materia = get_sesion_jornada_materia(slug)
        cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso', (materia, jornada, prof['id'])).fetchall()
        data = {'cursos':[],'promedio_institucional':0,'total_estudiantes':0,'activos':0}
        prom_sum=0; prom_count=0; total_al=0
        for c in cursos:
            curso_name = c['curso']
            alumnos = conn.execute('SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso_name, jornada)).fetchall()
            total_al += len(alumnos)
            act_ids = [r['id'] for r in conn.execute('SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?', (materia, jornada, curso_name, prof['id'])).fetchall()]
            promedios = []
            for al in alumnos:
                vals = []
                for act_id in act_ids:
                    r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                    if r and r['val'] is not None: vals.append(float(r['val']))
                if vals: promedios.append(round(sum(vals)/len(vals),2))
            prom_curso = round(sum(promedios)/len(promedios),2) if promedios else None
            if prom_curso is not None: prom_sum += prom_curso; prom_count += 1
            data['cursos'].append({'curso':curso_name,'promedio':prom_curso,'estudiantes':len(alumnos),'aprobados':sum(1 for p in promedios if p>=3) if promedios else 0,'perdiendo':sum(1 for p in promedios if p<3) if promedios else 0})
        data['promedio_institucional'] = round(prom_sum/prom_count,2) if prom_count else None
        data['total_estudiantes'] = total_al
        # Top students
        data['destacados'] = []
        # Activities taught by this prof
        return jsonify(data)
    finally:
        conn.close()

# ── FASE 7: Activities list with metadata ──────────────────────────────────────
@app.route('/<slug>/actividades/list')
def actividades_list(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        acts = conn.execute(
            'SELECT id, nombre, tipo, peso, fecha_limite, estado_act FROM actividades '
            'WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        result = []
        alunos_count = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso, jornada)).fetchone()
        total_al = alunos_count['c'] if alunos_count else 0
        for a in acts:
            notas_count = conn.execute('SELECT COUNT(*) as c FROM notas WHERE actividad_id=? AND val IS NOT NULL', (a['id'],)).fetchone()
            graded = notas_count['c'] if notas_count else 0
            pending = total_al - graded
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id=? AND val IS NOT NULL', (a['id'],)).fetchall()]
            prom = round(sum(vals)/len(vals),2) if vals else None
            result.append({
                'id':a['id'],'nombre':a['nombre'],'tipo':a['tipo'],'peso':a['peso'],
                'fecha_limite':a['fecha_limite'],'estado_act':a['estado_act'],
                'promedio':prom,'pendientes':pending,'total_estudiantes':total_al
            })
        return jsonify({'actividades':result})
    finally:
        conn.close()

# ── FASE 7: Mass activity management ──────────────────────────────────────────
@app.route('/<slug>/actividades/masiva', methods=['POST'])
def actividades_masiva(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    accion = data.get('accion')
    ids = data.get('ids', [])
    if not ids or not accion: return jsonify({'error':'Datos incompletos'}), 400
    conn = conectar(slug)
    try:
        placeholders = ','.join('?' for _ in ids)
        params = ids
        if accion == 'eliminar':
            for aid in ids:
                conn.execute('DELETE FROM notas WHERE actividad_id=?', (aid,))
            conn.execute(f'DELETE FROM actividades WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'publicar':
            conn.execute(f'UPDATE actividades SET estado_act=\'publicada\' WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'cerrar':
            conn.execute(f'UPDATE actividades SET estado_act=\'cerrada\' WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'archivar':
            conn.execute(f'UPDATE actividades SET estado_act=\'archivada\' WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'duplicar':
            for aid in ids:
                act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (aid, prof['id'])).fetchone()
                if act:
                    old_notas = conn.execute('SELECT aid, val FROM notas WHERE actividad_id=?', (aid,)).fetchall()
                    conn.execute(
                        'INSERT INTO actividades (nombre, tipo, peso, fecha_limite, estado_act, materia, jornada, curso, periodo, profesor_id, orden) '
                        'VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                        (act['nombre']+' (copia)', act['tipo'], act['peso'], act['fecha_limite'], 'borrador',
                         act['materia'], act['jornada'], act['curso'], act['periodo'], prof['id'], act.get('orden',0)))
                    new_id = conn.execute('SELECT last_insert_rowid() as lid').fetchone()['lid']
                    for n in old_notas:
                        conn.execute('INSERT INTO notas (actividad_id, aid, val) VALUES (?,?,?)', (new_id, n['aid'], n['val']))
        elif accion == 'cambiar_peso':
            peso = data.get('peso')
            if peso is not None:
                conn.execute(f'UPDATE actividades SET peso=? WHERE id IN ({placeholders}) AND profesor_id=?', [peso] + params + [prof['id']])
        elif accion == 'cambiar_tipo':
            tipo = data.get('tipo')
            if tipo:
                conn.execute(f'UPDATE actividades SET tipo=? WHERE id IN ({placeholders}) AND profesor_id=?', [tipo] + params + [prof['id']])
        elif accion == 'cambiar_fecha':
            fecha = data.get('fecha')
            if fecha:
                conn.execute(f'UPDATE actividades SET fecha_limite=? WHERE id IN ({placeholders}) AND profesor_id=?', [fecha] + params + [prof['id']])
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

# ── FASE 7: Pre-save validation ──────────────────────────────────────────────
@app.route('/<slug>/validar', methods=['POST'])
def curso_validar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = data.get('curso', '')
        periodo = data.get('periodo', 1)
        notas_data = data.get('notas', {})
        warnings = []
        # Check for empty notes
        empty_count = 0
        for aid, acts in notas_data.items():
            for act_id, val in acts.items():
                if val is None or val == '':
                    empty_count += 1
        if empty_count > 0:
            warnings.append({'tipo':'warning','mensaje':f'{empty_count} nota(s) vacía(s) que se ignorarán'})
        # Check for repeated identical values in same activity
        for aid, acts in notas_data.items():
            vals = [v for v in acts.values() if v is not None and v != '']
            if len(vals) > 1 and len(set(str(v) for v in vals)) == 1 and len(vals) > 2:
                warnings.append({'tipo':'info','mensaje':f'Estudiante #{aid}: todas las notas son iguales ({vals[0]})'})
                break
        # Check activities without weight
        acts = conn.execute(
            'SELECT id, nombre FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? AND (peso IS NULL OR peso=0)',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        for a in acts:
            warnings.append({'tipo':'warning','mensaje':f'Actividad "{a["nombre"]}" sin peso asignado'})
        # Check total weight > 100
        total_peso = conn.execute(
            'SELECT COALESCE(SUM(peso),0) as s FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? AND peso IS NOT NULL',
            (materia, jornada, curso, periodo, prof['id'])).fetchone()
        if total_peso and total_peso['s'] > 100:
            warnings.append({'tipo':'warning','mensaje':f'El peso total de actividades es {total_peso["s"]}% (máximo 100%)'})
        return jsonify({'warnings':warnings})
    finally:
        conn.close()

# ── FASE 7: AI Suggestions ──────────────────────────────────────────────────
@app.route('/<slug>/sugerencias')
def curso_sugerencias(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute('SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre', (curso, jornada)).fetchall()
        acts = conn.execute(
            'SELECT id, nombre, tipo FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        act_ids = [a['id'] for a in acts]
        sugerencias = []
        # Students needing recovery (prom < 3.0)
        for al in alumnos:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL', act_ids+[al['id']]).fetchall()]
            prom = round(sum(vals)/len(vals),2) if vals else None
            if prom is not None and prom < 3.0:
                sugerencias.append({'tipo':'recuperacion','alumno_id':al['id'],'nombre':al['nombre'],'promedio':prom,'accion':'Necesita recuperación o refuerzo'})
        # Who improved / worsened (compare first half vs second half)
        for al in alumnos:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL ORDER BY actividad_id', act_ids+[al['id']]).fetchall()]
            if len(vals) >= 4:
                half = len(vals)//2
                first_half = sum(vals[:half])/half
                second_half = sum(vals[half:])/half
                diff = second_half - first_half
                if diff > 0.5:
                    sugerencias.append({'tipo':'mejora','alumno_id':al['id'],'nombre':al['nombre'],'diferencia':round(diff,2),'accion':'Ha mejorado su rendimiento, reconocerlo'})
                elif diff < -0.5:
                    sugerencias.append({'tipo':'empeoro','alumno_id':al['id'],'nombre':al['nombre'],'diferencia':round(diff,2),'accion':'Ha disminuido su rendimiento, revisar causa'})
        # Activity difficulty analysis
        for a in acts:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id=? AND val IS NOT NULL', (a['id'],)).fetchall()]
            if vals:
                prom_act = sum(vals)/len(vals)
                if prom_act < 2.5:
                    sugerencias.append({'tipo':'dificil','actividad_id':a['id'],'nombre':a['nombre'],'promedio':round(prom_act,2),'accion':'Actividad demasiado difícil, considerar ajuste'})
                elif prom_act > 4.5:
                    sugerencias.append({'tipo':'facil','actividad_id':a['id'],'nombre':a['nombre'],'promedio':round(prom_act,2),'accion':'Actividad demasiado fácil, subir nivel'})
        # Students who should receive observation (sudden drop > 1.0 in any activity)
        for al in alumnos:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL ORDER BY actividad_id', act_ids+[al['id']]).fetchall()]
            for i in range(1, len(vals)):
                if vals[i-1] - vals[i] >= 1.0:
                    sugerencias.append({'tipo':'observacion','alumno_id':al['id'],'nombre':al['nombre'],'accion':f'Repentina caída de {vals[i-1]} a {vals[i]}, generar observación'})
                    break
        return jsonify({'sugerencias':sugerencias})
    finally:
        conn.close()

# ── FASE 7: Course comparison ────────────────────────────────────────────────
@app.route('/<slug>/comparar')
def curso_comparar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso', (materia, jornada, prof['id'])).fetchall()
        result = []
        for c in cursos:
            curso_name = c['curso']
            alumnos = conn.execute('SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre', (curso_name, jornada)).fetchall()
            acts = conn.execute('SELECT id, nombre FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=? ORDER BY orden', (materia, jornada, curso_name, prof['id'])).fetchall()
            act_ids = [a['id'] for a in acts]
            promedios = []
            total_vals = 0
            for al in alumnos:
                vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL', act_ids+[al['id']]).fetchall()]
                if vals:
                    promedios.append(sum(vals)/len(vals))
                    total_vals += len(vals)
            prom_curso = round(sum(promedios)/len(promedios),2) if promedios else 0
            approved = sum(1 for p in promedios if p >= 3)
            at_risk = sum(1 for p in promedios if p < 2.8)
            result.append({
                'curso':curso_name,
                'estudiantes':len(alumnos),
                'actividades':len(acts),
                'promedio':prom_curso,
                'aprobados':approved,
                'en_riesgo':at_risk,
                'total_notas':total_vals
            })
        return jsonify({'cursos':result})
    finally:
        conn.close()

# ── FASE 7: Timeline ─────────────────────────────────────────────────────────
@app.route('/<slug>/timeline')
def curso_timeline(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        events = []
        # Activity creation events (use audit_log if available)
        logs = conn.execute(
            "SELECT creado, accion, tabla, registro_id FROM audit_log WHERE tabla='actividades' AND "
            "registro_id IN (SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=?) "
            "ORDER BY creado DESC LIMIT 50",
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        for l in logs:
            act = conn.execute('SELECT nombre FROM actividades WHERE id=?', (l['registro_id'],)).fetchone()
            name = act['nombre'] if act else f'#{l["registro_id"]}'
            events.append({'tipo':'actividad_creada','titulo':f'Actividad "{name}" {l["accion"]}','fecha':l['creado'] or ''})
        # Recent note saves from auditoria
        notas_raw = conn.execute(
            'SELECT DISTINCT an.creado, an.actividad_id, ac.nombre as act_nombre FROM auditoria_notas an '
            'JOIN actividades ac ON ac.id=an.actividad_id '
            'WHERE ac.materia=? AND ac.jornada=? AND ac.curso=? AND ac.profesor_id=? AND an.creado IS NOT NULL '
            'ORDER BY an.creado DESC LIMIT 30',
            (materia, jornada, curso, prof['id'])).fetchall()
        for nr in notas_raw:
            events.append({'tipo':'notas_registradas','actividad_id':nr['actividad_id'],'titulo':f'Notas en "{nr["act_nombre"]}"','fecha':nr['creado'] or ''})
        # Sort by date descending, dates first
        events.sort(key=lambda e: e.get('fecha','') or '', reverse=True)
        return jsonify({'eventos':events[:30]})
    finally:
        conn.close()

# ── FASE 8/11: Institutional Control Center ────────────────────────────────
@app.route('/<slug>/institucional/centro-control')
def institucional_centro_control(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        # Gather all courses for this teacher
        cursos = conn.execute(
            'SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso',
            (materia, jornada, prof['id'])).fetchall()
        total_estudiantes = 0
        total_actividades = 0
        cursos_data = []
        prom_sum = 0.0
        prom_count = 0
        for c in cursos:
            curso_name = c['curso']
            alumnos = conn.execute('SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso_name, jornada)).fetchall()
            total_estudiantes += len(alumnos)
            acts = conn.execute('SELECT id, nombre, estado_act FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?', (materia, jornada, curso_name, prof['id'])).fetchall()
            total_actividades += len(acts)
            act_ids = [a['id'] for a in acts]
            promedios = []
            riesgo_count = 0
            sin_notas_count = 0
            for al in alumnos:
                vals = []
                for act_id in act_ids:
                    r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                    if r and r['val'] is not None: vals.append(float(r['val']))
                if vals:
                    p = sum(vals)/len(vals)
                    promedios.append(p)
                    if p < 2.8: riesgo_count += 1
                else:
                    sin_notas_count += 1
            prom_curso = round(sum(promedios)/len(promedios),2) if promedios else 0
            if promedios: prom_sum += prom_curso; prom_count += 1
            aprobados = sum(1 for p in promedios if p >= 3)
            inasistencia = conn.execute('SELECT COUNT(*) as c FROM asistencia a JOIN alumnos al ON al.id=a.aid WHERE al.curso=? AND al.jornada=? AND a.estado=?', (curso_name, jornada, 'A')).fetchone()
            cursos_data.append({
                'curso':curso_name,'estudiantes':len(alumnos),'actividades':len(acts),
                'promedio':prom_curso,'aprobados':aprobados,'riesgo':riesgo_count,
                'sin_notas':sin_notas_count,'inasistencia':inasistencia['c'] if inasistencia else 0
            })
        # Pending communications
        try:
            pend_comms = conn.execute("SELECT COUNT(*) as c FROM comunicaciones WHERE estado='pendiente'").fetchone()
        except:
            pend_comms = None
        # Today's events
        try:
            hoy_eventos = conn.execute("SELECT COUNT(*) as c FROM eventos_calendario WHERE fecha_inicio=date('now')").fetchone()
        except:
            hoy_eventos = None
        # Recient notifications (from audit)
        recientes = conn.execute("SELECT creado, accion FROM audit_log WHERE usuario_id=? ORDER BY creado DESC LIMIT 5", (prof['id'],)).fetchall()
        return jsonify({
            'promedio_institucional':round(prom_sum/prom_count,2) if prom_count else 0,
            'total_estudiantes':total_estudiantes,
            'total_actividades':total_actividades,
            'cursos':cursos_data,
            'comunicados_pendientes':pend_comms['c'] if pend_comms else 0,
            'eventos_hoy':hoy_eventos['c'] if hoy_eventos else 0,
            'actividades_sin_publicar':sum(1 for c in cursos_data for a in conn.execute('SELECT COUNT(*) as c FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=? AND estado_act=?', (materia, jornada, c['curso'], prof['id'], 'borrador')).fetchall() if a['c']>0),
            'alertas':{'total':sum(c['riesgo']+c['sin_notas'] for c in cursos_data),'riesgo':sum(c['riesgo'] for c in cursos_data),'sin_notas':sum(c['sin_notas'] for c in cursos_data)}
        })
    finally:
        conn.close()

# ── FASE 8: Communications ─────────────────────────────────────────────────
@app.route('/<slug>/comunicados')
def comunicados_list(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        comms = conn.execute(
            'SELECT id, titulo, contenido, destinatario_tipo, destinatario_valor, fecha_programada, fecha_creacion, estado FROM comunicaciones ORDER BY fecha_creacion DESC LIMIT 50',
            ()).fetchall()
        result = []
        for c in comms:
            result.append({
                'id':c['id'],'titulo':c['titulo'],'contenido':c['contenido'],
                'destinatario_tipo':c['destinatario_tipo'],'destinatario_valor':c['destinatario_valor'],
                'fecha_programada':c['fecha_programada'],'creado':c['fecha_creacion'],
                'estado':c['estado']
            })
        return jsonify({'comunicados':result})
    finally:
        conn.close()

@app.route('/<slug>/comunicados/crear', methods=['POST'])
def comunicados_crear(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    titulo = data.get('titulo','')
    contenido = data.get('contenido','')
    destinatario_tipo = data.get('destinatario_tipo','todos')
    destinatario_valor = data.get('destinatario_valor','')
    programada = data.get('programada')
    if not titulo: return jsonify({'error':'Titulo requerido'}), 400
    conn = conectar(slug)
    try:
        conn.execute(
            'INSERT INTO comunicaciones (rector_id, titulo, contenido, destinatario_tipo, destinatario_valor, fecha_programada, estado, fecha_creacion) VALUES (?,?,?,?,?,?,\'publicado\',datetime(\'now\'))',
            (prof['id'], titulo, contenido, destinatario_tipo, destinatario_valor, programada))
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

@app.route('/<slug>/comunicados/<int:cid>/leer', methods=['POST'])
def comunicados_leer(slug, cid):
    require_colegio(slug)
    conn = conectar(slug)
    try:
        conn.execute("UPDATE comunicaciones SET estado='leido' WHERE id=?", (cid,))
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

# ── FASE 10: Smart Calendar ────────────────────────────────────────────────
@app.route('/<slug>/calendario')
def calendario_list(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        events = []
        # Activities with due dates
        acts = conn.execute(
            'SELECT id, nombre, fecha_limite, tipo, curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? AND fecha_limite IS NOT NULL ORDER BY fecha_limite',
            (materia, jornada, prof['id'])).fetchall()
        for a in acts:
            events.append({'id':a['id'],'titulo':a['nombre'],'fecha':a['fecha_limite'][:10],'tipo':'evaluacion','curso':a['curso']})
        # Calendar events
        evs = conn.execute('SELECT id, titulo, fecha_inicio, tipo FROM eventos_calendario ORDER BY fecha_inicio', ()).fetchall()
        for e in evs:
            events.append({'id':e['id'],'titulo':e['titulo'],'fecha':e['fecha_inicio'],'tipo':e['tipo'] or 'evento','curso':''})
        events.sort(key=lambda e: e.get('fecha',''))
        return jsonify({'eventos':events})
    finally:
        conn.close()

# ── FASE 12: Extended student record ──────────────────────────────────────
@app.route('/<slug>/estudiante/<int:aid>/expediente')
def estudiante_expediente(slug, aid):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        al = conn.execute('SELECT * FROM alumnos WHERE id=?', (aid,)).fetchone()
        if not al: return jsonify({'error':'No encontrado'}), 404
        # Full academic history
        notas = conn.execute(
            'SELECT ac.nombre as actividad, ac.tipo, n.val, ac.fecha_limite FROM notas n JOIN actividades ac ON ac.id=n.actividad_id WHERE n.aid=? ORDER BY ac.fecha_limite',
            (aid,)).fetchall()
        asistencias = conn.execute('SELECT fecha, estado FROM asistencia WHERE aid=? ORDER BY fecha DESC LIMIT 30', (aid,)).fetchall()
        observaciones = conn.execute('SELECT texto, fecha FROM observaciones WHERE aid=? ORDER BY fecha DESC LIMIT 20', (aid,)).fetchall()
        # Sanctions and recognitions (from observador if available)
        sanciones = conn.execute('SELECT texto, fecha, tipo FROM observador_registros WHERE aid=? AND tipo IN (\'llamado_atencion\',\'sancion\') ORDER BY fecha DESC LIMIT 10', (aid,)).fetchall()
        reconocimientos = conn.execute('SELECT texto, fecha, tipo FROM observador_registros WHERE aid=? AND tipo=\'reconocimiento\' ORDER BY fecha DESC LIMIT 10', (aid,)).fetchall()
        return jsonify({
            'alumno':{'id':al['id'],'nombre':al['nombre'],'curso':al['curso'],'email_acudiente':al['email_acudiente']},
            'notas':[{'actividad':n['actividad'],'tipo':n['tipo'],'val':n['val'],'fecha':n['fecha_limite']} for n in notas],
            'asistencias':[{'fecha':a['fecha'],'presente':a['estado']=='P'} for a in asistencias],
            'observaciones':[{'texto':o['texto'],'fecha':o['fecha']} for o in observaciones],
            'sanciones':[{'texto':s['texto'],'fecha':s['fecha'],'tipo':s['tipo']} for s in sanciones] if sanciones else [],
            'reconocimientos':[{'texto':r['texto'],'fecha':r['fecha'],'tipo':r['tipo']} for r in reconocimientos] if reconocimientos else []
        })
    finally:
        conn.close()

# ── FASE 13: Smart Hub (role-aware dashboard) ───────────────────────────────
@app.route('/<slug>/smart-hub')
def smart_hub(slug):
    require_colegio(slug)
    user_type = session.get('user_type', 'profesor')
    user_id = session.get('user_id') or session.get('profesor_id')
    conn = conectar(slug)
    try:
        result = {'rol':user_type}
        if user_type == 'profesor':
            prof = get_profesor(slug)
            if not prof: return jsonify({'error':'No autorizado'}), 403
            jornada, materia = get_sesion_jornada_materia(slug)
            cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso', (materia, jornada, prof['id'])).fetchall()
            cursos_data = []
            total_riesgo = 0
            total_sin_notas = 0
            act_sin_publicar = 0
            for c in cursos:
                cn = c['curso']
                acts = conn.execute('SELECT id, estado_act FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?', (materia, jornada, cn, prof['id'])).fetchall()
                act_ids = [a['id'] for a in acts]
                for a in acts:
                    if a['estado_act'] == 'borrador': act_sin_publicar += 1
                alumnos = conn.execute('SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (cn, jornada)).fetchall()
                riesgo = 0
                sin_notas = 0
                for al in alumnos:
                    vals = [r['val'] for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL', act_ids+[al['id']]).fetchall()]
                    if vals:
                        p = sum(vals)/len(vals)
                        if p < 2.8: riesgo += 1
                    else:
                        sin_notas += 1
                total_riesgo += riesgo
                total_sin_notas += sin_notas
                cursos_data.append({'curso':cn,'riesgo':riesgo,'sin_notas':sin_notas})
            result['cursos'] = cursos_data
            result['total_riesgo'] = total_riesgo
            result['total_sin_notas'] = total_sin_notas
            result['actividades_sin_publicar'] = act_sin_publicar
        elif user_type == 'rector' or user_type == 'directora':
            # Rector view - institutional
            total_al = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE activo=1').fetchone()
            result['total_estudiantes'] = total_al['c'] if total_al else 0
            # Critical courses
            criticos = conn.execute("""
                SELECT a.curso, ROUND(AVG(CASE WHEN n.val IS NOT NULL THEN n.val ELSE 0 END),2) as prom,
                       COUNT(DISTINCT a.id) as estudiantes
                FROM alumnos a LEFT JOIN notas n ON n.aid=a.id
                WHERE a.activo=1
                GROUP BY a.curso ORDER BY prom ASC LIMIT 5
            """).fetchall()
            result['cursos_criticos'] = [{'curso':r['curso'],'promedio':r['prom'],'estudiantes':r['estudiantes']} for r in criticos]
            result['total_alertas'] = 0
        elif user_type == 'padre' or user_type == 'estudiante':
            # Parent/student view
            aid = request.args.get('aid', type=int) or user_id
            al = conn.execute('SELECT id, nombre, curso FROM alumnos WHERE id=?', (aid,)).fetchone()
            if al:
                result['alumno'] = {'id':al['id'],'nombre':al['nombre'],'curso':al['curso']}
        return jsonify(result)
    finally:
        conn.close()

# ── FASE 13: Paginated notes endpoint (optimized) ─────────────────────────
@app.route('/<slug>/notas/pagina')
def notas_pagina(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    per_page = min(per_page, 100)
    offset = (page - 1) * per_page
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        count = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso, jornada)).fetchone()
        total = count['c'] if count else 0
        alumnos = conn.execute(
            'SELECT id, nombre, num_curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre LIMIT ? OFFSET ?',
            (curso, jornada, per_page, offset)).fetchall()
        acts = conn.execute(
            'SELECT id, nombre, tipo, peso, estado_act, fecha_limite FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, request.args.get('periodo',1,type=int), prof['id'])).fetchall()
        act_ids = [a['id'] for a in acts]
        # Batch fetch all notes for these students and activities
        notas_map = {}
        if alumnos and act_ids:
            placeholders = ','.join('?'*len(act_ids))
            aids = [al['id'] for al in alumnos]
            rows = conn.execute(
                f'SELECT aid, actividad_id, val FROM notas WHERE aid IN ({",".join("?"*len(aids))}) AND actividad_id IN ({placeholders})',
                aids + act_ids).fetchall()
            for r in rows:
                key = (r['aid'], r['actividad_id'])
                notas_map[key] = r['val']
        return jsonify({
            'page':page,'per_page':per_page,'total':total,'total_pages':max(1,-(-total//per_page)),
            'actividades':[{'id':a['id'],'nombre':a['nombre'],'tipo':a['tipo'],'peso':a['peso'],'estado_act':a['estado_act'],'fecha_limite':a['fecha_limite']} for a in acts],
            'alumnos':[{'id':a['id'],'nombre':a['nombre'],'num_curso':a['num_curso']} for a in alumnos],
            'notas':{f"{r['aid']}_{r['actividad_id']}":{'aid':r['aid'],'actividad_id':r['actividad_id'],'val':r['val']} for r in rows} if 'rows' in dir() else {}
        })
    finally:
        conn.close()

# ── FASE 13: Enhanced audit log ────────────────────────────────────────────
@app.route('/<slug>/auditoria')
def auditoria_list(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 30, type=int)
    offset = (page - 1) * per_page
    conn = conectar(slug)
    try:
        count = conn.execute('SELECT COUNT(*) as c FROM audit_log WHERE usuario_id=?', (prof['id'],)).fetchone()
        total = count['c'] if count else 0
        logs = conn.execute(
            'SELECT id, accion, tabla, registro_id, valor_anterior, valor_nuevo, ip, creado FROM audit_log WHERE usuario_id=? ORDER BY creado DESC LIMIT ? OFFSET ?',
            (prof['id'], per_page, offset)).fetchall()
        return jsonify({
            'total':total,'page':page,'per_page':per_page,
            'items':[{'id':l['id'],'accion':l['accion'],'tabla':l['tabla'],'registro_id':l['registro_id'],
                      'valor_anterior':l['valor_anterior'],'valor_nuevo':l['valor_nuevo'],'ip':l['ip'],'creado':l['creado']} for l in logs]
        })
    finally:
        conn.close()

# ── FASE 14: School configuration ──────────────────────────────────────────
@app.route('/<slug>/config', methods=['GET','POST'])
def school_config(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        if request.method == 'GET':
            conf = conn.execute('SELECT * FROM config_institucion WHERE slug=?', (slug,)).fetchone()
            return jsonify({
                'nombre':conf['nombre_institucion'] if conf else '',
                'lema':conf['lema'] if conf else '',
                'logo':conf['logo'] if conf else '',
                'primary_color':conf['primary_color'] if conf else '#3b82f6',
                'secondary_color':conf['secondary_color'] if conf else '#6366f1'
            })
        else:
            data = request.get_json(silent=True) or {}
            if conf := conn.execute('SELECT id FROM config_institucion WHERE slug=?', (slug,)).fetchone():
                conn.execute('UPDATE config_institucion SET nombre_institucion=?, lema=?, logo=?, primary_color=?, secondary_color=? WHERE id=?',
                    (data.get('nombre',''), data.get('lema',''), data.get('logo',''), data.get('primary_color','#3b82f6'), data.get('secondary_color','#6366f1'), conf['id']))
            else:
                conn.execute('INSERT INTO config_institucion (slug, nombre_institucion, lema, logo, primary_color, secondary_color) VALUES (?,?,?,?,?,?)',
                    (slug, data.get('nombre',''), data.get('lema',''), data.get('logo',''), data.get('primary_color','#3b82f6'), data.get('secondary_color','#6366f1')))
            conn.commit()
            return jsonify({'status':'ok'})
    finally:
        conn.close()

# ── FASE 15: Activity Templates ────────────────────────────────────────────
@app.route('/<slug>/plantillas')
def plantillas_list(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        templates = conn.execute(
            'SELECT id, nombre, tipo, peso, descripcion, created_at FROM plantillas WHERE profesor_id=? ORDER BY nombre',
            (prof['id'],)).fetchall()
        return jsonify({'plantillas':[{'id':t['id'],'nombre':t['nombre'],'tipo':t['tipo'],'peso':t['peso'],'descripcion':t['descripcion'],'creado':t['created_at']} for t in templates]})
    finally:
        conn.close()

@app.route('/<slug>/plantillas/crear', methods=['POST'])
def plantillas_crear(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    nombre = data.get('nombre','')
    tipo = data.get('tipo','tarea')
    peso = data.get('peso',10)
    descripcion = data.get('descripcion','')
    if not nombre: return jsonify({'error':'Nombre requerido'}), 400
    conn = conectar(slug)
    try:
        conn.execute('INSERT INTO plantillas (profesor_id, nombre, tipo, peso, descripcion) VALUES (?,?,?,?,?)',
            (prof['id'], nombre, tipo, peso, descripcion))
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

@app.route('/<slug>/plantillas/aplicar', methods=['POST'])
def plantillas_aplicar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    tmpl_id = data.get('plantilla_id')
    curso = data.get('curso','')
    materia = data.get('materia','')
    if not tmpl_id or not curso: return jsonify({'error':'Datos incompletos'}), 400
    conn = conectar(slug)
    try:
        tmpl = conn.execute('SELECT * FROM plantillas WHERE id=? AND profesor_id=?', (tmpl_id, prof['id'])).fetchone()
        if not tmpl: return jsonify({'error':'Plantilla no encontrada'}), 404
        jornada = data.get('jornada') or 'mañana'
        periodo = data.get('periodo', 1)
        max_ord = conn.execute('SELECT COALESCE(MAX(orden),0) as mx FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?',
            (materia, jornada, curso, prof['id'])).fetchone()
        conn.execute(
            'INSERT INTO actividades (profesor_id, materia, jornada, curso, periodo, nombre, tipo, peso, estado_act, orden) VALUES (?,?,?,?,?,?,?,?,?,?)',
            (prof['id'], materia, jornada, curso, periodo, tmpl['nombre'], tmpl['tipo'], tmpl['peso'], 'borrador', (max_ord['mx']+1) if max_ord else 1))
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

@app.route('/<slug>/plantillas/eliminar/<int:tid>', methods=['POST'])
def plantillas_eliminar(slug, tid):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        conn.execute('DELETE FROM plantillas WHERE id=? AND profesor_id=?', (tid, prof['id']))
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

# ── FASE 15: Copy planning across courses ──────────────────────────────────
@app.route('/<slug>/planificacion/copiar', methods=['POST'])
def planificacion_copiar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    origen_curso = data.get('origen_curso')
    destino_cursos = data.get('destino_cursos', [])
    if not origen_curso or not destino_cursos: return jsonify({'error':'Datos incompletos'}), 400
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        acts = conn.execute(
            'SELECT nombre, tipo, peso, fecha_limite, estado_act FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, origen_curso, prof['id'])).fetchall()
        for dest in destino_cursos:
            for a in acts:
                conn.execute(
                    'INSERT INTO actividades (profesor_id, materia, jornada, curso, nombre, tipo, peso, fecha_limite, estado_act, orden) VALUES (?,?,?,?,?,?,?,?,?,'
                    '(SELECT COALESCE(MAX(orden),0)+1 FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?))',
                    (prof['id'], materia, jornada, dest, a['nombre'], a['tipo'], a['peso'], a['fecha_limite'], 'borrador', materia, jornada, dest, prof['id']))
        conn.commit()
        return jsonify({'status':'ok','copiadas':len(acts)*len(destino_cursos)})
    finally:
        conn.close()

# ── FASE 16: Analytics Comparison ──────────────────────────────────────────
@app.route('/<slug>/analitica/comparar')
def analitica_comparar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    tipo = request.args.get('tipo', 'periodos')  # periodos, docentes, sedes
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        result = {'tipo':tipo,'datos':[]}
        if tipo == 'periodos':
            for p in [1,2,3,4]:
                acts = conn.execute(
                    'SELECT id FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? AND COALESCE(periodo,1)=?',
                    (materia, jornada, prof['id'], p)).fetchall()
                act_ids = [a['id'] for a in acts]
                promedios = []
                if act_ids:
                    rows = conn.execute(
                        'SELECT n.val FROM notas n WHERE n.actividad_id IN ('+','.join('?'*len(act_ids))+') AND n.val IS NOT NULL',
                        act_ids).fetchall()
                    vals = [float(r['val']) for r in rows]
                    if vals: promedios.append(sum(vals)/len(vals))
                result['datos'].append({'periodo':p,'promedio':round(sum(promedios)/len(promedios),2) if promedios else 0,'actividades':len(acts)})
        elif tipo == 'docentes':
            profs = conn.execute('SELECT DISTINCT p.id, p.nombre FROM profesores p JOIN actividades a ON a.profesor_id=p.id WHERE a.materia=? AND a.jornada=?', (materia, jornada)).fetchall()
            for p in profs:
                vals = [float(r['val']) for r in conn.execute(
                    'SELECT n.val FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE a.profesor_id=? AND a.materia=? AND a.jornada=? AND n.val IS NOT NULL',
                    (p['id'], materia, jornada)).fetchall()]
                prom = round(sum(vals)/len(vals),2) if vals else 0
                result['datos'].append({'nombre':p['nombre'],'promedio':prom,'notas':len(vals)})
        return jsonify(result)
    finally:
        conn.close()

# ── FASE 15: Smart Migration (onboarding import) ──────────────────────────
@app.route('/<slug>/migrar/previsualizar', methods=['POST'])
def migrar_previsualizar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    import csv, io, json
    data = request.get_json(silent=True) or {}
    contenido = data.get('contenido','')
    tipo = data.get('tipo','estudiantes')  # estudiantes, notas, actividades
    if not contenido: return jsonify({'error':'Contenido requerido'}), 400
    try:
        lines = contenido.strip().split('\n')
        reader = csv.DictReader(io.StringIO(contenido))
        headers = reader.fieldnames
        rows = []
        for i, row in enumerate(reader):
            if i >= 10: break  # Preview first 10
            rows.append(dict(row))
        # Auto-detect column mapping
        sugg = {}
        if tipo == 'estudiantes':
            name_cols = [h for h in headers if any(x in h.lower() for x in ['nombre','name','alumno','estudiante'])]
            if name_cols: sugg['nombre'] = name_cols[0]
            curso_cols = [h for h in headers if any(x in h.lower() for x in ['curso','grado','grade','salon'])]
            if curso_cols: sugg['curso'] = curso_cols[0]
        elif tipo == 'notas':
            name_cols = [h for h in headers if any(x in h.lower() for x in ['nombre','name','alumno'])]
            if name_cols: sugg['nombre'] = name_cols[0]
            note_cols = [h for h in headers if any(x in h.lower() for x in ['nota','calif','grade','puntaje','val'])]
            if note_cols: sugg['nota'] = note_cols[0]
        return jsonify({'headers':headers,'filas':rows,'sugerencias':sugg,'total':len(lines)-1})
    except Exception as e:
        return jsonify({'error':str(e)}), 400

@app.route('/<slug>/migrar/ejecutar', methods=['POST'])
def migrar_ejecutar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    import csv, io
    data = request.get_json(silent=True) or {}
    contenido = data.get('contenido','')
    tipo = data.get('tipo','estudiantes')
    mapeo = data.get('mapeo', {})
    if not contenido: return jsonify({'error':'Contenido requerido'}), 400
    conn = conectar(slug)
    try:
        jornada, materia = get_sesion_jornada_materia(slug)
        reader = csv.DictReader(io.StringIO(contenido))
        count = 0
        if tipo == 'estudiantes':
            nombre_col = mapeo.get('nombre','nombre')
            curso_col = mapeo.get('curso','curso')
            for row in reader:
                nombre = row.get(nombre_col,'').strip()
                curso = row.get(curso_col,'')
                if nombre:
                    conn.execute('INSERT INTO alumnos (nombre, curso, jornada, activo) VALUES (?,?,?,1)', (nombre, curso, jornada))
                    count += 1
        elif tipo == 'actividades':
            nombre_col = mapeo.get('nombre','nombre')
            tipo_col = mapeo.get('tipo','tipo')
            peso_col = mapeo.get('peso','peso')
            curso_col = mapeo.get('curso','curso')
            for row in reader:
                nombre = row.get(nombre_col,'').strip()
                curso = row.get(curso_col,'')
                tipo_act = row.get(tipo_col,'tarea')
                peso = float(row.get(peso_col,10)) if row.get(peso_col) else 10
                if nombre:
                    conn.execute(
                        'INSERT INTO actividades (profesor_id, materia, jornada, curso, nombre, tipo, peso, estado_act, orden) VALUES (?,?,?,?,?,?,?,?,'
                        '(SELECT COALESCE(MAX(orden),0)+1 FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?))',
                        (prof['id'], materia, jornada, curso, nombre, tipo_act, peso, 'borrador', materia, jornada, curso, prof['id']))
                    count += 1
        conn.commit()
        return jsonify({'status':'ok','importados':count})
    except Exception as e:
        return jsonify({'error':str(e)}), 500
    finally:
        conn.close()

# ── FASE 17: Academic Management (Rector) ────────────────────────────────
@app.route('/<slug>/gestion-academica/promover', methods=['POST'])
def gestion_promover(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    curso_origen = data.get('curso_origen','')
    jornada = data.get('jornada','')
    curso_destino = data.get('curso_destino','')
    if not curso_origen or not curso_destino: return jsonify({'error':'Curso origen y destino requeridos'}), 400
    conn = conectar(slug)
    try:
        alumnos = conn.execute('SELECT id, nombre, curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso_origen, jornada)).fetchall()
        promovidos = 0
        for a in alumnos:
            conn.execute('INSERT INTO historial_academico (alumno_id, curso, jornada, promedio_final, estado) VALUES (?,?,?,?,?)',
                (a['id'], a['curso'], jornada, 0, 'promovido'))
            conn.execute('UPDATE alumnos SET curso=? WHERE id=?', (curso_destino, a['id']))
            promovidos += 1
        conn.commit()
        return jsonify({'status':'ok','promovidos':promovidos})
    finally:
        conn.close()

@app.route('/<slug>/gestion-academica/trasladar', methods=['POST'])
def gestion_trasladar(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    alumno_id = data.get('alumno_id')
    curso_nuevo = data.get('curso_nuevo','')
    tipo = data.get('tipo','curso')  # curso, jornada, sede
    if not alumno_id or not curso_nuevo: return jsonify({'error':'Datos incompletos'}), 400
    conn = conectar(slug)
    try:
        alumno = conn.execute('SELECT id, nombre, curso, jornada FROM alumnos WHERE id=?', (alumno_id,)).fetchone()
        if not alumno: return jsonify({'error':'Estudiante no encontrado'}), 404
        campo = 'curso' if tipo == 'curso' else 'jornada' if tipo == 'jornada' else 'sede'
        conn.execute('INSERT INTO historial_academico (alumno_id, curso, jornada, promedio_final, estado, observaciones) VALUES (?,?,?,?,?,?)',
            (alumno_id, alumno['curso'], alumno['jornada'], 0, 'trasladado', 'Trasladado a '+curso_nuevo))
        conn.execute(f'UPDATE alumnos SET {campo}=? WHERE id=?', (curso_nuevo, alumno_id))
        conn.commit()
        return jsonify({'status':'ok','nombre':alumno['nombre'],'tipo':tipo})
    finally:
        conn.close()

@app.route('/<slug>/gestion-academica/historial/<int:alumno_id>')
def gestion_historial(slug, alumno_id):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        rows = conn.execute('SELECT * FROM historial_academico WHERE alumno_id=? ORDER BY id DESC', (alumno_id,)).fetchall()
        alumno = conn.execute('SELECT id, nombre, curso FROM alumnos WHERE id=?', (alumno_id,)).fetchone()
        return jsonify({'alumno':{'id':alumno['id'],'nombre':alumno['nombre'],'curso':alumno['curso']} if alumno else None,
            'historial':[{'id':r['id'],'curso':r['curso'],'jornada':r['jornada'],'periodo':r['periodo'],'promedio':r['promedio_final'],'estado':r['estado'],'observaciones':r['observaciones'],'fecha':r['created_at']} for r in rows]})
    finally:
        conn.close()

@app.route('/<slug>/gestion-academica/alumnos')
def gestion_alumnos_list(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        rows = conn.execute('SELECT id, nombre, curso, jornada, activo FROM alumnos ORDER BY curso, nombre').fetchall()
        return jsonify({'alumnos':[dict(r) for r in rows]})
    finally:
        conn.close()

# ── FASE 21: Parent Portal ───────────────────────────────────────────────
# ── portal/parent routes migrated to app/routes/auth.py ──────────────────

@app.route('/<slug>/portal/dashboard')
def portal_padre_dashboard(slug):
    require_colegio(slug)
    pid = session.get(f'padre_id_{slug}')
    if not pid: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        hijos = conn.execute('SELECT a.id, a.nombre, a.curso, a.jornada FROM alumno_padre ap JOIN alumnos a ON a.id=ap.alumno_id WHERE ap.padre_id=?', (pid,)).fetchall()
        resultado = []
        for h in hijos:
            notas = conn.execute('SELECT COALESCE(AVG(n.val),0) as prom FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.alumno_id=? AND a.curso=?', (h['id'], h['curso'])).fetchone()
            asistencias = conn.execute('SELECT estado, COUNT(*) as cnt FROM asistencia_v2 WHERE alumno_id=? AND DATE(fecha)>=DATE("now","-30 days") GROUP BY estado', (h['id'],)).fetchall()
            act_count = conn.execute('SELECT COUNT(*) as cnt FROM actividades WHERE curso=? AND jornada=? AND estado_act="publicado"', (h['curso'], h['jornada'])).fetchone()
            resultado.append({'id':h['id'],'nombre':h['nombre'],'curso':h['curso'],'jornada':h['jornada'],
                'promedio':round(notas['prom'],2),'asistencia':[dict(a) for a in asistencias],'actividades':act_count['cnt']})
        return jsonify({'hijos':resultado})
    finally:
        conn.close()

@app.route('/<slug>/portal/notas/<int:alumno_id>')
def portal_padre_notas(slug, alumno_id):
    require_colegio(slug)
    pid = session.get(f'padre_id_{slug}')
    if not pid: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        rel = conn.execute('SELECT id FROM alumno_padre WHERE padre_id=? AND alumno_id=?', (pid, alumno_id)).fetchone()
        if not rel: return jsonify({'error':'No autorizado'}), 403
        acts = conn.execute('SELECT a.id, a.nombre, a.tipo, a.peso, COALESCE(ROUND(AVG(n.val),2),0) as prom FROM actividades a LEFT JOIN notas n ON n.actividad_id=a.id AND n.alumno_id=? WHERE a.curso=(SELECT curso FROM alumnos WHERE id=?) AND a.jornada=(SELECT jornada FROM alumnos WHERE id=?) GROUP BY a.id ORDER BY a.orden', (alumno_id, alumno_id, alumno_id)).fetchall()
        return jsonify({'actividades':[dict(a) for a in acts]})
    finally:
        conn.close()

@app.route('/<slug>/portal/asistencia/<int:alumno_id>')
def portal_padre_asistencia(slug, alumno_id):
    require_colegio(slug)
    pid = session.get(f'padre_id_{slug}')
    if not pid: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        rel = conn.execute('SELECT id FROM alumno_padre WHERE padre_id=? AND alumno_id=?', (pid, alumno_id)).fetchone()
        if not rel: return jsonify({'error':'No autorizado'}), 403
        rows = conn.execute('SELECT fecha, estado FROM asistencia_v2 WHERE alumno_id=? ORDER BY fecha DESC LIMIT 60', (alumno_id,)).fetchall()
        return jsonify({'asistencia':[dict(r) for r in rows]})
    finally:
        conn.close()

@app.route('/<slug>/portal/comunicados')
def portal_padre_comms(slug):
    require_colegio(slug)
    pid = session.get(f'padre_id_{slug}')
    if not pid: return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        hijo_ids = [r['alumno_id'] for r in conn.execute('SELECT alumno_id FROM alumno_padre WHERE padre_id=?', (pid,)).fetchall()]
        rows = conn.execute('SELECT id, titulo, contenido, created_at FROM comunicaciones WHERE destinatario_tipo IN ("todos","padres") AND estado="publicado" ORDER BY created_at DESC LIMIT 20').fetchall()
        return jsonify({'comunicados':[dict(r) for r in rows]})
    finally:
        conn.close()

# ── FASE 22: Institutional AI Assistant ──────────────────────────────────
@app.route('/<slug>/ai/ask', methods=['POST'])
def ai_ask(slug):
    require_colegio(slug)
    data = request.get_json(silent=True) or {}
    pregunta = data.get('pregunta','').lower().strip()
    if not pregunta: return jsonify({'error':'Pregunta requerida'}), 400
    conn = conectar(slug)
    try:
        # Pattern matching for common questions
        respuesta = {'pregunta':pregunta,'respuesta':'','datos':[]}
        # Risk students
        if any(p in pregunta for p in ['riesgo','perder','reprobar','bajo rendimiento','recuperacion']):
            curso = None
            for c in ['grado ','curso ']:
                idx = pregunta.find(c)
                if idx >= 0:
                    curso = pregunta[idx+len(c):].split()[0] if pregunta[idx+len(c):].split() else None
                    break
            q = 'SELECT a.id, a.nombre, a.curso, ROUND(AVG(n.val),2) as prom FROM alumnos a JOIN notas n ON n.alumno_id=a.id WHERE a.activo=1'
            params = []
            if curso:
                q += ' AND a.curso=?'
                params.append(curso)
            q += ' GROUP BY a.id HAVING prom < 3.0 ORDER BY prom ASC LIMIT 20'
            rows = conn.execute(q, params).fetchall()
            respuesta['respuesta'] = f'Se encontraron {len(rows)} estudiantes con riesgo académico (prom<3.0){" en "+curso if curso else ""}.'
            respuesta['datos'] = [{'nombre':r['nombre'],'curso':r['curso'],'promedio':r['prom']} for r in rows]
        # Lowest subjects
        elif any(p in pregunta for p in ['materia baja','promedio bajo','peor materia','materia más baja','nota baja']):
            rows = conn.execute('SELECT a.materia, ROUND(AVG(n.val),2) as prom FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY a.materia ORDER BY prom ASC LIMIT 10').fetchall()
            respuesta['respuesta'] = f"Materias con menor promedio: {rows[0]['materia']} ({rows[0]['prom']})" if rows else 'No hay datos suficientes.'
            respuesta['datos'] = [{'materia':r['materia'],'promedio':r['prom']} for r in rows]
        # Teacher without notes
        elif any(p in pregunta for p in ['docente sin','profesor sin','sin registrar','días sin','no ha registrado']):
            from datetime import datetime, timedelta
            days = 10
            for w in pregunta.split():
                if w.isdigit(): days = int(w); break
            cutoff = (datetime.now() - timedelta(days=days)).isoformat()
            rows = conn.execute('''SELECT p.id, p.nombre, MAX(n.created_at) as ultima_nota
                FROM profesores p JOIN actividades a ON a.profesor_id=p.id
                LEFT JOIN notas n ON n.actividad_id=a.id
                GROUP BY p.id HAVING ultima_nota IS NULL OR ultima_nota < ?''', (cutoff,)).fetchall()
            respuesta['respuesta'] = f'{len(rows)} docentes no han registrado notas en los últimos {days} días.'
            respuesta['datos'] = [{'nombre':r['nombre'],'ultima_nota':r['ultima_nota'] or 'Nunca'} for r in rows]
        # Attendance issues
        elif any(p in pregunta for p in ['inasistencia','ausencia','falta','asistencia baja','no asiste']):
            rows = conn.execute('SELECT a.nombre, a.curso, COUNT(*) as faltas FROM asistencia_v2 av JOIN alumnos a ON a.id=av.alumno_id WHERE av.estado IN ("X","E") AND a.activo=1 GROUP BY av.alumno_id HAVING faltas > 3 ORDER BY faltas DESC LIMIT 20').fetchall()
            respuesta['respuesta'] = f'{len(rows)} estudiantes tienen más de 3 inasistencias.'
            respuesta['datos'] = [{'nombre':r['nombre'],'curso':r['curso'],'faltas':r['faltas']} for r in rows]
        # Improvement
        elif any(p in pregunta for p in ['mejorado','mejora','progreso','subió','aumentó']):
            rows = conn.execute('''SELECT a.nombre, a.curso, ROUND(AVG(CASE WHEN n.created_at < date("now","-15 days") THEN n.val END),2) as antes,
                ROUND(AVG(CASE WHEN n.created_at >= date("now","-15 days") THEN n.val END),2) as despues
                FROM notas n JOIN actividades ac ON ac.id=n.actividad_id JOIN alumnos a ON a.id=n.alumno_id
                WHERE n.val IS NOT NULL GROUP BY n.alumno_id HAVING despues > antes AND antes > 0
                ORDER BY (despues-antes) DESC LIMIT 15''').fetchall()
            respuesta['respuesta'] = f'{len(rows)} estudiantes han mejorado su rendimiento recientemente.'
            respuesta['datos'] = [{'nombre':r['nombre'],'curso':r['curso'],'antes':r['antes'],'despues':r['despues']} for r in rows]
        # Critical courses
        elif any(p in pregunta for p in ['curso crítico','curso necesita','intervención','peor curso','curso bajo']):
            rows = conn.execute('SELECT a.curso, ROUND(AVG(n.val),2) as prom, COUNT(DISTINCT n.alumno_id) as estudiantes FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY a.curso ORDER BY prom ASC LIMIT 10').fetchall()
            respuesta['respuesta'] = f"Curso con menor rendimiento: {rows[0]['curso']} ({rows[0]['prom']})" if rows else 'No hay datos.'
            respuesta['datos'] = [{'curso':r['curso'],'promedio':r['prom'],'estudiantes':r['estudiantes']} for r in rows]
        # Summary by grade
        elif any(p in pregunta for p in ['resume','resumen','rendimiento de','grado ']):
            rows = conn.execute('SELECT a.curso, ROUND(AVG(n.val),2) as prom, COUNT(DISTINCT a.id) as acts FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY a.curso ORDER BY a.curso').fetchall()
            resp = '<div style="display:flex;flex-direction:column;gap:4px;">'
            for r in rows:
                color = 'var(--success)' if r['prom'] >= 3.5 else 'var(--warning)' if r['prom'] >= 3.0 else 'var(--danger)'
                resp += f'<div style="display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid var(--border);"><span>{r["curso"]}</span><span style="font-weight:700;color:{color};">{r["prom"]}</span></div>'
            resp += '</div>'
            respuesta['respuesta'] = f'Rendimiento por curso ({len(rows)} cursos):'
            respuesta['html'] = resp
            respuesta['datos'] = [dict(r) for r in rows]
        # Activities with most failing
        elif any(p in pregunta for p in ['actividad perdiendo','actividad más','actividad difícil','tarea difícil','examen difícil']):
            rows = conn.execute('SELECT ac.nombre, ac.tipo, ac.curso, ROUND(AVG(n.val),2) as prom, COUNT(*) as notas FROM notas n JOIN actividades ac ON ac.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY ac.id HAVING prom < 3.0 ORDER BY prom ASC LIMIT 15').fetchall()
            respuesta['respuesta'] = f'{len(rows)} actividades tienen promedio menor a 3.0.'
            respuesta['datos'] = [{'nombre':r['nombre'],'tipo':r['tipo'],'curso':r['curso'],'promedio':r['prom']} for r in rows]
        # Comparison with previous period
        elif any(p in pregunta for p in ['bajaron','disminuyó','respecto','anterior','comparado']):
            rows = conn.execute('''
                SELECT a.materia,
                    ROUND(AVG(CASE WHEN COALESCE(a.periodo,1) <= 2 THEN n.val END),2) as p1,
                    ROUND(AVG(CASE WHEN COALESCE(a.periodo,1) > 2 THEN n.val END),2) as p2
                FROM notas n JOIN actividades a ON a.id=n.actividad_id
                WHERE n.val IS NOT NULL GROUP BY a.materia HAVING p2 < p1
                ORDER BY (p1-p2) DESC LIMIT 10''').fetchall()
            respuesta['respuesta'] = f'{len(rows)} materias bajaron su promedio respecto al período anterior.' if rows else 'No se detectaron bajas significativas.'
            respuesta['datos'] = [{'materia':r['materia'],'periodo_anterior':r['p1'],'periodo_actual':r['p2'],'diferencia':round(r['p1']-r['p2'],2)} for r in rows]
        else:
            respuesta['respuesta'] = 'No entendí la pregunta. Intenta preguntar sobre: estudiantes en riesgo, materias con promedio bajo, docentes sin notas, inasistencias, o rendimiento por curso.'
        return jsonify(respuesta)
    except Exception as e:
        return jsonify({'error':str(e)}), 500
    finally:
        conn.close()

# ── FASE 23: Matrículas ───────────────────────────────────────────────────
@app.route('/<slug>/matriculas')
def matriculas_list(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        rows = conn.execute('SELECT m.*, a.nombre as alumno_nombre FROM matriculas m LEFT JOIN alumnos a ON a.id=m.alumno_id ORDER BY m.created_at DESC').fetchall()
        return jsonify({'matriculas':[dict(r) for r in rows]})
    finally:
        conn.close()

@app.route('/<slug>/matriculas/crear', methods=['POST'])
def matriculas_crear(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    nombre = data.get('nombre','').strip()
    if not nombre: return jsonify({'error':'Nombre requerido'}), 400
    conn = conectar(slug)
    try:
        conn.execute('INSERT INTO matriculas (nombre, documento, email, telefono, curso_solicitado, jornada, sede, estado) VALUES (?,?,?,?,?,?,?,?)',
            (nombre, data.get('documento',''), data.get('email',''), data.get('telefono',''), data.get('curso_solicitado',''), data.get('jornada','mañana'), data.get('sede',''), 'pendiente'))
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

@app.route('/<slug>/matriculas/<int:mid>/estado', methods=['POST'])
def matriculas_estado(slug, mid):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    estado = data.get('estado','')
    if estado not in ('aprobado','rechazado','pendiente'): return jsonify({'error':'Estado inválido'}), 400
    conn = conectar(slug)
    try:
        mat = conn.execute('SELECT * FROM matriculas WHERE id=?', (mid,)).fetchone()
        if not mat: return jsonify({'error':'No encontrado'}), 404
        conn.execute('UPDATE matriculas SET estado=?, updated_at=CURRENT_TIMESTAMP WHERE id=?', (estado, mid))
        if estado == 'aprobado' and not mat['alumno_id']:
            conn.execute('INSERT INTO alumnos (nombre, curso, jornada, activo) VALUES (?,?,?,1)',
                (mat['nombre'], mat['curso_solicitado'] or 'Sin asignar', mat['jornada']))
            conn.execute('UPDATE matriculas SET alumno_id=last_insert_rowid() WHERE id=?', (mid,))
        conn.commit()
        return jsonify({'status':'ok','estado':estado})
    finally:
        conn.close()

@app.route('/<slug>/matriculas/cupos')
def matriculas_cupos(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        cupos_por_curso = conn.execute('SELECT curso, jornada, COUNT(*) as inscritos FROM alumnos WHERE activo=1 GROUP BY curso, jornada').fetchall()
        return jsonify({'cupos':[dict(r) for r in cupos_por_curso]})
    finally:
        conn.close()

# ── FASE 24: Tesorería ───────────────────────────────────────────────────
@app.route('/<slug>/tesoreria/facturas')
def tesoreria_facturas(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        rows = conn.execute('SELECT f.*, a.nombre as alumno_nombre FROM tesoreria_facturas f LEFT JOIN alumnos a ON a.id=f.alumno_id ORDER BY f.created_at DESC').fetchall()
        return jsonify({'facturas':[dict(r) for r in rows]})
    finally:
        conn.close()

@app.route('/<slug>/tesoreria/facturas/crear', methods=['POST'])
def tesoreria_facturas_crear(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    alumno_id = data.get('alumno_id')
    concepto = data.get('concepto','').strip()
    monto = float(data.get('monto',0))
    if not concepto or monto <= 0: return jsonify({'error':'Concepto y monto requeridos'}), 400
    conn = conectar(slug)
    try:
        conn.execute('INSERT INTO tesoreria_facturas (alumno_id, concepto, monto, descuento, estado, fecha_vencimiento) VALUES (?,?,?,?,?,?)',
            (alumno_id, concepto, monto, float(data.get('descuento',0)), 'pendiente', data.get('fecha_vencimiento','')))
        conn.commit()
        return jsonify({'status':'ok'})
    finally:
        conn.close()

@app.route('/<slug>/tesoreria/facturas/<int:fid>/pagar', methods=['POST'])
def tesoreria_facturas_pagar(slug, fid):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    monto = float(data.get('monto',0))
    metodo = data.get('metodo','efectivo')
    if monto <= 0: return jsonify({'error':'Monto inválido'}), 400
    conn = conectar(slug)
    try:
        fact = conn.execute('SELECT * FROM tesoreria_facturas WHERE id=?', (fid,)).fetchone()
        if not fact: return jsonify({'error':'Factura no encontrada'}), 404
        conn.execute('INSERT INTO tesoreria_pagos (factura_id, monto, metodo, referencia) VALUES (?,?,?,?)',
            (fid, monto, metodo, data.get('referencia','')))
        total_pagado = conn.execute('SELECT COALESCE(SUM(monto),0) as total FROM tesoreria_pagos WHERE factura_id=?', (fid,)).fetchone()
        if total_pagado['total'] >= fact['monto'] - fact['descuento']:
            conn.execute('UPDATE tesoreria_facturas SET estado="pagado", fecha_pago=date("now") WHERE id=?', (fid,))
        conn.commit()
        return jsonify({'status':'ok','total_pagado':total_pagado['total']})
    finally:
        conn.close()

@app.route('/<slug>/tesoreria/facturas/<int:fid>/recibo')
def tesoreria_recibo(slug, fid):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    from datetime import datetime
    conn = conectar(slug)
    try:
        fact = conn.execute('SELECT f.*, a.nombre as alumno_nombre FROM tesoreria_facturas f LEFT JOIN alumnos a ON a.id=f.alumno_id WHERE f.id=?', (fid,)).fetchone()
        if not fact: return jsonify({'error':'No encontrada'}), 404
        pagos = conn.execute('SELECT * FROM tesoreria_pagos WHERE factura_id=?', (fid,)).fetchall()
        cfg = conn.execute('SELECT * FROM config_institucion LIMIT 1').fetchone()
        html = f'''<html><head><meta charset="utf-8"><style>
            body{{font-family:system-ui,sans-serif;max-width:600px;margin:40px auto;padding:20px;border:1px solid #ddd;border-radius:12px;}}
            h1{{font-size:22px;margin-bottom:4px;}} .sub{{color:#666;font-size:13px;margin-bottom:20px;}}
            table{{width:100%;border-collapse:collapse;margin:16px 0;}} th,td{{padding:8px 10px;text-align:left;border-bottom:1px solid #eee;font-size:13px;}}
            .total{{font-size:18px;font-weight:800;text-align:right;margin:12px 0;}}
            .footer{{margin-top:24px;font-size:11px;color:#999;text-align:center;border-top:1px solid #eee;padding-top:12px;}}
        </style></head><body>
        <h1>{(cfg and cfg.get('nombre','')) or 'Lumini'}</h1>
        <div class="sub">Recibo de pago · #{fact['id']}</div>
        <p style="font-size:13px;"><strong>Alumno:</strong> {fact.get('alumno_nombre') or 'N/A'}</p>
        <p style="font-size:13px;"><strong>Concepto:</strong> {fact['concepto']}</p>
        <p style="font-size:13px;"><strong>Emisión:</strong> {fact['fecha_emision']} · <strong>Vencimiento:</strong> {fact['fecha_vencimiento']}</p>
        <table><thead><tr><th>Fecha</th><th>Método</th><th>Monto</th></tr></thead><tbody>
        {"".join(f'<tr><td>{p["created_at"][:10]}</td><td>{p["metodo"]}</td><td>${p["monto"]:.2f}</td></tr>' for p in pagos)}
        </tbody></table>
        <div class="total">Total: ${sum(p['monto'] for p in pagos):.2f} / ${fact['monto']:.2f}</div>
        <div class="footer">Generado por Lumini · {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
        </body></html>'''
        return html
    finally:
        conn.close()

# ── FASE 28: Report Builder ──────────────────────────────────────────────
@app.route('/<slug>/reportes/tablas')
def reportes_tablas(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    conn = conectar(slug)
    try:
        tables = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND name NOT LIKE 'schema_%' AND name NOT LIKE 'password_%' ORDER BY name").fetchall()
        return jsonify({'tablas':[t['name'] for t in tables]})
    finally:
        conn.close()

@app.route('/<slug>/reportes/columnas')
def reportes_columnas(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    tabla = request.args.get('tabla','')
    if not tabla: return jsonify({'error':'Tabla requerida'}), 400
    conn = conectar(slug)
    try:
        cols = conn.execute(f"PRAGMA table_info(\"{tabla}\")").fetchall()
        return jsonify({'columnas':[{'name':c['name'],'type':c['type']} for c in cols]})
    finally:
        conn.close()

@app.route('/<slug>/reportes/ejecutar', methods=['POST'])
def reportes_ejecutar(slug):
    require_colegio(slug)
    if not get_rector(slug): return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    tabla = data.get('tabla','')
    campos = data.get('campos',[])
    filtros = data.get('filtros',[])
    orden = data.get('orden','')
    limite = min(int(data.get('limite',100)),500)
    if not tabla or not campos: return jsonify({'error':'Tabla y campos requeridos'}), 400
    conn = conectar(slug)
    try:
        safe_campos = []
        for c in campos:
            c2 = c.strip().replace('"','""')
            safe_campos.append(f'"{c2}"')
        q = f'SELECT {",".join(safe_campos)} FROM "{tabla}" WHERE 1=1'
        params = []
        for f in filtros:
            col = f.get('campo','').strip().replace('"','""')
            op = f.get('operador','contiene')
            val = f.get('valor','')
            if not col or not val: continue
            if op == 'contiene':
                q += f' AND "{col}" LIKE ?'
                params.append(f'%{val}%')
            elif op == 'igual':
                q += f' AND "{col}" = ?'
                params.append(val)
            elif op == 'mayor':
                q += f' AND "{col}" > ?'
                params.append(val)
            elif op == 'menor':
                q += f' AND "{col}" < ?'
                params.append(val)
        if orden:
            safe_ord = orden.strip().replace('"','""')
            q += f' ORDER BY "{safe_ord}"'
        q += f' LIMIT {limite}'
        rows = conn.execute(q, params).fetchall()
        return jsonify({'columnas':campos,'filas':[list(r) for r in rows],'total':len(rows)})
    except Exception as e:
        return jsonify({'error':str(e)}), 400
    finally:
        conn.close()

# ── FASE 23-33: Home Dashboard by role ───────────────────────────────────
@app.route('/<slug>/home')
def home_dashboard(slug):
    require_colegio(slug)
    colegio = get_colegio(slug)
    # Determine role
    prof = get_profesor(slug)
    rector = get_rector(slug)
    pid = session.get(f'padre_id_{slug}')
    aid = session.get(f'alumno_id_{slug}')
    conn = conectar(slug)
    try:
        if rector:
            total_est = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE activo=1').fetchone()['c']
            total_prof = conn.execute('SELECT COUNT(*) as c FROM profesores WHERE activo=1').fetchone()['c']
            riesgo = conn.execute('SELECT COUNT(DISTINCT n.alumno_id) as c FROM notas n JOIN actividades a ON a.id=n.actividad_id JOIN alumnos al ON al.id=n.alumno_id WHERE al.activo=1 GROUP BY n.alumno_id HAVING AVG(n.val) < 3.0').fetchall()
            pendientes = conn.execute('SELECT COUNT(*) as c FROM solicitudes_modificacion WHERE estado="pendiente"').fetchone()['c']
            cal_events = conn.execute("SELECT COUNT(*) as c FROM eventos_calendario WHERE DATE(fecha)>=DATE('now') AND DATE(fecha)<=DATE('now','+7 days')").fetchone()['c']
            return jsonify({'rol':'rector','total_estudiantes':total_est,'total_profesores':total_prof,'riesgo':len(riesgo),'pendientes':pendientes,'eventos':cal_events})
        elif prof:
            cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE profesor_id=?', (prof['id'],)).fetchall()
            acts_pend = conn.execute("SELECT COUNT(*) as c FROM actividades WHERE profesor_id=? AND estado_act='borrador'", (prof['id'],)).fetchone()['c']
            riesgo = conn.execute('SELECT COUNT(DISTINCT n.alumno_id) as c FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE a.profesor_id=? GROUP BY n.alumno_id HAVING AVG(n.val) < 3.0', (prof['id'],)).fetchall()
            return jsonify({'rol':'docente','nombre':prof['nombre'],'cursos':len(cursos),'actividades_pendientes':acts_pend,'riesgo':len(riesgo)})
        elif pid:
            hijos = conn.execute('SELECT a.id, a.nombre, a.curso FROM alumno_padre ap JOIN alumnos a ON a.id=ap.alumno_id WHERE ap.padre_id=?', (pid,)).fetchall()
            hijos_data = []
            for h in hijos:
                prom = conn.execute('SELECT COALESCE(AVG(n.val),0) as p FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.alumno_id=?', (h['id'],)).fetchone()
                hijos_data.append({'id':h['id'],'nombre':h['nombre'],'curso':h['curso'],'promedio':round(prom['p'],2)})
            return jsonify({'rol':'padre','hijos':hijos_data})
        elif aid:
            al = conn.execute('SELECT * FROM alumnos WHERE id=?', (aid,)).fetchone()
            return jsonify({'rol':'estudiante','nombre':al['nombre'],'curso':al['curso']})
        else:
            return jsonify({'rol':'anonimo'})
    finally:
        conn.close()

# ── NOTAS ─────────────────────────────────────────────────────────────────────
@app.route('/<slug>/guardar_nota', methods=['POST'])
def guardar_nota(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    aid          = request.form.get('aid', type=int)
    actividad_id = request.form.get('actividad_id', type=int)
    val          = request.form.get('val', type=float)
    if None in (aid, actividad_id, val): return ('', 400)
    conn = conectar(slug)
    act = conn.execute(
        'SELECT a.id, a.profesor_id, a.curso, COALESCE(a.periodo,1) as p FROM actividades a WHERE a.id=?',
        (actividad_id,)).fetchone()
    if not act:
        conn.close()
        return ('', 404)
    if act['profesor_id'] != prof['id']:
        conn.close()
        return ('', 403)
    alumno = conn.execute('SELECT id FROM alumnos WHERE id=? AND curso=? AND activo=1',
                          (aid, act['curso'])).fetchone()
    if not alumno:
        conn.close()
        return ('', 403)
    if periodo_cerrado(slug, act['p']):
        conn.close()
        return jsonify({'status':'error','codigo':'PERIODO_CERRADO','mensaje':'El per\u00edodo est\u00e1 cerrado.'}), 403
    old = conn.execute(
        'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
        (aid, actividad_id)).fetchone()
    old_val = old['val'] if old else None
    conn.execute(
        '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
           ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
        (aid, actividad_id, val))
    conn.commit()
    audit_log(slug, prof['id'], 'nota_editada', 'notas', registro_id=None,
              valor_anterior={'aid': aid, 'actividad_id': actividad_id, 'val': old_val},
              valor_nuevo={'aid': aid, 'actividad_id': actividad_id, 'val': val})
    tipo_nota = 'creacion' if old_val is None else 'modificacion'
    jornada, materia = get_sesion_jornada_materia(slug)
    auditar_nota(slug, prof['id'], 'profesor', tipo_nota, 'notas', aid,
                 act['curso'], materia, act['p'],
                 campo='nota', actividad_id=actividad_id,
                 valor_anterior=old_val, valor_nuevo=val)
    prom_est = calcular_stats_estudiante(conn, slug, aid, act['curso'], materia, jornada, act['p'], prof['id'])
    nf = calcular_nota_final_estudiante(conn, slug, aid, act['curso'], materia, jornada, act['p'], prof['id'])
    curso_stats = calcular_stats_curso(conn, slug, act['curso'], materia, jornada, act['p'], prof['id'])
    conn.close()
    logger.info('guardar_nota: aid=%d actividad_id=%d val=%s prom_est=%s nf=%s', aid, actividad_id, val, prom_est, nf)
    return jsonify({'status':'ok','promedio':prom_est,'nota_final':nf,'promedio_curso':curso_stats['promedio_curso'],'notas_pendientes':curso_stats['notas_pendientes']})

# ── HISTORIAL NOTAS ──────────────────────────────────────────────────────────
@app.route('/<slug>/historial_notas/<int:aid>')
def historial_notas(slug, aid):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    conn = conectar(slug)
    rows = conn.execute(
        '''SELECT a.id, a.tipo_accion, a.tabla, a.campo, a.valor_anterior, a.valor_nuevo,
                  a.creado, a.materia, a.periodo, a.motivo, a.aid,
                  COALESCE(ac.nombre, '') as actividad_nombre
           FROM auditoria_notas a
           LEFT JOIN actividades ac ON ac.id = a.actividad_id
           WHERE a.aid = ?
           ORDER BY a.creado DESC
           LIMIT 200''',
        (aid,)).fetchall()
    conn.close()
    return jsonify([{
        'id': r['id'],
        'tipo_accion': r['tipo_accion'],
        'tabla': r['tabla'],
        'campo': r['campo'],
        'valor_anterior': r['valor_anterior'],
        'valor_nuevo': r['valor_nuevo'],
        'creado': r['creado'],
        'materia': r['materia'],
        'periodo': r['periodo'],
        'motivo': r['motivo'],
        'actividad_nombre': r['actividad_nombre'],
        'aid': r['aid'],
    } for r in rows])

@app.route('/<slug>/historial_curso')
def historial_curso(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    curso = request.args.get('curso', '')
    periodo = request.args.get('periodo', 1, type=int)
    if not curso: return jsonify([])
    conn = conectar(slug)
    rows = conn.execute(
        '''SELECT a.id, a.tipo_accion, a.tabla, a.campo, a.valor_anterior, a.valor_nuevo,
                  a.creado, a.materia, a.periodo, a.motivo, a.aid,
                  COALESCE(ac.nombre, '') as actividad_nombre
           FROM auditoria_notas a
           LEFT JOIN actividades ac ON ac.id = a.actividad_id
           WHERE a.curso = ? AND a.materia = ? AND a.periodo = ? AND a.profesor_id = ?
           ORDER BY a.creado DESC
           LIMIT 500''',
        (curso, request.args.get('materia', ''), periodo, prof['id'])).fetchall()
    conn.close()
    return jsonify([{
        'id': r['id'], 'tipo_accion': r['tipo_accion'], 'tabla': r['tabla'],
        'campo': r['campo'], 'valor_anterior': r['valor_anterior'],
        'valor_nuevo': r['valor_nuevo'], 'creado': r['creado'],
        'materia': r['materia'], 'periodo': r['periodo'], 'motivo': r['motivo'],
        'actividad_nombre': r['actividad_nombre'], 'aid': r['aid'],
    } for r in rows])

# ── EVALUACIONES ──────────────────────────────────────────────────────────────
@app.route('/<slug>/guardar_evaluacion', methods=['POST'])
def guardar_evaluacion(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    aid     = request.form.get('aid', type=int)
    ev      = request.form.get('evaluacion', type=float)
    au      = request.form.get('autoevaluacion', type=float)
    periodo = request.form.get('periodo', 1, type=int)
    curso   = request.form.get('curso', '')
    if aid is None: return ('', 400)
    conn = conectar(slug)
    if periodo_cerrado(slug, periodo):
        conn.close()
        return jsonify({'status':'error','codigo':'PERIODO_CERRADO','mensaje':'El per\u00edodo est\u00e1 cerrado.'}), 403
    if not curso:
        cursos_prof = get_cursos_profesor(slug, prof['id'], materia, jornada)
        curso = cursos_prof[0] if cursos_prof else ''
    existing = conn.execute(
        '''SELECT evaluacion, autoevaluacion FROM evaluaciones
           WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?''',
        (aid, prof['id'], materia, jornada, periodo)
    ).fetchone()
    old_eval = existing['evaluacion'] if existing else None
    old_auto = existing['autoevaluacion'] if existing else None
    ev_final = ev if ev is not None else old_eval
    au_final = au if au is not None else old_auto
    try:
        conn.execute(
            '''INSERT INTO evaluaciones
               (aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
               VALUES (?,?,?,?,?,?,?)
               ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
               DO UPDATE SET evaluacion=excluded.evaluacion, autoevaluacion=excluded.autoevaluacion''',
            (aid, prof['id'], materia, jornada, ev_final, au_final, periodo))
        conn.commit()
    except sqlite3.OperationalError as e:
        conn.rollback()
        if 'ON CONFLICT clause does not match' in str(e):
            logger.warning(f'[{slug}] ON CONFLICT falló en guardar_evaluacion, reparando...')
            _recrear_si_unique_incorrecto(conn, slug, 'evaluaciones',
                '(aid,profesor_id,materia,jornada,periodo)',
                '''CREATE TABLE evaluaciones (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
                    materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
                    evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
                    UNIQUE(aid,profesor_id,materia,jornada,periodo))''',
                '''(id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
                   SELECT id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,
                          COALESCE(periodo,1) FROM evaluaciones_old''')
            conn.execute(
                '''INSERT INTO evaluaciones
                   (aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                   DO UPDATE SET evaluacion=excluded.evaluacion, autoevaluacion=excluded.autoevaluacion''',
                (aid, prof['id'], materia, jornada, ev_final, au_final, periodo))
            conn.commit()
        else:
            raise
    audit_log(slug, prof['id'], 'evaluacion_editada', 'evaluaciones', registro_id=None,
              valor_anterior={'aid': aid, 'evaluacion': old_eval, 'autoevaluacion': old_auto},
              valor_nuevo={'aid': aid, 'evaluacion': ev_final, 'autoevaluacion': au_final})
    if ev is not None:
        tipo_ev = 'creacion' if old_eval is None else 'modificacion'
        auditar_nota(slug, prof['id'], 'profesor', tipo_ev, 'evaluaciones', aid,
                     curso, materia, periodo, campo='evaluacion',
                     valor_anterior=old_eval, valor_nuevo=ev_final)
    if au is not None:
        tipo_au = 'creacion' if old_auto is None else 'modificacion'
        auditar_nota(slug, prof['id'], 'profesor', tipo_au, 'evaluaciones', aid,
                     curso, materia, periodo, campo='autoevaluacion',
                     valor_anterior=old_auto, valor_nuevo=au_final)
    prom_est = calcular_stats_estudiante(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
    nf = calcular_nota_final_estudiante(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
    curso_stats = calcular_stats_curso(conn, slug, curso, materia, jornada, periodo, prof['id'])
    conn.close()
    return jsonify({'status':'ok','promedio':prom_est,'nota_final':nf,'promedio_curso':curso_stats['promedio_curso'],'notas_pendientes':curso_stats['notas_pendientes']})

# ── BATCH SAVE ────────────────────────────────────────────────────────────────
@app.route('/<slug>/guardar_nota_batch', methods=['POST'])
def guardar_nota_batch(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error'}), 403
    if not validar_csrf(): return jsonify({'status':'error'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error'}), 400
    curso = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    raw = request.form.get('batch', '')
    if not raw: return jsonify({'status':'error'}), 400
    try: items = json.loads(raw)
    except: return jsonify({'status':'error'}), 400
    conn = conectar(slug)
    resultados = {}
    try:
        for item in items:
            aid = item.get('aid')
            actividad_id = item.get('actividad_id')
            val = item.get('val')
            if None in (aid, actividad_id, val):
                continue
            act = conn.execute(
                'SELECT a.id, a.profesor_id, a.curso, COALESCE(a.periodo,1) as p FROM actividades a WHERE a.id=?',
                (actividad_id,)).fetchone()
            if not act or act['profesor_id'] != prof['id']: continue
            if periodo_cerrado(slug, act['p']):
                continue
            old = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (aid, actividad_id)).fetchone()
            old_val = old['val'] if old else None
            conn.execute(
                '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
                   ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
                (aid, actividad_id, val))
            audit_log(slug, prof['id'], 'nota_editada', 'notas', registro_id=None,
                      valor_anterior={'aid': aid, 'actividad_id': actividad_id, 'val': old_val},
                      valor_nuevo={'aid': aid, 'actividad_id': actividad_id, 'val': val})
            tipo_nota = 'creacion' if old_val is None else 'modificacion'
            auditar_nota(slug, prof['id'], 'profesor', tipo_nota, 'notas', aid,
                         act['curso'], materia, act['p'],
                         campo='nota', actividad_id=actividad_id,
                         valor_anterior=old_val, valor_nuevo=val)
        conn.commit()
        # Recalculate all affected students
        aids = set(item.get('aid') for item in items if item.get('aid'))
        for aid in aids:
            prom_est = calcular_stats_estudiante(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
            nf = calcular_nota_final_estudiante(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
            resultados[aid] = {'promedio': prom_est, 'nota_final': nf}
    except Exception as e:
        conn.close()
        logger.error(f'Error en guardar_nota_batch: {e}')
        return jsonify({'status':'error'}), 500
    curso_stats = calcular_stats_curso(conn, slug, curso, materia, jornada, periodo, prof['id'])
    conn.close()
    return jsonify({'status':'ok', 'resultados': resultados, 'stats_curso': curso_stats})

@app.route('/<slug>/guardar_evaluacion_batch', methods=['POST'])
def guardar_evaluacion_batch(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error'}), 403
    if not validar_csrf(): return jsonify({'status':'error'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error'}), 400
    curso = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    raw = request.form.get('batch', '')
    if not raw: return jsonify({'status':'error'}), 400
    try: items = json.loads(raw)
    except: return jsonify({'status':'error'}), 400
    conn = conectar(slug)
    resultados = {}
    try:
        for item in items:
            aid = item.get('aid')
            if not aid: continue
            ev = item.get('evaluacion', type=float)
            au = item.get('autoevaluacion', type=float)
            existing = conn.execute(
                '''SELECT evaluacion, autoevaluacion FROM evaluaciones
                   WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?''',
                (aid, prof['id'], materia, jornada, periodo)).fetchone()
            old_eval = existing['evaluacion'] if existing else None
            old_auto = existing['autoevaluacion'] if existing else None
            ev_final = ev if ev is not None else old_eval
            au_final = au if au is not None else old_auto
            conn.execute(
                '''INSERT INTO evaluaciones
                   (aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                   DO UPDATE SET evaluacion=excluded.evaluacion, autoevaluacion=excluded.autoevaluacion''',
                (aid, prof['id'], materia, jornada, ev_final, au_final, periodo))
            if ev is not None:
                tipo_ev = 'creacion' if old_eval is None else 'modificacion'
                auditar_nota(slug, prof['id'], 'profesor', tipo_ev, 'evaluaciones', aid,
                             curso, materia, periodo, campo='evaluacion',
                             valor_anterior=old_eval, valor_nuevo=ev_final)
            if au is not None:
                tipo_au = 'creacion' if old_auto is None else 'modificacion'
                auditar_nota(slug, prof['id'], 'profesor', tipo_au, 'evaluaciones', aid,
                             curso, materia, periodo, campo='autoevaluacion',
                             valor_anterior=old_auto, valor_nuevo=au_final)
        conn.commit()
        aids = set(item.get('aid') for item in items if item.get('aid'))
        for aid in aids:
            prom_est = calcular_stats_estudiante(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
            nf = calcular_nota_final_estudiante(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
            resultados[aid] = {'promedio': prom_est, 'nota_final': nf}
    except Exception as e:
        conn.close()
        logger.error(f'Error en guardar_evaluacion_batch: {e}')
        return jsonify({'status':'error'}), 500
    curso_stats = calcular_stats_curso(conn, slug, curso, materia, jornada, periodo, prof['id'])
    conn.close()
    return jsonify({'status':'ok', 'resultados': resultados, 'stats_curso': curso_stats})

# ── SOLICITUDES DE MODIFICACION ──────────────────────────────────────────────
@app.route('/<slug>/solicitar_modificacion', methods=['POST'])
def solicitar_modificacion(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    aid             = request.form.get('aid', type=int)
    actividad_id    = request.form.get('actividad_id', type=int)
    tipo            = request.form.get('tipo', '').strip()
    valor_solicitado = request.form.get('valor_solicitado', type=float)
    motivo          = request.form.get('motivo', '').strip()
    periodo         = request.form.get('periodo', 1, type=int)
    if None in (aid, valor_solicitado) or not motivo or not tipo:
        return jsonify({'status':'error','mensaje':'Datos incompletos'}), 400
    if valor_solicitado < 0 or valor_solicitado > 5:
        return jsonify({'status':'error','mensaje':'Nota debe estar entre 0 y 5'}), 400
    if tipo not in ('actividad', 'evaluacion', 'autoevaluacion'):
        return jsonify({'status':'error','mensaje':'Tipo inv\u00e1lido'}), 400
    if not periodo_cerrado(slug, periodo):
        return jsonify({'status':'error','mensaje':'El per\u00edodo no est\u00e1 cerrado'}), 400
    conn = conectar(slug)
    jornada_ctx, materia_ctx = get_sesion_jornada_materia(slug)
    materia = materia_ctx or ''
    jornada = jornada_ctx or ''
    curso = ''
    valor_actual = None
    if tipo == 'actividad':
        if actividad_id is None:
            conn.close()
            return jsonify({'status':'error','mensaje':'actividad_id requerido para tipo actividad'}), 400
        act = conn.execute(
            'SELECT id, profesor_id, curso, materia, COALESCE(periodo,1) as p FROM actividades WHERE id=?',
            (actividad_id,)).fetchone()
        if not act:
            conn.close()
            return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
        if act['profesor_id'] != prof['id']:
            conn.close()
            return jsonify({'status':'error','mensaje':'No eres el propietario de esta actividad'}), 403
        materia = act['materia']
        curso = act['curso']
        nota_db = conn.execute(
            'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
            (aid, actividad_id)).fetchone()
        valor_actual = nota_db['val'] if nota_db else None
    elif tipo == 'evaluacion':
        curso = request.form.get('curso', '')
        alumno = conn.execute(
            'SELECT evaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND COALESCE(periodo,1)=?',
            (aid, prof['id'], materia, periodo)).fetchone()
        valor_actual = alumno['evaluacion'] if alumno else None
    elif tipo == 'autoevaluacion':
        curso = request.form.get('curso', '')
        alumno = conn.execute(
            'SELECT autoevaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND COALESCE(periodo,1)=?',
            (aid, prof['id'], materia, periodo)).fetchone()
        valor_actual = alumno['autoevaluacion'] if alumno else None
    if not curso:
        curso = request.form.get('curso', '')
    conn.execute(
        '''INSERT INTO solicitudes_modificacion
           (slug, aid, profesor_id, materia, curso, jornada, periodo, tipo, actividad_id,
            valor_actual, valor_solicitado, motivo, estado)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'pendiente')''',
        (slug, aid, prof['id'], materia, curso, jornada, periodo, tipo, actividad_id,
         str(valor_actual) if valor_actual is not None else None, str(valor_solicitado), motivo))
    sid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.commit()
    auditar_nota(slug, prof['id'], 'profesor', 'solicitud_creada', 'solicitudes_modificacion', aid,
                 curso, materia, periodo, campo=tipo, actividad_id=actividad_id,
                 valor_anterior=valor_actual, valor_nuevo=valor_solicitado,
                 motivo='Solicitud #%d: %s' % (sid, motivo))
    # Notify all rectores
    rectores = conn.execute('SELECT id FROM rectores WHERE activo=1').fetchall()
    for r in rectores:
        crear_notificacion(slug, 'rector', r['id'],
            'Nueva solicitud de modificaci\u00f3n de %s' % prof['nombre'],
            'El profesor %s solicita cambiar %s de %s a %s. Motivo: %s' % (
                prof['nombre'], tipo, valor_actual or 'sin nota', valor_solicitado, motivo),
            link=url_for('rector_solicitudes', slug=slug))
    conn.close()
    return jsonify({'status':'ok','mensaje':'Solicitud enviada correctamente.','id':sid})

    conn.close()
    return jsonify({'status':'ok','mensaje':'Solicitud enviada correctamente.','id':sid})

# ── EXCEL IMPORT / EXPORT (Fase 5) ───────────────────────────────────────
def _excel_armar_wb(slug, prof, materia, jornada, curso_sel, periodo, actividades, alumnos):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = 'Notas'
    header_fill = PatternFill('solid', fgColor='6D28D9')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    headers = ['N°', 'Estudiante', 'AID'] + [a['nombre'] for a in actividades] + ['Evaluación', 'Autoevaluación', 'Promedio']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    conn = conectar(slug)
    try:
        aid_list = [a['id'] for a in alumnos]
        if aid_list:
            placeholders = ','.join('?' * len(aid_list))
            notas_all = conn.execute(
                f'''SELECT n.aid, n.actividad_id, n.val FROM notas n
                    JOIN actividades ac ON ac.id=n.actividad_id
                    WHERE n.aid IN ({placeholders}) AND ac.materia=? AND ac.jornada=? AND ac.curso=?
                    AND COALESCE(ac.periodo,1)=? AND ac.profesor_id=?''',
                (*aid_list, materia, jornada, curso_sel, periodo, prof['id'])).fetchall()
            notas_by_aid = {}
            for r in notas_all:
                notas_by_aid.setdefault(r['aid'], {})[r['actividad_id']] = r['val']
            evals_all = conn.execute(
                f'''SELECT aid, evaluacion, autoevaluacion FROM evaluaciones
                    WHERE aid IN ({placeholders}) AND profesor_id=? AND materia=? AND jornada=?
                    AND COALESCE(periodo,1)=?''',
                (*aid_list, prof['id'], materia, jornada, periodo)).fetchall()
            evals_by_aid = {r['aid']: {'ev': r['evaluacion'], 'auto': r['autoevaluacion']} for r in evals_all}
        else:
            notas_by_aid, evals_by_aid = {}, {}
        for i, a in enumerate(alumnos, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).alignment = center
            ws.cell(row=row, column=2, value=a['nombre'])
            ws.cell(row=row, column=3, value=a['id']).alignment = center
            for j, act in enumerate(actividades, 4):
                val = None
                if a['id'] in notas_by_aid:
                    val = notas_by_aid[a['id']].get(act['id'])
                if val is not None:
                    ws.cell(row=row, column=j, value=float(val)).alignment = center
            ev = evals_by_aid.get(a['id'], {})
            ev_val = ev.get('ev')
            auto_val = ev.get('auto')
            ecol = 4 + len(actividades)
            if ev_val is not None:
                ws.cell(row=row, column=ecol, value=float(ev_val)).alignment = center
            if auto_val is not None:
                ws.cell(row=row, column=ecol + 1, value=float(auto_val)).alignment = center
            notas_dict = notas_by_aid.get(a['id'], {})
            prom = _promedio_ponderado([notas_dict.get(act['id']) for act in actividades], ev_val, auto_val)
            ws.cell(row=row, column=ecol + 2, value=round(prom, 2) if prom is not None else '').alignment = center
    finally:
        conn.close()
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    for j in range(len(actividades)):
        ws.column_dimensions[chr(68 + j)].width = 14
    return wb

@app.route('/<slug>/plantilla_notas')
def plantilla_notas(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('seleccionar_jornada', slug=slug))
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    alumnos = []
    if curso_sel:
        conn = conectar(slug)
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
            (curso_sel, jornada)).fetchall()
        conn.close()
    wb = _excel_armar_wb(slug, prof, materia, jornada, curso_sel, periodo, actividades, alumnos)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'plantilla_lumini_{slug}_{curso_sel}_{periodo}.xlsx'
    return Response(bio.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})

@app.route('/<slug>/exportar_notas')
def exportar_notas(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('seleccionar_jornada', slug=slug))
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    alumnos = []
    if curso_sel:
        conn = conectar(slug)
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
            (curso_sel, jornada)).fetchall()
        conn.close()
    wb = _excel_armar_wb(slug, prof, materia, jornada, curso_sel, periodo, actividades, alumnos)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'notas_{slug}_{curso_sel}_{periodo}.xlsx'
    return Response(bio.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})

@app.route('/<slug>/importar_notas', methods=['GET'])
def importar_notas(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('seleccionar_jornada', slug=slug))
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    return render_template('importar_notas.html', slug=slug, colegio=get_colegio(slug), profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel, periodo=periodo,
                           materia=materia, jornada=jornada, actividades=actividades)

@app.route('/<slug>/importar_notas/preview', methods=['POST'])
def importar_notas_preview(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    if 'archivo' not in request.files:
        return jsonify({'status':'error','mensaje':'No se envio ningun archivo'}), 400
    f = request.files['archivo']
    if not f.filename or not f.filename.lower().endswith('.xlsx'):
        return jsonify({'status':'error','mensaje':'El archivo debe ser .xlsx'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f'Error al leer archivo de notas: {e}')
        return jsonify({'status':'error','mensaje':'Error al leer el archivo. Verifica el formato.'}), 400
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    if not header_row or str(header_row[0]).strip() != 'N°':
        return jsonify({'status':'error','mensaje':'Formato de archivo invalido. La primera columna debe ser N°'}), 400
    rows_data = list(ws.iter_rows(min_row=2, values_only=False))
    if not rows_data:
        return jsonify({'status':'error','mensaje':'El archivo no contiene datos'}), 400
    conn = conectar(slug)
    try:
        actividades_existentes = conn.execute(
            '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
               AND COALESCE(periodo,1)=? ORDER BY orden''',
            (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
        existing_act_names = {a['nombre']: a for a in actividades_existentes}
        act_cols = []
        eval_col = auto_col = None
        for col_idx, h in enumerate(header_row):
            if col_idx <= 2:
                continue
            h_str = str(h).strip() if h is not None else ''
            if h_str == 'Evaluación':
                eval_col = col_idx
            elif h_str == 'Autoevaluación':
                auto_col = col_idx
            elif h_str == 'Promedio':
                continue
            elif h_str:
                act_cols.append((col_idx, h_str))
        new_activities = []
        col_map = {}
        max_orden = max([a['orden'] for a in actividades_existentes], default=0)
        for col_idx, act_name in act_cols:
            if act_name in existing_act_names:
                col_map[col_idx] = {'tipo': 'actividad', 'nombre': act_name, 'actividad_id': existing_act_names[act_name]['id']}
            else:
                max_orden += 1
                col_map[col_idx] = {'tipo': 'actividad', 'nombre': act_name, 'actividad_id': None, 'orden': max_orden}
                new_activities.append({'nombre': act_name, 'orden': max_orden})
        all_ok = True
        preview_rows = []
        aid_set = set()
        for row_cells in rows_data:
            cells = [c.value for c in row_cells]
            if not any(c is not None for c in cells):
                continue
            raw_nombre = str(cells[1]).strip() if cells[1] is not None else ''
            raw_aid = cells[2]
            row_errors = []
            aid = None
            if raw_aid is not None:
                try:
                    aid = int(raw_aid)
                except (ValueError, TypeError):
                    pass
            alumno = None
            if aid:
                al = conn.execute('SELECT * FROM alumnos WHERE id=? AND jornada=? AND activo=1', (aid, jornada)).fetchone()
                if al and al['curso'] == curso_sel:
                    alumno = al
            if not alumno and raw_nombre:
                al = conn.execute(
                    'SELECT * FROM alumnos WHERE nombre=? AND curso=? AND jornada=? AND activo=1',
                    (raw_nombre, curso_sel, jornada)).fetchone()
                if al:
                    alumno = al
                    aid = al['id']
            if not alumno:
                row_errors.append('Estudiante no encontrado en este curso')
                all_ok = False
            if aid:
                if aid in aid_set:
                    row_errors.append('Estudiante duplicado en el archivo')
                    all_ok = False
                aid_set.add(aid)
            changes = {}
            for col_idx, cinfo in col_map.items():
                raw_val = cells[col_idx] if col_idx < len(cells) else None
                val = None
                if raw_val is not None:
                    try:
                        val = float(str(raw_val).replace(',', '.'))
                        if val < 0 or val > 5:
                            row_errors.append(f'{cinfo["nombre"]}: nota fuera de rango (0-5)')
                            all_ok = False
                            continue
                    except (ValueError, TypeError):
                        row_errors.append(f'{cinfo["nombre"]}: valor invalido')
                        all_ok = False
                        continue
                changes[f'act_{col_idx}'] = {'tipo': 'actividad', 'actividad_id': cinfo.get('actividad_id'),
                                               'valor': val, 'nombre_col': cinfo['nombre']}
            if alumno:
                if eval_col is not None and eval_col < len(cells):
                    raw_val = cells[eval_col]
                    if raw_val is not None:
                        try:
                            val = float(str(raw_val).replace(',', '.'))
                            if val < 0 or val > 5:
                                row_errors.append('Evaluación fuera de rango')
                                all_ok = False
                            else:
                                changes['eval'] = {'tipo': 'evaluacion', 'valor': val}
                        except (ValueError, TypeError):
                            row_errors.append('Evaluación invalida')
                            all_ok = False
                if auto_col is not None and auto_col < len(cells):
                    raw_val = cells[auto_col]
                    if raw_val is not None:
                        try:
                            val = float(str(raw_val).replace(',', '.'))
                            if val < 0 or val > 5:
                                row_errors.append('Autoevaluación fuera de rango')
                                all_ok = False
                            else:
                                changes['auto'] = {'tipo': 'autoevaluacion', 'valor': val}
                        except (ValueError, TypeError):
                            row_errors.append('Autoevaluación invalida')
                            all_ok = False
            preview_rows.append({
                'fila': row_cells[0].row,
                'aid': aid,
                'nombre': raw_nombre,
                'alumno': dict(alumno) if alumno else None,
                'errors': row_errors,
                'changes': changes,
                'ok': len(row_errors) == 0,
            })
    finally:
        conn.close()
    return jsonify({
        'status': 'ok' if all_ok else 'error',
        'total': len(preview_rows),
        'validos': sum(1 for r in preview_rows if r['ok']),
        'errores': sum(1 for r in preview_rows if not r['ok']),
        'filas': preview_rows,
        'nuevas_actividades': new_activities,
        'curso': curso_sel,
        'periodo': periodo,
        'all_ok': all_ok,
    })

@app.route('/<slug>/importar_notas/confirmar', methods=['POST'])
def importar_notas_confirmar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    data_json = request.form.get('data', '')
    if not data_json:
        return jsonify({'status':'error','mensaje':'No hay datos para guardar'}), 400
    try:
        data = json.loads(data_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'status':'error','mensaje':'Datos invalidos'}), 400
    if not data.get('all_ok'):
        return jsonify({'status':'error','mensaje':'Hay errores que deben corregirse primero'}), 400
    conn = conectar(slug)
    try:
        new_act_names = {}
        for na in data.get('nuevas_actividades', []):
            conn.execute(
                'INSERT INTO actividades (nombre, profesor_id, materia, jornada, curso, orden, periodo) VALUES (?,?,?,?,?,?,?)',
                (na['nombre'], prof['id'], materia, jornada, curso_sel, na['orden'], periodo))
            act_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            new_act_names[na['nombre']] = act_id
        updated = 0
        for f in data['filas']:
            if not f.get('ok') or not f.get('alumno'):
                continue
            aid = f['aid']
            for key, ch in f.get('changes', {}).items():
                if ch['tipo'] == 'actividad':
                    act_id = ch.get('actividad_id')
                    if act_id is None:
                        act_id = new_act_names.get(ch.get('nombre_col', ''))
                    if act_id is None:
                        continue
                    existing = conn.execute(
                        'SELECT val FROM notas WHERE aid=? AND actividad_id=?', (aid, act_id)).fetchone()
                    old_val = existing['val'] if existing else None
                    if old_val != ch['valor']:
                        if ch['valor'] is not None:
                            conn.execute(
                                '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
                                   ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
                                (aid, act_id, ch['valor']))
                        elif existing:
                            conn.execute('DELETE FROM notas WHERE aid=? AND actividad_id=?', (aid, act_id))
                        auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'notas', aid,
                                     curso_sel, materia, periodo,
                                     campo='nota', actividad_id=act_id,
                                     valor_anterior=old_val, valor_nuevo=ch['valor'],
                                     motivo='Importacion masiva Excel')
                        updated += 1
                elif ch['tipo'] == 'evaluacion':
                    existing = conn.execute(
                        'SELECT evaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['evaluacion'] if existing else None
                    if old_val != ch['valor']:
                        if ch['valor'] is not None:
                            conn.execute(
                                '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,evaluacion,periodo)
                                   VALUES (?,?,?,?,?,?)
                                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                                   DO UPDATE SET evaluacion=excluded.evaluacion''',
                                (aid, prof['id'], materia, jornada, ch['valor'], periodo))
                        elif existing:
                            conn.execute(
                                'UPDATE evaluaciones SET evaluacion=NULL WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                                (aid, prof['id'], materia, jornada, periodo))
                        auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='evaluacion',
                                     valor_anterior=old_val, valor_nuevo=ch['valor'],
                                     motivo='Importacion masiva Excel')
                        updated += 1
                elif ch['tipo'] == 'autoevaluacion':
                    existing = conn.execute(
                        'SELECT autoevaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['autoevaluacion'] if existing else None
                    if old_val != ch['valor']:
                        if ch['valor'] is not None:
                            conn.execute(
                                '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,autoevaluacion,periodo)
                                   VALUES (?,?,?,?,?,?)
                                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                                   DO UPDATE SET autoevaluacion=excluded.autoevaluacion''',
                                (aid, prof['id'], materia, jornada, ch['valor'], periodo))
                        elif existing:
                            conn.execute(
                                'UPDATE evaluaciones SET autoevaluacion=NULL WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                                (aid, prof['id'], materia, jornada, periodo))
                        auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='autoevaluacion',
                                     valor_anterior=old_val, valor_nuevo=ch['valor'],
                                     motivo='Importacion masiva Excel')
                        updated += 1
        conn.commit()
    except Exception as e:
        conn.close()
        logger.error(f'Error al guardar: {e}')
        return jsonify({'status':'error','mensaje':'Error al guardar. Intenta de nuevo.'}), 500
    conn.close()
    return jsonify({'status':'ok', 'mensaje': f'Importacion completada. {updated} valores actualizados.', 'updated': updated})

# ── MIGRAR DESDE EXCEL (WIZARD) ───────────────────────────────────────────────

@app.route('/<slug>/migrar-excel', methods=['GET'])
def migrar_excel(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('seleccionar_jornada', slug=slug))
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    return render_template('migrar_excel.html', slug=slug, colegio=get_colegio(slug), profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel, periodo=periodo,
                           materia=materia, jornada=jornada, actividades=actividades)

@app.route('/<slug>/migrar-excel/analizar', methods=['POST'])
def migrar_excel_analizar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    if 'archivo' not in request.files:
        return jsonify({'status':'error','mensaje':'No se envio ningun archivo'}), 400
    f = request.files['archivo']
    if not f.filename or not extension_permitida(f.filename):
        return jsonify({'status':'error','mensaje':'Formato no valido. Usa .xlsx'}), 400
    import tempfile, os
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        f.save(tmp.name)
        import openpyxl
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return jsonify({'status':'error','mensaje':'El archivo debe tener al menos una fila de encabezados y una fila de datos'}), 400
        encabezados = [str(c).strip() if c is not None else '' for c in rows[0]]
        conn = conectar(slug)
        alumnos = conn.execute(
            'SELECT id, CONCAT(apellidos,\' \',nombres) as nombre, documento FROM alumnos WHERE curso=? AND activo=1 ORDER BY apellidos,nombres',
            (curso_sel,)).fetchall()
        actividades = conn.execute(
            '''SELECT id, nombre, orden FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
               AND COALESCE(periodo,1)=? ORDER BY orden''',
            (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
        conn.close()
        acts_map = {a['nombre'].strip().lower(): a for a in actividades}
        col_detectadas = []
        col_pendientes = []
        for i, h in enumerate(encabezados):
            hl = h.strip().lower()
            if not hl:
                col_pendientes.append({'indice': i, 'nombre': h, 'tipo': 'skip'})
                continue
            act_match = acts_map.get(hl)
            if act_match:
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'actividad_id': act_match['id'], 'actividad_nombre': act_match['nombre']})
            elif hl in ('nombre', 'nombre del estudiante', 'estudiante', 'alumno', 'alumno(a)'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'nombre'})
            elif hl in ('documento', 'id', 'identificacion', 'codigo', 'código', 'cedula', 'cédula'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'documento'})
            elif hl in ('evaluacion', 'evaluación', 'eva', 'nota evaluacion', 'nota evaluación', 'eval'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'evaluacion'})
            elif hl in ('autoevaluacion', 'autoevaluación', 'auto-evaluacion', 'auto', 'auto-evaluación'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'autoevaluacion'})
            elif hl in ('proyecto', 'nota proyecto'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'proyecto'})
            elif hl in ('recuperacion', 'recuperación', 'recu', 'nota recuperacion', 'nota recuperación'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'recuperacion'})
            elif hl in ('nota final', 'nota definitiva', 'definitiva', 'promedio', 'final'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'nota_final'})
            else:
                col_pendientes.append({'indice': i, 'nombre': h, 'tipo': 'manual'})
        filas = []
        nuevos_map = {}
        for idx, row in enumerate(rows[1:], start=2):
            vals = [str(v).strip() if v is not None else '' for v in row]
            nombre_al = ''
            documento_al = ''
            errores = []
            for col in col_detectadas:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                if col.get('campo') == 'nombre' and v:
                    nombre_al = v
                elif col.get('campo') == 'documento' and v:
                    documento_al = v
            for col in col_pendientes:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                if v:
                    errores.append(f'Columna "{col["nombre"]}" sin asignar: "{v}"')
            alumno_match = None
            if documento_al:
                for a in alumnos:
                    if str(a['documento']).strip() == documento_al:
                        alumno_match = a
                        break
            if not alumno_match and nombre_al:
                na = nombre_al.strip().lower()
                for a in alumnos:
                    if a['nombre'].strip().lower() == na:
                        alumno_match = a
                        break
            if not alumno_match and nombre_al:
                partes = nombre_al.strip().lower().split()
                for a in alumnos:
                    an = a['nombre'].strip().lower()
                    if all(p in an for p in partes):
                        alumno_match = a
                        break
            ok = alumno_match is not None
            if not ok and not errores:
                errores.append('Estudiante no encontrado en el curso')
            valores = {}
            for col in col_detectadas:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                if col.get('actividad_id') and v:
                    try:
                        vn = v.replace(',', '.')
                        valores[col['indice']] = round(float(vn), 2)
                    except ValueError:
                        valores[col['indice']] = v
                elif col.get('campo') in ('evaluacion','autoevaluacion','proyecto','recuperacion','nota_final') and v:
                    try:
                        vn = v.replace(',', '.')
                        valores[col['indice']] = round(float(vn), 2)
                    except ValueError:
                        valores[col['indice']] = v
                else:
                    valores[col['indice']] = v
            for col in col_pendientes:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                valores[col['indice']] = v if v else None
            filas.append({
                'fila': idx, 'alumno': alumno_match['id'] if alumno_match else None,
                'nombre': alumno_match['nombre'] if alumno_match else (nombre_al or f'Fila {idx}'),
                'ok': ok, 'aid': alumno_match['id'] if alumno_match else None,
                'errores': errores, 'valores': valores
            })
        nuevas_actividades = []
        for col in col_pendientes:
            hl = col['nombre'].strip().lower()
            if hl and hl not in acts_map:
                next_orden = (max((a['orden'] for a in actividades), default=0) + 1 + len(nuevas_actividades))
                nuevas_actividades.append({'nombre': col['nombre'].strip(), 'orden': next_orden})
                col['tipo'] = 'auto'
                col['actividad_nombre'] = col['nombre'].strip()
        validos = sum(1 for f in filas if f['ok'])
        errores_count = sum(1 for f in filas if not f['ok'])
        status = 'ok' if errores_count == 0 else 'warning'
        return jsonify({
            'status': status,
            'columnas': col_detectadas + col_pendientes,
            'columnas_pendientes': col_pendientes,
            'filas': filas,
            'total': len(filas),
            'validos': validos,
            'errores': errores_count,
            'nuevas_actividades': nuevas_actividades,
            'actividades_existentes': [{'id': a['id'], 'nombre': a['nombre']} for a in actividades]
        })
    except Exception as e:
        logger.exception(f'Error analizando archivo: {e}')
        return jsonify({'status':'error','mensaje':'Error al procesar el archivo: '+str(e)}), 500
    finally:
        os.unlink(tmp.name)

@app.route('/<slug>/migrar-excel/confirmar', methods=['POST'])
def migrar_excel_confirmar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    data_json = request.form.get('data', '')
    if not data_json:
        return jsonify({'status':'error','mensaje':'No hay datos para guardar'}), 400
    try:
        data = json.loads(data_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'status':'error','mensaje':'Datos invalidos'}), 400
    conn = conectar(slug)
    try:
        new_act_ids = {}
        for na in data.get('nuevas_actividades', []):
            conn.execute(
                'INSERT INTO actividades (nombre, profesor_id, materia, jornada, curso, orden, periodo) VALUES (?,?,?,?,?,?,?)',
                (na['nombre'], prof['id'], materia, jornada, curso_sel, na['orden'], periodo))
            act_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            new_act_ids[na['nombre']] = act_id
        mapping = data.get('column_mapping', {})
        updated = 0
        for f in data.get('filas', []):
            if not f.get('ok') or not f.get('aid'):
                continue
            aid = f['aid']
            for col in data.get('columnas', []):
                col_idx = str(col['indice'])
                val = f.get('valores', {}).get(col['indice'])
                if val is None or val == '':
                    continue
                try:
                    val_num = round(float(str(val).replace(',', '.')), 2)
                except (ValueError, TypeError):
                    continue
                if col.get('actividad_id'):
                    act_id = col['actividad_id']
                    existing = conn.execute(
                        'SELECT val FROM notas WHERE aid=? AND actividad_id=?', (aid, act_id)).fetchone()
                    old_val = existing['val'] if existing else None
                    if old_val != val_num:
                        conn.execute(
                            '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
                               ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
                            (aid, act_id, val_num))
                        auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'notas', aid,
                                     curso_sel, materia, periodo,
                                     campo='nota', actividad_id=act_id,
                                     valor_anterior=old_val, valor_nuevo=val_num,
                                     motivo='Migracion desde Excel')
                        updated += 1
                elif col.get('campo') in ('evaluacion',):
                    existing = conn.execute(
                        'SELECT evaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['evaluacion'] if existing else None
                    if old_val != val_num:
                        conn.execute(
                            '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,evaluacion,periodo)
                               VALUES (?,?,?,?,?,?)
                               ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                               DO UPDATE SET evaluacion=excluded.evaluacion''',
                            (aid, prof['id'], materia, jornada, val_num, periodo))
                        auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='evaluacion',
                                     valor_anterior=old_val, valor_nuevo=val_num,
                                     motivo='Migracion desde Excel')
                        updated += 1
                elif col.get('campo') in ('autoevaluacion',):
                    existing = conn.execute(
                        'SELECT autoevaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['autoevaluacion'] if existing else None
                    if old_val != val_num:
                        if old_val is not None:
                            conn.execute(
                                '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,autoevaluacion,periodo)
                                   VALUES (?,?,?,?,?,?)
                                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                                   DO UPDATE SET autoevaluacion=excluded.autoevaluacion''',
                                (aid, prof['id'], materia, jornada, val_num, periodo))
                        else:
                            conn.execute(
                                'UPDATE evaluaciones SET autoevaluacion=? WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                                (val_num, aid, prof['id'], materia, jornada, periodo))
                        auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='autoevaluacion',
                                     valor_anterior=old_val, valor_nuevo=val_num,
                                     motivo='Migracion desde Excel')
                        updated += 1
        conn.commit()
    except Exception as e:
        conn.close()
        logger.error(f'Error al confirmar migracion: {e}')
        return jsonify({'status':'error','mensaje':'Error al guardar. Intenta de nuevo.'}), 500
    conn.close()
    return jsonify({'status':'ok', 'mensaje': f'Migracion completada. {updated} valores guardados.', 'updated': updated})

# ── DASHBOARD ──────────────────────────────────────────────────────────────────

def _estadisticas_desc(vals):
    """Compute descriptive statistics for a list of numeric values.
    Returns {media, mediana, moda, desviacion, maximo, minimo, q1, q2, q3, p10, p90} or None for empty."""
    if not vals: return None
    clean = [v for v in vals if v is not None]
    if not clean: return None
    n = len(clean)
    s = sorted(clean)
    media = round(sum(clean) / n, 2)
    # mediana
    if n % 2 == 0:
        mediana = (s[n // 2 - 1] + s[n // 2]) / 2
    else:
        mediana = s[n // 2]
    # moda
    from collections import Counter
    freq = Counter(clean)
    max_f = max(freq.values())
    moda = [k for k, v in freq.items() if v == max_f]
    moda = moda[0] if len(moda) == 1 else None
    # desviacion estandar (poblacional)
    var = sum((x - media) ** 2 for x in clean) / n
    desv = round(var ** 0.5, 2)
    maximo = max(clean)
    minimo = min(clean)
    # cuartiles / percentiles
    def pct(p):
        idx = max(0, min(n - 1, round(n * p / 100)))
        return s[idx]
    q1 = pct(25)
    q2 = mediana
    q3 = pct(75)
    return {
        'media': round(media, 2), 'mediana': round(mediana, 2), 'moda': round(moda, 2) if moda is not None else None,
        'desviacion': desv, 'maximo': round(maximo, 2), 'minimo': round(minimo, 2),
        'q1': round(q1, 2), 'q2': round(q2, 2), 'q3': round(q3, 2),
        'p10': round(pct(10), 2), 'p90': round(pct(90), 2),
    }

def _dashboard_student_grades(conn, slug, profesor_id, materia, jornada, curso=None, periodo=None):
    """Batch compute weighted final grades for all students in scope. Returns list of dicts."""
    if curso:
        alumnos = conn.execute(
            'SELECT id, nombre, curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
    else:
        alumnos = conn.execute(
            '''SELECT a.id, a.nombre, a.curso FROM alumnos a
               JOIN asignaciones_curso ac ON ac.curso=a.curso
               WHERE ac.profesor_id=? AND ac.materia=? AND ac.jornada=?
                 AND a.jornada=? AND a.activo=1 ORDER BY a.nombre''',
            (profesor_id, materia, jornada, jornada)).fetchall()
    if not alumnos: return []
    aids = [a['id'] for a in alumnos]
    ph = ','.join('?' * len(aids))
    notas_rows = conn.execute(
        f'''SELECT n.aid, n.val FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
            WHERE n.aid IN ({ph}) AND ac.profesor_id=? AND ac.materia=? AND ac.jornada=?
            AND (? IS NULL OR ac.periodo=?)''',
        (*aids, profesor_id, materia, jornada, periodo, periodo)).fetchall()
    notas_by_aid = {}
    for r in notas_rows:
        notas_by_aid.setdefault(r['aid'], []).append(r['val'])
    ev_rows = conn.execute(
        f'''SELECT aid, evaluacion, autoevaluacion FROM evaluaciones
            WHERE aid IN ({ph}) AND profesor_id=? AND materia=? AND jornada=?
            AND (? IS NULL OR periodo=?)''',
        (*aids, profesor_id, materia, jornada, periodo, periodo)).fetchall()
    ev_by_aid = {r['aid']: r for r in ev_rows}
    res = []
    for a in alumnos:
        vals = notas_by_aid.get(a['id'], [])
        ev = ev_by_aid.get(a['id'])
        ev_v = ev['evaluacion'] if ev and ev['evaluacion'] is not None else None
        au_v = ev['autoevaluacion'] if ev and ev['autoevaluacion'] is not None else None
        final = _promedio_ponderado(vals, ev_v, au_v)
        res.append({'id': a['id'], 'nombre': a['nombre'], 'curso': a['curso'],
                     'nota_final': final, 'actividades': vals})
    return res

def _dashboard_profesor_data(conn, slug, prof, curso=None, materia=None, jornada=None, periodo=None):
    """Compute full dashboard JSON for a profesor."""
    m = materia or ''
    j = jornada or ''
    cursos_q = [curso] if curso else [r['curso'] for r in conn.execute(
        'SELECT DISTINCT curso FROM asignaciones_curso WHERE profesor_id=? AND materia=? AND jornada=?',
        (prof['id'], m, j)).fetchall()]
    scoped = lambda c: conn.execute(
        'SELECT id, nombre, curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
        (c, j)).fetchall()
    all_alumnos = []
    for c in cursos_q:
        all_alumnos.extend(scoped(c))
    aids = [a['id'] for a in all_alumnos]
    # cards
    total_estudiantes = len(all_alumnos)
    total_actividades = conn.execute(
        'SELECT COUNT(*) FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND (? IS NULL OR curso=?) AND (? IS NULL OR periodo=?)',
        (prof['id'], m, j, curso, curso, periodo, periodo)).fetchone()[0]
    calificadas = conn.execute(
        '''SELECT COUNT(*) FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
           WHERE ac.profesor_id=? AND ac.materia=? AND ac.jornada=?
           AND (? IS NULL OR ac.curso=?) AND (? IS NULL OR ac.periodo=?)''',
        (prof['id'], m, j, curso, curso, periodo, periodo)).fetchone()[0] if aids else 0
    pendientes = max(0, total_actividades * total_estudiantes - calificadas)
    students = _dashboard_student_grades(conn, slug, prof['id'], m, j, curso, periodo) if total_estudiantes else []
    finals = [s['nota_final'] for s in students if s['nota_final'] is not None]
    # config threshold
    cfg = config_get(slug)
    escala_max = float(cfg.get('escala_max', 5.0))
    nota_min_aprobar = float(cfg.get('nota_minima_aprobar', 3.0))
    if escala_max > 5.0:
        nota_min_aprobar = nota_min_aprobar / 2.0  # normalize 1-10 -> 1-5
    aprobados = sum(1 for f in finals if f >= nota_min_aprobar)
    reprobados = sum(1 for f in finals if f < nota_min_aprobar)
    nota_max = max(finals) if finals else None
    nota_min = min(finals) if finals else None
    # grade distribution (raw grades, not weighted)
    dist = {'0-1': 0, '1-2': 0, '2-3': 0, '3-4': 0, '4-5': 0}
    all_vals = conn.execute(
        f'''SELECT n.val FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
            WHERE ac.profesor_id=? AND ac.materia=? AND ac.jornada=?
            AND (? IS NULL OR ac.curso=?) AND (? IS NULL OR ac.periodo=?)''',
        (prof['id'], m, j, curso, curso, periodo, periodo)).fetchall()
    for r in all_vals:
        v = r['val']
        if v < 1: dist['0-1'] += 1
        elif v < 2: dist['1-2'] += 1
        elif v < 3: dist['2-3'] += 1
        elif v < 4: dist['3-4'] += 1
        else: dist['4-5'] += 1
    distribucion = [{'label': k, 'count': v} for k, v in dist.items()]
    # batch-fetch all notas + evaluaciones for ALL periods (used by course avg + evolution)
    _batch_aids = aids or []
    if _batch_aids:
        ph_b = ','.join('?' * len(_batch_aids))
        all_notas_periodos = conn.execute(
            f'''SELECT n.aid, n.val, ac.periodo FROM notas n
                JOIN actividades ac ON ac.id=n.actividad_id
                WHERE n.aid IN ({ph_b}) AND ac.profesor_id=? AND ac.materia=? AND ac.jornada=?''',
            (*_batch_aids, prof['id'], m, j)).fetchall()
        all_ev_periodos = conn.execute(
            f'''SELECT aid, evaluacion, periodo FROM evaluaciones
                WHERE aid IN ({ph_b}) AND profesor_id=? AND materia=? AND jornada=?''',
            (*_batch_aids, prof['id'], m, j)).fetchall()
    else:
        all_notas_periodos = []; all_ev_periodos = []
    # average by course (from batch data, 0 extra queries)
    notas_by_aid_c = {}
    for r in all_notas_periodos:
        notas_by_aid_c.setdefault(r['aid'], []).append(r['val'])
    ev_by_aid_c = {}
    for r in all_ev_periodos:
        ev_by_aid_c[r['aid']] = r['evaluacion']
    prom_curso = []
    for c in cursos_q:
        cur_finals = []
        for a in all_alumnos:
            if a['curso'] != c: continue
            v = notas_by_aid_c.get(a['id'], [])
            e = ev_by_aid_c.get(a['id'])
            ff = _promedio_ponderado(v, e, None)
            if ff is not None: cur_finals.append(ff)
        prom_curso.append({'curso': c, 'promedio': round(sum(cur_finals) / len(cur_finals), 2) if cur_finals else None, 'count': len(cur_finals)})
    # promedio por materia (only current materia for teacher)
    prom_materia = [{'materia': m, 'promedio': round(sum(finals) / len(finals), 2) if finals else None, 'count': len(finals)}]
    # evolution by period (from batch data, 0 extra queries)
    notas_by_aid_p = {}
    for r in all_notas_periodos:
        notas_by_aid_p.setdefault((r['aid'], r['periodo']), []).append(r['val'])
    ev_by_aid_p = {}
    for r in all_ev_periodos:
        ev_by_aid_p[(r['aid'], r['periodo'])] = r['evaluacion']
    evol = []
    for p in range(1, 5):
        finals_p = []
        for a in all_alumnos:
            vals_p = notas_by_aid_p.get((a['id'], p), [])
            ev_p = ev_by_aid_p.get((a['id'], p))
            ff = _promedio_ponderado(vals_p, ev_p, None)
            if ff is not None: finals_p.append(ff)
        evol.append({'periodo': p, 'promedio': round(sum(finals_p) / len(finals_p), 2) if finals_p else None, 'count': len(finals_p)})
    # rendimiento por actividad (batch all grades in 1 query instead of N)
    acts = conn.execute(
        'SELECT id, nombre FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND (? IS NULL OR curso=?) AND (? IS NULL OR periodo=?) ORDER BY orden',
        (prof['id'], m, j, curso, curso, periodo, periodo)).fetchall()
    rend_acts = []
    if acts:
        act_ids = [a['id'] for a in acts]
        ph = ','.join('?' * len(act_ids))
        all_grades = conn.execute(
            f'SELECT actividad_id, val FROM notas WHERE actividad_id IN ({ph})', act_ids).fetchall()
        grades_by_act = {}
        for r in all_grades:
            grades_by_act.setdefault(r['actividad_id'], []).append(r['val'])
        for act in acts:
            vals = grades_by_act.get(act['id'], [])
            cnt = len(vals)
            prom = round(sum(vals) / cnt, 2) if cnt else None
            aprob = sum(1 for v in vals if v >= nota_min_aprobar) if vals else 0
            pct_aprob = round(aprob / cnt * 100, 1) if cnt else None
            rend_acts.append({'actividad': act['nombre'], 'promedio': prom, 'calificadas': cnt, 'porcentaje_aprobacion': pct_aprob})
    # rankings
    top_students = sorted(students, key=lambda s: s['nota_final'] or 0, reverse=True)[:10]
    top_cursos = sorted(prom_curso, key=lambda c: c['promedio'] or 0, reverse=True)
    # alerts
    threshold_bajo = 3.0
    bajo_est = [s for s in students if s['nota_final'] is not None and s['nota_final'] < threshold_bajo]
    bajo_cursos = [c for c in prom_curso if c['promedio'] is not None and c['promedio'] < 3.2]
    bajo_acts = [a for a in rend_acts if a['promedio'] is not None and a['promedio'] < 2.5]
    destacados = [s for s in students if s['nota_final'] is not None and s['nota_final'] > 4.5]
    # statistics
    stats = _estadisticas_desc(finals)
    return {
        'cards': {
            'promedio_curso': round(sum(finals) / len(finals), 2) if finals else None,
            'promedio_materia': round(sum(finals) / len(finals), 2) if finals else None,
            'total_estudiantes': total_estudiantes,
            'total_actividades': total_actividades,
            'actividades_calificadas': calificadas,
            'actividades_pendientes': pendientes,
            'aprobados': aprobados, 'reprobados': reprobados,
            'nota_max': nota_max, 'nota_min': nota_min,
        },
        'charts': {
            'distribucion': distribucion,
            'promedio_por_curso': prom_curso,
            'promedio_por_materia': prom_materia,
            'evolucion_periodos': evol,
            'rendimiento_actividades': rend_acts,
        },
        'rankings': {
            'top_estudiantes': [{'nombre': s['nombre'], 'promedio': s['nota_final']} for s in top_students],
            'top_cursos': top_cursos,
        },
        'alerts': {
            'estudiantes_bajo': [{'nombre': s['nombre'], 'promedio': s['nota_final'], 'curso': s['curso']} for s in bajo_est],
            'cursos_bajo': bajo_cursos,
            'actividades_bajo': bajo_acts,
            'destacados': [{'nombre': s['nombre'], 'promedio': s['nota_final'], 'curso': s['curso']} for s in destacados],
        },
        'estadisticas': stats,
    }

def _dashboard_rector_data(conn, slug, rector):
    """Compute full dashboard JSON for rector using batch queries."""
    # ── card-level aggregates (6 single-row queries) ──
    total_estudiantes = conn.execute('SELECT COUNT(*) FROM alumnos WHERE activo=1').fetchone()[0]
    total_profesores = conn.execute('SELECT COUNT(*) FROM profesores WHERE activo=1').fetchone()[0]
    total_cursos = conn.execute('SELECT COUNT(DISTINCT curso) FROM alumnos WHERE activo=1').fetchone()[0]
    total_materias = conn.execute('SELECT COUNT(DISTINCT materia) FROM asignaciones_materia').fetchone()[0]
    total_actividades = conn.execute('SELECT COUNT(*) FROM actividades').fetchone()[0]
    solicitudes_pend = conn.execute("SELECT COUNT(*) FROM solicitudes_modificacion WHERE estado='pendiente' AND slug=?", (slug,)).fetchone()[0]
    periodos = conn.execute('SELECT periodo, estado FROM periodos_estado').fetchall()
    periodos_abiertos = sum(1 for p in periodos if p['estado'] == 'abierto')
    periodos_cerrados = sum(1 for p in periodos if p['estado'] == 'cerrado')
    cfg = config_get(slug)
    escala_max = float(cfg.get('escala_max', 5.0))
    nota_min_aprobar = float(cfg.get('nota_minima_aprobar', 3.0))
    if escala_max > 5.0:
        nota_min_aprobar /= 2.0

    # ── batch 1: all active students ──
    alumnos = conn.execute(
        'SELECT id, nombre, curso, jornada FROM alumnos WHERE activo=1 ORDER BY id'
    ).fetchall()
    alumno_map = {a['id']: a for a in alumnos}

    # ── batch 2: all active teachers with their subject assignments ──
    profes = conn.execute('SELECT id, nombre FROM profesores WHERE activo=1').fetchall()
    prof_map = {p['id']: p for p in profes}

    asignaciones = conn.execute(
        'SELECT profesor_id, materia, jornada FROM asignaciones_materia'
    ).fetchall()
    prof_subjects = {}
    for a in asignaciones:
        prof_subjects.setdefault(a['profesor_id'], []).append((a['materia'], a['jornada']))

    # ── batch 3: all notas with their subject context ──
    notas_all = conn.execute('''
        SELECT n.aid, n.val, ac.materia, ac.jornada, ac.profesor_id, ac.curso
        FROM notas n
        JOIN actividades ac ON ac.id = n.actividad_id
    ''').fetchall()

    # ── batch 4: all evaluaciones ──
    ev_all = conn.execute(
        'SELECT aid, materia, jornada, evaluacion, autoevaluacion FROM evaluaciones'
    ).fetchall()

    # ── compute final grade for every (student, materia, jornada) pair ──
    notas_idx = {}
    for r in notas_all:
        key = (r['aid'], r['materia'], r['jornada'])
        notas_idx.setdefault(key, []).append(r['val'])
    ev_idx = {}
    for r in ev_all:
        key = (r['aid'], r['materia'], r['jornada'])
        ev_idx[key] = r

    student_subject_grades = {}
    all_keys = set(notas_idx) | set(ev_idx)
    for aid, materia, jornada in all_keys:
        if aid not in alumno_map:
            continue
        vals = notas_idx.get((aid, materia, jornada), [])
        ev = ev_idx.get((aid, materia, jornada))
        ev_v = ev['evaluacion'] if ev and ev['evaluacion'] is not None else None
        au_v = ev['autoevaluacion'] if ev and ev['autoevaluacion'] is not None else None
        final = _promedio_ponderado(vals, ev_v, au_v)
        key = (aid, materia, jornada)
        student_subject_grades[key] = final

    # ── aggregate: per-student average across all subjects ──
    student_avgs = {}
    for (aid, materia, jornada), final in student_subject_grades.items():
        if final is not None:
            student_avgs.setdefault(aid, []).append(final)
    student_overall = {}
    for aid, vals in student_avgs.items():
        student_overall[aid] = round(sum(vals) / len(vals), 2)
    all_finals = list(student_overall.values())

    # ── teacher-level averages ──
    prof_avgs = {}
    for p in profes:
        p_vals = []
        for (aid, materia, jornada), final in student_subject_grades.items():
            if final is not None:
                is_teacher_subject = any(
                    m == materia and j == jornada
                    for m, j in prof_subjects.get(p['id'], [])
                )
                if is_teacher_subject:
                    p_vals.append(final)
        if p_vals:
            prof_avgs[p['nombre']] = round(sum(p_vals) / len(p_vals), 2)

    prom_institucional = round(sum(all_finals) / len(all_finals), 2) if all_finals else None

    # ── course averages ──
    curso_avgs = {}
    for a in alumnos:
        avg = student_overall.get(a['id'])
        if avg is not None:
            curso_avgs.setdefault(a['curso'], []).append(avg)
    curso_avgs = {k: round(sum(v) / len(v), 2) for k, v in curso_avgs.items()}
    mejor_curso = max(curso_avgs, key=curso_avgs.get) if curso_avgs else None
    peor_curso = min(curso_avgs, key=curso_avgs.get) if curso_avgs else None

    # ── subject averages ──
    subj_vals = {}
    for (aid, materia, jornada), final in student_subject_grades.items():
        if final is not None:
            subj_vals.setdefault(materia, []).append(final)
    subj_avgs = {k: round(sum(v) / len(v), 2) for k, v in subj_vals.items()}
    mejor_materia = max(subj_avgs, key=subj_avgs.get) if subj_avgs else None
    peor_materia = min(subj_avgs, key=subj_avgs.get) if subj_avgs else None

    # ── grade distribution (single query with CASE) ──
    dist = {'0-1': 0, '1-2': 0, '2-3': 0, '3-4': 0, '4-5': 0}
    for row in conn.execute('SELECT val FROM notas').fetchall():
        v = row['val']
        if v < 1: dist['0-1'] += 1
        elif v < 2: dist['1-2'] += 1
        elif v < 3: dist['2-3'] += 1
        elif v < 4: dist['3-4'] += 1
        else: dist['4-5'] += 1

    top_docentes = sorted(prof_avgs.items(), key=lambda x: x[1], reverse=True)[:10]

    # ── low-performing students (using computed averages) ──
    bajo_list = []
    for aid, avg in sorted(student_overall.items(), key=lambda x: x[1]):
        if avg < nota_min_aprobar:
            a = alumno_map.get(aid)
            if a:
                bajo_list.append({'nombre': a['nombre'], 'promedio': avg, 'curso': a['curso']})
                if len(bajo_list) >= 20:
                    break

    stats = _estadisticas_desc(all_finals)
    return {
        'cards': {
            'total_estudiantes': total_estudiantes, 'total_profesores': total_profesores,
            'total_cursos': total_cursos, 'total_materias': total_materias,
            'total_actividades': total_actividades,
            'promedio_institucional': prom_institucional,
            'mejor_curso': mejor_curso, 'peor_curso': peor_curso,
            'mejor_materia': mejor_materia, 'peor_materia': peor_materia,
            'solicitudes_pendientes': solicitudes_pend,
            'periodos_abiertos': periodos_abiertos, 'periodos_cerrados': periodos_cerrados,
        },
        'charts': {
            'distribucion': [{'label': k, 'count': v} for k, v in dist.items()],
            'promedio_por_curso': [{'curso': k, 'promedio': v} for k, v in sorted(curso_avgs.items(), key=lambda x: x[1], reverse=True)],
            'promedio_por_materia': [{'materia': k, 'promedio': v} for k, v in sorted(subj_avgs.items(), key=lambda x: x[1], reverse=True)],
            'rendimiento_actividades': [],
        },
        'rankings': {
            'top_estudiantes': [],
            'top_cursos': [{'curso': k, 'promedio': v} for k, v in sorted(curso_avgs.items(), key=lambda x: x[1], reverse=True)[:10]],
            'top_docentes': [{'nombre': n, 'promedio': v} for n, v in top_docentes],
        },
        'alerts': {
            'estudiantes_bajo': bajo_list,
            'cursos_bajo': [{'curso': k, 'promedio': v} for k, v in sorted(curso_avgs.items(), key=lambda x: x[1]) if v < 3.2],
            'destacados': [],
        },
        'estadisticas': stats,
    }

@app.route('/<slug>/dashboard')
def dashboard(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    rector = get_rector(slug)
    if not prof and not rector:
        return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    num_periodos = colegio['num_periodos'] if colegio and 'num_periodos' in colegio.keys() else 4
    conn = conectar(slug)
    if prof:
        jornada, materia = get_sesion_jornada_materia(slug)
        mis_cursos = get_cursos_profesor(slug, prof['id'], materia or '', jornada or '')
        instance = 'profesor'
        nombre = prof['nombre']
    elif rector:
        jornada = ''
        materia = ''
        mis_cursos = [r['curso'] for r in conn.execute(
            'SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()]
        materias_list = [r['materia'] for r in conn.execute(
            'SELECT DISTINCT materia FROM asignaciones_materia ORDER BY materia').fetchall()]
        instance = 'rector'
        nombre = rector['nombre']
    conn.close()
    colegio_dash = get_colegio(slug)
    return render_template('dashboard.html', slug=slug, colegio=colegio_dash, instance=instance, nombre=nombre,
                           num_periodos=num_periodos, mis_cursos=mis_cursos,
                           materias_list=materias_list if rector else [materia],
                           jornada=jornada, materia=materia)

@app.route('/<slug>/dashboard_data')
def dashboard_data(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    rector = get_rector(slug)
    if not prof and not rector:
        return jsonify({'error': 'no_auth'}), 401
    curso = request.args.get('curso') or None
    materia = request.args.get('materia') or None
    jornada_sel = request.args.get('jornada') or None
    periodo = request.args.get('periodo', type=int) or None
    conn = conectar(slug)
    try:
        if prof:
            sess_jornada, sess_materia = get_sesion_jornada_materia(slug)
            m = materia or sess_materia or ''
            j = jornada_sel or sess_jornada or ''
            data = _dashboard_profesor_data(conn, slug, prof, curso, m, j, periodo)
        else:
            data = _dashboard_rector_data(conn, slug, rector)
    finally:
        conn.close()
    return jsonify(data)

# ── AGENDA ────────────────────────────────────────────────────────────────────
@app.route('/<slug>/nuevo_trabajo', methods=['POST'])
def nuevo_trabajo(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso_sel', '')
    conn = conectar(slug)
    conn.execute('INSERT INTO compromisos (titulo,fecha,materia,curso,jornada) VALUES (?,?,?,?,?)',
                 (request.form.get('titulo'), request.form.get('fecha'), materia, curso_sel, jornada))
    conn.commit(); conn.close()
    return redirect(url_for('home', slug=slug, curso=curso_sel))

@app.route('/<slug>/borrar_trabajo/<int:id_t>', methods=['POST'])
def borrar_trabajo(slug, id_t):
    if not validar_csrf(): return redirect(url_for('home', slug=slug))
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    conn = conectar(slug)
    c = conn.execute('SELECT curso FROM compromisos WHERE id=?', (id_t,)).fetchone()
    curso = c['curso'] if c else ''
    conn.execute('DELETE FROM compromisos WHERE id=? AND materia=?', (id_t, materia))
    conn.commit(); conn.close()
    return redirect(url_for('home', slug=slug, curso=curso))

# ── ALUMNOS ───────────────────────────────────────────────────────────────────
@app.route('/<slug>/registrar', methods=['POST'])
def registrar(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    nom       = request.form.get('nombre', '').strip()
    cur       = request.form.get('curso', '').strip()
    curso_sel = request.form.get('curso_sel', cur)
    if nom and cur and jornada:
        conn = conectar(slug)
        with conn:
            conn.execute(
                'INSERT INTO alumnos (nombre,curso,jornada,num_curso,activo) VALUES (?,?,?,0,1)',
                (nom, cur, jornada))
            todos = conn.execute(
                'SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
                (cur, jornada)).fetchall()
            with conn:
                for i, a in enumerate(todos, 1):
                    conn.execute('UPDATE alumnos SET num_curso=? WHERE id=?', (i, a['id']))
        audit_log(slug, prof['id'], 'crear', 'alumnos')
        conn.close()
    return redirect(url_for('home', slug=slug, curso=curso_sel))

@app.route('/<slug>/archivar_alumno/<int:id>', methods=['POST'])
def archivar_alumno(slug, id):
    if not validar_csrf(): return redirect(url_for('home', slug=slug))
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso', '')
    conn = conectar(slug)
    alumno = conn.execute('SELECT curso FROM alumnos WHERE id=?', (id,)).fetchone()
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not alumno or alumno['curso'] not in mis_cursos:
        conn.close(); return ('No autorizado', 403)
    conn.execute('UPDATE alumnos SET activo=0 WHERE id=?', (id,))
    conn.commit(); audit_log(slug, prof['id'], 'archivar', 'alumnos', id)
    conn.close()
    return redirect(url_for('home', slug=slug, curso=curso_sel))

@app.route('/<slug>/reactivar_alumno/<int:id>', methods=['POST'])
def reactivar_alumno(slug, id):
    if not validar_csrf(): return redirect(url_for('archivados', slug=slug))
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso', '')
    conn = conectar(slug)
    alumno = conn.execute('SELECT curso FROM alumnos WHERE id=?', (id,)).fetchone()
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not alumno or alumno['curso'] not in mis_cursos:
        conn.close(); return ('No autorizado', 403)
    conn.execute('UPDATE alumnos SET activo=1 WHERE id=?', (id,))
    conn.commit(); audit_log(slug, prof['id'], 'reactivar', 'alumnos', id)
    conn.close()
    return redirect(url_for('archivados', slug=slug, curso=curso_sel))

@app.route('/<slug>/eliminar_alumno/<int:id>', methods=['POST'])
def eliminar_alumno(slug, id):
    if not validar_csrf(): return redirect(url_for('archivados', slug=slug))
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso', '')
    conn = conectar(slug)
    alumno = conn.execute('SELECT curso FROM alumnos WHERE id=?', (id,)).fetchone()
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not alumno or alumno['curso'] not in mis_cursos:
        conn.close(); return ('No autorizado', 403)
    conn.execute('DELETE FROM alumnos WHERE id=?', (id,))
    conn.execute('DELETE FROM notas WHERE aid=?', (id,))
    conn.execute('DELETE FROM evaluaciones WHERE aid=?', (id,))
    conn.execute('DELETE FROM asistencia WHERE aid=?', (id,))
    conn.execute('DELETE FROM observaciones WHERE aid=?', (id,))
    conn.commit(); audit_log(slug, prof['id'], 'eliminar', 'alumnos', id)
    conn.close()
    return redirect(url_for('archivados', slug=slug, curso=curso_sel))

# ── ARCHIVADOS ────────────────────────────────────────────────────────────────
@app.route('/<slug>/archivados')
def archivados(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    colegio    = get_colegio(slug)
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel  = request.args.get('curso', mis_cursos[0] if mis_cursos else None)
    conn = conectar(slug)
    alumnos_arch = []
    if curso_sel:
        alumnos_arch = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=0 ORDER BY nombre COLLATE NOCASE',
            (curso_sel, jornada)).fetchall()
    profs_arch = conn.execute(
        'SELECT * FROM profesores WHERE activo=0 ORDER BY nombre COLLATE NOCASE').fetchall()
    profs_raw = conn.execute(
        'SELECT * FROM profesores WHERE activo=1 ORDER BY nombre COLLATE NOCASE').fetchall()
    # batch-fetch all assignments in 2 queries instead of N*3
    all_mat = conn.execute(
        'SELECT id, profesor_id, materia, jornada FROM asignaciones_materia ORDER BY jornada, materia').fetchall()
    all_cur = conn.execute(
        'SELECT profesor_id, materia, jornada, curso FROM asignaciones_curso').fetchall()
    # index by profesor_id
    mat_by_prof = {}
    for r in all_mat:
        mat_by_prof.setdefault(r['profesor_id'], []).append(r)
    cur_by_prof_mat_jor = {}
    for r in all_cur:
        cur_by_prof_mat_jor[(r['profesor_id'], r['materia'], r['jornada'])] = r['curso']
    # batch-fetch all other active profs by materia/jornada
    other_profs_raw = conn.execute(
        '''SELECT p2.id, p2.nombre, am.materia, am.jornada
           FROM profesores p2
           JOIN asignaciones_materia am ON am.profesor_id=p2.id
           WHERE p2.activo=1''').fetchall()
    other_by_mat_jor = {}
    for r in other_profs_raw:
        other_by_mat_jor.setdefault((r['materia'], r['jornada']), []).append(r)
    profesores_activos = []
    for p in profs_raw:
        mjs = mat_by_prof.get(p['id'], [])
        cursos_info = []
        for mj in mjs:
            curso_val = cur_by_prof_mat_jor.get((p['id'], mj['materia'], mj['jornada']))
            if curso_val:
                cursos_info.append({'curso': curso_val, 'materia': mj['materia'], 'jornada': mj['jornada']})
        otros_profesores = []
        seen_otros = set()
        for mj in mjs:
            for o in other_by_mat_jor.get((mj['materia'], mj['jornada']), []):
                if o['id'] == p['id']: continue
                entry_key = (o['id'], o['materia'], o['jornada'])
                if entry_key not in seen_otros:
                    seen_otros.add(entry_key)
                    entry = {'id': o['id'], 'nombre': o['nombre'], 'materia': o['materia'], 'jornada': o['jornada']}
                    if entry not in otros_profesores:
                        otros_profesores.append(entry)
        profesores_activos.append({
            'id': p['id'], 'nombre': p['nombre'], 'usuario': p['usuario'],
            'email': p['email'] or '',
            'materias_jornadas': [dict(mj) for mj in mjs],
            'cursos_info': cursos_info,
            'otros_profesores': otros_profesores,
        })
    conn.close()
    return render_template('archivados.html',
                           slug=slug, colegio=colegio, profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel,
                           alumnos_archivados=alumnos_arch,
                           profesores_archivados=profs_arch,
                           profesores_activos=profesores_activos)

@app.route('/<slug>/archivar_profesor/<int:id>', methods=['POST'])
def archivar_profesor(slug, id):
    if not validar_csrf(): return jsonify({'ok': False, 'mensaje': 'Error CSRF'}), 403
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'ok': False, 'mensaje': 'Solo el rector puede archivar profesores'}), 403
    conn = conectar(slug)
    conn.execute('UPDATE profesores SET activo=0 WHERE id=?', (id,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/<slug>/archivar_profesor_con_reasignacion', methods=['POST'])
def archivar_profesor_con_reasignacion(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'ok': False, 'mensaje': 'Solo el rector puede archivar profesores'})
    if not validar_csrf(): return jsonify({'ok': False, 'mensaje': 'Error CSRF'})
    profesor_id      = request.form.get('profesor_id', type=int)
    prof_destino_id  = request.form.get('prof_destino_id', type=int)
    cursos_reasignar = request.form.getlist('cursos_reasignar')
    if not profesor_id:
        return jsonify({'ok': False, 'mensaje': 'Datos incompletos.'})
    conn = conectar(slug)
    try:
        if prof_destino_id and cursos_reasignar:
            for item in cursos_reasignar:
                partes = item.split('|')
                if len(partes) != 3: continue
                curso, mat, jor = partes
                conn.execute(
                    '''UPDATE actividades SET profesor_id=?
                       WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?''',
                    (prof_destino_id, profesor_id, mat, jor, curso))
                conn.execute(
                    '''UPDATE evaluaciones SET profesor_id=?
                       WHERE profesor_id=? AND materia=? AND jornada=?
                       AND aid IN (SELECT id FROM alumnos WHERE curso=? AND jornada=?)''',
                    (prof_destino_id, profesor_id, mat, jor, curso, jor))
                conn.execute(
                    'INSERT OR IGNORE INTO asignaciones_curso (profesor_id,materia,jornada,curso) VALUES (?,?,?,?)',
                    (prof_destino_id, mat, jor, curso))
                conn.execute(
                    'DELETE FROM asignaciones_curso WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?',
                    (profesor_id, mat, jor, curso))
                conn.execute(
                    'INSERT OR IGNORE INTO asignaciones_materia (profesor_id,materia,jornada) VALUES (?,?,?)',
                    (prof_destino_id, mat, jor))
        conn.execute('UPDATE profesores SET activo=0 WHERE id=?', (profesor_id,))
        conn.commit()
        return jsonify({'ok': True, 'mensaje': 'Profesor archivado correctamente.'})
    except Exception as e:
        conn.rollback()
        logger.error(f'Error al archivar profesor {profesor_id} en {slug}: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al archivar. Intenta de nuevo.'})
    finally:
        conn.close()

@app.route('/<slug>/reactivar_profesor/<int:id>', methods=['POST'])
def reactivar_profesor(slug, id):
    if not validar_csrf(): return jsonify({'ok': False, 'mensaje': 'Error CSRF'}), 403
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'ok': False, 'mensaje': 'Solo el rector puede reactivar profesores'}), 403
    conn = conectar(slug)
    conn.execute('UPDATE profesores SET activo=1 WHERE id=?', (id,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/<slug>/eliminar_profesor/<int:id>', methods=['POST'])
def eliminar_profesor(slug, id):
    if not validar_csrf(): return jsonify({'ok': False, 'mensaje': 'Error CSRF'}), 403
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'ok': False, 'mensaje': 'Solo el rector puede eliminar profesores'}), 403
    conn = conectar(slug)
    conn.execute('DELETE FROM profesores WHERE id=?', (id,))
    conn.execute('DELETE FROM asignaciones_materia WHERE profesor_id=?', (id,))
    conn.execute('DELETE FROM asignaciones_curso WHERE profesor_id=?', (id,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

# ── ASISTENCIA HELPERS ──────────────────────────────────────────────────────────
ESTADOS_ASISTENCIA = {'P': 'Presente', 'A': 'Ausente', 'T': 'Tardanza', 'E': 'Excusa', 'X': 'Permiso', 'S': 'Salida anticipada'}
COLORES_ASISTENCIA = {'P': 'green', 'A': 'red', 'T': 'yellow', 'E': 'blue', 'X': 'purple', 'S': 'orange'}

def _asistencia_stats(conn, curso=None, jornada=None, aid=None):
    where = 'WHERE activo=1'
    params = []
    if curso:
        where += ' AND curso=?'; params.append(curso)
    if jornada:
        where += ' AND jornada=?'; params.append(jornada)
    if aid:
        where += ' AND id=?'; params.append(aid)
    stats = {k: 0 for k in ESTADOS_ASISTENCIA}
    stats['total'] = 0
    rows = conn.execute(
        f'SELECT a.id FROM alumnos a {where} ORDER BY a.id', params).fetchall()
    if not rows:
        stats['porcentaje_asistencia'] = 0
        stats['porcentaje_inasistencia'] = 0
        stats['porcentaje_tardanzas'] = 0
        return stats
    aids = [r['id'] for r in rows]
    placeholders = ','.join('?' * len(aids))
    asis_rows = conn.execute(
        f'SELECT estado, COUNT(*) as c FROM asistencia WHERE aid IN ({placeholders}) GROUP BY estado',
        aids).fetchall()
    total = 0
    for ar in asis_rows:
        stats[ar['estado']] = ar['c']
        total += ar['c']
    stats['total'] = total
    stats['porcentaje_asistencia'] = round(stats['P'] / total * 100, 1) if total else 0
    stats['porcentaje_inasistencia'] = round((stats['A'] + stats['E'] + stats['X'] + stats['S']) / total * 100, 1) if total else 0
    stats['porcentaje_tardanzas'] = round(stats['T'] / total * 100, 1) if total else 0
    return stats

def _asistencia_alertas(conn, slug, curso, jornada):
    from collections import defaultdict
    alertas = []
    alumnos = conn.execute(
        'SELECT id, nombre, num_curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
        (curso, jornada)).fetchall()
    if not alumnos:
        return alertas
    aids = [a['id'] for a in alumnos]
    placeholders = ','.join('?' * len(aids))

    # >3 consecutive absences
    abs_consec = conn.execute(
        f'''SELECT aid, fecha FROM asistencia
            WHERE aid IN ({placeholders}) AND estado='A' AND fecha >= date('now','-30 days')
            ORDER BY aid, fecha''', aids).fetchall()
    por_alumno = defaultdict(list)
    for r in abs_consec:
        por_alumno[r['aid']].append(r['fecha'])
    for aid, fechas in por_alumno.items():
        fechas = sorted(set(fechas))
        streak = 1
        max_streak = 1
        for i in range(1, len(fechas)):
            from datetime import datetime as _dt
            diff = (_dt.strptime(fechas[i], '%Y-%m-%d') - _dt.strptime(fechas[i-1], '%Y-%m-%d')).days
            if diff == 1:
                streak += 1
                max_streak = max(max_streak, streak)
            else:
                streak = 1
        if max_streak >= 3:
            alumno = next((a for a in alumnos if a['id'] == aid), None)
            if alumno:
                alertas.append({'aid': aid, 'nombre': alumno['nombre'], 'tipo': 'ausencias_consecutivas', 'detalle': f'{max_streak} ausencias consecutivas', 'severidad': 'alta'})

    # >5 tardanzas
    tardanzas = conn.execute(
        f'SELECT aid, COUNT(*) as c FROM asistencia WHERE aid IN ({placeholders}) AND estado="T" GROUP BY aid',
        aids).fetchall()
    tard_map = {r['aid']: r['c'] for r in tardanzas}
    for aid, c in tard_map.items():
        if c > 5:
            alumno = next((a for a in alumnos if a['id'] == aid), None)
            if alumno:
                alertas.append({'aid': aid, 'nombre': alumno['nombre'], 'tipo': 'tardanzas_excesivas', 'detalle': f'{c} tardanzas registradas', 'severidad': 'media' if c <= 10 else 'alta'})

    # <80% attendance (batch query)
    asis_stats = conn.execute(
        f'SELECT aid, estado, COUNT(*) as c FROM asistencia WHERE aid IN ({placeholders}) GROUP BY aid, estado',
        aids).fetchall()
    stats_por_aid = {}
    for r in asis_stats:
        stats_por_aid.setdefault(r['aid'], {})[r['estado']] = r['c']
    for alumno in alumnos:
        s = stats_por_aid.get(alumno['id'], {})
        total = sum(s.values())
        if total > 0:
            pct = round((s.get('P', 0) + s.get('X', 0)) / total * 100)
            if pct < 80:
                alertas.append({'aid': alumno['id'], 'nombre': alumno['nombre'], 'tipo': 'baja_asistencia', 'detalle': f'{pct}% asistencia', 'severidad': 'alta'})

    return alertas

# ── ASISTENCIA (PROFESOR) ──────────────────────────────────────────────────────
@app.route('/<slug>/asistencia', methods=['GET'])
def asistencia(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('seleccionar_jornada', slug=slug))
    colegio    = get_colegio(slug)
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel  = request.args.get('curso', mis_cursos[0] if mis_cursos else None)
    fecha_sel  = request.args.get('fecha', datetime.today().strftime('%Y-%m-%d'))
    try:
        fecha_dt = datetime.strptime(fecha_sel, '%Y-%m-%d') if fecha_sel else datetime.today()
    except ValueError:
        fecha_sel = datetime.today().strftime('%Y-%m-%d')
        fecha_dt  = datetime.today()
    fecha_sel_dia_anterior  = (fecha_dt - timedelta(days=1)).strftime('%Y-%m-%d')
    fecha_sel_dia_siguiente = (fecha_dt + timedelta(days=1)).strftime('%Y-%m-%d')
    hoy_fecha  = datetime.today().strftime('%Y-%m-%d')
    hoy_hora   = datetime.today().strftime('%H:%M')
    if not curso_sel:
        return render_template('asistencia.html', profesor=prof, slug=slug, colegio=colegio,
                               materia=materia, jornada=jornada, mis_cursos=mis_cursos,
                               curso_sel=None, estudiantes=[], fecha_sel=fecha_sel,
                               fecha_sel_dia_anterior=fecha_sel_dia_anterior,
                               fecha_sel_dia_siguiente=fecha_sel_dia_siguiente,
                               hoy_fecha=hoy_fecha, hoy_hora=hoy_hora,
                               estados_asistencia=ESTADOS_ASISTENCIA,
                               colores_asistencia=COLORES_ASISTENCIA,
                               materias_jornadas=get_materias_profesor(slug, prof['id']))
    conn = conectar(slug)
    try:
        alumnos = conn.execute(
            'SELECT id, nombre, num_curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
            (curso_sel, jornada)).fetchall()
        asis_rows = []
        if alumnos:
            placeholders = ','.join('?' * len(alumnos))
            aid_tuple = tuple(a['id'] for a in alumnos)
            asis_rows = conn.execute(
                f'SELECT aid, estado, observacion, hora FROM asistencia WHERE fecha=? AND aid IN ({placeholders})',
                (fecha_sel,) + aid_tuple).fetchall()
        asis_map = {r['aid']: {'estado': r['estado'], 'observacion': r['observacion'] or '', 'hora': r['hora'] or ''} for r in asis_rows}
        datos = []
        for a in alumnos:
            info = asis_map.get(a['id'], {})
            datos.append({
                'id': a['id'], 'nombre': a['nombre'],
                'num_curso': a['num_curso'],
                'asistencia': info.get('estado', ''),
                'observacion': info.get('observacion', ''),
                'hora': info.get('hora', ''),
            })
        stats = _asistencia_stats(conn, curso=curso_sel, jornada=jornada)
    finally:
        conn.close()
    return render_template('asistencia.html', profesor=prof, slug=slug, colegio=colegio,
                           materia=materia, jornada=jornada, mis_cursos=mis_cursos,
                           curso_sel=curso_sel, estudiantes=datos, fecha_sel=fecha_sel,
                           fecha_sel_dia_anterior=fecha_sel_dia_anterior,
                           fecha_sel_dia_siguiente=fecha_sel_dia_siguiente,
                           hoy_fecha=hoy_fecha, hoy_hora=hoy_hora,
                           stats=stats,
                           estados_asistencia=ESTADOS_ASISTENCIA,
                           colores_asistencia=COLORES_ASISTENCIA,
                           materias_jornadas=get_materias_profesor(slug, prof['id']))

@app.route('/<slug>/marcar_asistencia', methods=['POST'])
def marcar_asistencia(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    aid    = request.form.get('aid', type=int)
    estado = request.form.get('estado')
    fecha  = request.form.get('fecha', '')
    observacion = request.form.get('observacion', '').strip()
    hora  = request.form.get('hora', '')
    if aid is None or not estado: return ('', 400)
    if estado not in ESTADOS_ASISTENCIA: return ('', 400)
    if fecha:
        try:
            datetime.strptime(fecha, '%Y-%m-%d')
        except ValueError:
            return ('', 400)
    jornada, materia = get_sesion_jornada_materia(slug)
    conn = conectar(slug)
    cursos_prof = get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not cursos_prof:
        conn.close(); return ('', 403)
    alumno = conn.execute(
        'SELECT id FROM alumnos WHERE id=? AND curso IN ({}) AND jornada=? AND activo=1'.format(
            ','.join('?' * len(cursos_prof))),
        (aid, *cursos_prof, jornada)).fetchone()
    if not alumno:
        conn.close(); return ('', 403)
    if fecha:
        conn.execute('''INSERT INTO asistencia (aid,fecha,estado,observacion,hora,usuario_tipo,usuario_id)
                        VALUES (?,?,?,?,?,?,?)
                        ON CONFLICT(aid,fecha) DO UPDATE SET estado=excluded.estado,
                                                             observacion=excluded.observacion,
                                                             hora=excluded.hora,
                                                             usuario_tipo=excluded.usuario_tipo,
                                                             usuario_id=excluded.usuario_id''',
                     (aid, fecha, estado, observacion, hora, 'profesor', prof['id']))
    else:
        conn.execute('''INSERT INTO asistencia (aid,fecha,estado,observacion,hora,usuario_tipo,usuario_id)
                        VALUES (?,date("now"),?,?,?,?,?)
                        ON CONFLICT(aid,fecha) DO UPDATE SET estado=excluded.estado,
                                                             observacion=excluded.observacion,
                                                             hora=excluded.hora,
                                                             usuario_tipo=excluded.usuario_tipo,
                                                             usuario_id=excluded.usuario_id''',
                     (aid, estado, observacion, hora, 'profesor', prof['id']))
    conn.commit()
    audit_log(slug, prof['id'], 'asistencia_editada', 'asistencia', aid,
              None, {'estado': estado, 'observacion': observacion, 'hora': hora})
    conn.close()
    return jsonify({'status':'ok'})

# ── ASISTENCIA DATA (AJAX) ──────────────────────────────────────────────────────
@app.route('/<slug>/asistencia_data')
def asistencia_data(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof or not validar_csrf(): return jsonify({'error': 'No autorizado'}), 403
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia: return jsonify({'error': 'Sin jornada/materia'}), 400
    conn = conectar(slug)
    try:
        curso = request.args.get('curso', '')
        if not curso:
            conn.close(); return jsonify({'error': 'Curso requerido'}), 400
        stats = _asistencia_stats(conn, curso=curso, jornada=jornada)
        alertas = _asistencia_alertas(conn, slug, curso, jornada)
        from collections import defaultdict
        MESES = {'01':'Ene','02':'Feb','03':'Mar','04':'Abr','05':'May','06':'Jun',
                 '07':'Jul','08':'Ago','09':'Sep','10':'Oct','11':'Nov','12':'Dic'}
        alumnos = conn.execute(
            'SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1',
            (curso, jornada)).fetchall()
        aids = tuple(a['id'] for a in alumnos)
        calendario = defaultdict(lambda: defaultdict(int))
        if aids:
            placeholders = ','.join('?' * len(aids))
            rows = conn.execute(
                f'SELECT fecha, estado FROM asistencia WHERE aid IN ({placeholders}) ORDER BY fecha',
                aids).fetchall()
            for r in rows:
                calendario[r['fecha']][r['estado']] += 1
    finally:
        conn.close()
    return jsonify({
        'stats': stats,
        'alertas': alertas,
        'calendario': {k: dict(v) for k, v in calendario.items()},
        'estados': dict(ESTADOS_ASISTENCIA),
        'colores': COLORES_ASISTENCIA,
    })

# ── ASISTENCIA REPORTE EXCEL ──────────────────────────────────────────────────
@app.route('/<slug>/asistencia_reporte_excel')
def asistencia_reporte_excel(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia: return ('', 400)
    conn = conectar(slug)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = 'Asistencia'
        hd_font = Font(bold=True, color='FFFFFF', size=11)
        hd_fill = PatternFill('solid', fgColor='1E293B')
        thin = Side(style='thin', color='334155')
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        curso = request.args.get('curso', '')
        if not curso:
            conn.close(); return ('Curso requerido', 400)
        alumnos = conn.execute(
            'SELECT id, nombre, num_curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
        if not alumnos:
            conn.close(); return ('Sin estudiantes', 404)
        aids = [a['id'] for a in alumnos]
        placeholder = ','.join('?' * len(aids))
        asis_rows = conn.execute(
            f'SELECT aid, fecha, estado, observacion FROM asistencia WHERE aid IN ({placeholder}) ORDER BY aid, fecha',
            aids).fetchall()
        fechas = sorted(set(r['fecha'] for r in asis_rows))
        header = ['#', 'Estudiante'] + fechas
        ws.append(header)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hd_font; cell.fill = hd_fill; cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        asis_map = {}
        for r in asis_rows:
            asis_map.setdefault(r['aid'], {})[r['fecha']] = {'estado': r['estado'], 'obs': r['observacion'] or ''}
        for i, a in enumerate(alumnos, start=2):
            ws.append([a['num_curso'], a['nombre']] + [asis_map.get(a['id'], {}).get(f, {}).get('estado', '') for f in fechas])
            for c in range(1, len(header) + 1):
                ws.cell(row=i, column=c).border = border
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for ci in range(3, len(header) + 1):
            ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = 7
    finally:
        conn.close()
    from flask import Response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=asistencia_{slug}_{curso}.xlsx'})

# ── OBSERVACIONES ─────────────────────────────────────────────────────────────
@app.route('/<slug>/agregar_observacion', methods=['POST'])
def agregar_observacion(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    texto = request.form.get('texto', '').strip()
    aid   = request.form.get('aid', type=int)
    if not texto or aid is None: return ('', 400)
    conn = conectar(slug)
    cursos_prof = get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not cursos_prof:
        conn.close(); return ('', 403)
    alumno = conn.execute(
        'SELECT id FROM alumnos WHERE id=? AND curso IN ({}) AND jornada=? AND activo=1'.format(
            ','.join('?' * len(cursos_prof))),
        (aid, *cursos_prof, jornada)).fetchone()
    if not alumno:
        conn.close(); return ('', 403)
    conn.execute('INSERT INTO observaciones (aid,materia,texto,fecha) VALUES (?,?,?,date("now"))',
                 (aid, materia, texto))
    conn.commit()
    obs = conn.execute(
        'SELECT id, materia, texto, fecha FROM observaciones WHERE aid=? AND materia=? ORDER BY id DESC LIMIT 1',
        (aid, materia)).fetchone()
    audit_log(slug, prof['id'], 'observacion_creada', 'observaciones', registro_id=obs['id'],
              valor_anterior=None, valor_nuevo={'aid': aid, 'texto': texto})
    conn.close()
    return jsonify({'id': obs['id'], 'materia': obs['materia'],
                    'texto': obs['texto'], 'fecha': obs['fecha']})

@app.route('/<slug>/editar_observacion/<int:id_o>', methods=['POST'])
def editar_observacion(slug, id_o):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    texto = request.form.get('texto', '').strip()
    if not texto: return ('', 400)
    conn = conectar(slug)
    obs = conn.execute(
        'SELECT id, aid, materia, texto, fecha FROM observaciones WHERE id=? AND materia=?',
        (id_o, materia)).fetchone()
    if not obs:
        conn.close(); return ('', 404)
    old_text = obs['texto']
    conn.execute('UPDATE observaciones SET texto=? WHERE id=?', (texto, id_o))
    conn.commit()
    audit_log(slug, prof['id'], 'observacion_editada', 'observaciones', registro_id=id_o,
              valor_anterior={'texto': old_text}, valor_nuevo={'texto': texto})
    conn.close()
    return jsonify({'id': obs['id'], 'aid': obs['aid'], 'materia': obs['materia'],
                    'texto': texto, 'fecha': obs['fecha']})

@app.route('/<slug>/borrar_observacion/<int:id_o>', methods=['POST'])
def borrar_observacion(slug, id_o):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = get_sesion_jornada_materia(slug)
    conn = conectar(slug)
    obs = conn.execute('SELECT id, aid, materia, texto FROM observaciones WHERE id=?', (id_o,)).fetchone()
    if obs and obs['materia'] == materia:
        conn.execute('DELETE FROM observaciones WHERE id=?', (id_o,))
        conn.commit()
        audit_log(slug, prof['id'], 'observacion_eliminada', 'observaciones', registro_id=id_o,
                  valor_anterior={'aid': obs['aid'], 'texto': obs['texto']}, valor_nuevo=None)
    conn.close()
    return jsonify({'ok': True})

# ── PERFIL / CURSOS ───────────────────────────────────────────────────────────
# ── cambiar_password migrated to app/routes/auth.py ──────────────────────

@app.route('/<slug>/agregar_cursos', methods=['POST'])
def agregar_cursos(slug):
    if not validar_csrf():
        return 'Error de seguridad', 400
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    cursos = request.form.getlist('cursos')
    extra  = request.form.get('cursos_extra', '').strip()
    if extra: cursos += [c.strip() for c in extra.split(',') if c.strip()]
    conn = conectar(slug)
    for c in cursos:
        if c:
            conn.execute(
                'INSERT OR IGNORE INTO asignaciones_curso (profesor_id,materia,jornada,curso) VALUES (?,?,?,?)',
                (prof['id'], materia, jornada, c))
    conn.commit(); conn.close()
    return redirect(url_for('cambiar_password', slug=slug))

@app.route('/<slug>/quitar_curso/<curso>', methods=['POST'])
def quitar_curso(slug, curso):
    if not validar_csrf(): return redirect(url_for('cambiar_password', slug=slug))
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    conn = conectar(slug)
    conn.execute(
        'DELETE FROM asignaciones_curso WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?',
        (prof['id'], materia, jornada, curso))
    conn.commit(); conn.close()
    return redirect(url_for('cambiar_password', slug=slug))

# ── TRANSFERIR CURSO ──────────────────────────────────────────────────────────
@app.route('/<slug>/transferir_curso', methods=['GET', 'POST'])
def transferir_curso(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('seleccionar_jornada', slug=slug))
    colegio    = get_colegio(slug)
    error = exito = None
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
    conn = conectar(slug)
    profesores_destino = conn.execute(
        '''SELECT p.id, p.nombre FROM profesores p
           JOIN asignaciones_materia am ON am.profesor_id=p.id
           WHERE am.materia=? AND am.jornada=? AND p.id!=? AND p.activo=1
           ORDER BY p.nombre''',
        (materia, jornada, prof['id'])).fetchall()
    num_periodos = int(colegio['num_periodos']) if colegio and colegio['num_periodos'] else 4
    conn.close()

    if request.method == 'POST':
        if not validar_csrf():
            error = 'Error de seguridad.'
        else:
            accion           = request.form.get('accion', '')
            curso_transferir = request.form.get('curso', '').strip()
            periodos_str     = request.form.get('periodos', '')
            periodos         = [p for p in periodos_str.split(',') if p.strip()]
            if not curso_transferir:
                error = 'Selecciona un curso.'
            elif not periodos:
                error = 'Selecciona al menos un periodo.'
            elif curso_transferir not in mis_cursos:
                error = 'Ese curso no te pertenece.'
            else:
                conn = conectar(slug)
                if accion == 'transferir':
                    prof_destino_id = request.form.get('profesor_destino_id', type=int)
                    if not prof_destino_id:
                        error = 'Selecciona un profesor destino.'; conn.close()
                    else:
                        for p in periodos:
                            p = int(p)
                            conn.execute(
                                '''UPDATE actividades SET profesor_id=?
                                   WHERE profesor_id=? AND materia=? AND jornada=?
                                   AND curso=? AND COALESCE(periodo,1)=?''',
                                (prof_destino_id, prof['id'], materia, jornada, curso_transferir, p))
                            conn.execute(
                                '''UPDATE evaluaciones SET profesor_id=?
                                   WHERE profesor_id=? AND materia=? AND jornada=?
                                   AND COALESCE(periodo,1)=?
                                   AND aid IN (SELECT id FROM alumnos WHERE curso=? AND jornada=?)''',
                                (prof_destino_id, prof['id'], materia, jornada,
                                 p, curso_transferir, jornada))
                        conn.execute(
                            'INSERT OR IGNORE INTO asignaciones_curso (profesor_id,materia,jornada,curso) VALUES (?,?,?,?)',
                            (prof_destino_id, materia, jornada, curso_transferir))
                        conn.execute(
                            'DELETE FROM asignaciones_curso WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?',
                            (prof['id'], materia, jornada, curso_transferir))
                        conn.commit(); conn.close()
                        exito = f'✅ Curso {curso_transferir} transferido correctamente.'
                        mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
                elif accion == 'archivar_curso':
                    conn.execute(
                        'DELETE FROM asignaciones_curso WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?',
                        (prof['id'], materia, jornada, curso_transferir))
                    conn.commit(); conn.close()
                    exito = f'✅ Curso {curso_transferir} archivado.'
                    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada)
                else:
                    conn.close(); error = 'Acción no reconocida.'

    return render_template('transferir_curso.html',
                           slug=slug, colegio=colegio, profesor=prof,
                           mis_cursos=mis_cursos, profesores_destino=profesores_destino,
                           error=error, exito=exito, materia=materia, jornada=jornada,
                           num_periodos=num_periodos)

# ── HORARIOS ──────────────────────────────────────────────────────────────────
DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']

@app.route('/<slug>/horarios', methods=['GET', 'POST'])
def horarios(slug):
    require_colegio(slug)
    if not session.get(f'rol_{slug}'): return redirect(url_for('auth.login', slug=slug))
    prof    = get_profesor(slug)
    colegio = get_colegio(slug)
    jornada, materia = get_sesion_jornada_materia(slug)
    if prof and (not jornada or not materia):
        return redirect(url_for('seleccionar_jornada', slug=slug))
    mis_cursos = get_cursos_profesor(slug, prof['id'], materia, jornada) if prof else []
    curso_sel  = request.args.get('curso', mis_cursos[0] if mis_cursos else None)

    if request.method == 'POST':
        if not validar_csrf():
            return redirect(url_for('horarios', slug=slug))
        c = conectar(slug)
        try:
            dia      = request.form.get('dia', '')
            franja   = request.form.get('franja', '')
            num      = request.form.get('num', '').strip()
            mat      = request.form.get('materia', '').strip()
            profesor = request.form.get('profesor', '').strip()
            curso_p  = request.form.get('curso', curso_sel)
            if mat or profesor:
                c.execute(
                    '''INSERT INTO horarios_curso (curso,jornada,dia,franja,num,materia,profesor)
                       VALUES (?,?,?,?,?,?,?)
                       ON CONFLICT(curso,jornada,dia,franja) DO UPDATE SET
                           num=excluded.num, materia=excluded.materia, profesor=excluded.profesor''',
                    (curso_p, jornada, dia, franja, num, mat, profesor))
            else:
                c.execute(
                    'DELETE FROM horarios_curso WHERE curso=? AND jornada=? AND dia=? AND franja=?',
                    (curso_p, jornada, dia, franja))
            c.commit()
        finally:
            c.close()
        return ('', 204)
    c = conectar(slug)
    filas = []
    if curso_sel:
        filas = c.execute(
            'SELECT dia, franja, num, materia, profesor FROM horarios_curso WHERE curso=? AND jornada=?',
            (curso_sel, jornada)).fetchall()
    c.close()
    horario_map = {
        f"{r['dia']}_{r['franja']}": {'num': r['num'], 'materia': r['materia'], 'profesor': r['profesor']}
        for r in filas}
    return render_template('horarios.html', slug=slug, colegio=colegio, profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel, dias=DIAS_SEMANA,
                           horario_map=horario_map, jornada=jornada)

# ── ESTUDIANTE ────────────────────────────────────────────────────────────────
@app.route('/<slug>/estudiante')
def vista_estudiante(slug):
    require_colegio(slug)
    if session.get(f'rol_{slug}') != 'estudiante':
        return redirect(url_for('auth.login', slug=slug))
    aid     = session.get(f'alumno_id_{slug}')
    colegio = get_colegio(slug)
    conn    = conectar(slug)
    alumno  = conn.execute('SELECT * FROM alumnos WHERE id=? AND activo=1', (aid,)).fetchone()
    if not alumno:
        conn.close()
        session.pop(f'rol_{slug}', None)
        session.pop(f'alumno_id_{slug}', None)
        return redirect(url_for('auth.login', slug=slug))
    agenda = conn.execute(
        'SELECT * FROM compromisos WHERE curso=? AND jornada=? ORDER BY fecha, materia',
        (alumno['curso'], alumno['jornada'])).fetchall()
    periodo = request.args.get('periodo', 1, type=int)
    notas_raw = conn.execute(
        '''SELECT ac.materia, ac.nombre as act_nombre, n.val
           FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
           WHERE n.aid=? AND ac.curso=? AND ac.jornada=?
           AND COALESCE(ac.periodo,1)=?
           ORDER BY ac.materia, ac.orden''',
        (aid, alumno['curso'], alumno['jornada'], periodo)).fetchall()
    evals_raw = conn.execute(
        'SELECT materia, evaluacion, autoevaluacion FROM evaluaciones WHERE aid=? AND COALESCE(periodo,1)=?',
        (aid, periodo)).fetchall()
    evals_map = {e['materia']: dict(e) for e in evals_raw}
    notas_pm = {}
    for nr in notas_raw:
        notas_pm.setdefault(nr['materia'], []).append({'actividad': nr['act_nombre'], 'val': nr['val']})
    for mat in evals_map:
        if mat not in notas_pm: notas_pm[mat] = []
    proms_pm = {}
    todos_finales = []
    for mat, notas in notas_pm.items():
        notas_vals = [n['val'] for n in notas]
        ev = evals_map.get(mat, {})
        eval_v = ev.get('evaluacion') if ev.get('evaluacion') is not None else None
        auto_v = ev.get('autoevaluacion') if ev.get('autoevaluacion') is not None else None
        prom = _promedio_ponderado(notas_vals, eval_v, auto_v)
        proms_pm[mat] = prom
        if prom is not None: todos_finales.append(prom)
    promedio_general = round(sum(todos_finales) / len(todos_finales), 2) if todos_finales else None
    asist_raw   = conn.execute(
        'SELECT fecha, estado, observacion FROM asistencia WHERE aid=? ORDER BY fecha', (aid,)).fetchall()
    asist_stats = {k: 0 for k in ESTADOS_ASISTENCIA}
    asist_stats['total'] = 0
    MESES = {'01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr', '05': 'May', '06': 'Jun',
             '07': 'Jul', '08': 'Ago', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'}
    historial_meses = {}
    for h in asist_raw:
        asist_stats[h['estado']] = asist_stats.get(h['estado'], 0) + 1
        asist_stats['total'] += 1
        if h['fecha']:
            p = h['fecha'].split('-')
            if len(p) >= 2:
                label = f"{MESES.get(p[1], p[1])} {p[0]}"
                historial_meses.setdefault(label, []).append({'fecha': h['fecha'], 'estado': h['estado'], 'observacion': h['observacion'] or ''})
    total = asist_stats['total']
    asist_stats['porcentaje_asistencia'] = round(asist_stats['P'] / total * 100, 1) if total else 0
    asist_stats['porcentaje_inasistencia'] = round((asist_stats['A'] + asist_stats['E'] + asist_stats['X'] + asist_stats['S']) / total * 100, 1) if total else 0
    asist_stats['porcentaje_tardanzas'] = round(asist_stats['T'] / total * 100, 1) if total else 0
    observaciones = conn.execute(
        'SELECT materia, texto, fecha FROM observaciones WHERE aid=? ORDER BY fecha DESC', (aid,)).fetchall()
    horario_raw = conn.execute(
        'SELECT dia, franja, num, materia, profesor FROM horarios_curso WHERE curso=? AND jornada=?',
        (alumno['curso'], alumno['jornada'])).fetchall()
    horario_map = {
        f"{r['dia']}_{r['franja']}": {'num': r['num'], 'materia': r['materia'], 'profesor': r['profesor']}
        for r in horario_raw}
    conn.close()
    pendientes = comunicaciones_pendientes(slug, 'estudiante', aid)
    return render_template('estudiante.html',
                           alumno=alumno, slug=slug, colegio=colegio, agenda=agenda,
                           notas_por_materia=notas_pm, evals_map=evals_map,
                           proms_por_materia=proms_pm,
                           promedio_general=promedio_general,
                           asist_stats=asist_stats, historial_meses=historial_meses,
                           observaciones=observaciones, horario_map=horario_map,
                           comunicaciones_pendientes=pendientes)

# ── DIRECTORA ─────────────────────────────────────────────────────────────────
def get_directora(slug):
    cache_key = f'_direc_{slug}'
    if hasattr(g, cache_key): return getattr(g, cache_key)
    did = session.get(f'directora_id_{slug}')
    if not did: setattr(g, cache_key, None); return None
    conn = conectar(slug)
    d = conn.execute('SELECT * FROM directoras WHERE id=? AND activo=1', (did,)).fetchone()
    conn.close()
    if not d: session.pop(f'directora_id_{slug}', None)
    setattr(g, cache_key, d)
    return d

def get_rector(slug):
    cache_key = f'_rector_{slug}'
    if hasattr(g, cache_key): return getattr(g, cache_key)
    rid = session.get(f'rector_id_{slug}')
    if not rid: setattr(g, cache_key, None); return None
    conn = conectar(slug)
    r = conn.execute('SELECT * FROM rectores WHERE id=? AND activo=1', (rid,)).fetchone()
    conn.close()
    if not r: session.pop(f'rector_id_{slug}', None)
    setattr(g, cache_key, r)
    return r

# ── Rector auth routes migrated to app/routes/auth.py ───────────────────

@app.route('/<slug>/rector')
@app.route('/<slug>/rector/panel')
def rector_panel(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    conn = conectar(slug)
    total_est = conn.execute(
        'SELECT COUNT(*) as c FROM alumnos WHERE activo=1').fetchone()['c']
    total_prof = conn.execute(
        'SELECT COUNT(*) as c FROM profesores WHERE activo=1').fetchone()['c']
    total_cursos = conn.execute(
        'SELECT COUNT(DISTINCT curso) as c FROM alumnos WHERE activo=1').fetchone()['c']
    total_materias = conn.execute(
        'SELECT COUNT(DISTINCT materia) as c FROM asignaciones_materia').fetchone()['c']
    total_directoras = conn.execute(
        'SELECT COUNT(*) as c FROM directoras WHERE activo=1').fetchone()['c']
    hoy = datetime.today().strftime('%Y-%m-%d')
    asistencia_hoy = conn.execute(
        "SELECT COUNT(DISTINCT aid) as c FROM asistencia WHERE fecha=?", (hoy,)).fetchone()['c']
    comunicaciones = conn.execute(
        '''SELECT * FROM comunicaciones WHERE rector_id=? AND activo=1
           ORDER BY fecha_creacion DESC LIMIT 5''',
        (rector['id'],)).fetchall()
    notif_count = conn.execute(
        'SELECT COUNT(*) as c FROM notificaciones WHERE usuario_tipo=? AND usuario_id=? AND leida=0',
        ('rector', rector['id'])).fetchone()['c']

    # ── Dashboard extras ─────────────────────────────────────────────────────
    actividad_reciente = conn.execute(
        '''SELECT accion, tabla, creado
           FROM audit_log ORDER BY creado DESC LIMIT 8''').fetchall()
    actividad_reciente = [dict(r) for r in actividad_reciente]

    ultimos_estudiantes = conn.execute(
        '''SELECT id, nombre, curso, jornada FROM alumnos
           WHERE activo=1 ORDER BY id DESC LIMIT 5''').fetchall()
    ultimos_estudiantes = [dict(r) for r in ultimos_estudiantes]

    ultimos_profesores = conn.execute(
        '''SELECT id, nombre, email FROM profesores
           WHERE activo=1 ORDER BY id DESC LIMIT 5''').fetchall()
    ultimos_profesores = [dict(r) for r in ultimos_profesores]

    proximos_eventos = conn.execute(
        '''SELECT titulo, fecha, materia, curso, jornada
           FROM compromisos WHERE fecha >= ?
           ORDER BY fecha LIMIT 5''', (hoy,)).fetchall()
    proximos_eventos = [dict(r) for r in proximos_eventos]

    conn.close()
    prom_general = 0
    return render_template('rector_panel.html',
                           slug=slug, colegio=colegio, rector=rector,
                           total_estudiantes=total_est,
                           total_profesores=total_prof,
                           total_cursos=total_cursos,
                           total_materias=total_materias,
                           total_directoras=total_directoras,
                           asistencia_hoy=asistencia_hoy,
                           comunicaciones=comunicaciones,
                           notif_count=notif_count,
                           actividad_reciente=actividad_reciente,
                           ultimos_estudiantes=ultimos_estudiantes,
                           ultimos_profesores=ultimos_profesores,
                           proximos_eventos=proximos_eventos)

# ── Rector logout migrated to app/routes/auth.py ────────────────────────

# ── RECTOR: HORARIOS ───────────────────────────────────────────────────────────
@app.route('/<slug>/rector/horarios')
def rector_horarios(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    conn = conectar(slug)
    cursos = [r['curso'] for r in conn.execute(
        'SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()]
    jornadas = JORNADAS
    conn.close()
    return render_template('rector_horarios.html',
                           slug=slug, colegio=colegio, rector=rector,
                           cursos=cursos, jornadas=jornadas,
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

@app.route('/<slug>/rector/horarios/datos')
def rector_horarios_datos(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({})
    curso = request.args.get('curso', '')
    jornada = request.args.get('jornada', '')
    if not curso: return jsonify({})
    conn = conectar(slug)
    filas = conn.execute(
        'SELECT dia, franja, num, materia, profesor FROM horarios_curso WHERE curso=? AND jornada=?',
        (curso, jornada)).fetchall()
    conn.close()
    mapa = {}
    for r in filas:
        mapa[f"{r['dia']}_{r['franja']}"] = {'num': r['num'], 'materia': r['materia'], 'profesor': r['profesor']}
    return jsonify(mapa)

# ── RECTOR: PROFESORES ─────────────────────────────────────────────────────────
@app.route('/<slug>/rector/profesores')
def rector_profesores(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    per_page = min(max(per_page, 10), 200)
    conn = conectar(slug)
    total = conn.execute(
        'SELECT COUNT(*) as c FROM profesores').fetchone()['c']
    profesores = [dict(r) for r in conn.execute(
        'SELECT id, nombre, email, activo FROM profesores ORDER BY nombre LIMIT ? OFFSET ?',
        (per_page, (page - 1) * per_page)).fetchall()]
    conn.close()
    return render_template('rector_profesores.html',
                           slug=slug, colegio=colegio, rector=rector,
                           profesores=profesores, page=page, per_page=per_page,
                           total=total, total_pages=(total + per_page - 1) // per_page,
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

# ── RECTOR: ESTUDIANTES ────────────────────────────────────────────────────────
@app.route('/<slug>/rector/estudiantes')
def rector_estudiantes(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    per_page = min(max(per_page, 10), 200)
    conn = conectar(slug)
    total = conn.execute(
        'SELECT COUNT(*) as c FROM alumnos WHERE activo=1').fetchone()['c']
    estudiantes = [dict(r) for r in conn.execute(
        '''SELECT id, nombre, curso, jornada FROM alumnos WHERE activo=1
           ORDER BY curso, nombre LIMIT ? OFFSET ?''',
        (per_page, (page - 1) * per_page)).fetchall()]
    conn.close()
    return render_template('rector_estudiantes.html',
                           slug=slug, colegio=colegio, rector=rector,
                           estudiantes=estudiantes, page=page, per_page=per_page,
                           total=total, total_pages=(total + per_page - 1) // per_page,
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

# ── RECTOR: CURSOS ─────────────────────────────────────────────────────────────
@app.route('/<slug>/rector/cursos')
def rector_cursos(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 12, type=int)
    conn = conectar(slug)
    rows = conn.execute(
        '''SELECT curso, jornada, COUNT(*) as total,
                  SUM(CASE WHEN activo=1 THEN 1 ELSE 0 END) as activos
           FROM alumnos GROUP BY curso, jornada ORDER BY curso''').fetchall()
    cursos = [dict(r) for r in rows]
    total = len(cursos)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    cursos_page = cursos[start:start + per_page]
    conn.close()
    return render_template('rector_cursos.html',
                           slug=slug, colegio=colegio, rector=rector,
                           cursos=cursos_page, total=total,
                           page=page, per_page=per_page, total_pages=total_pages,
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

# ── RECTOR: REPORTES ───────────────────────────────────────────────────────────
@app.route('/<slug>/rector/reportes')
def rector_reportes(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    conn = conectar(slug)
    total_est = conn.execute(
        'SELECT COUNT(*) as c FROM alumnos WHERE activo=1').fetchone()['c']
    total_prof = conn.execute(
        'SELECT COUNT(*) as c FROM profesores WHERE activo=1').fetchone()['c']
    total_cursos = conn.execute(
        'SELECT COUNT(DISTINCT curso) as c FROM alumnos WHERE activo=1').fetchone()['c']
    total_directoras = conn.execute(
        'SELECT COUNT(*) as c FROM directoras WHERE activo=1').fetchone()['c']
    conn.close()
    return render_template('rector_reportes.html',
                           slug=slug, colegio=colegio, rector=rector,
                           total_est=total_est, total_prof=total_prof,
                           total_cursos=total_cursos,
                           total_directoras=total_directoras,
                            notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

# ── RECTOR: ASISTENCIA ──────────────────────────────────────────────────────────
@app.route('/<slug>/rector/asistencia')
def rector_asistencia(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    conn = conectar(slug)
    cursos = [r['curso'] for r in conn.execute(
        'SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()]
    jornadas = [r['jornada'] for r in conn.execute(
        'SELECT DISTINCT jornada FROM alumnos WHERE activo=1 ORDER BY jornada').fetchall()]
    profesores = conn.execute(
        'SELECT id, nombre FROM profesores WHERE activo=1 ORDER BY nombre').fetchall()
    conn.close()
    return render_template('rector_asistencia.html',
                           slug=slug, colegio=colegio, rector=rector,
                           cursos=cursos, jornadas=jornadas,
                           profesores=profesores,
                           estados_asistencia=ESTADOS_ASISTENCIA,
                           hoy_fecha=datetime.today().strftime('%Y-%m-%d'),
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

@app.route('/<slug>/rector/asistencia_data')
def rector_asistencia_data(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'error': 'No autorizado'}), 403
    conn = conectar(slug)
    try:
        curso   = request.args.get('curso', '')
        jornada = request.args.get('jornada', '')
        materia = request.args.get('materia', '')
        profesor_id = request.args.get('profesor_id', type=int)
        fecha   = request.args.get('fecha', datetime.today().strftime('%Y-%m-%d'))
        try:
            datetime.strptime(fecha, '%Y-%m-%d')
        except ValueError:
            fecha = datetime.today().strftime('%Y-%m-%d')
        where = 'a.activo=1'
        params = []
        if curso:
            where += ' AND a.curso=?'; params.append(curso)
        if jornada:
            where += ' AND a.jornada=?'; params.append(jornada)
        alumnos = conn.execute(
            f'SELECT a.id, a.nombre, a.num_curso, a.curso, a.jornada FROM alumnos a WHERE {where} ORDER BY a.curso, a.nombre',
            params).fetchall()
        for a in alumnos:
            a = dict(a)
        if not alumnos:
            conn.close(); return jsonify({'estudiantes': [], 'stats': _asistencia_stats(conn, curso, jornada)})

        aids = [a['id'] for a in alumnos]
        placeholders = ','.join('?' * len(aids))
        asis_rows = conn.execute(
            f'SELECT aid, estado, observacion, hora FROM asistencia WHERE fecha=? AND aid IN ({placeholders})',
            (fecha,) + tuple(aids)).fetchall()
        asis_map = {r['aid']: {'estado': r['estado'], 'observacion': r['observacion'] or '', 'hora': r['hora'] or ''} for r in asis_rows}
        estudiantes = []
        for a in alumnos:
            info = asis_map.get(a['id'], {})
            estudiantes.append({
                'id': a['id'], 'nombre': a['nombre'],
                'num_curso': a['num_curso'], 'curso': a['curso'],
                'asistencia': info.get('estado', ''),
                'observacion': info.get('observacion', ''),
                'hora': info.get('hora', ''),
            })
        stats = _asistencia_stats(conn, curso=curso, jornada=jornada)
        alertas = _asistencia_alertas(conn, slug, curso or '', jornada or '') if curso and jornada else []
    finally:
        conn.close()
    return jsonify({'estudiantes': estudiantes, 'stats': stats, 'alertas': alertas,
                    'estados': dict(ESTADOS_ASISTENCIA)})

# ── RECTOR: CONFIGURACIÓN ──────────────────────────────────────────────────────
@app.route('/<slug>/rector/configuracion', methods=['GET', 'POST'])
def rector_configuracion(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    error = exito = None
    conn = conectar(slug)

    accion = request.form.get('accion', '')
    if request.method == 'POST':
        if not validar_csrf():
            return 'Error de seguridad', 400

        if accion == 'perfil':
            nombre = request.form.get('nombre', '').strip()
            email = request.form.get('email', '').strip()
            pw_actual = request.form.get('password_actual', '').strip()
            pw_nueva = request.form.get('password_nueva', '').strip()
            if not nombre:
                error = 'El nombre es obligatorio.'
            elif pw_actual and not verificar_pw(pw_actual, rector['password']):
                error = 'La contraseña actual no es correcta.'
            elif pw_nueva and len(pw_nueva) < 6:
                error = 'Mínimo 6 caracteres para la nueva contraseña.'
            else:
                if pw_nueva:
                    conn.execute('UPDATE rectores SET nombre=?, email=?, password=? WHERE id=?',
                                 (nombre, email, hash_pw(pw_nueva), rector['id']))
                else:
                    conn.execute('UPDATE rectores SET nombre=?, email=? WHERE id=?',
                                 (nombre, email, rector['id']))
                conn.commit()
                exito = 'Perfil actualizado correctamente.'
                rector = conn.execute('SELECT * FROM rectores WHERE id=?', (rector['id'],)).fetchone()

        elif accion == 'institucion':
            tipo_ev = request.form.get('tipo_evaluacion', 'numerica')
            esc_min = request.form.get('escala_min', 1, type=float)
            esc_max = request.form.get('escala_max', 10, type=float)
            nota_min = request.form.get('nota_minima_aprobar', 6, type=float)
            decimales = request.form.get('decimales_notas', 1, type=int)
            num_per = request.form.get('num_periodos', 4, type=int)
            acuse = 1 if request.form.get('acuse_recibo') else 0

            # Nombres personalizados de roles
            roles_json = json.dumps({
                'rector': request.form.get('rol_rector', 'Rector'),
                'authority': request.form.get('rol_authority', 'Coordinador'),
                'teacher': request.form.get('rol_teacher', 'Docente'),
                'student': request.form.get('rol_student', 'Estudiante'),
                'guardian': request.form.get('rol_guardian', 'Acudiente'),
            })

            # Jornadas
            jornadas_raw = request.form.get('jornadas', 'Mañana, Tarde, Nocturna')
            jornadas_list = [j.strip() for j in jornadas_raw.split(',') if j.strip()]
            jornadas_json = json.dumps(jornadas_list)

            conn.execute('''UPDATE config_institucion SET
                tipo_evaluacion=?, escala_min=?, escala_max=?, nota_minima_aprobar=?,
                decimales_notas=?, num_periodos=?, acuse_recibo=?,
                roles_json=?, jornadas_json=?, updated_at=datetime('now','localtime')
                WHERE slug=?''',
                (tipo_ev, esc_min, esc_max, nota_min, decimales, num_per,
                 acuse, roles_json, jornadas_json, slug))
            conn.commit()
            _cache_invalidate(slug)
            exito = 'Configuración institucional guardada.'

    config = config_get(slug)
    periodos_estado = conn.execute(
        'SELECT * FROM periodos_estado ORDER BY periodo').fetchall()
    conn.close()
    return render_template('rector_configuracion.html',
                           slug=slug, colegio=colegio, rector=rector,
                           config=config, error=error, exito=exito,
                           periodos_estado={r['periodo']: dict(r) for r in periodos_estado},
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

@app.route('/<slug>/rector/periodos/<int:periodo>/<accion>', methods=['POST'])
def rector_periodo_accion(slug, periodo, accion):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return ('No autorizado', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    if accion not in ('abrir', 'cerrar'): return ('Accion invalida', 400)
    conn = conectar(slug)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if accion == 'cerrar':
        conn.execute('''INSERT INTO periodos_estado (periodo, estado, fecha_cierre, cerrado_por)
                        VALUES (?, 'cerrado', ?, ?)
                        ON CONFLICT(periodo) DO UPDATE SET estado='cerrado', fecha_cierre=?, cerrado_por=?''',
                     (periodo, now, rector['id'], now, rector['id']))
        audit_log(slug, rector['id'], 'periodo_cerrado', 'periodos_estado', registro_id=periodo)
    else:
        conn.execute('''INSERT INTO periodos_estado (periodo, estado, fecha_apertura, abierto_por)
                        VALUES (?, 'abierto', ?, ?)
                        ON CONFLICT(periodo) DO UPDATE SET estado='abierto', fecha_apertura=?, abierto_por=?''',
                     (periodo, now, rector['id'], now, rector['id']))
        audit_log(slug, rector['id'], 'periodo_abierto', 'periodos_estado', registro_id=periodo)
    conn.commit()
    conn.close()
    return redirect(url_for('rector_configuracion', slug=slug, _anchor='periodos'))

@app.route('/<slug>/rector/solicitudes')
def rector_solicitudes(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    conn = conectar(slug)
    solicitudes = conn.execute(
        '''SELECT s.*, a.nombre as alumno_nombre, p.nombre as profesor_nombre,
                  COALESCE(ac.nombre, s.tipo) as actividad_nombre
           FROM solicitudes_modificacion s
           JOIN alumnos a ON a.id=s.aid
           LEFT JOIN actividades ac ON ac.id=s.actividad_id
           JOIN profesores p ON p.id=s.profesor_id
           WHERE s.slug=?
           ORDER BY s.fecha_solicitud DESC''', (slug,)).fetchall()
    conn.close()
    return render_template('rector_solicitudes.html',
                           slug=slug, colegio=get_colegio(slug), rector=rector,
                           solicitudes=[dict(s) for s in solicitudes],
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

@app.route('/<slug>/rector/solicitudes/<int:sid>/<accion>', methods=['POST'])
def rector_solicitud_accion(slug, sid, accion):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    if accion not in ('aprobar', 'rechazar'): return jsonify({'status':'error','mensaje':'Accion invalida'}), 400
    conn = conectar(slug)
    sol = conn.execute(
        'SELECT * FROM solicitudes_modificacion WHERE id=? AND slug=?', (sid, slug)).fetchone()
    if not sol:
        conn.close()
        return jsonify({'status':'error','mensaje':'Solicitud no encontrada'}), 404
    if sol['estado'] != 'pendiente':
        conn.close()
        return jsonify({'status':'error','mensaje':'La solicitud ya fue ' + sol['estado']}), 400
    if sol['profesor_id'] == rector['id']:
        conn.close()
        return jsonify({'status':'error','mensaje':'No puedes aprobar tu propia solicitud'}), 403
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if accion == 'aprobar':
        valor_sol = float(sol['valor_solicitado']) if sol['valor_solicitado'] else None
        curso_ctx = sol['curso']
        materia_ctx = sol['materia']
        if sol['tipo'] == 'actividad' and sol['actividad_id'] is not None:
            conn.execute(
                '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
                   ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
                (sol['aid'], sol['actividad_id'], valor_sol))
            conn.commit()  # commit before auditar_nota opens its own connection
            auditar_nota(slug, rector['id'], 'rector', 'modificacion', 'notas', sol['aid'],
                         curso_ctx, materia_ctx, sol['periodo'],
                         campo='nota', actividad_id=sol['actividad_id'],
                         valor_anterior=sol['valor_actual'], valor_nuevo=valor_sol,
                         motivo='Aprobado por rector (solicitud #%d)' % sid)
        elif sol['tipo'] in ('evaluacion', 'autoevaluacion'):
            jornada_eval = sol['jornada']
            if sol['tipo'] == 'evaluacion':
                conn.execute(
                    '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,evaluacion,periodo)
                       VALUES (?,?,?,?,?,?)
                       ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                       DO UPDATE SET evaluacion=excluded.evaluacion''',
                    (sol['aid'], sol['profesor_id'], materia_ctx,
                     jornada_eval, valor_sol, sol['periodo']))
            else:
                conn.execute(
                    '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,autoevaluacion,periodo)
                       VALUES (?,?,?,?,?,?)
                       ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                       DO UPDATE SET autoevaluacion=excluded.autoevaluacion''',
                    (sol['aid'], sol['profesor_id'], materia_ctx,
                     jornada_eval, valor_sol, sol['periodo']))
            conn.commit()  # commit before auditar_nota opens its own connection
            auditar_nota(slug, rector['id'], 'rector', 'modificacion', 'evaluaciones', sol['aid'],
                         curso_ctx, materia_ctx, sol['periodo'],
                         campo=sol['tipo'],
                         valor_anterior=sol['valor_actual'], valor_nuevo=valor_sol,
                         motivo='Aprobado por rector (solicitud #%d)' % sid)
        conn.execute(
            '''UPDATE solicitudes_modificacion
               SET estado='aprobada', aprobado_por=?, fecha_respuesta=?
               WHERE id=?''',
            (rector['id'], now, sid))
        conn.commit()
        # Notify teacher
        crear_notificacion(slug, 'profesor', sol['profesor_id'],
            'Solicitud aprobada',
            'Tu solicitud #%d fue aprobada por el rector.' % sid)
    else:
        conn.execute(
            '''UPDATE solicitudes_modificacion
               SET estado='rechazada', aprobado_por=?, fecha_respuesta=?
               WHERE id=?''',
            (rector['id'], now, sid))
        conn.commit()  # commit before auditar_nota opens its own connection
        auditar_nota(slug, rector['id'], 'rector', 'solicitud_rechazada', 'solicitudes_modificacion', sol['aid'],
                     sol['curso'], sol['materia'], sol['periodo'],
                     campo=sol['tipo'], actividad_id=sol['actividad_id'],
                     valor_anterior=sol['valor_actual'], valor_nuevo=sol['valor_solicitado'],
                     motivo='Rechazado por rector (solicitud #%d)' % sid)
        crear_notificacion(slug, 'profesor', sol['profesor_id'],
            'Solicitud rechazada',
            'Tu solicitud #%d fue rechazada por el rector.' % sid)
    conn.close()
    return jsonify({'status':'ok','mensaje':'Solicitud ' + ('aprobada' if accion == 'aprobar' else 'rechazada')})

@app.route('/<slug>/rector/auditoria')
def rector_auditoria(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    conn = conectar(slug)

    tabla = request.args.get('tabla', '')
    page = max(1, int(request.args.get('page', 1)))
    limit = 50
    offset = (page - 1) * limit

    where = []
    params = []
    if tabla:
        where.append('a.tabla = ?')
        params.append(tabla)

    where_clause = ('WHERE ' + ' AND '.join(where)) if where else ''
    total = conn.execute(f'SELECT COUNT(*) as c FROM audit_log a {where_clause}', params).fetchone()['c']
    registros = conn.execute(f'''
        SELECT a.*, u.nombre as usuario_nombre
        FROM audit_log a
        LEFT JOIN usuarios u ON a.usuario_id = u.id
        {where_clause}
        ORDER BY a.creado DESC LIMIT ? OFFSET ?
    ''', params + [limit, offset]).fetchall()

    tablas = [r['tabla'] for r in conn.execute(
        "SELECT DISTINCT tabla FROM audit_log ORDER BY tabla"
    ).fetchall()]
    conn.close()

    total_pages = max(1, (total + limit - 1) // limit)
    return render_template('rector_auditoria.html',
                         slug=slug, colegio=colegio, rector=rector,
                         registros=[dict(r) for r in registros],
                         tabla=tabla, tablas=tablas,
                         page=page, total_pages=total_pages, total=total,
                         notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

# ── NOTIFICATION HELPERS ───────────────────────────────────────────────────────
def crear_notificacion(slug, usuario_tipo, usuario_id, titulo, mensaje='', tipo='info', link=''):
    conn = conectar(slug)
    conn.execute(
        'INSERT INTO notificaciones (usuario_tipo,usuario_id,titulo,mensaje,tipo,link) VALUES (?,?,?,?,?,?)',
        (usuario_tipo, usuario_id, titulo, mensaje, tipo, link))
    conn.commit()
    conn.close()

def notificaciones_no_leidas(slug, usuario_tipo, usuario_id, conn=None):
    cerrar = conn or conectar(slug)
    c = cerrar.execute(
        'SELECT COUNT(*) as c FROM notificaciones WHERE usuario_tipo=? AND usuario_id=? AND leida=0',
        (usuario_tipo, usuario_id)).fetchone()['c']
    if not conn: cerrar.close()
    return c

def generar_destinatarios(slug, comunicacion_id):
    try:
        conn = conectar(slug)
    except Exception as e:
        app.logger.error(f'generar_destinatarios: error conectando DB {slug}: {e}')
        return
    try:
        cols_cl = [r[1] for r in conn.execute('PRAGMA table_info(comunicaciones_leidas)').fetchall()]
        if 'leido' not in cols_cl:
            conn.execute('ALTER TABLE comunicaciones_leidas ADD COLUMN leido INTEGER DEFAULT 0')
            conn.commit()
            cols_cl = [r[1] for r in conn.execute('PRAGMA table_info(comunicaciones_leidas)').fetchall()]
    except Exception as e:
        app.logger.error(f'generar_destinatarios: error migrando columna leido: {e}')
    if 'leido' not in cols_cl:
        conn.close()
        app.logger.error('generar_destinatarios: columna leido no disponible en comunicaciones_leidas, abortando')
        return
    com = conn.execute('SELECT * FROM comunicaciones WHERE id=?', (comunicacion_id,)).fetchone()
    if not com:
        conn.close()
        app.logger.warning(f'generar_destinatarios: comunicacion {comunicacion_id} no encontrada')
        return
    if com['estado'] != 'publicado':
        conn.close()
        return
    dest_tipo = com['destinatario_tipo']
    try:
        val_arr = json.loads(com['destinatario_valor']) if com['destinatario_valor'] else []
    except (json.JSONDecodeError, TypeError) as e:
        app.logger.warning(f'generar_destinatarios: error parseando destinatario_valor="{com["destinatario_valor"]}": {e}')
        val_arr = []
    if not isinstance(val_arr, list):
        app.logger.warning(f'generar_destinatarios: destinatario_valor no es un array, ignorando: {type(val_arr).__name__}')
        val_arr = []
    destinatarios = []
    if dest_tipo == 'todo_colegio':
        for r in conn.execute('SELECT id FROM profesores WHERE activo=1').fetchall():
            destinatarios.append(('profesor', r['id']))
        for r in conn.execute('SELECT id FROM directoras WHERE activo=1').fetchall():
            destinatarios.append(('directora', r['id']))
        for r in conn.execute('SELECT id FROM alumnos WHERE activo=1').fetchall():
            destinatarios.append(('estudiante', r['id']))
    elif dest_tipo == 'profesores':
        if val_arr:
            for v in val_arr:
                if isinstance(v, str) and v.startswith('prof_'):
                    try:
                        destinatarios.append(('profesor', int(v.split('_')[1])))
                    except (ValueError, IndexError):
                        app.logger.warning(f'generar_destinatarios: valor prof_ invalido: {v}')
        else:
            for r in conn.execute('SELECT id FROM profesores WHERE activo=1').fetchall():
                destinatarios.append(('profesor', r['id']))
    elif dest_tipo == 'directores':
        if val_arr:
            for v in val_arr:
                if isinstance(v, str) and v.startswith('dir_'):
                    try:
                        destinatarios.append(('directora', int(v.split('_')[1])))
                    except (ValueError, IndexError):
                        app.logger.warning(f'generar_destinatarios: valor dir_ invalido: {v}')
        else:
            for r in conn.execute('SELECT id FROM directoras WHERE activo=1').fetchall():
                destinatarios.append(('directora', r['id']))
    elif dest_tipo == 'estudiantes':
        for r in conn.execute('SELECT id FROM alumnos WHERE activo=1').fetchall():
            destinatarios.append(('estudiante', r['id']))
    elif dest_tipo == 'grado':
        if val_arr:
            cursos_grado = {}
            for row in conn.execute('SELECT DISTINCT curso FROM alumnos WHERE activo=1').fetchall():
                c = row['curso']
                grade_num = ''.join(filter(str.isdigit, c))
                cursos_grado.setdefault(grade_num, []).append(c)
            for grado in val_arr:
                for curso in cursos_grado.get(str(grado), []):
                    for r in conn.execute('SELECT id FROM alumnos WHERE activo=1 AND curso=?', (curso,)).fetchall():
                        destinatarios.append(('estudiante', r['id']))
    elif dest_tipo == 'cursos':
        if val_arr:
            for curso in val_arr:
                for r in conn.execute('SELECT id FROM alumnos WHERE activo=1 AND curso=?', (curso,)).fetchall():
                    destinatarios.append(('estudiante', r['id']))
    for tipo, uid in destinatarios:
        try:
            conn.execute(
                'INSERT OR IGNORE INTO comunicaciones_leidas (comunicacion_id,usuario_tipo,usuario_id,leido) VALUES (?,?,?,0)',
                (comunicacion_id, tipo, uid))
        except Exception as e_insert:
            app.logger.error(f'generar_destinatarios: error insertando destinatario tipo={tipo} uid={uid}: {e_insert}')
    conn.commit()
    conn.close()

def comunicaciones_pendientes(slug, usuario_tipo, usuario_id, conn=None):
    cerrar = conn or conectar(slug)
    cols_cl = [r[1] for r in cerrar.execute('PRAGMA table_info(comunicaciones_leidas)').fetchall()]
    if 'leido' not in cols_cl:
        if not conn: cerrar.close()
        return []
    rows = cerrar.execute(
        '''SELECT c.*, cl.leido, cl.fecha_lectura
           FROM comunicaciones c
           JOIN comunicaciones_leidas cl ON cl.comunicacion_id=c.id
           WHERE cl.usuario_tipo=? AND cl.usuario_id=? AND COALESCE(cl.leido,0)=0
           AND c.estado='publicado' AND c.activo=1
           ORDER BY c.fecha_publicacion DESC''',
        (usuario_tipo, usuario_id)).fetchall()
    if not conn: cerrar.close()
    return [dict(r) for r in rows]

# ── COMUNICACIONES (RECTOR) ────────────────────────────────────────────────────
@app.route('/<slug>/rector/comunicaciones')
def rector_comunicaciones(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    estado_filtro = request.args.get('estado', '')
    conn = conectar(slug)
    if estado_filtro:
        comunicaciones = conn.execute(
            '''SELECT * FROM comunicaciones WHERE rector_id=? AND activo=1 AND estado=?
               ORDER BY fecha_creacion DESC''',
            (rector['id'], estado_filtro)).fetchall()
    else:
        comunicaciones = conn.execute(
            '''SELECT * FROM comunicaciones WHERE rector_id=? AND activo=1
               ORDER BY fecha_creacion DESC''',
            (rector['id'],)).fetchall()
    notif_count = notificaciones_no_leidas(slug, 'rector', rector['id'])
    conn.close()
    return render_template('rector_comunicaciones.html',
                           slug=slug, colegio=colegio, rector=rector,
                           comunicaciones=comunicaciones,
                           estado_filtro=estado_filtro,
                           notif_count=notif_count)

@app.route('/<slug>/rector/comunicaciones/nueva', methods=['GET', 'POST'])
def rector_comunicacion_nueva(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    error = exito = None
    conn = conectar(slug)
    cursos = [r['curso'] for r in conn.execute(
        'SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()]
    profesores = [dict(r) for r in conn.execute(
        'SELECT id, nombre FROM profesores WHERE activo=1 ORDER BY nombre').fetchall()]
    directoras = [dict(r) for r in conn.execute(
        'SELECT id, nombre, curso FROM directoras WHERE activo=1 ORDER BY nombre').fetchall()]
    conn.close()
    if request.method == 'POST':
        if not validar_csrf():
            return 'Error de seguridad', 400
        titulo = request.form.get('titulo', '').strip()
        contenido = request.form.get('contenido', '').strip()
        dest_tipo = request.form.get('destinatario_tipo', '').strip()
        dest_valor = request.form.get('destinatario_valor', '').strip()
        prioridad = request.form.get('prioridad', 'normal').strip()
        programar = request.form.get('fecha_programada', '').strip()
        publicar_ahora = request.form.get('publicar_ahora', '0').strip()
        if not titulo or not contenido or not dest_tipo:
            error = 'Completa todos los campos.'
        else:
            conn = conectar(slug)
            cursor = conn.execute(
                '''INSERT INTO comunicaciones (rector_id,titulo,contenido,destinatario_tipo,destinatario_valor,prioridad,estado,fecha_programada,fecha_publicacion)
                   VALUES (?,?,?,?,?,?,?,?,?)''',
                (rector['id'], titulo, contenido, dest_tipo, dest_valor, prioridad,
                 'publicado' if publicar_ahora == '1' else ('programado' if programar else 'borrador'),
                 programar if programar else None,
                 datetime.today().strftime('%Y-%m-%d %H:%M:%S') if publicar_ahora == '1' else None))
            new_id = cursor.lastrowid
            conn.commit()
            conn.close()
            if publicar_ahora == '1':
                try:
                    generar_destinatarios(slug, new_id)
                except Exception as e:
                    app.logger.error(f'Error en generar_destinatarios (nueva): {e}')
            exito = 'Comunicación creada correctamente.'
    return render_template('rector_comunicacion_form.html',
                           slug=slug, colegio=colegio, rector=rector,
                           error=error, exito=exito, comunicacion=None,
                           cursos=cursos, profesores=profesores,
                           directoras=directoras,
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

@app.route('/<slug>/rector/comunicaciones/<int:cid>/editar', methods=['GET', 'POST'])
def rector_comunicacion_editar(slug, cid):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    conn = conectar(slug)
    com = conn.execute(
        'SELECT * FROM comunicaciones WHERE id=? AND rector_id=? AND activo=1',
        (cid, rector['id'])).fetchone()
    if not com: conn.close(); return 'Comunicación no encontrada', 404
    cursos = [r['curso'] for r in conn.execute(
        'SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()]
    profesores = [dict(r) for r in conn.execute(
        'SELECT id, nombre FROM profesores WHERE activo=1 ORDER BY nombre').fetchall()]
    directoras = [dict(r) for r in conn.execute(
        'SELECT id, nombre, curso FROM directoras WHERE activo=1 ORDER BY nombre').fetchall()]
    error = exito = None
    if request.method == 'POST':
        if not validar_csrf():
            return 'Error de seguridad', 400
        titulo = request.form.get('titulo', '').strip()
        contenido = request.form.get('contenido', '').strip()
        dest_tipo = request.form.get('destinatario_tipo', '').strip()
        dest_valor = request.form.get('destinatario_valor', '').strip()
        prioridad = request.form.get('prioridad', 'normal').strip()
        programar = request.form.get('fecha_programada', '').strip()
        publicar_ahora = request.form.get('publicar_ahora', '0').strip()
        if not titulo or not contenido or not dest_tipo:
            error = 'Completa todos los campos.'
        else:
            conn.execute(
                '''UPDATE comunicaciones SET titulo=?,contenido=?,destinatario_tipo=?,destinatario_valor=?,
                   prioridad=?,estado=?,fecha_programada=?,fecha_publicacion=?
                   WHERE id=? AND rector_id=?''',
                (titulo, contenido, dest_tipo, dest_valor, prioridad,
                 'publicado' if publicar_ahora == '1' else ('programado' if programar else com['estado']),
                 programar if programar else None,
                 datetime.today().strftime('%Y-%m-%d %H:%M:%S') if publicar_ahora == '1' else (com['fecha_publicacion'] if com['fecha_publicacion'] else None),
                 cid, rector['id']))
            conn.commit()
            if publicar_ahora == '1':
                conn.close()
                try:
                    generar_destinatarios(slug, cid)
                except Exception as e:
                    app.logger.error(f'Error en generar_destinatarios: {e}')
                conn = conectar(slug)
            exito = 'Comunicación actualizada correctamente.'
            com = conn.execute(
                'SELECT * FROM comunicaciones WHERE id=? AND activo=1', (cid,)).fetchone()
    conn.close()
    return render_template('rector_comunicacion_form.html',
                           slug=slug, colegio=colegio, rector=rector,
                           error=error, exito=exito, comunicacion=com,
                           cursos=cursos, profesores=profesores,
                           directoras=directoras,
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

@app.route('/<slug>/rector/comunicaciones/<int:cid>')
def rector_comunicacion_detalle(slug, cid):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    colegio = get_colegio(slug)
    conn = conectar(slug)
    com = conn.execute(
        'SELECT * FROM comunicaciones WHERE id=? AND rector_id=? AND activo=1',
        (cid, rector['id'])).fetchone()
    if not com: conn.close(); return 'Comunicación no encontrada', 404
    total_dest = conn.execute(
        'SELECT COUNT(*) as c FROM comunicaciones_leidas WHERE comunicacion_id=?', (cid,)).fetchone()['c']
    leidas_count = conn.execute(
        'SELECT COUNT(*) as c FROM comunicaciones_leidas WHERE comunicacion_id=? AND leido=1', (cid,)).fetchone()['c']
    no_leidas = total_dest - leidas_count
    conn.close()
    return render_template('rector_comunicacion_detail.html',
                           slug=slug, colegio=colegio, rector=rector,
                           com=com, total_dest=total_dest, leidas=leidas_count, no_leidas=no_leidas,
                           notif_count=notificaciones_no_leidas(slug, 'rector', rector['id']))

@app.route('/<slug>/rector/comunicaciones/<int:cid>/publicar', methods=['POST'])
def rector_comunicacion_publicar(slug, cid):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf(): return 'Error de seguridad', 400
    conn = conectar(slug)
    conn.execute(
        '''UPDATE comunicaciones SET estado='publicado',fecha_publicacion=datetime('now','localtime')
           WHERE id=? AND rector_id=? AND activo=1''',
        (cid, rector['id']))
    conn.commit()
    conn.close()
    try:
        generar_destinatarios(slug, cid)
    except Exception as e:
        app.logger.error(f'Error en generar_destinatarios (publicar): {e}')
    return redirect(url_for('rector_comunicaciones', slug=slug))

@app.route('/<slug>/rector/comunicaciones/<int:cid>/archivar', methods=['POST'])
def rector_comunicacion_archivar(slug, cid):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf(): return 'Error de seguridad', 400
    conn = conectar(slug)
    conn.execute(
        "UPDATE comunicaciones SET estado='archivado' WHERE id=? AND rector_id=? AND activo=1",
        (cid, rector['id']))
    conn.commit()
    conn.close()
    return redirect(url_for('rector_comunicaciones', slug=slug))

@app.route('/<slug>/rector/comunicaciones/<int:cid>/eliminar', methods=['POST'])
def rector_comunicacion_eliminar(slug, cid):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf(): return 'Error de seguridad', 400
    conn = conectar(slug)
    conn.execute(
        'DELETE FROM comunicaciones WHERE id=? AND rector_id=?',
        (cid, rector['id']))
    conn.execute('DELETE FROM comunicaciones_leidas WHERE comunicacion_id=?', (cid,))
    conn.commit()
    conn.close()
    return redirect(url_for('rector_comunicaciones', slug=slug))

# ── NOTIFICACIONES (TODOS LOS ROLES) ───────────────────────────────────────────
@app.route('/<slug>/notificaciones')
def notificaciones(slug):
    require_colegio(slug)
    colegio = get_colegio(slug)
    usuario_tipo = None
    usuario_id = None
    rector = get_rector(slug)
    if rector: usuario_tipo, usuario_id = 'rector', rector['id']
    if not usuario_id:
        prof = get_profesor(slug)
        if prof: usuario_tipo, usuario_id = 'profesor', prof['id']
    if not usuario_id:
        directora = get_directora(slug)
        if directora: usuario_tipo, usuario_id = 'directora', directora['id']
    if not usuario_id:
        aid = session.get(f'alumno_id_{slug}')
        if aid: usuario_tipo, usuario_id = 'estudiante', aid
    if not usuario_id:
        return redirect(url_for('auth.login', slug=slug))
    conn = conectar(slug)
    notifs = conn.execute(
        'SELECT * FROM notificaciones WHERE usuario_tipo=? AND usuario_id=? ORDER BY fecha_creacion DESC LIMIT 100',
        (usuario_tipo, usuario_id)).fetchall()
    conn.close()
    return render_template('notificaciones.html',
                           slug=slug, colegio=colegio,
                           notificaciones=notifs,
                           usuario_tipo=usuario_tipo)

@app.route('/<slug>/notificaciones/<int:nid>/leer', methods=['POST'])
def notificacion_leer(slug, nid):
    require_colegio(slug)
    if not validar_csrf(): return 'Error de seguridad', 400
    usuario_tipo = None; usuario_id = None
    rector = get_rector(slug)
    if rector: usuario_tipo, usuario_id = 'rector', rector['id']
    if not usuario_id:
        prof = get_profesor(slug)
        if prof: usuario_tipo, usuario_id = 'profesor', prof['id']
    if not usuario_id:
        directora = get_directora(slug)
        if directora: usuario_tipo, usuario_id = 'directora', directora['id']
    if not usuario_id:
        aid = session.get(f'alumno_id_{slug}')
        if aid: usuario_tipo, usuario_id = 'estudiante', aid
    if not usuario_id:
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 403
    conn = conectar(slug)
    conn.execute('UPDATE notificaciones SET leida=1 WHERE id=? AND usuario_tipo=? AND usuario_id=?',
                 (nid, usuario_tipo, usuario_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/<slug>/notificaciones/contar')
def notificaciones_contar(slug):
    require_colegio(slug)
    usuario_tipo = None
    usuario_id = None
    rector = get_rector(slug)
    if rector: usuario_tipo, usuario_id = 'rector', rector['id']
    if not usuario_id:
        prof = get_profesor(slug)
        if prof: usuario_tipo, usuario_id = 'profesor', prof['id']
    if not usuario_id:
        directora = get_directora(slug)
        if directora: usuario_tipo, usuario_id = 'directora', directora['id']
    if not usuario_id:
        aid = session.get(f'alumno_id_{slug}')
        if aid: usuario_tipo, usuario_id = 'estudiante', aid
    if not usuario_id:
        return jsonify({'count': 0})
    c = notificaciones_no_leidas(slug, usuario_tipo, usuario_id)
    return jsonify({'count': c})

@app.route('/<slug>/comunicaciones/<int:cid>/leer', methods=['POST'])
def comunicacion_leer(slug, cid):
    if not validar_csrf(): return jsonify({'error': 'Error CSRF'}), 403
    require_colegio(slug)
    usuario_tipo = None
    usuario_id = None
    rector = get_rector(slug)
    if rector: usuario_tipo, usuario_id = 'rector', rector['id']
    if not usuario_id:
        prof = get_profesor(slug)
        if prof: usuario_tipo, usuario_id = 'profesor', prof['id']
    if not usuario_id:
        directora = get_directora(slug)
        if directora: usuario_tipo, usuario_id = 'directora', directora['id']
    if not usuario_id:
        aid = session.get(f'alumno_id_{slug}')
        if aid: usuario_tipo, usuario_id = 'estudiante', aid
    if not usuario_id:
        return jsonify({'error': 'No autorizado'}), 403
    conn = conectar(slug)
    try:
        cols_cl = [r[1] for r in conn.execute('PRAGMA table_info(comunicaciones_leidas)').fetchall()]
        if 'leido' not in cols_cl:
            conn.execute('ALTER TABLE comunicaciones_leidas ADD COLUMN leido INTEGER DEFAULT 0')
            conn.commit()
            cols_cl = [r[1] for r in conn.execute('PRAGMA table_info(comunicaciones_leidas)').fetchall()]
        if 'leido' not in cols_cl:
            conn.close()
            return jsonify({'error': 'Error de migración'}), 500
    except Exception as e:
        app.logger.error(f'comunicacion_leer: error migrando columna leido: {e}')
    existing = conn.execute(
        'SELECT 1 FROM comunicaciones_leidas WHERE comunicacion_id=? AND usuario_tipo=? AND usuario_id=?',
        (cid, usuario_tipo, usuario_id)).fetchone()
    if existing:
        conn.execute(
            'UPDATE comunicaciones_leidas SET leido=1, fecha_lectura=datetime(\'now\',\'localtime\') WHERE comunicacion_id=? AND usuario_tipo=? AND usuario_id=?',
            (cid, usuario_tipo, usuario_id))
    else:
        conn.execute(
            'INSERT INTO comunicaciones_leidas (comunicacion_id,usuario_tipo,usuario_id,leido,fecha_lectura) VALUES (?,?,?,1,datetime(\'now\',\'localtime\'))',
            (cid, usuario_tipo, usuario_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── CANALES CRUD (Rector) ──────────────────────────────────────────────────────
@app.route('/<slug>/rector/canales')
def rector_canales(slug):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return redirect(url_for('auth.rector_login', slug=slug))
    conn = conectar(slug)
    canales = conn.execute('SELECT * FROM canales WHERE slug=? ORDER BY fecha_creacion DESC', (slug,)).fetchall()
    cursos = conn.execute('SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()
    materias_rows = conn.execute('SELECT DISTINCT materia FROM actividades').fetchall()
    if not materias_rows:
        materias_rows = conn.execute('SELECT DISTINCT materia FROM asignaciones_materia').fetchall()
    materias = list(set(r['materia'] for r in materias_rows))
    conn.close()
    colegio = get_colegio(slug)
    return render_template('rector_canales.html', slug=slug, rector=rector, canales=canales, colegio=colegio,
                          cursos=[r['curso'] for r in cursos],
                          materias=materias)

@app.route('/<slug>/rector/canales/crear', methods=['POST'])
def rector_canales_crear(slug):
    if not validar_csrf(): return jsonify({'ok': False, 'error': 'Error CSRF'}), 403
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'ok':False,'error':'No autorizado'})
    tipo = request.form.get('tipo')
    nombre = request.form.get('nombre','').strip()
    curso = request.form.get('curso','')
    materia = request.form.get('materia','')
    descripcion = request.form.get('descripcion','')
    if not nombre:
        nombres = {'institucional':'Institucional','rectoria':'Rectoría','profesores':'Profesores',
                   'director_curso':f'Directores {curso}','curso':f'Curso {curso}','materia':f'Materia {materia}'}
        nombre = nombres.get(tipo, tipo)
    conn = conectar(slug)
    cid = conn.execute('INSERT INTO canales (slug,rector_id,tipo,nombre,descripcion,curso,materia) VALUES (?,?,?,?,?,?,?)',
                       (slug,rector['id'],tipo,nombre,descripcion,curso,materia)).lastrowid
    asignar_miembros_auto(conn, slug, cid, tipo, curso, materia)
    conn.commit()
    conn.close()
    return jsonify({'ok':True, 'canal_id':cid})

@app.route('/<slug>/rector/canales/<int:cid>/eliminar', methods=['POST'])
def rector_canales_eliminar(slug, cid):
    if not validar_csrf(): return jsonify({'ok': False, 'error': 'Error CSRF'}), 403
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'ok':False,'error':'No autorizado'})
    conn = conectar(slug)
    conn.execute('UPDATE canales SET activo=0 WHERE id=? AND slug=?', (cid, slug))
    conn.commit()
    conn.close()
    return jsonify({'ok':True})

@app.route('/<slug>/rector/canales/<int:cid>/miembros')
def rector_canales_miembros(slug, cid):
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'ok':False,'error':'No autorizado'})
    conn = conectar(slug)
    miembros = conn.execute('SELECT * FROM canal_miembros WHERE canal_id=?', (cid,)).fetchall()
    canal = conn.execute('SELECT * FROM canales WHERE id=?', (cid,)).fetchone()
    conn.close()
    data = [dict(m) for m in miembros]
    conn2 = conectar(slug)
    for m in data:
        m['nombre_usuario'] = nombre_usuario_canal(conn2, m['usuario_tipo'], m['usuario_id'])
    conn2.close()
    return jsonify({'ok':True, 'miembros':data, 'canal':dict(canal) if canal else None})

# ── CANALES API (usuarios) ─────────────────────────────────────────────────────
@app.route('/<slug>/api/canales')
def api_canales(slug):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autenticado'}), 401
    if tipo == 'rector':
        rector = get_rector(slug)
        conn = conectar(slug)
        rows = conn.execute('''
            SELECT c.*,
                (SELECT mensaje FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultimo_mensaje,
                (SELECT usuario_tipo FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultimo_autor_tipo,
                (SELECT usuario_id FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultimo_autor_id,
                (SELECT fecha FROM mensajes_canal WHERE canal_id=c.id ORDER BY id DESC LIMIT 1) as ultima_fecha,
                (SELECT COUNT(*) FROM mensajes_canal mc
                 LEFT JOIN mensajes_leidos ml ON ml.mensaje_id=mc.id AND ml.usuario_tipo='rector' AND ml.usuario_id=?
                 WHERE mc.canal_id=c.id AND ml.id IS NULL) as no_leidos
            FROM canales c WHERE c.activo=1 ORDER BY ultima_fecha DESC''', (rector['id'],)).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    return jsonify(canales_usuario(slug, tipo, uid))

def _enriquecer_mensaje(conn, d):
    d['archivos'] = archivos_por_mensaje(conn, d['id'])
    d['reacciones'] = reacciones_por_mensaje(conn, d['id'])
    if d.get('responde_a'):
        padre = conn.execute('SELECT id, mensaje, usuario_tipo, usuario_id FROM mensajes_canal WHERE id=?', (d['responde_a'],)).fetchone()
        if padre:
            d['responde_a_info'] = {
                'id': padre['id'],
                'mensaje': padre['mensaje'][:120],
                'autor_nombre': nombre_usuario_canal(conn, padre['usuario_tipo'], padre['usuario_id'])
            }
    return d

def _enriquecer_mensajes_batch(conn, mensajes):
    """Batch enrich a list of message dicts: 4 queries instead of 3*N."""
    if not mensajes: return
    mids = [m['id'] for m in mensajes]
    ph = ','.join('?' * len(mids))
    # batch files
    arch_rows = conn.execute(
        f'SELECT * FROM mensajes_archivos WHERE mensaje_id IN ({ph}) ORDER BY id', mids).fetchall()
    arch_by_mid = {}
    for r in arch_rows:
        arch_by_mid.setdefault(r['mensaje_id'], []).append(dict(r))
    # batch reactions
    reac_rows = conn.execute(
        f'SELECT mensaje_id, reaccion, usuario_tipo, usuario_id FROM mensajes_reacciones WHERE mensaje_id IN ({ph})',
        mids).fetchall()
    reac_by_mid = {}
    for r in reac_rows:
        reac_by_mid.setdefault(r['mensaje_id'], {}).setdefault(r['reaccion'], []).append({'tipo': r['usuario_tipo'], 'id': r['usuario_id']})
    # batch author names (collect unique tipo+uid pairs)
    seen = set()
    tipo_ids = {'profesor': set(), 'estudiante': set(), 'rector': set(), 'directora': set()}
    for m in mensajes:
        key = (m['usuario_tipo'], m['usuario_id'])
        if key not in seen:
            seen.add(key)
            if m['usuario_tipo'] in tipo_ids:
                tipo_ids[m['usuario_tipo']].add(m['usuario_id'])
    name_map = {}
    for t, ids in tipo_ids.items():
        if not ids: continue
        ph2 = ','.join('?' * len(ids))
        table_map = {'profesor': 'profesores', 'estudiante': 'alumnos', 'rector': 'rectores', 'directora': 'directoras'}
        rows = conn.execute(f'SELECT id, nombre FROM {table_map[t]} WHERE id IN ({ph2})', list(ids)).fetchall()
        for r in rows:
            name_map[(t, r['id'])] = r['nombre']
    # batch reply info
    reply_ids = set(m['responde_a'] for m in mensajes if m.get('responde_a'))
    reply_info = {}
    if reply_ids:
        ph3 = ','.join('?' * len(reply_ids))
        padres = conn.execute(
            f'SELECT id, mensaje, usuario_tipo, usuario_id FROM mensajes_canal WHERE id IN ({ph3})', list(reply_ids)).fetchall()
        for p in padres:
            reply_info[p['id']] = {
                'id': p['id'],
                'mensaje': p['mensaje'][:120],
                'autor_nombre': name_map.get((p['usuario_tipo'], p['usuario_id']), nombre_usuario_canal(conn, p['usuario_tipo'], p['usuario_id']))
            }
    # apply to messages
    for m in mensajes:
        m['archivos'] = arch_by_mid.get(m['id'], [])
        m['reacciones'] = reac_by_mid.get(m['id'], {})
        m['autor_nombre'] = name_map.get((m['usuario_tipo'], m['usuario_id']), 'Desconocido')
        if m.get('responde_a') and m['responde_a'] in reply_info:
            m['responde_a_info'] = reply_info[m['responde_a']]

@app.route('/<slug>/api/canales/<int:cid>/mensajes')
def api_canales_mensajes(slug, cid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify([])
    conn = conectar(slug)
    canal = conn.execute('SELECT * FROM canales WHERE id=? AND activo=1', (cid,)).fetchone()
    if not canal: conn.close(); return jsonify([])
    if tipo != 'rector':
        miembro = conn.execute('SELECT 1 FROM canal_miembros WHERE canal_id=? AND usuario_tipo=? AND usuario_id=?',
                              (cid, tipo, uid)).fetchone()
        if not miembro: conn.close(); return jsonify([])
    mensajes = conn.execute('''
        SELECT m.*, COALESCE(ml.id,0) as leido
        FROM mensajes_canal m
        LEFT JOIN mensajes_leidos ml ON ml.mensaje_id=m.id AND ml.usuario_tipo=? AND ml.usuario_id=?
        WHERE m.canal_id=? AND m.eliminado=0 ORDER BY m.id ASC''', (tipo, uid, cid)).fetchall()
    result = [dict(r) for r in mensajes]
    _enriquecer_mensajes_batch(conn, result)
    conn.close()
    return jsonify(result)

@app.route('/<slug>/api/canales/<int:cid>/mensajes/nuevos')
def api_canales_mensajes_nuevos(slug, cid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autenticado'}), 401
    ultimo_id = request.args.get('ultimo_id', 0, type=int)
    conn = conectar(slug)
    canal = conn.execute('SELECT * FROM canales WHERE id=? AND activo=1', (cid,)).fetchone()
    if not canal: conn.close(); return jsonify({'ok':False,'error':'Canal no encontrado'})
    if tipo != 'rector':
        miembro = conn.execute('SELECT 1 FROM canal_miembros WHERE canal_id=? AND usuario_tipo=? AND usuario_id=?',
                              (cid, tipo, uid)).fetchone()
        if not miembro: conn.close(); return jsonify({'ok':False,'error':'No eres miembro'})
    mensajes = conn.execute('''
        SELECT m.*, COALESCE(ml.id,0) as leido
        FROM mensajes_canal m
        LEFT JOIN mensajes_leidos ml ON ml.mensaje_id=m.id AND ml.usuario_tipo=? AND ml.usuario_id=?
        WHERE m.canal_id=? AND m.id > ? AND m.eliminado=0 ORDER BY m.id ASC''',
        (tipo, uid, cid, ultimo_id)).fetchall()
    result = [dict(r) for r in mensajes]
    _enriquecer_mensajes_batch(conn, result)
    conn.close()
    return jsonify({'ok':True, 'mensajes':result})

@app.route('/<slug>/api/canales/<int:cid>/enviar', methods=['POST'])
def api_canales_enviar(slug, cid):
    if not validar_csrf(): return jsonify({'ok': False, 'error': 'Error CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'})
    mensaje = request.form.get('mensaje','').strip()
    responde_a = request.form.get('responde_a', type=int)
    tiene_archivos = 0
    conn = conectar(slug)
    canal = conn.execute('SELECT * FROM canales WHERE id=? AND activo=1', (cid,)).fetchone()
    if not canal: conn.close(); return jsonify({'ok':False,'error':'Canal no encontrado'})
    if tipo != 'rector':
        miembro = conn.execute('SELECT 1 FROM canal_miembros WHERE canal_id=? AND usuario_tipo=? AND usuario_id=?',
                              (cid, tipo, uid)).fetchone()
        if not miembro: conn.close(); return jsonify({'ok':False,'error':'No eres miembro'})
    mid = conn.execute(
        'INSERT INTO mensajes_canal (canal_id,usuario_tipo,usuario_id,mensaje,responde_a,tiene_archivos) VALUES (?,?,?,?,?,?)',
        (cid, tipo, uid, mensaje, responde_a, tiene_archivos)).lastrowid
    archivos_subidos = []
    if request.files:
        for key in request.files:
            f = request.files[key]
            if f and f.filename:
                fid, err = guardar_archivo_mensaje(slug, cid, f, tipo, uid)
                if fid:
                    conn.execute('UPDATE mensajes_archivos SET mensaje_id=? WHERE id=?', (mid, fid))
                    archivos_subidos.append(fid)
                    tiene_archivos = 1
    if tiene_archivos:
        conn.execute('UPDATE mensajes_canal SET tiene_archivos=1 WHERE id=?', (mid,))
    # Actualizar canal_actividad
    conn.execute('INSERT OR REPLACE INTO canal_actividad (canal_id, usuario_tipo, usuario_id, estado, ultima_vista) VALUES (?,?,?,?,?)',
                (cid, tipo, uid, 'online', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()
    return jsonify({'ok':True, 'mensaje_id':mid, 'archivos': archivos_subidos})

@app.route('/<slug>/api/canales/<int:cid>/leer', methods=['POST'])
def api_canales_leer(slug, cid):
    if not validar_csrf(): return jsonify({'ok': False, 'error': 'Error CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False})
    conn = conectar(slug)
    mids = [r['id'] for r in conn.execute('SELECT id FROM mensajes_canal WHERE canal_id=?', (cid,)).fetchall()]
    if mids:
        ph = ','.join('?' * len(mids))
        conn.execute(f'INSERT OR IGNORE INTO mensajes_leidos (mensaje_id,usuario_tipo,usuario_id) SELECT id,?,? FROM mensajes_canal WHERE canal_id=? AND id IN ({ph})',
                    (tipo, uid, cid) + tuple(mids))
    conn.commit()
    conn.close()
    return jsonify({'ok':True})

# ── FASE 5 – ARCHIVOS ─────────────────────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/archivos/subir', methods=['POST'])
def api_canales_subir_archivos(slug, cid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'}), 401
    if 'archivo' not in request.files: return jsonify({'ok':False,'error':'No hay archivo'}), 400
    f = request.files['archivo']
    if not f.filename: return jsonify({'ok':False,'error':'Archivo vacío'}), 400
    fid, err = guardar_archivo_mensaje(slug, cid, f, tipo, uid)
    if err: return jsonify({'ok':False,'error':err}), 400
    return jsonify({'ok':True, 'archivo_id': fid})

@app.route('/<slug>/api/archivos/<int:fid>/descargar')
def api_archivo_descargar(slug, fid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return 'No autorizado', 401
    conn = conectar(slug)
    arch = conn.execute('SELECT * FROM mensajes_archivos WHERE id=?', (fid,)).fetchone()
    conn.close()
    if not arch: return 'No encontrado', 404
    ruta = os.path.join(app.root_path, 'static', 'uploads', slug, arch['nombre_archivo'])
    if not os.path.exists(ruta): return 'No encontrado', 404
    return send_file(ruta, mimetype=arch['tipo_mime'], as_attachment=True,
                     download_name=arch['nombre_original'])

@app.route('/<slug>/api/archivos/<int:fid>/previsualizar')
def api_archivo_previsualizar(slug, fid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return 'No autorizado', 401
    conn = conectar(slug)
    arch = conn.execute('SELECT * FROM mensajes_archivos WHERE id=?', (fid,)).fetchone()
    conn.close()
    if not arch: return 'No encontrado', 404
    ruta = os.path.join(app.root_path, 'static', 'uploads', slug, arch['nombre_archivo'])
    if not os.path.exists(ruta): return 'No encontrado', 404
    if arch['es_imagen']:
        return send_file(ruta, mimetype=arch['tipo_mime'])
    if arch['tipo_mime'] == 'application/pdf':
        return send_file(ruta, mimetype='application/pdf')
    return jsonify({'ok':False,'error':'Vista previa no disponible'})

@app.route('/<slug>/api/archivos/<int:fid>/eliminar', methods=['DELETE','POST'])
def api_archivo_eliminar(slug, fid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'}), 401
    conn = conectar(slug)
    arch = conn.execute('SELECT * FROM mensajes_archivos WHERE id=?', (fid,)).fetchone()
    if not arch: conn.close(); return jsonify({'ok':False,'error':'No encontrado'}), 404
    if arch['usuario_tipo'] != tipo or arch['usuario_id'] != uid:
        if tipo != 'rector':
            conn.close(); return jsonify({'ok':False,'error':'No puedes eliminar este archivo'}), 403
    conn.execute('DELETE FROM mensajes_archivos WHERE id=?', (fid,))
    conn.commit()
    audit_log(slug, uid, 'delete', 'mensajes_archivos', fid,
              valor_anterior={'nombre_original': arch['nombre_original']})
    conn.close()
    ruta = os.path.join(app.root_path, 'static', 'uploads', slug, arch['nombre_archivo'])
    try: os.remove(ruta)
    except Exception: pass
    return jsonify({'ok':True})

# ── FASE 5 – REACCIONES ───────────────────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/reaccionar', methods=['POST'])
def api_canales_reaccionar(slug, cid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'}), 401
    mensaje_id = request.form.get('mensaje_id', type=int)
    reaccion = request.form.get('reaccion','').strip()
    if not mensaje_id or reaccion not in ('👍','✅','❓','📌','❤'):
        return jsonify({'ok':False,'error':'Reacción inválida'}), 400
    conn = conectar(slug)
    existing = conn.execute(
        'SELECT id FROM mensajes_reacciones WHERE mensaje_id=? AND usuario_tipo=? AND usuario_id=? AND reaccion=?',
        (mensaje_id, tipo, uid, reaccion)).fetchone()
    if existing:
        conn.execute('DELETE FROM mensajes_reacciones WHERE id=?', (existing['id'],))
        conn.commit(); conn.close()
        return jsonify({'ok':True, 'activo':False})
    conn.execute('INSERT OR IGNORE INTO mensajes_reacciones (mensaje_id,usuario_tipo,usuario_id,reaccion) VALUES (?,?,?,?)',
                (mensaje_id, tipo, uid, reaccion))
    conn.commit(); conn.close()
    return jsonify({'ok':True, 'activo':True})

# ── FASE 5 – MENSAJES FIJADOS ─────────────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/fijar', methods=['POST'])
def api_canales_fijar(slug, cid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'}), 401
    mensaje_id = request.form.get('mensaje_id', type=int)
    if not mensaje_id: return jsonify({'ok':False,'error':'mensaje_id requerido'}), 400
    conn = conectar(slug)
    existing = conn.execute('SELECT id FROM mensajes_fijados WHERE canal_id=? AND mensaje_id=?',
                           (cid, mensaje_id)).fetchone()
    if existing:
        conn.execute('DELETE FROM mensajes_fijados WHERE id=?', (existing['id'],))
        conn.commit(); conn.close()
        return jsonify({'ok':True, 'fijado':False})
    conn.execute('INSERT INTO mensajes_fijados (canal_id,mensaje_id,fijado_por_tipo,fijado_por_id) VALUES (?,?,?,?)',
                (cid, mensaje_id, tipo, uid))
    conn.commit(); conn.close()
    return jsonify({'ok':True, 'fijado':True})

@app.route('/<slug>/api/canales/<int:cid>/fijados')
def api_canales_fijados(slug, cid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify([])
    conn = conectar(slug)
    rows = conn.execute(
        '''SELECT m.id, m.mensaje, m.usuario_tipo, m.usuario_id, m.fecha, m.editado,
                  f.fecha as fijado_en, f.fijado_por_tipo, f.fijado_por_id
           FROM mensajes_fijados f
           JOIN mensajes_canal m ON m.id=f.mensaje_id
           WHERE f.canal_id=? AND m.eliminado=0
           ORDER BY f.id DESC''', (cid,)).fetchall()
    result = [dict(r) for r in rows]
    for m in result:
        m['autor_nombre'] = nombre_usuario_canal(conn, m['usuario_tipo'], m['usuario_id'])
    conn.close()
    return jsonify(result)

# ── FASE 5 – BIBLIOTECA DEL CANAL ─────────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/biblioteca')
def api_canales_biblioteca(slug, cid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({})
    conn = conectar(slug)
    archivos = [dict(r) for r in conn.execute(
        'SELECT * FROM mensajes_archivos WHERE canal_id=? ORDER BY fecha DESC LIMIT 50', (cid,)).fetchall()]
    enlaces = [dict(r) for r in conn.execute(
        'SELECT * FROM canal_enlaces WHERE canal_id=? ORDER BY fecha DESC LIMIT 50', (cid,)).fetchall()]
    conn.close()
    return jsonify({'archivos': archivos, 'enlaces': enlaces})

# ── FASE 5 – BUSCAR EN EL CANAL ───────────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/buscar')
def api_canales_buscar(slug, cid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify([])
    q = request.args.get('q','').strip()
    autor = request.args.get('autor','').strip()
    desde = request.args.get('desde','').strip()
    hasta = request.args.get('hasta','').strip()
    conn = conectar(slug)
    sql = 'SELECT m.* FROM mensajes_canal m WHERE m.canal_id=? AND m.eliminado=0'
    params = [cid]
    if q:
        sql += ' AND m.mensaje LIKE ?'
        params.append(f'%{q}%')
    if autor:
        sql += ' AND (SELECT nombre FROM profesores WHERE id=m.usuario_id AND m.usuario_tipo=\'profesor\') LIKE ?'
        params.append(f'%{autor}%')
    if desde:
        sql += ' AND m.fecha >= ?'
        params.append(desde)
    if hasta:
        sql += ' AND m.fecha <= ?'
        params.append(hasta + ' 23:59:59')
    sql += ' ORDER BY m.id DESC LIMIT 100'
    rows = conn.execute(sql, params).fetchall()
    result = [dict(r) for r in rows]
    _enriquecer_mensajes_batch(conn, result)
    conn.close()
    return jsonify(result)

# ── FASE 5 – EDITAR / ELIMINAR MENSAJES ────────────────────────────────────────
TIEMPO_EDICION_SEGUNDOS = 300  # 5 minutos, configurable

@app.route('/<slug>/api/canales/<int:cid>/editar/<int:mid>', methods=['POST'])
def api_canales_editar(slug, cid, mid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'}), 401
    conn = conectar(slug)
    msg = conn.execute('SELECT * FROM mensajes_canal WHERE id=? AND canal_id=?', (mid, cid)).fetchone()
    if not msg: conn.close(); return jsonify({'ok':False,'error':'No encontrado'}), 404
    if msg['usuario_tipo'] != tipo or msg['usuario_id'] != uid:
        conn.close(); return jsonify({'ok':False,'error':'No puedes editar este mensaje'}), 403
    if msg['eliminado']:
        conn.close(); return jsonify({'ok':False,'error':'Mensaje eliminado'}), 400
    try:
        creado = datetime.strptime(msg['fecha'], '%Y-%m-%d %H:%M:%S') if msg['fecha'] else datetime.min
    except (ValueError, TypeError):
        creado = datetime.min
    if (datetime.now() - creado).total_seconds() > TIEMPO_EDICION_SEGUNDOS:
        conn.close(); return jsonify({'ok':False,'error':'Tiempo de edición expirado'}), 400
    nuevo_texto = request.form.get('mensaje','').strip()
    if not nuevo_texto: conn.close(); return jsonify({'ok':False,'error':'Mensaje vacío'}), 400
    viejo_texto = msg['mensaje']
    conn.execute('UPDATE mensajes_canal SET mensaje=?, editado=editado+1, editado_en=? WHERE id=?',
                (nuevo_texto, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), mid))
    conn.commit()
    audit_log(slug, uid, 'update', 'mensajes_canal', mid,
              valor_anterior={'mensaje': viejo_texto},
              valor_nuevo={'mensaje': nuevo_texto})
    conn.close()
    return jsonify({'ok':True})

@app.route('/<slug>/api/canales/<int:cid>/eliminar/<int:mid>', methods=['DELETE','POST'])
def api_canales_eliminar(slug, cid, mid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'}), 401
    conn = conectar(slug)
    msg = conn.execute('SELECT * FROM mensajes_canal WHERE id=? AND canal_id=?', (mid, cid)).fetchone()
    if not msg: conn.close(); return jsonify({'ok':False,'error':'No encontrado'}), 404
    if msg['usuario_tipo'] != tipo or msg['usuario_id'] != uid:
        conn.close(); return jsonify({'ok':False,'error':'No puedes eliminar este mensaje'}), 403
    conn.execute('UPDATE mensajes_canal SET eliminado=1 WHERE id=?', (mid,))
    conn.commit()
    audit_log(slug, uid, 'delete', 'mensajes_canal', mid,
              valor_anterior={'mensaje': msg['mensaje'][:200]})
    conn.close()
    return jsonify({'ok':True})

# ── FASE 5 – LECTURAS ──────────────────────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/lecturas')
def api_canales_lecturas(slug, cid):
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify([])
    conn = conectar(slug)
    miembros = conn.execute(
        'SELECT usuario_tipo, usuario_id FROM canal_miembros WHERE canal_id=?', (cid,)).fetchall()
    total_msg = conn.execute(
        'SELECT COUNT(*) as c FROM mensajes_canal WHERE canal_id=? AND eliminado=0', (cid,)).fetchone()['c']
    ult_vistas = {}
    for row in conn.execute(
        'SELECT usuario_tipo, usuario_id, ultima_vista FROM canal_actividad WHERE canal_id=?',
        (cid,)).fetchall():
        ult_vistas[f"{row['usuario_tipo']}_{row['usuario_id']}"] = row['ultima_vista']
    leidos_por_miembro = {}
    for row in conn.execute(
        '''SELECT ml.usuario_tipo, ml.usuario_id, COUNT(DISTINCT mc.id) as c
           FROM mensajes_leidos ml
           JOIN mensajes_canal mc ON ml.mensaje_id=mc.id
           WHERE mc.canal_id=? AND mc.eliminado=0
           GROUP BY ml.usuario_tipo, ml.usuario_id''',
        (cid,)).fetchall():
        leidos_por_miembro[f"{row['usuario_tipo']}_{row['usuario_id']}"] = row['c']
    # batch name lookups
    seen = set()
    tipo_ids = {'profesor': set(), 'estudiante': set(), 'rector': set(), 'directora': set()}
    for m in miembros:
        key = (m['usuario_tipo'], m['usuario_id'])
        if key not in seen:
            seen.add(key)
            if m['usuario_tipo'] in tipo_ids:
                tipo_ids[m['usuario_tipo']].add(m['usuario_id'])
    name_map = {}
    table_map = {'profesor': 'profesores', 'estudiante': 'alumnos', 'rector': 'rectores', 'directora': 'directoras'}
    for t, ids in tipo_ids.items():
        if not ids: continue
        ph2 = ','.join('?' * len(ids))
        rows = conn.execute(f'SELECT id, nombre FROM {table_map[t]} WHERE id IN ({ph2})', list(ids)).fetchall()
        for r in rows:
            name_map[(t, r['id'])] = r['nombre']
    result = {}
    for m in miembros:
        key = f"{m['usuario_tipo']}_{m['usuario_id']}"
        result[key] = {
            'nombre': name_map.get((m['usuario_tipo'], m['usuario_id']), 'Desconocido'),
            'tipo': m['usuario_tipo'],
            'total': total_msg,
            'leidos': leidos_por_miembro.get(key, 0),
            'ultima_vista': ult_vistas.get(key),
        }
    conn.close()
    return jsonify(result)

# ── FASE 5 – ESCRIBIENDO / ACTIVIDAD ──────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/escribiendo', methods=['POST'])
def api_canales_escribiendo(slug, cid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False})
    conn = conectar(slug)
    conn.execute('INSERT OR REPLACE INTO canal_actividad (canal_id,usuario_tipo,usuario_id,estado,ultima_vista) VALUES (?,?,?,?,?)',
                (cid, tipo, uid, 'typing', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/<slug>/api/canales/<int:cid>/actividad')
def api_canales_actividad(slug, cid):
    require_colegio(slug)
    conn = conectar(slug)
    rows = conn.execute(
        '''SELECT ca.*
           FROM canal_actividad ca
           WHERE ca.canal_id=?''', (cid,)).fetchall()
    result = {}
    now = datetime.now()
    for r in rows:
        estado = r['estado']
        ult_vista = datetime.strptime(r['ultima_vista'], '%Y-%m-%d %H:%M:%S') if r['ultima_vista'] else None
        if estado == 'typing' and ult_vista and (now - ult_vista).total_seconds() > 8:
            estado = 'online'
        if ult_vista and (now - ult_vista).total_seconds() > 120:
            estado = 'offline'
        nombre = nombre_usuario_canal(conn, r['usuario_tipo'], r['usuario_id'])
        key = f"{r['usuario_tipo']}_{r['usuario_id']}"
        result[key] = {'estado': estado, 'nombre': nombre, 'ultima_vista': r['ultima_vista']}
    conn.close()
    return jsonify(result)

# ── FASE 5 – GUARDAR ENLACE ───────────────────────────────────────────────────
@app.route('/<slug>/api/canales/<int:cid>/enlaces', methods=['POST'])
def api_canales_guardar_enlace(slug, cid):
    if not validar_csrf(): return jsonify({'ok':False,'error':'CSRF'}), 403
    require_colegio(slug)
    tipo, uid = get_usuario_actual(slug)
    if not tipo: return jsonify({'ok':False,'error':'No autorizado'}), 401
    url = request.form.get('url','').strip()
    titulo = request.form.get('titulo','').strip()
    if not url: return jsonify({'ok':False,'error':'URL requerida'}), 400
    conn = conectar(slug)
    conn.execute('INSERT INTO canal_enlaces (canal_id,titulo,url,agregado_por_tipo,agregado_por_id) VALUES (?,?,?,?,?)',
                (cid, titulo or url, url, tipo, uid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ── API COMUNICACIONES (Fase 2 — polling) ─────────────────────────────────────
@app.route('/<slug>/api/comunicaciones')
def api_comunicaciones(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if prof:
        return jsonify(comunicaciones_pendientes(slug,'profesor',prof['id']))
    aid = session.get(f'alumno_id_{slug}')
    if aid:
        return jsonify(comunicaciones_pendientes(slug,'estudiante',aid))
    return jsonify([])

@app.route('/<slug>/api/comunicaciones/count')
def api_comunicaciones_count(slug):
    require_colegio(slug)
    prof = get_profesor(slug)
    if prof:
        return jsonify({'pendientes':len(comunicaciones_pendientes(slug,'profesor',prof['id']))})
    aid = session.get(f'alumno_id_{slug}')
    if aid:
        return jsonify({'pendientes':len(comunicaciones_pendientes(slug,'estudiante',aid))})
    return jsonify({'pendientes':0})

# ── EVENTOS CALENDARIO ────────────────────────────────────────────────────────
@app.route('/<slug>/rector/comunicaciones/<int:cid>/evento')
def rector_comunicacion_evento(slug, cid):
    """Return JSON for calendar integration."""
    require_colegio(slug)
    rector = get_rector(slug)
    if not rector: return jsonify({'error': 'No autorizado'}), 403
    conn = conectar(slug)
    com = conn.execute(
        'SELECT id, titulo, fecha_programada, prioridad FROM comunicaciones WHERE id=? AND rector_id=? AND activo=1',
        (cid, rector['id'])).fetchone()
    conn.close()
    if not com: return jsonify({'error': 'No encontrada'}), 404
    return jsonify({
        'id': com['id'],
        'title': com['titulo'],
        'start': com['fecha_programada'] or datetime.today().strftime('%Y-%m-%d'),
        'className': 'event-' + com['prioridad']
    })

# ── GESTIÓN DE RECTORES (SÓLO PRINCIPAL) ───────────────────────────────────────
def require_rector_principal(slug):
    r = get_rector(slug)
    if not r: abort(401)
    if not r['es_principal']: abort(403)
    return r

@app.route('/<slug>/rector/gestion-rectores')
def rector_gestion(slug):
    r = require_rector_principal(slug)
    colegio = get_colegio(slug)
    conn = conectar(slug)
    rectores = conn.execute(
        'SELECT id, nombre, usuario, email, activo, es_principal FROM rectores ORDER BY es_principal DESC, id').fetchall()
    notif_count = notificaciones_no_leidas(slug, 'rector', r['id'])
    conn.close()
    return render_template('rector_gestion.html',
                           slug=slug, colegio=colegio, rector=r,
                           rectores=rectores, notif_count=notif_count)

@app.route('/<slug>/rector/gestion-rectores/crear', methods=['GET', 'POST'])
def rector_gestion_crear(slug):
    r = require_rector_principal(slug)
    colegio = get_colegio(slug)
    error = exito = None
    if request.method == 'POST':
        if not validar_csrf(): return 'Error de seguridad', 400
        nombre = request.form.get('nombre', '').strip()
        usuario = request.form.get('usuario', '').strip()
        password = request.form.get('password', '').strip()
        confirmar = request.form.get('confirmar_password', '').strip()
        email = request.form.get('email', '').strip()
        if not nombre or not usuario or not password:
            error = 'Completa todos los campos obligatorios.'
        elif len(password) < 6:
            error = 'Mínimo 6 caracteres para la contraseña.'
        elif password != confirmar:
            error = 'Las contraseñas no coinciden.'
        else:
            conn = conectar(slug)
            if conn.execute('SELECT 1 FROM rectores WHERE usuario=?', (usuario,)).fetchone():
                error = 'Ese usuario ya existe.'
            else:
                conn.execute(
                    'INSERT INTO rectores (nombre, usuario, password, email) VALUES (?, ?, ?, ?)',
                    (nombre, usuario, hash_pw(password), email))
                conn.commit()
                exito = f'Rector "{nombre}" creado correctamente.'
                crear_notificacion(slug, 'rector', r['id'],
                    'Nuevo rector creado', f'Se creó el rector {nombre} ({usuario}).', 'success')
            conn.close()
    return render_template('rector_gestion.html',
                           slug=slug, colegio=colegio, rector=r,
                           error=error, exito=exito, crear=True,
                           notif_count=notificaciones_no_leidas(slug, 'rector', r['id']))

@app.route('/<slug>/rector/gestion-rectores/<int:rid>/editar', methods=['GET', 'POST'])
def rector_gestion_editar(slug, rid):
    r = require_rector_principal(slug)
    colegio = get_colegio(slug)
    conn = conectar(slug)
    target = conn.execute('SELECT * FROM rectores WHERE id=?', (rid,)).fetchone()
    if not target: conn.close(); return 'Rector no encontrado', 404
    error = exito = None
    if request.method == 'POST':
        if not validar_csrf(): return 'Error de seguridad', 400
        nombre = request.form.get('nombre', '').strip()
        usuario = request.form.get('usuario', '').strip()
        password = request.form.get('password', '').strip()
        confirmar = request.form.get('confirmar_password', '').strip()
        email = request.form.get('email', '').strip()
        if not nombre or not usuario:
            error = 'Nombre y usuario son obligatorios.'
        elif password and len(password) < 6:
            error = 'Mínimo 6 caracteres.'
        elif password and password != confirmar:
            error = 'Las contraseñas no coinciden.'
        else:
            existing = conn.execute('SELECT 1 FROM rectores WHERE usuario=? AND id!=?', (usuario, rid)).fetchone()
            if existing:
                error = 'Ese nombre de usuario ya está en uso.'
            else:
                if password:
                    conn.execute(
                        'UPDATE rectores SET nombre=?, usuario=?, password=?, email=? WHERE id=?',
                        (nombre, usuario, hash_pw(password), email, rid))
                else:
                    conn.execute(
                        'UPDATE rectores SET nombre=?, usuario=?, email=? WHERE id=?',
                        (nombre, usuario, email, rid))
                conn.commit()
                exito = 'Rector actualizado correctamente.'
                target = conn.execute('SELECT * FROM rectores WHERE id=?', (rid,)).fetchone()
    conn.close()
    return render_template('rector_gestion.html',
                           slug=slug, colegio=colegio, rector=r,
                           error=error, exito=exito, editar=target,
                           notif_count=notificaciones_no_leidas(slug, 'rector', r['id']))

@app.route('/<slug>/rector/gestion-rectores/<int:rid>/toggle', methods=['POST'])
def rector_gestion_toggle(slug, rid):
    r = require_rector_principal(slug)
    if not validar_csrf(): return 'Error de seguridad', 400
    if rid == r['id']: return 'No puedes desactivarte a ti mismo.', 400
    conn = conectar(slug)
    conn.execute('UPDATE rectores SET activo = CASE WHEN activo=1 THEN 0 ELSE 1 END WHERE id=?', (rid,))
    conn.commit()
    conn.close()
    return redirect(url_for('rector_gestion', slug=slug))

@app.route('/<slug>/rector/gestion-rectores/<int:rid>/eliminar', methods=['POST'])
def rector_gestion_eliminar(slug, rid):
    r = require_rector_principal(slug)
    if not validar_csrf(): return 'Error de seguridad', 400
    if rid == r['id']: return 'No puedes eliminar tu propia cuenta.', 400
    conn = conectar(slug)
    target = conn.execute('SELECT es_principal FROM rectores WHERE id=?', (rid,)).fetchone()
    if not target: conn.close(); return 'Rector no encontrado', 404
    if target['es_principal']:
        conn.close(); return 'No puedes eliminar al Rector Principal. Transfiere el rol primero.', 400
    conn.execute('DELETE FROM rectores WHERE id=?', (rid,))
    conn.commit()
    conn.close()
    return redirect(url_for('rector_gestion', slug=slug))

@app.route('/<slug>/rector/gestion-rectores/<int:rid>/hacer-principal', methods=['POST'])
def rector_gestion_hacer_principal(slug, rid):
    r = require_rector_principal(slug)
    if not validar_csrf(): return 'Error de seguridad', 400
    if rid == r['id']: return 'Ya eres el Rector Principal.', 400
    conn = conectar(slug)
    target = conn.execute('SELECT id, activo FROM rectores WHERE id=?', (rid,)).fetchone()
    if not target: conn.close(); return 'Rector no encontrado', 404
    if not target['activo']:
        conn.close(); return 'No puedes transferir el rol a un rector inactivo.', 400
    conn.execute('UPDATE rectores SET es_principal=0 WHERE id=?', (r['id'],))
    conn.execute('UPDATE rectores SET es_principal=1 WHERE id=?', (rid,))
    conn.commit()
    conn.close()
    session[f'rector_id_{slug}'] = rid
    crear_notificacion(slug, 'rector', rid,
        'Rector Principal transferido', f'{r["nombre"]} te ha transferido el rol de Rector Principal.', 'warning')
    return redirect(url_for('rector_panel', slug=slug))

# ── Directora auth routes migrated to app/routes/auth.py ────────────────

@app.route('/<slug>/directora')
@app.route('/<slug>/directora/panel')
def directora_panel(slug):
    require_colegio(slug)
    directora = get_directora(slug)
    if not directora: return redirect(url_for('auth.directora_login', slug=slug))
    colegio  = get_colegio(slug)
    curso    = directora['curso']
    jornada  = directora['jornada']
    periodo  = request.args.get('periodo', 1, type=int)
    num_periodos = int(colegio['num_periodos']) if colegio and colegio['num_periodos'] else 4
    conn = conectar(slug)
    alumnos = conn.execute(
        'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
        (curso, jornada)).fetchall()
    lista_materias = [r['materia'] for r in conn.execute(
        'SELECT DISTINCT materia FROM actividades WHERE curso=? AND jornada=? AND COALESCE(periodo,1)=? ORDER BY materia',
        (curso, jornada, periodo)).fetchall()]
    profs_raw = conn.execute(
        '''SELECT DISTINCT am.materia, p.nombre,
           (SELECT COUNT(*) FROM actividades a
            WHERE a.profesor_id=p.id AND a.curso=? AND a.jornada=?
            AND COALESCE(a.periodo,1)=?) as cnt
           FROM profesores p
           JOIN asignaciones_curso ac ON ac.profesor_id=p.id
           JOIN asignaciones_materia am ON am.profesor_id=p.id AND am.jornada=ac.jornada AND am.materia=ac.materia
           WHERE ac.curso=? AND ac.jornada=? AND p.activo=1''',
        (curso, jornada, periodo, curso, jornada)).fetchall()
    materias_enviadas = set()
    profesores = []
    for p in profs_raw:
        enviado = p['cnt'] > 0
        if enviado: materias_enviadas.add(p['materia'])
        profesores.append({'materia': p['materia'], 'nombre': p['nombre'],
                           'enviado': enviado, 'fecha_envio': None})
    # Pre-fetch all notas and evaluaciones for this course/period (avoids N+1)
    aid_alumno = {a['id'] for a in alumnos}
    notas_all = conn.execute(
        '''SELECT n.aid, ac.materia, n.val FROM notas n
           JOIN actividades ac ON ac.id=n.actividad_id
           WHERE ac.curso=? AND ac.jornada=? AND COALESCE(ac.periodo,1)=?
           ORDER BY n.aid, ac.materia''',
        (curso, jornada, periodo)).fetchall()
    notas_by = {}
    for r in notas_all:
        notas_by.setdefault((r['aid'], r['materia']), []).append(r['val'])
    if aid_alumno:
        evals_all = conn.execute(
            '''SELECT aid, materia, evaluacion, autoevaluacion FROM evaluaciones
               WHERE aid IN ({}) AND COALESCE(periodo,1)=?'''.format(
                   ','.join('?' * len(aid_alumno))),
            (*aid_alumno, periodo)).fetchall()
    else:
        evals_all = []
    evals_by = {}
    for r in evals_all:
        evals_by[(r['aid'], r['materia'])] = r

    tabla = []
    for a in alumnos:
        fila = {'id': a['id'], 'nombre': a['nombre'],
                'email': a['email_acudiente'] or '', 'materias': {}, 'promedio': None}
        todos_finales = []
        for mat in lista_materias:
            notas_vals = notas_by.get((a['id'], mat), [])
            ev = evals_by.get((a['id'], mat))
            eval_v   = ev['evaluacion']     if ev and ev['evaluacion']     is not None else None
            auto_v   = ev['autoevaluacion'] if ev and ev['autoevaluacion'] is not None else None
            final = _promedio_ponderado(notas_vals, eval_v, auto_v)
            act_prom = round(sum(notas_vals) / len(notas_vals), 2) if notas_vals else None
            fila['materias'][mat] = {'act': act_prom, 'eval': eval_v, 'auto': auto_v, 'final': final}
            if final is not None: todos_finales.append(final)
        fila['promedio'] = round(sum(todos_finales) / len(todos_finales), 2) if todos_finales else None
        tabla.append(fila)

    # ── Dashboard extras ────────────────────────────────────────────────
    actividad_reciente = conn.execute(
        '''SELECT accion, tabla, creado
           FROM audit_log ORDER BY creado DESC LIMIT 6''').fetchall()
    actividad_reciente = [dict(r) for r in actividad_reciente]

    notif_count = conn.execute(
        'SELECT COUNT(*) as c FROM notificaciones WHERE usuario_tipo=? AND usuario_id=? AND leida=0',
        ('directora', directora['id'])).fetchone()['c']

    aprobados = sum(1 for f in tabla if f['promedio'] is not None and f['promedio'] >= 3.0)
    reprobados = sum(1 for f in tabla if f['promedio'] is not None and f['promedio'] < 3.0)
    sin_notas = sum(1 for f in tabla if f['promedio'] is None)

    conn.close()
    return render_template('directora_panel.html',
                           slug=slug, colegio=colegio, directora=directora,
                           curso=curso, jornada=jornada, periodo=periodo,
                           num_periodos=num_periodos,
                           lista_materias=lista_materias,
                           materias_enviadas=materias_enviadas,
                           profesores=profesores, tabla=tabla,
                           actividad_reciente=actividad_reciente,
                           notif_count=notif_count,
                           aprobados=aprobados, reprobados=reprobados,
                           sin_notas=sin_notas)

@app.route('/<slug>/directora/boletin_pdf')
def directora_boletin_pdf(slug):
    require_colegio(slug)
    directora = get_directora(slug)
    if not directora: return redirect(url_for('auth.directora_login', slug=slug))
    from flask import Response
    colegio  = get_colegio(slug)
    curso    = directora['curso']
    jornada  = directora['jornada']
    periodo  = request.args.get('periodo', 1, type=int)
    aid_solo = request.args.get('aid', type=int)
    conn = conectar(slug)
    if aid_solo:
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE id=? AND curso=?', (aid_solo, curso)).fetchall()
    else:
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
            (curso, jornada)).fetchall()
    all_pdfs = []
    for alumno in alumnos:
        try:
            pdf_bytes, _ = generar_pdf_alumno(alumno, slug, colegio, curso, jornada, periodo, conn)
        except ImportError:
            return render_template('error.html',
                                   codigo=501,
                                   mensaje='La generación de PDF requiere la librería <strong>reportlab</strong>. '
                                           'Consulte al administrador del sistema para instalarla.')
        all_pdfs.append(pdf_bytes)
    conn.close()
    if not all_pdfs: return ('Sin alumnos', 404)
    if len(all_pdfs) == 1:
        return Response(all_pdfs[0], mimetype='application/pdf',
                        headers={'Content-Disposition':
                                 f'attachment;filename=boletin_{curso}_{jornada}_P{periodo}.pdf'})
    try:
        from pypdf import PdfWriter, PdfReader
        writer = PdfWriter()
        for pdf_bytes in all_pdfs:
            reader = PdfReader(BytesIO(pdf_bytes))
            for page in reader.pages:
                writer.add_page(page)
        out = BytesIO()
        writer.write(out); out.seek(0)
        return Response(out, mimetype='application/pdf',
                        headers={'Content-Disposition':
                                 f'attachment;filename=boletin_{curso}_{jornada}_P{periodo}.pdf'})
    except ImportError:
        return Response(all_pdfs[0], mimetype='application/pdf',
                        headers={'Content-Disposition':
                                 f'attachment;filename=boletin_{curso}_{jornada}_P{periodo}.pdf'})

# ── Directora logout migrated to app/routes/auth.py ────────────────────

@app.route('/<slug>/directora/enviar_correos', methods=['POST'])
def directora_enviar_correos(slug):
    require_colegio(slug)
    directora = get_directora(slug)
    if not directora: return jsonify({'ok': False, 'mensaje': 'No autorizado'})
    if not validar_csrf(): return jsonify({'ok': False, 'mensaje': 'Error CSRF'})
    if not SENDGRID_API_KEY:
        return jsonify({'ok': False, 'mensaje': 'Envío de correos no configurado (falta SENDGRID_API_KEY).'})
    import base64
    colegio  = get_colegio(slug)
    curso    = directora['curso']
    jornada  = directora['jornada']
    periodo  = int(request.form.get('periodo', 1))
    aid_solo = request.form.get('aid', type=int)
    conn     = conectar(slug)
    if aid_solo:
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE id=? AND curso=?', (aid_solo, curso)).fetchall()
    else:
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
            (curso, jornada)).fetchall()
    enviados = fallidos = sin_correo = 0
    for alumno in alumnos:
        email_dest = alumno['email_acudiente'] if alumno['email_acudiente'] else None
        if not email_dest: sin_correo += 1; continue
        try:
            pdf_bytes, prom_general = generar_pdf_alumno(
                alumno, slug, colegio, curso, jornada, periodo, conn)
        except Exception as e:
            logger.error(f'Error generando PDF para {alumno["nombre"]}: {e}')
            fallidos += 1; continue
        asunto = f'Boletín de Notas — {alumno["nombre"]} · Periodo {periodo}'
        try:
            pri_hex = colegio['primary_color'] if colegio and colegio['primary_color'] else '#6c63ff'
        except (KeyError, AttributeError, TypeError):
            pri_hex = '#6c63ff'
        cuerpo = f'''<div style="font-family:sans-serif;max-width:500px;margin:0 auto;">
            <h2 style="color:{pri_hex};">LUMINI — Boletín de Notas</h2>
            <p>Estimado acudiente,</p>
            <p>Adjunto encontrará el boletín de notas de <strong>{html.escape(str(alumno['nombre']))}</strong>
               correspondiente al <strong>Periodo {periodo}</strong>.</p>
            <p><strong>Promedio general: {prom_general}</strong></p>
            <p style="color:#888;font-size:12px;">
               {html.escape(str(colegio['nombre'] if colegio else slug))} · {curso} · {jornada}</p>
        </div>'''
        adj_nombre = f'boletin_{alumno["nombre"].replace(" ", "_")}_P{periodo}.pdf'
        if enviar_correo(email_dest, asunto, cuerpo, pdf_bytes, adj_nombre, 'application/pdf'):
            enviados += 1
            logger.info(f'Boletín enviado a {email_dest} para {alumno["nombre"]}')
        else:
            fallidos += 1
    conn.close()
    partes = []
    if enviados:   partes.append(f'✅ {enviados} enviado(s)')
    if fallidos:   partes.append(f'❌ {fallidos} fallido(s)')
    if sin_correo: partes.append(f'⚠️ {sin_correo} sin correo registrado')
    return jsonify({'ok': fallidos == 0, 'mensaje': ' · '.join(partes) or 'Sin destinatarios'})

@app.route('/<slug>/directora/guardar_email', methods=['POST'])
def directora_guardar_email(slug):
    require_colegio(slug)
    directora = get_directora(slug)
    if not directora: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    aid   = request.form.get('aid', type=int)
    email = request.form.get('email', '').strip()
    conn  = conectar(slug)
    conn.execute('UPDATE alumnos SET email_acudiente=? WHERE id=?', (email, aid))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/<slug>/directora/crear_desde_panel', methods=['POST'])
def directora_crear_desde_panel(slug):
    require_colegio(slug)
    directora = get_directora(slug)
    if not directora: return jsonify({'ok': False, 'mensaje': 'No autorizado'})
    if not validar_csrf(): return jsonify({'ok': False, 'mensaje': 'Error CSRF'})
    migrar_db(slug)
    nombre  = request.form.get('nombre', '').strip()
    usuario = request.form.get('usuario', '').strip()
    pw      = request.form.get('password', '').strip()
    curso   = request.form.get('curso', '').strip()
    email   = request.form.get('email', '').strip()
    jornada = directora['jornada']
    if not nombre or not usuario or not pw or not curso:
        return jsonify({'ok': False, 'mensaje': 'Completa todos los campos.'})
    if len(pw) < 6:
        return jsonify({'ok': False, 'mensaje': 'Mínimo 6 caracteres.'})
    conn = conectar(slug)
    if conn.execute('SELECT 1 FROM directoras WHERE usuario=?', (usuario,)).fetchone():
        conn.close()
        return jsonify({'ok': False, 'mensaje': 'Ese usuario ya existe.'})
    conn.execute(
        'INSERT INTO directoras (nombre,usuario,password,curso,jornada,email) VALUES (?,?,?,?,?,?)',
        (nombre, usuario, hash_pw(pw), curso, jornada, email))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'mensaje': f'Cuenta creada para {nombre}.'})

# ── STATIC / ROOT / ERRORS ────────────────────────────────────────────────────
@app.route('/static/<path:filename>')
def static_files(filename):
    resp = send_from_directory(
        os.path.join(os.path.dirname(__file__), 'static'), filename)
    resp.headers['Cache-Control'] = 'public, max-age=604800, immutable'
    return resp

@app.route("/offline")
def offline():
    return render_template("offline.html")

@app.route("/")
def index():
    conn = conectar_master()
    colegios = conn.execute("SELECT slug, nombre, logo FROM colegios WHERE activo=1 ORDER BY nombre").fetchall()
    conn.close()
    return render_template("index_root.html", colegios=colegios)

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    if app.config.get('SESSION_COOKIE_SECURE'):
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://unpkg.com https://www.datadoghq-browser-agent.com; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; font-src 'self' https://fonts.gstatic.com; img-src 'self' data:; connect-src 'self'"
    return response

@app.errorhandler(400)
def bad_request(e):
    return render_template('error.html', codigo=400, mensaje='Solicitud inválida.'), 400

@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', codigo=403, mensaje='Acceso denegado.'), 403

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', codigo=404, mensaje='Página no encontrada.'), 404

@app.errorhandler(405)
def method_not_allowed(e):
    return render_template('error.html', codigo=405, mensaje='Método no permitido.'), 405

@app.errorhandler(413)
def too_large(e):
    return render_template('error.html', codigo=413,
                           mensaje='El archivo es demasiado grande. Máximo permitido: 2 MB.'), 413

@app.errorhandler(429)
def too_many_requests(e):
    return render_template('error.html', codigo=429,
                           mensaje='Demasiadas solicitudes. Intenta de nuevo más tarde.'), 429

@app.errorhandler(500)
def server_error(e):
    logger.error(f'Error interno: {e}')
    return render_template('error.html', codigo=500,
                           mensaje='Error interno del servidor. Intenta de nuevo más tarde.'), 500

@app.errorhandler(502)
def bad_gateway(e):
    return render_template('error.html', codigo=502, mensaje='Servicio temporalmente no disponible.'), 502

@app.errorhandler(503)
def service_unavailable(e):
    return render_template('error.html', codigo=503, mensaje='Servicio en mantenimiento.'), 503

# ── FILTRO DÍAS RESTANTES ─────────────────────────────────────────────────────
from datetime import date as _date

@app.template_filter('dias_restantes')
def dias_restantes(fecha_str):
    try:
        fecha = _date.fromisoformat(str(fecha_str))
        return (fecha - _date.today()).days
    except Exception:
        return None

@app.template_filter('hex_to_rgb')
def hex_to_rgb(hex_color):
    """Converts #RRGGBB to 'r,g,b' string for use in rgba()."""
    h = hex_color.lstrip('#')
    if len(h) != 6:
        return '108,99,255'
    return f'{int(h[0:2], 16)},{int(h[2:4], 16)},{int(h[4:6], 16)}'

# ── ADMIN CÓDIGOS DE INVITACIÓN ──────────────────────────────────────────────
@app.route('/admin/codigos', methods=['GET', 'POST'])
@app.route('/admin/codigos/<slug>', methods=['GET', 'POST'])
def admin_codigos(slug=None):
    if not session.get('admin_auth'):
        return redirect(url_for('auth.admin'))
    cm = conectar_master()
    error = exito = None

    if request.method == 'POST':
        if not validar_csrf():
            return 'Error de seguridad', 400
        accion = request.form.get('accion')
        if accion == 'actualizar_codigos':
            s = request.form.get('slug', '').strip()
            cod_prof = request.form.get('codigo_profesores', '').strip()
            cod_dir  = request.form.get('codigo_directoras', '').strip()
            cod_rec  = request.form.get('codigo_rectores', '').strip()
            cm.execute(
                'UPDATE colegios SET codigo_profesores=?, codigo_directoras=?, codigo_rectores=? WHERE slug=?',
                (cod_prof, cod_dir, cod_rec, s))
            cm.commit()
            exito = 'Códigos actualizados correctamente.'
            slug = s
        elif accion == 'generar_codigos':
            s = request.form.get('slug', '').strip()
            prefijo = request.form.get('prefijo', '').strip()
            if not prefijo:
                error = 'Elige un prefijo para los códigos.'
            else:
                import secrets as sec
                new_prof = f'{prefijo}_prof_{sec.token_hex(4)}'
                new_dir  = f'{prefijo}_dir_{sec.token_hex(4)}'
                new_rec  = f'{prefijo}_rec_{sec.token_hex(4)}'
                cm.execute(
                    'UPDATE colegios SET codigo_profesores=?, codigo_directoras=?, codigo_rectores=? WHERE slug=?',
                    (new_prof, new_dir, new_rec, s))
                cm.commit()
                exito = f'Códigos generados para {s}: Profesores={new_prof}, Directoras={new_dir}, Rectores={new_rec}'
                slug = s

    colegios = cm.execute('SELECT * FROM colegios ORDER BY nombre').fetchall()
    colegio_selected = None
    if slug:
        colegio_selected = cm.execute('SELECT * FROM colegios WHERE slug=?', (slug,)).fetchone()
    cm.close()
    return render_template('admin_codigos.html',
                           colegios=colegios, colegio=colegio_selected,
                           error=error, exito=exito)

# ── API PROFESORES PARA ADMIN ─────────────────────────────────────────────────
@app.route('/admin/profesores/<slug>')
def admin_ver_profesores(slug):
    if not session.get('admin_auth'):
        return jsonify({'error': 'No autorizado'}), 403
    if not get_colegio(slug):
        return jsonify({'error': 'Colegio no encontrado'}), 404
    init_db(slug)
    conn = conectar(slug)
    profs = conn.execute(
        'SELECT id, nombre, usuario, activo FROM profesores ORDER BY nombre').fetchall()
    resultado = []
    for p in profs:
        mats = conn.execute(
            'SELECT materia, jornada FROM asignaciones_materia WHERE profesor_id=? ORDER BY jornada, materia',
            (p['id'],)).fetchall()
        resultado.append({
            'nombre': p['nombre'], 'usuario': p['usuario'], 'activo': p['activo'],
            'materias': [{'materia': m['materia'], 'jornada': m['jornada']} for m in mats]
        })
    conn.close()
    return jsonify({'profesores': resultado})

# ── BACKUP AUTOMÁTICO ─────────────────────────────────────────────────────────
import threading, shutil
from datetime import timedelta, datetime as _dt

def hacer_backup():
    try:
        hoy = _dt.now().strftime('%Y-%m-%d')
        backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        shutil.copy2(MASTER_DB, os.path.join(backup_dir, f'master_{hoy}.db'))
        for f in os.listdir(DB_FOLDER):
            if f.endswith('.db'):
                shutil.copy2(os.path.join(DB_FOLDER, f),
                             os.path.join(backup_dir, f'{f[:-3]}_{hoy}.db'))
        logger.info(f'Backup automático completado: {hoy}')
    except Exception as e:
        logger.error(f'Error en backup: {e}')

def programar_backup():
    hacer_backup()
    t = threading.Timer(86400, programar_backup)
    t.daemon = True
    t.start()

# ── ENTERPRISE ROUTES (Rector / Observador / Certificados) ─────────────────
@app.route('/<slug>/rector/expediente')
def rector_expediente(slug):
    conn = conectar(slug)
    colegio = get_colegio(slug)
    rector = conn.execute('SELECT * FROM rectores WHERE activo=1 ORDER BY es_principal DESC LIMIT 1').fetchone()
    aid = request.args.get('aid', type=int) or request.args.get('alumno_id', type=int)
    alumno = None
    notas_por_materia = {}
    asistencia = []
    observaciones = []
    cursos_raw = conn.execute('SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()
    cursos = [r['curso'] for r in cursos_raw]
    notif_count = notificaciones_no_leidas(slug, 'rector', 0)
    if aid:
        alumno = conn.execute('SELECT * FROM alumnos WHERE id=?', (aid,)).fetchone()
        if alumno:
            notas_raw = conn.execute('''
                SELECT a.materia,
                       ROUND(AVG(n.val), 1) AS promedio,
                       COUNT(n.id) AS evaluaciones
                FROM notas n
                JOIN actividades a ON a.id = n.actividad_id
                WHERE n.aid=?
                GROUP BY a.materia ORDER BY promedio DESC
            ''', (aid,)).fetchall()
            notas_por_materia = {r['materia']: {'promedio': r['promedio'], 'evaluaciones': r['evaluaciones']} for r in notas_raw}
            asistencia = conn.execute('''
                SELECT fecha, estado, observacion FROM asistencia WHERE aid=? ORDER BY fecha DESC LIMIT 20
            ''', (aid,)).fetchall()
            observaciones = conn.execute('''
                SELECT o.*
                FROM observador_registros o
                WHERE o.aid=?
                ORDER BY o.fecha DESC LIMIT 50
            ''', (aid,)).fetchall()
    return render_template('rector/expediente.html', slug=slug, colegio=colegio, rector=rector,
                          alumno=alumno, notas_por_materia=notas_por_materia, asistencia=asistencia,
                          observaciones=observaciones, cursos=cursos, notif_count=notif_count)

@app.route('/<slug>/rector/observador')
def rector_observador(slug):
    conn = conectar(slug)
    colegio = get_colegio(slug)
    rector = conn.execute('SELECT * FROM rectores WHERE activo=1 ORDER BY es_principal DESC LIMIT 1').fetchone()
    notif_count = notificaciones_no_leidas(slug, 'rector', 0)
    return render_template('rector/observador.html', slug=slug, colegio=colegio, rector=rector, notif_count=notif_count)

@app.route('/<slug>/rector/certificados')
def rector_certificados(slug):
    conn = conectar(slug)
    colegio = get_colegio(slug)
    rector = conn.execute('SELECT * FROM rectores WHERE activo=1 ORDER BY es_principal DESC LIMIT 1').fetchone()
    notif_count = notificaciones_no_leidas(slug, 'rector', 0)
    cursos_raw = conn.execute('SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()
    cursos = [r['curso'] for r in cursos_raw]
    return render_template('rector/certificados.html', slug=slug, colegio=colegio, rector=rector, cursos=cursos, notif_count=notif_count)

@app.route('/<slug>/rector/calendario')
def rector_calendario(slug):
    conn = conectar(slug)
    colegio = get_colegio(slug)
    rector = conn.execute('SELECT * FROM rectores WHERE activo=1 ORDER BY es_principal DESC LIMIT 1').fetchone()
    notif_count = notificaciones_no_leidas(slug, 'rector', 0)
    return render_template('rector/calendario.html', slug=slug, colegio=colegio, rector=rector, notif_count=notif_count)

@app.route('/<slug>/rector/mensajes')
def rector_mensajes(slug):
    conn = conectar(slug)
    colegio = get_colegio(slug)
    rector = conn.execute('SELECT * FROM rectores WHERE activo=1 ORDER BY es_principal DESC LIMIT 1').fetchone()
    notif_count = notificaciones_no_leidas(slug, 'rector', 0)
    return render_template('rector/mensajes.html', slug=slug, colegio=colegio, rector=rector, notif_count=notif_count)

# ── RECTOR API ──────────────────────────────────────────────────────────────────
@app.route('/<slug>/api/rector/estudiantes')
def api_rector_estudiantes(slug):
    conn = conectar(slug)
    q = request.args.get('q', '').strip()
    curso = request.args.get('curso', '').strip()
    if curso:
        rows = conn.execute('SELECT a.id, a.nombre, a.curso FROM alumnos a WHERE a.curso=? AND a.activo=1 ORDER BY a.nombre', (curso,)).fetchall()
        return jsonify({'estudiantes': [dict(r) for r in rows]})
    if len(q) < 2:
        return jsonify({'ok': False, 'data': []})
    rows = conn.execute('''
        SELECT a.id, a.nombre, a.curso
        FROM alumnos a
        WHERE a.nombre LIKE ?
        ORDER BY a.nombre LIMIT 15
    ''', (f'%{q}%',)).fetchall()
    return jsonify({'ok': True, 'data': [dict(r) for r in rows]})

@app.route('/<slug>/api/rector/observador/<int:aid>', methods=['GET', 'POST'])
def api_rector_observador(slug, aid):
    conn = conectar(slug)
    if request.method == 'POST':
        if not validar_csrf():
            return jsonify({'ok': False, 'error': 'CSRF inválido'}), 400
        data = request.get_json(silent=True) or {}
        tipo = data.get('tipo', 'llamado')
        texto = data.get('texto', '').strip()
        if not texto:
            return jsonify({'ok': False, 'error': 'Texto requerido'}), 400
        conn.execute('''INSERT INTO observador_registros (slug, aid, tipo, texto, docente, estado)
                        VALUES (?,?,?,?,?,?)''',
                     (slug, aid, tipo, texto, session.get('nombre', ''), 'pendiente'))
        conn.commit()
        return jsonify({'ok': True})
    rows = conn.execute('''
        SELECT o.*, CASE o.tipo
            WHEN 'positivo' THEN 'Positivo'
            WHEN 'llamado' THEN 'Llamado de atención'
            WHEN 'compromiso' THEN 'Compromiso'
            WHEN 'seguimiento' THEN 'Seguimiento'
        END AS tipo_label
        FROM observador_registros o
        WHERE o.aid=? AND o.slug=?
        ORDER BY o.fecha DESC LIMIT 50
    ''', (aid, slug)).fetchall()
    return jsonify({'ok': True, 'data': [dict(r) for r in rows]})

# ── INIT ────────────────────────────────────────────────────────────────────────
init_master_db()
t = threading.Timer(30, programar_backup)
t.daemon = True
t.start()

if __name__ == '__main__':
    _port = int(os.environ.get('PORT', 8000 if ENV == 'production' else 5000))
    try:
        from waitress import serve
        logger.info(f'Servidor Waitress en http://0.0.0.0:{_port}')
        serve(app, host='0.0.0.0', port=_port, threads=8)
    except ImportError:
        logger.warning('waitress no instalado. Usando Flask dev server (sin reloader).')
        app.run(host='127.0.0.1', port=_port, debug=(ENV != 'production'), use_reloader=False)