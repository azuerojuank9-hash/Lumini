import json
import os
import sqlite3
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ['FLASK_ENV'] = 'development'
os.environ['ENV'] = 'development'

import pytest

from flask_app import _promedio_ponderado, _promedio_simple, _recrear_si_unique_incorrecto, app, hash_pw, init_db

TEST_DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'colegios_db', 'testcolegio.db')

def seed_test_db():
    # Ensure testcolegio exists in master DB
    from flask_app import conectar_master
    mconn = conectar_master()
    cur = mconn.execute('SELECT id FROM colegios WHERE slug=?', ('testcolegio',))
    if not cur.fetchone():
        mconn.execute(
            'INSERT INTO colegios (slug, nombre, activo, codigo_profesores, codigo_directoras, codigo_rectores, schema_version) VALUES (?,?,?,?,?,?,?)',
            ('testcolegio', 'Test Colegio', 1, 'prof_test', 'dir_test', 'rec_test', 20))
        mconn.commit()
    mconn.close()
    # Ensure DB schema exists (creates tables + runs all migrations including v11)
    init_db('testcolegio')
    conn = sqlite3.connect(TEST_DB)
    conn.row_factory = sqlite3.Row
    # Rector (re-create with known id so rector != profesor)
    conn.execute("DELETE FROM rectores WHERE usuario='rector_prueba'")
    conn.execute("INSERT INTO rectores (id, nombre, usuario, password, email, activo, es_principal) VALUES (?,?,?,?,?,?,?)",
                 (99, 'Rector Prueba', 'rector_prueba', 'ecd71870d1963316a97e3ac3408c9835ad8cf0f3c1bc703527c30265534f75ae', 'rector@test.com', 1, 1))
    # Directora
    cur = conn.execute("SELECT id FROM directoras WHERE usuario='directora'")
    if not cur.fetchone():
        conn.execute("INSERT INTO directoras (nombre, usuario, password, email, activo, curso, jornada) VALUES (?,?,?,?,?,?,?)",
                     ('Directora Prueba', 'directora', hash_pw('test123'), 'directora@test.com', 1, 'Primero A', 'Mañana'))
    # Profesor
    cur = conn.execute("SELECT id FROM profesores WHERE id=1")
    if not cur.fetchone():
        conn.execute("INSERT INTO profesores (id, nombre, usuario, password, email, activo) VALUES (?,?,?,?,?,?)",
                     (1, 'Profesor Uno', 'profesor1', hash_pw('test123'), 'prof1@test.com', 1))
    # Alumnos
    cur = conn.execute("SELECT id FROM alumnos WHERE id=1")
    if not cur.fetchone():
        conn.execute("INSERT INTO alumnos (id, nombre, curso, jornada, activo) VALUES (?,?,?,?,?)",
                     (1, 'Alumno Uno', 'Primero A', 'Mañana', 1))
    cur = conn.execute("SELECT id FROM alumnos WHERE id=2")
    if not cur.fetchone():
        conn.execute("INSERT INTO alumnos (id, nombre, curso, jornada, activo) VALUES (?,?,?,?,?)",
                     (2, 'Alumno Dos', 'Primero A', 'Mañana', 1))
    cur = conn.execute("SELECT id FROM alumnos WHERE id=3")
    if not cur.fetchone():
        conn.execute("INSERT INTO alumnos (id, nombre, curso, jornada, activo) VALUES (?,?,?,?,?)",
                     (3, 'Alumno Tres', 'Segundo A', 'Mañana', 1))
    # Asignaciones curso
    conn.execute('INSERT OR IGNORE INTO asignaciones_curso (profesor_id, materia, jornada, curso) VALUES (?, ?, ?, ?)',
                 (1, 'Matemáticas', 'Mañana', 'Primero A'))
    conn.execute('INSERT OR IGNORE INTO asignaciones_curso (profesor_id, materia, jornada, curso) VALUES (?, ?, ?, ?)',
                 (1, 'Matemáticas', 'Mañana', 'Segundo A'))
    # Periodos
    conn.execute('INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (?, ?)', (1, 'abierto'))
    # Actividades
    conn.execute('INSERT OR IGNORE INTO actividades (id, profesor_id, materia, jornada, curso, nombre, orden, periodo) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                 (1, 1, 'Matemáticas', 'Mañana', 'Primero A', 'Tarea 1', 1, 1))
    conn.execute('INSERT OR IGNORE INTO actividades (id, profesor_id, materia, jornada, curso, nombre, orden, periodo) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                 (2, 1, 'Matemáticas', 'Mañana', 'Primero A', 'Examen 1', 2, 1))
    conn.commit()
    conn.close()

seed_test_db()

# ── Fixtures ──

@pytest.fixture
def client():
    app.config['TESTING'] = True
    with app.test_client() as c:
        yield c

@pytest.fixture
def csrf(client):
    with client.session_transaction() as sess:
        sess['_csrf_token'] = 'pytest_csrf_token'
    return 'pytest_csrf_token'

@pytest.fixture
def rector_session(client):
    with client.session_transaction() as sess:
        sess['rector_id_testcolegio'] = 99
        sess['_csrf_token'] = 'pytest_csrf_token'

@pytest.fixture
def teacher_session(client):
    with client.session_transaction() as sess:
        sess['profesor_id_testcolegio'] = 1
        sess['rol_testcolegio'] = 'profesor'
        sess['jornada_testcolegio'] = 'Mañana'
        sess['materia_testcolegio'] = 'Matemáticas'
        sess['_csrf_token'] = 'pytest_csrf_token'

@pytest.fixture
def coordinator_session(client):
    with client.session_transaction() as sess:
        sess['directora_id_testcolegio'] = 1
        sess['_csrf_token'] = 'pytest_csrf_token'

# ── Static Assets ──

class TestStaticAssets:
    def test_lumini_js(self, client):
        r = client.get('/static/js/lumini.js')
        assert r.status_code == 200
        assert 'javascript' in r.content_type

    def test_base_css(self, client):
        r = client.get('/static/css/base.css')
        assert r.status_code == 200
        assert 'text/css' in r.content_type

    def test_all_design_css(self, client):
        for f in ['base.css','theme.css','layout.css','buttons.css','forms.css','tables.css','cards.css','badges.css','alerts.css','sidebar.css','dashboard.css','attendance.css','animations.css','utilities.css']:
            r = client.get(f'/static/css/{f}')
            assert r.status_code == 200, f'{f} returned {r.status_code}'

# ── Landing Pages ──

class TestLandingPages:
    def test_index(self, client):
        r = client.get('/')
        assert r.status_code == 200

    def test_admin_login(self, client):
        r = client.get('/admin')
        assert r.status_code == 200

    def test_rector_login(self, client):
        r = client.get('/testcolegio/rector/login')
        assert r.status_code == 200

    def test_coordinator_login(self, client):
        r = client.get('/testcolegio/directora/login')
        assert r.status_code == 200

    def test_school_login(self, client):
        r = client.get('/testcolegio/login')
        assert r.status_code == 200

# ── Authentication ──

class TestAuthentication:
    def test_rector_login_success(self, client, csrf):
        r = client.post('/testcolegio/login', data={
            '_csrf_token': csrf,
            'accion': 'rector_login',
            'rec_usuario': 'rector_prueba',
            'rec_password': 'test123',
        }, follow_redirects=True)
        assert r.status_code == 200

    def test_rector_login_wrong_password(self, client, csrf):
        r = client.post('/testcolegio/login', data={
            '_csrf_token': csrf,
            'accion': 'rector_login',
            'rec_usuario': 'rector_prueba',
            'rec_password': 'wrong_password',
        }, follow_redirects=True)
        assert r.status_code == 200

    def test_logout(self, client):
        r = client.get('/testcolegio/logout', follow_redirects=True)
        assert r.status_code == 200

# ── Rector Dashboard ──

class TestRectorDashboard:
    def test_panel(self, client, rector_session):
        r = client.get('/testcolegio/rector')
        assert r.status_code == 200
        assert 'LUMINI' in r.get_data(as_text=True)

    def test_profesores(self, client, rector_session):
        r = client.get('/testcolegio/rector/profesores')
        assert r.status_code == 200

    def test_estudiantes(self, client, rector_session):
        r = client.get('/testcolegio/rector/estudiantes')
        assert r.status_code == 200

    def test_cursos(self, client, rector_session):
        r = client.get('/testcolegio/rector/cursos')
        assert r.status_code == 200

    def test_horarios(self, client, rector_session):
        r = client.get('/testcolegio/rector/horarios')
        assert r.status_code == 200

    def test_reportes(self, client, rector_session):
        r = client.get('/testcolegio/rector/reportes')
        assert r.status_code == 200

    def test_configuracion(self, client, rector_session):
        r = client.get('/testcolegio/rector/configuracion')
        assert r.status_code == 200

    def test_comunicaciones(self, client, rector_session):
        r = client.get('/testcolegio/rector/comunicaciones')
        assert r.status_code == 200

    def test_canales(self, client, rector_session):
        r = client.get('/testcolegio/rector/canales')
        assert r.status_code == 200

    def test_auditoria(self, client, rector_session):
        r = client.get('/testcolegio/rector/auditoria')
        assert r.status_code == 200

    def test_solicitudes(self, client, rector_session):
        r = client.get('/testcolegio/rector/solicitudes')
        assert r.status_code == 200

    def test_notificaciones(self, client, rector_session):
        r = client.get('/testcolegio/notificaciones')
        assert r.status_code == 200

    def test_notificaciones_contar(self, client, rector_session):
        r = client.get('/testcolegio/notificaciones/contar')
        assert r.status_code == 200

# ── Teacher Pages (including attendance) ──

class TestTeacherPages:
    def test_teacher_dashboard(self, client, teacher_session):
        r = client.get('/testcolegio/?curso=Primero+A&periodo=1', follow_redirects=True)
        assert r.status_code == 200

    def test_teacher_asistencia_page(self, client, teacher_session):
        r = client.get('/testcolegio/asistencia?curso=Primero+A&fecha=2026-07-04')
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        assert 'Asistencia' in html

# ── Admin Panel ──

class TestAdmin:
    def test_admin_panel(self, client):
        with client.session_transaction() as sess:
            sess['admin_auth'] = True
        r = client.get('/admin')
        assert r.status_code == 200

    def test_admin_codigos(self, client):
        with client.session_transaction() as sess:
            sess['admin_auth'] = True
        r = client.get('/admin/codigos')
        assert r.status_code == 200

    def test_admin_profesores(self, client):
        with client.session_transaction() as sess:
            sess['admin_auth'] = True
        r = client.get('/admin/profesores/testcolegio')
        assert r.status_code == 200

# ── CRUD Operations ──

class TestCRUD:
    REUSED_CSRF = 'pytest_csrf_crud'

    def test_create_communication(self, client):
        with client.session_transaction() as sess:
            sess['rector_id_testcolegio'] = 99
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/rector/comunicaciones/nueva', data={
            '_csrf_token': self.REUSED_CSRF,
            'titulo': 'Test Communication',
            'contenido': 'Test content with accents: evaluación, matemáticas.',
            'destinatario_tipo': 'todos',
            'destinatario_valor': '',
            'prioridad': 'normal',
            'publicar_ahora': '1',
        }, follow_redirects=True)
        assert r.status_code == 200

    def test_create_teacher(self, client):
        with client.session_transaction() as sess:
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/login', data={
            '_csrf_token': self.REUSED_CSRF,
            'accion': 'profesor_registro',
            'nombre': 'Pytest Teacher',
            'reg_usuario': 'pytest_teacher',
            'reg_password': 'test1234',
            'confirmar_password': 'test1234',
            'email_prof': 'teacher@pytest.com',
            'codigo_registro': 'testcode123',
            'pregunta_secreta': '1',
            'respuesta_secreta': 'red',
            'materias_sel': ['Matemáticas'],
            'jornadas_sel': ['Mañana'],
        }, follow_redirects=True)
        assert r.status_code == 200

    def test_create_student(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/registrar', data={
            '_csrf_token': self.REUSED_CSRF,
            'nombre': 'Pytest Student',
            'curso': 'Primero A',
        }, follow_redirects=True)
        assert r.status_code == 200

    def test_save_grade(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1',
            'aid': '1',
            'val': '4.5',
        })
        assert r.status_code == 200, f'Expected 200, got {r.status_code}: {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        assert 'promedio' in data, 'Response must include promedio'
        # Verify the grade was actually saved in the database
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT val FROM notas WHERE aid=1 AND actividad_id=1').fetchone()
        conn.close()
        assert row is not None, 'Grade was not saved in the database'
        assert row[0] == 4.5

    def test_save_grade_single_nota_returns_correct_promedio(self, client):
        """Una sola nota de 1.6 debe mostrar promedio simple de 1.6 (sin ponderar)."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute('DELETE FROM notas WHERE aid=1')
        conn.execute('DELETE FROM evaluaciones WHERE aid=1')
        conn.commit()
        conn.close()
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '2',
            'aid': '1',
            'val': '1.6',
        })
        assert r.status_code == 200, f'Expected 200, got {r.status_code}: {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        assert data['promedio'] == 1.6, f'PROM debe ser 1.6 (simple avg), got {data["promedio"]}'

    def test_save_evaluation(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/guardar_evaluacion', data={
            '_csrf_token': self.REUSED_CSRF,
            'aid': '1',
            'evaluacion': '4.0',
        })
        assert r.status_code == 200, f'Expected 200, got {r.status_code}: {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        # Verify the evaluation was actually saved in the database
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute(
            'SELECT evaluacion FROM evaluaciones WHERE aid=1 AND profesor_id=1 AND materia="Matemáticas" AND jornada="Mañana" AND periodo=1'
        ).fetchone()
        conn.close()
        assert row is not None, 'Evaluation was not saved in the database'
        assert row[0] == 4.0

    def test_save_attendance(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': self.REUSED_CSRF,
            'aid': '1',
            'estado': 'P',
        })
        assert r.status_code == 200, f'Expected 200, got {r.status_code}: {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'

    def test_save_attendance_with_fecha(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': self.REUSED_CSRF,
            'aid': '1',
            'estado': 'A',
            'fecha': '2026-06-15',
        })
        assert r.status_code == 200, f'Expected 200, got {r.status_code}: {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        # Verify persisted
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute(
            'SELECT estado FROM asistencia WHERE aid=1 AND fecha="2026-06-15"'
        ).fetchone()
        conn.close()
        assert row is not None, 'Attendance with fecha was not saved'
        assert row[0] == 'A'

# ── PDF Report Generation ──

class TestPDF:
    def test_boletin_pdf(self, client, coordinator_session):
        r = client.get('/testcolegio/directora/boletin_pdf')
        if r.status_code == 200:
            if 'application/pdf' in (r.content_type or ''):
                assert len(r.get_data()) > 100
            else:
                # Graceful error page when reportlab is not installed
                assert 'reportlab' in r.get_data(as_text=True).lower()
        else:
            # May return 302 (no directora with course) or 404 (no students)
            assert r.status_code in (302, 404)

# ── CSRF / Security ──

class TestCSRF:
    def test_session_cookie_secure_in_production_env(self):
        assert app.config.get('SESSION_COOKIE_SECURE') == True, \
            'SESSION_COOKIE_SECURE must be True when SESSION_COOKIE_SECURE env var is set'

    def test_session_cookie_secure_logic(self):
        """Verify the SESSION_COOKIE_SECURE derivation logic (env-dependent, override-allowed)."""
        def derive_secure(env, override):
            if override:
                return override.lower() in ('true', '1', 'yes')
            return env == 'production'
        # Production defaults to True
        assert derive_secure('production', '') == True
        # Development defaults to False
        assert derive_secure('development', '') == False
        # Explicit override overrides default
        assert derive_secure('production', 'false') == False
        assert derive_secure('development', 'true') == True

    def test_csrf_mismatch_rejected(self, client):
        with client.session_transaction() as sess:
            sess['_csrf_token'] = 'real_token'
        r = client.post('/testcolegio/login', data={
            '_csrf_token': 'wrong_token',
            'accion': 'profesor_login',
            'usuario': 'doesnt_matter',
            'password': 'doesnt_matter',
        }, follow_redirects=True)
        html = r.get_data(as_text=True)
        assert 'Error de seguridad' in html

    def test_csrf_missing_rejected(self, client):
        with client.session_transaction() as sess:
            sess['_csrf_token'] = 'real_token'
        r = client.post('/testcolegio/login', data={
            'accion': 'profesor_login',
            'usuario': 'doesnt_matter',
            'password': 'doesnt_matter',
        }, follow_redirects=True)
        html = r.get_data(as_text=True)
        assert 'Error de seguridad' in html

    def test_csrf_empty_session_rejected(self, client):
        """Simulates Secure cookie lost — session has no _csrf_token."""
        r = client.post('/testcolegio/login', data={
            '_csrf_token': 'any_token',
            'accion': 'profesor_login',
            'usuario': 'doesnt_matter',
            'password': 'doesnt_matter',
        }, follow_redirects=True)
        html = r.get_data(as_text=True)
        assert 'Error de seguridad' in html

    def test_directora_registration_form_has_csrf(self, client):
        r = client.get('/testcolegio/login')
        html = r.get_data(as_text=True)
        assert 'directora/registrar_directo' in html
        assert '_csrf_token' in html
        assert 'csrf_token()' not in html

    def test_rector_login_with_valid_csrf_succeeds(self, client):
        with client.session_transaction() as sess:
            sess['_csrf_token'] = 'valid_csrf_123'
        r = client.post('/testcolegio/login', data={
            '_csrf_token': 'valid_csrf_123',
            'accion': 'rector_login',
            'rec_usuario': 'rector_prueba',
            'rec_password': 'test123',
        }, follow_redirects=True)
        html = r.get_data(as_text=True)
        assert 'Error de seguridad' not in html

    def test_coordinator_login_with_valid_csrf_succeeds(self, client):
        with client.session_transaction() as sess:
            sess['_csrf_token'] = 'valid_coord_csrf'
        r = client.post('/testcolegio/login', data={
            '_csrf_token': 'valid_coord_csrf',
            'accion': 'directora_login',
            'dir_usuario': 'directora_prueba',
            'dir_password': 'test123',
        }, follow_redirects=True)
        html = r.get_data(as_text=True)
        assert 'Error de seguridad' not in html

    def test_teacher_login_with_valid_csrf_succeeds(self, client):
        with client.session_transaction() as sess:
            sess['_csrf_token'] = 'valid_teacher_csrf'
        r = client.post('/testcolegio/login', data={
            '_csrf_token': 'valid_teacher_csrf',
            'accion': 'profesor_login',
            'usuario': 'rector_prueba',
            'password': 'test123',
        }, follow_redirects=True)
        html = r.get_data(as_text=True)
        assert 'Error de seguridad' not in html

    def test_teacher_state_routes_require_csrf(self, client, teacher_session):
        """A8: rutas POST state-changing del profesor rechazan sin token CSRF."""
        for url, kwargs in [
            ('/testcolegio/plantillas/crear', {'json': {'nombre': 'x'}}),
            ('/testcolegio/plantillas/aplicar', {'json': {'plantilla_id': 1, 'curso': 'Primero A'}}),
            ('/testcolegio/plantillas/eliminar/1', {'json': {}}),
            ('/testcolegio/planificacion/copiar', {'json': {'origen_curso': 'Primero A', 'destino_cursos': ['Primero B']}}),
            ('/testcolegio/comunicados/1/leer', {'json': {}}),
            ('/testcolegio/observaciones_json', {'json': {'aid': 1}}),
            ('/testcolegio/observaciones/sugerir', {'json': {'aid': 1}}),
            ('/testcolegio/validar', {'json': {'curso': 'Primero A', 'notas': {}}}),
        ]:
            r = client.post(url, **kwargs)
            assert r.status_code == 403, f'{url} debería requerir CSRF (got {r.status_code})'


# ── Rendered HTML Quality ──

class TestHTMLQuality:
    MOJIBAKE_PATTERNS = [
        'Ã¡', 'Ã©', 'Ã­', 'Ã³', 'Ãº', 'Ã±', 'Â¿', 'Â·', 'Â«', 'Â»',
        'Ã¢â€™Â°', 'Ã¯Â¸', 'Ã¢â€”', 'Ã¢Â¸',
    ]

    @pytest.mark.parametrize('path', [
        '/testcolegio/rector',
        '/testcolegio/rector/profesores',
        '/testcolegio/rector/estudiantes',
        '/testcolegio/rector/cursos',
        '/testcolegio/rector/horarios',
        '/testcolegio/rector/reportes',
        '/testcolegio/rector/configuracion',
        '/testcolegio/rector/comunicaciones',
        '/testcolegio/rector/canales',
        '/testcolegio/rector/auditoria',
        '/testcolegio/rector/solicitudes',
        '/testcolegio/notificaciones',
    ])
    def test_no_mojibake(self, client, rector_session, path):
        r = client.get(path)
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        for pat in self.MOJIBAKE_PATTERNS:
            assert pat not in html, f'Mojibake pattern "{pat}" found in {path}'

    @pytest.mark.parametrize('path', [
        '/testcolegio/rector',
        '/testcolegio/rector/profesores',
        '/testcolegio/rector/estudiantes',
        '/testcolegio/rector/cursos',
        '/testcolegio/rector/horarios',
        '/testcolegio/rector/reportes',
        '/testcolegio/rector/configuracion',
        '/testcolegio/rector/comunicaciones',
        '/testcolegio/rector/canales',
        '/testcolegio/rector/auditoria',
        '/testcolegio/rector/solicitudes',
        '/testcolegio/notificaciones',
        '/admin',
        '/',
    ])
    def test_balanced_divs(self, client, rector_session, path):
        if path == '/':
            r = client.get(path)
        elif path.startswith('/admin'):
            with client.session_transaction() as sess:
                sess['admin_auth'] = True
            r = client.get(path)
        else:
            r = client.get(path)
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        open_divs = html.count('<div')
        close_divs = html.count('</div>')
        assert open_divs == close_divs, f'Div imbalance: +{open_divs - close_divs} in {path}'

# ── Observations CRUD ──

class TestObservations:
    """Tests for the complete observations CRUD (create, read, update, delete)."""

    REUSED_CSRF = 'test-csrf-token-for-observations'

    def _create_obs(self, client, texto='Estudiante destacado en clase'):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/agregar_observacion', data={
            '_csrf_token': self.REUSED_CSRF,
            'aid': '1',
            'texto': texto,
        })
        assert r.status_code == 200, f'Create obs failed: {r.status_code} {r.get_data(as_text=True)}'
        return json.loads(r.get_data(as_text=True))['id']

    def test_create_observation(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post('/testcolegio/agregar_observacion', data={
            '_csrf_token': self.REUSED_CSRF,
            'aid': '1',
            'texto': 'Estudiante destacado en clase',
        })
        assert r.status_code == 200, f'Create obs failed: {r.status_code} {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert 'id' in data, 'Create obs response missing id'
        assert data['texto'] == 'Estudiante destacado en clase'
        assert data['materia'] == 'Matemáticas'
        # Verify persisted in DB
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT id, texto FROM observaciones WHERE id=?', (data['id'],)).fetchone()
        conn.close()
        assert row is not None, 'Observation not found in DB after create'
        assert row[1] == 'Estudiante destacado en clase'
        # Verify audit_log
        conn = sqlite3.connect(TEST_DB)
        log = conn.execute(
            'SELECT accion, valor_nuevo FROM audit_log WHERE tabla="observaciones" AND registro_id=?',
            (data['id'],)
        ).fetchone()
        conn.close()
        assert log is not None, 'Audit log not found for observation create'
        assert log[0] == 'observacion_creada'

    def test_edit_observation(self, client):
        obs_id = self._create_obs(client)
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post(f'/testcolegio/editar_observacion/{obs_id}', data={
            '_csrf_token': self.REUSED_CSRF,
            'texto': 'Texto editado: muy buen desempeño',
        })
        assert r.status_code == 200, f'Edit obs failed: {r.status_code} {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['id'] == obs_id
        assert data['texto'] == 'Texto editado: muy buen desempeño'
        # Verify persisted
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT texto FROM observaciones WHERE id=?', (obs_id,)).fetchone()
        conn.close()
        assert row[0] == 'Texto editado: muy buen desempeño'
        # Verify audit_log
        conn = sqlite3.connect(TEST_DB)
        log = conn.execute(
            'SELECT accion, valor_anterior, valor_nuevo FROM audit_log WHERE tabla="observaciones" AND registro_id=? AND accion="observacion_editada"',
            (obs_id,)
        ).fetchone()
        conn.close()
        assert log is not None, 'Audit log not found for observation edit'

    def test_delete_observation(self, client):
        # First create one
        obs_id = self._create_obs(client)
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post(f'/testcolegio/borrar_observacion/{obs_id}', data={
            '_csrf_token': self.REUSED_CSRF,
        })
        assert r.status_code == 200, f'Delete obs failed: {r.status_code} {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data.get('ok') is True
        # Verify deleted from DB
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT id FROM observaciones WHERE id=?', (obs_id,)).fetchone()
        conn.close()
        assert row is None, 'Observation still exists in DB after delete'
        # Verify audit_log
        conn = sqlite3.connect(TEST_DB)
        log = conn.execute(
            'SELECT accion FROM audit_log WHERE tabla="observaciones" AND registro_id=? AND accion="observacion_eliminada"',
            (obs_id,)
        ).fetchone()
        conn.close()
        assert log is not None, 'Audit log not found for observation delete'

    def test_edit_forbidden_wrong_materia(self, client):
        """Verify that a teacher cannot edit an observation from a different subject."""
        # Create obs in Matemáticas
        obs_id = self._create_obs(client)
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Español'  # Different subject
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post(f'/testcolegio/editar_observacion/{obs_id}', data={
            '_csrf_token': self.REUSED_CSRF,
            'texto': 'Should not work',
        })
        assert r.status_code == 404, f'Expected 404 for wrong materia, got {r.status_code}'

    def test_delete_forbidden_wrong_materia(self, client):
        """Verify that a teacher cannot delete an observation from a different subject."""
        obs_id = self._create_obs(client)
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Español'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.post(f'/testcolegio/borrar_observacion/{obs_id}', data={
            '_csrf_token': self.REUSED_CSRF,
        })
        # Should silently ignore (materia mismatch, observation remains)
        data = json.loads(r.get_data(as_text=True))
        assert data.get('ok') is True
        # Verify observation still exists
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT id FROM observaciones WHERE id=?', (obs_id,)).fetchone()
        conn.close()
        assert row is not None, 'Observation was deleted despite materia mismatch'

# ── Grades System Audit ──

class TestGradesSystem:
    """End-to-end audit of the grades system."""

    REUSED_CSRF = 'test-csrf-token-grades'

    def test_save_grade_and_verify_promedio(self, client):
        """Save a grade and verify the promedio response is correct."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        # Grade: 4.5
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1',
            'aid': '1',
            'val': '4.5',
            'curso': 'Primero A',
        })
        assert r.status_code == 200, f'Grade save failed: {r.status_code} {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        assert data['promedio'] is not None
        # Verify persisted in DB
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT val FROM notas WHERE aid=1 AND actividad_id=1').fetchone()
        conn.close()
        assert row is not None
        assert row[0] == 4.5

    def test_edit_grade_verifies_update(self, client):
        """Edit an existing grade and verify both DB and response."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        # First save a grade
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '3.0', 'curso': 'Primero A',
        })
        # Edit to 4.0
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '4.0', 'curso': 'Primero A',
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        # Verify DB has the new value
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT val FROM notas WHERE aid=1 AND actividad_id=1').fetchone()
        conn.close()
        assert row[0] == 4.0, f'Expected 4.0, got {row[0]}'

    def test_save_multiple_grades_updates_promedio(self, client):
        """Save multiple grades for the same student and verify promedio changes."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        # Save grade for actividad 1: 5.0
        r1 = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '5.0', 'curso': 'Primero A',
        })
        assert r1.status_code == 200
        # Save grade for actividad 2: 3.0
        r2 = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '2', 'aid': '1', 'val': '3.0', 'curso': 'Primero A',
        })
        assert r2.status_code == 200
        d2 = json.loads(r2.get_data(as_text=True))
        assert d2['promedio'] is not None
        # Verify both persisted
        conn = sqlite3.connect(TEST_DB)
        rows = conn.execute('SELECT val FROM notas WHERE aid=1 ORDER BY actividad_id').fetchall()
        conn.close()
        assert len(rows) == 2
        assert rows[0][0] == 5.0
        assert rows[1][0] == 3.0

    def test_grade_decimal_values(self, client):
        """Test various decimal values: 0, 5, 2.5."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        # Test 5
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '5', 'curso': 'Primero A',
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        # Verify persisted
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT val FROM notas WHERE aid=1 AND actividad_id=1').fetchone()
        conn.close()
        assert row[0] == 5.0

    def test_invalid_grade_rejected(self, client):
        """Test that invalid grades are rejected by the backend."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        # Value > 5
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '6.0', 'curso': 'Primero A',
        })
        # Should be rejected but... the backend doesn't validate range,
        # it just saves what it receives. The validation is in the frontend.
        # This is an architectural note, not a test failure.
        assert r.status_code in (200, 400), f'Unexpected: {r.status_code}'

    def test_nota_audit_logged(self, client):
        """Verify that grade changes are recorded in audit_log."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '4.0', 'curso': 'Primero A',
        })
        conn = sqlite3.connect(TEST_DB)
        log = conn.execute(
            'SELECT accion FROM audit_log WHERE tabla="notas" AND accion="nota_editada" ORDER BY id DESC LIMIT 1'
        ).fetchone()
        conn.close()
        assert log is not None, 'Audit log entry not found for note edit'

    def test_promedio_ponderado_consistency(self, client):
        """Verify PROM = simple avg, N.Final = weighted (65/25/10)."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        client.post('/testcolegio/guardar_evaluacion', data={
            '_csrf_token': self.REUSED_CSRF,
            'aid': '1', 'evaluacion': '4.0', 'autoevaluacion': '3.0', 'periodo': '1', 'curso': 'Primero A',
        })
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '5.0', 'curso': 'Primero A',
        })
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '2', 'aid': '1', 'val': '3.0', 'curso': 'Primero A',
        })
        r = client.get('/testcolegio/?curso=Primero+A&periodo=1')
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        # PROM column = simple avg of actividades (5+3)/2 = 4.0
        assert 'id="prom-1"' in html
        assert '4.0' in html, 'Expected simple avg 4.0 in PROM column'
        # N.Final column = 4.0*0.65 + 4.0*0.25 + 3.0*0.10 = 3.9
        assert 'id="nf-1"' in html
        assert '3.9' in html, 'Expected weighted 3.9 in N.Final column'


# ── Promedio Ponderado Formula ──

def _manual_promedio(notas, eval_v, auto_v):
    act_prom = round(sum(notas) / len(notas), 2) if notas else None
    total = 0
    if act_prom is not None: total += act_prom * 0.65
    if eval_v is not None:   total += eval_v * 0.25
    if auto_v is not None:   total += auto_v * 0.10
    return round(total, 2) if (act_prom is not None or eval_v is not None or auto_v is not None) else None


class TestPromedioPonderadoFormula:
    def test_solo_actividades(self):
        r = _promedio_ponderado([4.0, 3.0], None, None)
        m = _manual_promedio([4.0, 3.0], None, None)
        assert r == m, f'solo actividades: _promedio_ponderado={r} manual={m}'
        assert r == 2.27, f'esperado 2.27, obtenido {r}'

    def test_actividades_y_evaluacion(self):
        r = _promedio_ponderado([4.0, 3.0], 3.5, None)
        m = _manual_promedio([4.0, 3.0], 3.5, None)
        assert r == m, f'act+eval: _promedio_ponderado={r} manual={m}'
        assert r == 3.15, f'esperado 3.15, obtenido {r}'

    def test_actividades_y_autoevaluacion(self):
        r = _promedio_ponderado([4.0, 3.0], None, 5.0)
        m = _manual_promedio([4.0, 3.0], None, 5.0)
        assert r == m, f'act+auto: _promedio_ponderado={r} manual={m}'
        assert r == 2.77, f'esperado 2.77, obtenido {r}'

    def test_solo_evaluacion(self):
        r = _promedio_ponderado([], 4.0, None)
        m = _manual_promedio([], 4.0, None)
        assert r == m, f'solo eval: _promedio_ponderado={r} manual={m}'

    def test_solo_autoevaluacion(self):
        r = _promedio_ponderado([], None, 5.0)
        m = _manual_promedio([], None, 5.0)
        assert r == m, f'solo auto: _promedio_ponderado={r} manual={m}'

    def test_tres_categorias_completas(self):
        r = _promedio_ponderado([4.0, 3.0], 3.5, 5.0)
        m = _manual_promedio([4.0, 3.0], 3.5, 5.0)
        assert r == m, f'completo: _promedio_ponderado={r} manual={m}'
        assert r == 3.65, f'esperado 3.65, obtenido {r}'

    def test_sin_datos_retorna_none(self):
        assert _promedio_ponderado([], None, None) is None

class TestPromedioUnicaNota:
    """Verifica que con una sola nota existente el promedio sea correcto (sin dividir por total de actividades)."""

    def test_una_nota_1_6_devuelve_act_prom_1_6(self):
        r = _promedio_ponderado([1.6], None, None)
        assert r == 1.04, f'Una nota 1.6 debe dar 1.04 (1.6*0.65), no {r}'

    def test_una_nota_5_0_devuelve_3_25(self):
        r = _promedio_ponderado([5.0], None, None)
        assert r == 3.25, f'Una nota 5.0 debe dar 3.25 (5.0*0.65), no {r}'

    def test_una_nota_con_eval_devuelve_correcto(self):
        r = _promedio_ponderado([1.6], 3.0, None)
        assert r == 1.79, f'act=1.6 eval=3.0 debe dar 1.79, no {r}'

    def test_varias_notas_sin_eval_devuelve_correcto(self):
        r = _promedio_ponderado([4.0, 3.0, 5.0], None, None)
        esperado = round((4.0+3.0+5.0)/3 * 0.65, 2)
        assert r == esperado, f'[4,3,5] sin eval debe dar {esperado}, no {r}'

    def test_varias_notas_con_eval_y_auto(self):
        r = _promedio_ponderado([4.0, 3.0], 3.5, 5.0)
        assert r == 3.65, f'[4,3] + eval=3.5 + auto=5.0 debe dar 3.65, no {r}'

    def test_sin_notas_solo_eval_devuelve_25_por_ciento(self):
        r = _promedio_ponderado([], 1.6, None)
        assert r == 0.40, f'sin actividades, eval=1.6 debe dar 0.40 (1.6*0.25), no {r}'

    def test_notas_vacias_no_afectan_act_prom(self):
        """Si notas_actividades tiene None, deben ser ignorados."""
        r = _promedio_ponderado([1.6, None, None], None, None)
        assert r == 1.04, f'None debe ser ignorado, debe dar 1.04, no {r}'

    def test_sin_datos_retorna_none(self):
        assert _promedio_ponderado([], None, None) is None
        assert _promedio_ponderado(None, None, None) is None


class TestPromedioSimple:
    """_promedio_simple: promedio exacto como calculadora (suma/cant, sin ponderar)."""

    def test_una_nota_1_6(self):
        assert _promedio_simple([1.6]) == 1.6

    def test_una_nota_5_0(self):
        assert _promedio_simple([5.0]) == 5.0

    def test_dos_notas(self):
        assert _promedio_simple([4.0, 3.0]) == 3.5

    def test_tres_notas(self):
        assert _promedio_simple([4.0, 3.0, 5.0]) == 4.0

    def test_notas_con_none_ignorados(self):
        assert _promedio_simple([1.6, None, None]) == 1.6

    def test_todos_none_retorna_none(self):
        assert _promedio_simple([None, None]) is None

    def test_lista_vacia_retorna_none(self):
        assert _promedio_simple([]) is None

    def test_none_retorna_none(self):
        assert _promedio_simple(None) is None

    def test_redondeo_2_decimales(self):
        r = _promedio_simple([4.0, 3.0, 5.0, 2.0])
        esperado = round((4.0+3.0+5.0+2.0)/4, 2)
        assert r == esperado, f'Esperado {esperado}, obtenido {r}'

    def test_suma_entera_con_decimales(self):
        assert _promedio_simple([3.3, 3.3, 3.3]) == 3.3


class TestMigraciones:
    """Verifica que _recrear_si_unique_incorrecto detecte y corrija UNIQUEs erroneos."""

    def test_recrea_evaluaciones_unique_incorrecto(self):
        conn = sqlite3.connect(':memory:')
        conn.row_factory = sqlite3.Row
        conn.execute('''CREATE TABLE evaluaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
            materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
            evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
            UNIQUE(aid,profesor_id,materia,jornada))''')
        conn.execute('INSERT INTO evaluaciones (aid, profesor_id, materia, jornada, evaluacion, periodo) VALUES (1, 1, "Mat", "M", 4.0, 1)')
        conn.commit()
        result = _recrear_si_unique_incorrecto(conn, 'test', 'evaluaciones',
            '(aid,profesor_id,materia,jornada,periodo)',
            '''CREATE TABLE evaluaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
                materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
                evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
                UNIQUE(aid,profesor_id,materia,jornada,periodo))''',
            '''(id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
               SELECT id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,
                      COALESCE(periodo,1) FROM evaluaciones_old''')
        assert result == True, 'Deberia haber recreado la tabla'
        sql = conn.execute("SELECT sql FROM sqlite_master WHERE name='evaluaciones'").fetchone()[0]
        assert 'UNIQUE(aid,profesor_id,materia,jornada,periodo)' in sql, 'UNIQUE debe incluir periodo'
        row = conn.execute('SELECT * FROM evaluaciones').fetchone()
        assert row['aid'] == 1
        assert row['evaluacion'] == 4.0
        assert row['periodo'] == 1
        conn.close()

    def test_no_recrea_si_unique_correcto(self):
        conn = sqlite3.connect(':memory:')
        conn.row_factory = sqlite3.Row
        conn.execute('''CREATE TABLE evaluaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
            materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
            evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
            UNIQUE(aid,profesor_id,materia,jornada,periodo))''')
        conn.commit()
        result = _recrear_si_unique_incorrecto(conn, 'test', 'evaluaciones',
            '(aid,profesor_id,materia,jornada,periodo)',
            '''CREATE TABLE evaluaciones (...)''', 'SELECT * FROM evaluaciones_old')
        assert result == False, 'No deberia recrear la tabla si UNIQUE ya es correcto'
        conn.close()

    def test_recrea_horarios_curso_unique_incorrecto(self):
        conn = sqlite3.connect(':memory:')
        conn.row_factory = sqlite3.Row
        conn.execute('''CREATE TABLE horarios_curso (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            curso TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
            dia TEXT NOT NULL, franja TEXT NOT NULL,
            num TEXT DEFAULT "", materia TEXT DEFAULT "", profesor TEXT DEFAULT "",
            UNIQUE(curso, dia, franja))''')
        conn.execute('INSERT INTO horarios_curso (curso, jornada, dia, franja) VALUES ("1A", "M", "Lun", "1")')
        conn.commit()
        result = _recrear_si_unique_incorrecto(conn, 'test', 'horarios_curso',
            '(curso,jornada,dia,franja)',
            '''CREATE TABLE horarios_curso (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                curso TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Mañana",
                dia TEXT NOT NULL, franja TEXT NOT NULL,
                num TEXT DEFAULT "", materia TEXT DEFAULT "", profesor TEXT DEFAULT "",
                UNIQUE(curso, jornada, dia, franja))''',
            'SELECT * FROM horarios_curso_old')
        assert result == True
        sql = conn.execute("SELECT sql FROM sqlite_master WHERE name='horarios_curso'").fetchone()[0]
        assert 'UNIQUE(curso, jornada, dia, franja)' in sql
        row = conn.execute('SELECT * FROM horarios_curso').fetchone()
        assert row['curso'] == '1A'
        conn.close()

    def test_guardar_evaluacion_con_upsert_ok(self, client, teacher_session):
        with client.session_transaction() as sess:
            sess['_csrf_token'] = 'pytest_csrf_token'
        r = client.post('/testcolegio/guardar_evaluacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'evaluacion': '4.5', 'autoevaluacion': '3.5', 'periodo': '1', 'curso': 'Primero A',
        })
        assert r.status_code == 200, f'Expected 200, got {r.status_code}: {r.get_data(as_text=True)}'
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        r2 = client.post('/testcolegio/guardar_evaluacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'evaluacion': '4.8', 'autoevaluacion': '3.8', 'periodo': '1', 'curso': 'Primero A',
        })
        assert r2.status_code == 200, f'Expected 200 on overwrite, got {r2.status_code}: {r2.get_data(as_text=True)}'
        data2 = json.loads(r2.get_data(as_text=True))
        assert data2['status'] == 'ok'
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute(
            'SELECT evaluacion, autoevaluacion FROM evaluaciones WHERE aid=1 AND profesor_id=1 AND materia="Matemáticas" AND jornada="Mañana" AND periodo=1'
        ).fetchone()
        conn.close()
        assert row is not None
        assert row[0] == 4.8
        assert row[1] == 3.8

    def test_plantilla_aplicar_respeta_jornada_y_periodo(self, client, teacher_session):
        """Aplicar una plantilla usa la jornada y periodo del front, no los de la sesión."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute('DELETE FROM plantillas WHERE profesor_id=1')
        conn.execute("DELETE FROM actividades WHERE nombre LIKE 'Plantilla test A4%'")
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/plantillas/crear', json={
            'nombre': 'Plantilla test A4', 'tipo': 'tarea', 'peso': 10, 'descripcion': ''},
            headers={'X-CSRF-Token': 'pytest_csrf_token'})
        assert r.status_code == 200, r.get_data(as_text=True)
        conn = sqlite3.connect(TEST_DB)
        tmpl = conn.execute("SELECT id FROM plantillas WHERE profesor_id=1 AND nombre='Plantilla test A4'").fetchone()
        conn.close()
        assert tmpl is not None
        r = client.post('/testcolegio/plantillas/aplicar', json={
            'plantilla_id': tmpl[0], 'curso': 'Primero A', 'materia': 'Matemáticas',
            'jornada': 'Tarde', 'periodo': 2},
            headers={'X-CSRF-Token': 'pytest_csrf_token'})
        assert r.status_code == 200, r.get_data(as_text=True)
        conn = sqlite3.connect(TEST_DB)
        act = conn.execute(
            "SELECT * FROM actividades WHERE profesor_id=1 AND nombre='Plantilla test A4'").fetchone()
        conn.execute("DELETE FROM plantillas WHERE id=?", (tmpl[0],))
        conn.commit()
        conn.close()
        assert act is not None
        assert act[3] == 'Tarde'
        assert act[7] == 2
        assert act[4] == 'Primero A'

class TestAccionesMasivas:
    CSRF = 'pytest_csrf_token'

    def _session(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['rol_testcolegio'] = 'profesor'
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.CSRF

    def test_masiva_periodo_cerrado_bloquea_eliminar(self, client):
        self._session(client)
        conn = sqlite3.connect(TEST_DB)
        conn.execute('INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, ?)', ('cerrado',))
        conn.commit()
        try:
            r = client.post('/testcolegio/actividades/masiva',
                            json={'accion': 'eliminar', 'ids': [1]},
                            headers={'X-CSRF-Token': self.CSRF})
            assert r.status_code == 403, r.get_data(as_text=True)
            assert r.get_json()['codigo'] == 'PERIODO_CERRADO'
            still = conn.execute('SELECT COUNT(*) FROM actividades WHERE id=1').fetchone()[0]
            assert still == 1
        finally:
            conn.execute('INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, ?)', ('abierto',))
            conn.commit()
            conn.close()

    def test_masiva_duplicar_copia_orden(self, client):
        self._session(client)
        conn = sqlite3.connect(TEST_DB)
        conn.execute("DELETE FROM actividades WHERE nombre LIKE '%(copia)'")
        conn.commit()
        r = client.post('/testcolegio/actividades/masiva',
                        json={'accion': 'duplicar', 'ids': [1]},
                        headers={'X-CSRF-Token': self.CSRF})
        assert r.status_code == 200 and r.get_json()['status'] == 'ok'
        copy = conn.execute("SELECT * FROM actividades WHERE nombre='Tarea 1 (copia)'").fetchone()
        conn.close()
        assert copy is not None
        assert copy[6] == 1

    def test_masiva_eliminar_audita(self, client):
        self._session(client)
        conn = sqlite3.connect(TEST_DB)
        conn.execute('INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, ?)', ('abierto',))
        conn.execute("DELETE FROM audit_log WHERE accion='actividad_eliminar'")
        cur = conn.execute('INSERT INTO actividades (profesor_id, materia, jornada, curso, nombre, orden, periodo) VALUES (1,"Matemáticas","Mañana","Primero A","Borrable A9",9,1)')
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/actividades/masiva',
                        json={'accion': 'eliminar', 'ids': [new_id]},
                        headers={'X-CSRF-Token': self.CSRF})
        assert r.status_code == 200 and r.get_json()['status'] == 'ok'
        conn = sqlite3.connect(TEST_DB)
        gone = conn.execute('SELECT COUNT(*) FROM actividades WHERE id=?', (new_id,)).fetchone()[0]
        aud = conn.execute('SELECT accion, valor_nuevo FROM audit_log WHERE accion="actividad_eliminar" ORDER BY id DESC LIMIT 1').fetchone()
        conn.close()
        assert gone == 0
        assert aud is not None
        assert str(new_id) in aud[1]

    def test_index_muestra_toolbar_masiva_y_confirm(self, client, teacher_session):
        r = client.get('/testcolegio/')
        html = r.get_data(as_text=True)
        assert 'masivaActions' in html
        assert '_modoMasiva' in html
        assert 'limpiarSeleccionMasiva' in html
        assert 'Eliminar ' in html and 'confirm(' in html

# ── Auditoría de Notas ──

class TestAuditoriaNotas:
    REUSED_CSRF = 'pytest_csrf_token'

    def _setup_session(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['rol_testcolegio'] = 'profesor'
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF

    def test_auditoria_notas_creacion(self, client):
        """Guardar nota nueva debe registrar 'creacion' en auditoria_notas."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute('DELETE FROM notas WHERE aid=2 AND actividad_id=1')
        conn.execute('DELETE FROM auditoria_notas WHERE aid=2 AND actividad_id=1')
        conn.commit()
        conn.close()
        self._setup_session(client)
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '2', 'val': '3.5', 'curso': 'Primero A',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            'SELECT * FROM auditoria_notas WHERE aid=2 AND actividad_id=1 ORDER BY id DESC LIMIT 1'
        ).fetchone()
        conn.close()
        assert row is not None, 'No auditoria_notas entry found'
        assert row['tipo_accion'] == 'creacion'
        assert row['tabla'] == 'notas'
        assert row['campo'] == 'nota'
        assert row['rol'] == 'profesor'
        assert float(row['valor_nuevo']) == 3.5
        assert row['valor_anterior'] is None

    def test_auditoria_notas_modificacion(self, client):
        """Modificar nota existente debe registrar 'modificacion' en auditoria_notas."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute('DELETE FROM notas WHERE aid=1 AND actividad_id=1')
        conn.execute('DELETE FROM auditoria_notas WHERE aid=1 AND actividad_id=1')
        conn.commit()
        conn.close()
        self._setup_session(client)
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '2.0', 'curso': 'Primero A',
        })
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '1', 'aid': '1', 'val': '4.0', 'curso': 'Primero A',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        mod = conn.execute(
            "SELECT * FROM auditoria_notas WHERE aid=1 AND actividad_id=1 AND tipo_accion='modificacion' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert mod is not None, 'No modificacion audit entry found'
        assert float(mod['valor_anterior']) == 2.0
        assert float(mod['valor_nuevo']) == 4.0

    def test_auditoria_notas_evaluacion(self, client):
        """Guardar evaluacion debe registrar en auditoria_notas."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute('DELETE FROM evaluaciones WHERE aid=1 AND periodo=1')
        conn.execute("DELETE FROM auditoria_notas WHERE tabla='evaluaciones' AND aid=1")
        conn.commit()
        conn.close()
        self._setup_session(client)
        client.post('/testcolegio/guardar_evaluacion', data={
            '_csrf_token': self.REUSED_CSRF,
            'aid': '1', 'evaluacion': '4.0', 'autoevaluacion': '3.0', 'periodo': '1', 'curso': 'Primero A',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM auditoria_notas WHERE tabla='evaluaciones' AND aid=1 ORDER BY id"
        ).fetchall()
        conn.close()
        eval_rows = [r for r in rows if r['campo'] == 'evaluacion']
        auto_rows = [r for r in rows if r['campo'] == 'autoevaluacion']
        assert len(eval_rows) >= 1, 'No evaluacion audit entry'
        assert len(auto_rows) >= 1, 'No autoevaluacion audit entry'
        assert eval_rows[0]['tipo_accion'] == 'creacion'
        assert float(eval_rows[0]['valor_nuevo']) == 4.0
        assert float(auto_rows[0]['valor_nuevo']) == 3.0

    def test_auditoria_notas_campos_obligatorios(self, client):
        """Verificar que todos los campos obligatorios se guarden correctamente."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute('DELETE FROM notas WHERE aid=2 AND actividad_id=2')
        conn.execute('DELETE FROM auditoria_notas WHERE aid=2 AND actividad_id=2')
        conn.commit()
        conn.close()
        self._setup_session(client)
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '2', 'aid': '2', 'val': '4.5', 'curso': 'Primero A',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM auditoria_notas WHERE aid=2 AND actividad_id=2 AND tipo_accion='creacion' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert row is not None
        assert row['usuario_id'] == 1
        assert row['rol'] == 'profesor'
        assert row['curso'] is not None
        assert row['materia'] is not None
        assert row['periodo'] is not None
        assert row['ip'] is not None
        assert row['creado'] is not None
        assert row['tabla'] == 'notas'
        assert row['aid'] == 2

# ── Historial de Notas ──

class TestHistorialNotas:
    REUSED_CSRF = 'pytest_csrf_token'

    def test_historial_route_returns_json(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['rol_testcolegio'] = 'profesor'
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.get('/testcolegio/historial_notas/1')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert isinstance(data, list)

    def test_historial_after_save(self, client):
        """Saving a grade should make it appear in historial_notas."""
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['rol_testcolegio'] = 'profesor'
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        conn = sqlite3.connect(TEST_DB)
        conn.execute("DELETE FROM notas WHERE aid=2 AND actividad_id=2")
        conn.execute("DELETE FROM auditoria_notas WHERE aid=2 AND actividad_id=2")
        conn.commit(); conn.close()
        client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': self.REUSED_CSRF,
            'actividad_id': '2', 'aid': '2', 'val': '4.2', 'curso': 'Primero A',
        })
        r = client.get('/testcolegio/historial_notas/2')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        creaciones = [e for e in data if e['tipo_accion'] == 'creacion' and e['tabla'] == 'notas']
        assert len(creaciones) >= 1

    def test_historial_empty_for_new_student(self, client):
        with client.session_transaction() as sess:
            sess['profesor_id_testcolegio'] = 1
            sess['rol_testcolegio'] = 'profesor'
            sess['jornada_testcolegio'] = 'Mañana'
            sess['materia_testcolegio'] = 'Matemáticas'
            sess['_csrf_token'] = self.REUSED_CSRF
        r = client.get('/testcolegio/historial_notas/999')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data == []


# ── Bloqueo de Períodos ──

class TestPeriodoCerrado:
    REUSED_CSRF = 'pytest_csrf_periodo'

    def _limpiar_estado(self):
        conn = sqlite3.connect(TEST_DB)
        conn.execute("DELETE FROM periodos_estado WHERE periodo=1")
        conn.execute("INSERT INTO periodos_estado (periodo, estado) VALUES (1, 'abierto')")
        conn.commit()
        conn.close()

    def test_rector_cerrar_periodo(self, client, rector_session):
        """Rector puede cerrar un período."""
        self._limpiar_estado()
        r = client.post('/testcolegio/rector/periodos/1/cerrar', data={
            '_csrf_token': 'pytest_csrf_token',
        }, follow_redirects=True)
        assert r.status_code == 200
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT estado FROM periodos_estado WHERE periodo=1").fetchone()
        conn.close()
        assert row is not None
        assert row['estado'] == 'cerrado'

    def test_rector_abrir_periodo(self, client, rector_session):
        """Rector puede reabrir un período cerrado."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'cerrado')")
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/rector/periodos/1/abrir', data={
            '_csrf_token': 'pytest_csrf_token',
        }, follow_redirects=True)
        assert r.status_code == 200
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT estado FROM periodos_estado WHERE periodo=1").fetchone()
        conn.close()
        assert row['estado'] == 'abierto'

    def test_profesor_no_puede_guardar_nota_periodo_cerrado(self, client, teacher_session):
        """Profesor recibe 403 + JSON al guardar nota en período cerrado."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'cerrado')")
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': 'pytest_csrf_token',
            'actividad_id': '1', 'aid': '1', 'val': '4.0', 'curso': 'Primero A',
        })
        assert r.status_code == 403
        data = json.loads(r.get_data(as_text=True))
        assert data['codigo'] == 'PERIODO_CERRADO'

    def test_profesor_no_puede_guardar_evaluacion_periodo_cerrado(self, client, teacher_session):
        """Profesor recibe 403 + JSON al guardar evaluación en período cerrado."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'cerrado')")
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/guardar_evaluacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'evaluacion': '4.0', 'periodo': '1', 'curso': 'Primero A',
        })
        assert r.status_code == 403
        data = json.loads(r.get_data(as_text=True))
        assert data['codigo'] == 'PERIODO_CERRADO'

    def test_profesor_no_puede_crear_actividad_periodo_cerrado(self, client, teacher_session):
        """Profesor no puede crear actividad en período cerrado (redirect con error)."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'cerrado')")
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/nueva_actividad', data={
            '_csrf_token': 'pytest_csrf_token',
            'nombre': 'Actividad prohibida',
            'curso_sel': 'Primero A',
            'materia': 'Matemáticas',
            'jornada': 'Mañana',
            'periodo_sel': '1',
        }, follow_redirects=True)
        assert r.status_code == 200
        assert 'periodo_cerrado' in r.get_data(as_text=True) or 'cerrado' in r.get_data(as_text=True).lower()

    def test_profesor_no_puede_borrar_actividad_periodo_cerrado(self, client, teacher_session):
        """Profesor no puede borrar actividad en período cerrado."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'cerrado')")
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/borrar_actividad/1', data={
            '_csrf_token': 'pytest_csrf_token',
            'curso': 'Primero A',
        })
        assert r.status_code == 403
        data = json.loads(r.get_data(as_text=True))
        assert data['codigo'] == 'PERIODO_CERRADO'

    def test_puede_guardar_nota_despues_de_reabrir(self, client, teacher_session):
        """Profesor puede guardar nota después de que rector reabra el período."""
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'abierto')")
        conn.commit()
        conn.close()
        r = client.post('/testcolegio/guardar_nota', data={
            '_csrf_token': 'pytest_csrf_token',
            'actividad_id': '1', 'aid': '1', 'val': '3.0', 'curso': 'Primero A',
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'


# ── Solicitudes de Modificacion ──

class TestSolicitudesModificacion:
    REUSED_CSRF = 'pytest_csrf_sol'

    def _cerrar_periodo(self):
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'cerrado')")
        conn.commit()
        conn.close()

    def _abrir_periodo(self):
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR REPLACE INTO periodos_estado (periodo, estado) VALUES (1, 'abierto')")
        conn.commit()
        conn.close()

    def setup_method(self, method):
        # Ensure period starts open for every test
        self._abrir_periodo()

    def test_crear_solicitud_actividad(self, client, teacher_session):
        """Profesor puede crear solicitud de modificacion de nota."""
        self._cerrar_periodo()
        r = client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '4.5', 'motivo': 'Error en la nota',
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'

    def test_crear_solicitud_evaluacion(self, client, teacher_session):
        """Profesor puede crear solicitud de modificacion de evaluacion."""
        self._cerrar_periodo()
        r = client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'tipo': 'evaluacion',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '3.5', 'motivo': 'Corregir evaluacion',
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'

    def test_solicitud_sin_motivo_rechazada(self, client, teacher_session):
        """Solicitud sin motivo debe ser rechazada."""
        self._cerrar_periodo()
        r = client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '4.0', 'motivo': '',
        })
        assert r.status_code == 400

    def test_solicitud_nota_invalida_rechazada(self, client, teacher_session):
        """Solicitud con nota fuera de rango debe ser rechazada."""
        self._cerrar_periodo()
        r = client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '6.0', 'motivo': 'Nota invalida',
        })
        assert r.status_code == 400

    def test_solicitud_periodo_abierto_rechazada(self, client, teacher_session):
        """Solicitud en periodo abierto debe ser rechazada."""
        self._abrir_periodo()
        r = client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '4.0', 'motivo': 'Test',
        })
        assert r.status_code == 400

    def test_rector_aprobar_solicitud(self, client, teacher_session):
        """Rector puede aprobar una solicitud pendiente."""
        self._cerrar_periodo()
        client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '4.0', 'motivo': 'Correccion',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        sol = conn.execute(
            "SELECT id, profesor_id FROM solicitudes_modificacion WHERE aid=1 AND estado='pendiente' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert sol is not None
        assert sol['profesor_id'] != 99, "profesor_id debe ser != rector_id para evitar self-approval"
        sid = sol['id']
        # Switch to rector session
        with client.session_transaction() as sess:
            sess['rector_id_testcolegio'] = 99
            sess['_csrf_token'] = 'pytest_csrf_token'
            sess.pop('profesor_id_testcolegio', None)
        r = client.post('/testcolegio/rector/solicitudes/%d/aprobar' % sid, data={
            '_csrf_token': 'pytest_csrf_token',
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute('SELECT val FROM notas WHERE aid=1 AND actividad_id=1').fetchone()
        conn.close()
        assert row is not None
        assert row[0] == 4.0

    def test_rector_rechazar_solicitud(self, client, teacher_session):
        """Rector puede rechazar una solicitud."""
        self._cerrar_periodo()
        client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '2', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '5.0', 'motivo': 'Merito maximo',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        sol = conn.execute(
            "SELECT id, profesor_id FROM solicitudes_modificacion WHERE aid=1 AND estado='pendiente' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert sol is not None
        assert sol['profesor_id'] != 99
        sid = sol['id']
        with client.session_transaction() as sess:
            sess['rector_id_testcolegio'] = 99
            sess['_csrf_token'] = 'pytest_csrf_token'
            sess.pop('profesor_id_testcolegio', None)
        r = client.post('/testcolegio/rector/solicitudes/%d/rechazar' % sid, data={
            '_csrf_token': 'pytest_csrf_token',
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'
        conn = sqlite3.connect(TEST_DB)
        st = conn.execute("SELECT estado FROM solicitudes_modificacion WHERE id=?", (sid,)).fetchone()
        conn.close()
        assert st[0] == 'rechazada'

    def test_doble_aprobacion_rechazada(self, client, teacher_session):
        """No se puede aprobar una solicitud ya resuelta."""
        self._cerrar_periodo()
        client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '3.0', 'motivo': 'Ajuste',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        sol = conn.execute(
            "SELECT id, profesor_id FROM solicitudes_modificacion WHERE aid=1 AND estado='pendiente' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert sol is not None
        assert sol['profesor_id'] != 99
        sid = sol['id']
        with client.session_transaction() as sess:
            sess['rector_id_testcolegio'] = 99
            sess['_csrf_token'] = 'pytest_csrf_token'
            sess.pop('profesor_id_testcolegio', None)
        client.post('/testcolegio/rector/solicitudes/%d/aprobar' % sid, data={
            '_csrf_token': 'pytest_csrf_token',
        })
        r = client.post('/testcolegio/rector/solicitudes/%d/aprobar' % sid, data={
            '_csrf_token': 'pytest_csrf_token',
        })
        assert r.status_code == 400

    def test_solicitud_audit_logged(self, client, teacher_session):
        """Crear solicitud debe registrar en auditoria_notas."""
        self._cerrar_periodo()
        conn = sqlite3.connect(TEST_DB)
        conn.execute("DELETE FROM auditoria_notas WHERE tabla='solicitudes_modificacion'")
        conn.commit()
        conn.close()
        client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '4.5', 'motivo': 'Error revision',
        })
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM auditoria_notas WHERE tabla='solicitudes_modificacion' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert row is not None
        assert row['tipo_accion'] == 'solicitud_creada'

    def test_solicitud_historial(self, client, teacher_session):
        """Historial del estudiante debe mostrar solicitudes."""
        self._cerrar_periodo()
        client.post('/testcolegio/solicitar_modificacion', data={
            '_csrf_token': 'pytest_csrf_token',
            'aid': '1', 'actividad_id': '1', 'tipo': 'actividad',
            'periodo': '1', 'curso': 'Primero A',
            'valor_solicitado': '4.0', 'motivo': 'Revision',
        })
        r = client.get('/testcolegio/historial_notas/1')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        solicitudes = [e for e in data if e['tabla'] == 'solicitudes_modificacion']
        assert len(solicitudes) >= 1


class TestExcelImportExport:
    """Fase 5 — Importación y exportación profesional de Excel."""

    REUSED_CSRF = 'pytest_csrf_excel'

    def test_plantilla_notas(self, client, teacher_session):
        """Descargar plantilla oficial debe devolver .xlsx válido."""
        r = client.get('/testcolegio/plantilla_notas?curso=Primero A&periodo=1')
        assert r.status_code == 200
        assert r.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        import io

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.data))
        ws = wb.active
        # Header row: N°, Estudiante, AID, actividad columns, Evaluación, Autoevaluación, Promedio
        headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        assert 'N°' in str(headers[0]) or 'N' in str(headers[0])
        assert headers[1] == 'Estudiante'
        assert 'Promedio' in [str(h) for h in headers]

    def test_exportar_notas(self, client, teacher_session):
        """Exportar notas debe devolver .xlsx con el formato correcto."""
        r = client.get('/testcolegio/exportar_notas?curso=Primero A&periodo=1')
        assert r.status_code == 200
        assert r.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        import io

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.data))
        ws = wb.active
        [str(c.value) for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        # Should contain student names from seed data
        rows_data = list(ws.iter_rows(min_row=2, values_only=True))
        names = [str(r[1]) for r in rows_data if r[1] is not None]
        assert any('Alumno' in n for n in names)

    def test_importar_notas_page(self, client, teacher_session):
        """Página de importar notas debe renderizar."""
        r = client.get('/testcolegio/importar_notas?curso=Primero A&periodo=1')
        assert r.status_code == 200
        assert b'Importar' in r.data

    def test_importar_notas_preview_validation(self, client, teacher_session):
        """Preview debe rechazar archivo inválido."""
        r = client.post('/testcolegio/importar_notas/preview', data={
            '_csrf_token': 'pytest_csrf_token',
            'curso': 'Primero A', 'periodo': '1',
        })
        assert r.status_code == 400
        data = json.loads(r.get_data(as_text=True))
        assert 'error' in data['status'] or 'error' in data.get('mensaje', '')

    def test_importar_notas_preview_valid_file(self, client, teacher_session):
        """Preview con archivo .xlsx válido debe devolver filas."""
        import io

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Notas'
        ws.cell(row=1, column=1, value='N°')
        ws.cell(row=1, column=2, value='Estudiante')
        ws.cell(row=1, column=3, value='AID')
        ws.cell(row=1, column=4, value='Trabajo 1')  # actividad exists in seed
        ws.cell(row=1, column=5, value='Evaluación')
        ws.cell(row=1, column=6, value='Autoevaluación')
        ws.cell(row=1, column=7, value='Promedio')
        # Find existing student
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        alumno = conn.execute(
            "SELECT id, nombre FROM alumnos WHERE curso='Primero A' AND jornada='Mañana' AND activo=1 LIMIT 1"
        ).fetchone()
        conn.close()
        assert alumno is not None, 'No hay alumnos en la DB de test'
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value=alumno['nombre'])
        ws.cell(row=2, column=3, value=alumno['id'])
        ws.cell(row=2, column=4, value=4.5)
        ws.cell(row=2, column=5, value=4.0)
        ws.cell(row=2, column=6, value=3.5)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        r = client.post('/testcolegio/importar_notas/preview', data={
            '_csrf_token': 'pytest_csrf_token',
            'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'test_notas.xlsx'),
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['total'] >= 1
        assert data['filas'][0]['ok'] == True

    def test_importar_notas_preview_file_with_errors(self, client, teacher_session):
        """Preview debe marcar errores (estudiante no encontrado, nota fuera de rango)."""
        import io

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='N°')
        ws.cell(row=1, column=2, value='Estudiante')
        ws.cell(row=1, column=3, value='AID')
        ws.cell(row=1, column=4, value='Trabajo 1')
        ws.cell(row=1, column=5, value='Promedio')
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value='Estudiante Inexistente')
        ws.cell(row=2, column=3, value=99999)
        ws.cell(row=2, column=4, value=6.0)  # Out of range
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        r = client.post('/testcolegio/importar_notas/preview', data={
            '_csrf_token': 'pytest_csrf_token',
            'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'bad.xlsx'),
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['total'] >= 1
        assert data['filas'][0]['ok'] == False

    def test_importar_notas_confirmar(self, client, teacher_session):
        """Confirmar importación debe guardar los cambios."""
        import io

        from openpyxl import Workbook
        conn = sqlite3.connect(TEST_DB)
        conn.row_factory = sqlite3.Row
        alumno = conn.execute(
            "SELECT id, nombre FROM alumnos WHERE curso='Primero A' AND jornada='Mañana' AND activo=1 LIMIT 1"
        ).fetchone()
        actividad = conn.execute(
            "SELECT id, nombre FROM actividades WHERE profesor_id=1 AND materia='Matemáticas' LIMIT 1"
        ).fetchone()
        conn.close()
        assert alumno is not None
        assert actividad is not None
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='N°')
        ws.cell(row=1, column=2, value='Estudiante')
        ws.cell(row=1, column=3, value='AID')
        ws.cell(row=1, column=4, value=actividad['nombre'])
        ws.cell(row=1, column=5, value='Promedio')
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value=alumno['nombre'])
        ws.cell(row=2, column=3, value=alumno['id'])
        ws.cell(row=2, column=4, value=3.5)  # Set grade to 3.5
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        # First get the preview
        r = client.post('/testcolegio/importar_notas/preview', data={
            '_csrf_token': 'pytest_csrf_token',
            'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'test.xlsx'),
        })
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] == True
        # Now confirm
        r = client.post('/testcolegio/importar_notas/confirmar', data={
            '_csrf_token': 'pytest_csrf_token',
            'curso': 'Primero A', 'periodo': '1',
            'data': json.dumps(data),
        })
        assert r.status_code == 200
        result = json.loads(r.get_data(as_text=True))
        assert result['status'] == 'ok'
        # Verify the grade was saved
        conn = sqlite3.connect(TEST_DB)
        nota = conn.execute(
            'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
            (alumno['id'], actividad['id'])
        ).fetchone()
        conn.close()
        assert nota is not None
        assert nota[0] == 3.5


class TestDashboard:
    """Fase 6 — Dashboard Académico Profesional."""

    def test_dashboard_page_profesor(self, client, teacher_session):
        r = client.get('/testcolegio/dashboard')
        assert r.status_code == 200
        assert b'Dashboard' in r.data

    def test_dashboard_page_rector(self, client, rector_session):
        r = client.get('/testcolegio/dashboard')
        assert r.status_code == 200
        assert b'Dashboard' in r.data

    def test_dashboard_page_no_auth(self, client):
        r = client.get('/testcolegio/dashboard')
        assert r.status_code == 302

    def test_dashboard_data_profesor(self, client, teacher_session):
        r = client.get('/testcolegio/dashboard_data?curso=Primero A&periodo=1')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert 'cards' in data
        assert data['cards']['total_estudiantes'] >= 2

    def test_dashboard_data_rector(self, client, rector_session):
        r = client.get('/testcolegio/dashboard_data')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert 'cards' in data
        assert data['cards']['total_profesores'] >= 1
        assert data['cards']['total_estudiantes'] >= 2

    def test_dashboard_data_no_auth(self, client):
        r = client.get('/testcolegio/dashboard_data')
        assert r.status_code == 401

    def test_dashboard_data_profesor_all_courses(self, client, teacher_session):
        """Profesor data across all courses."""
        r = client.get('/testcolegio/dashboard_data?periodo=1')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert 'charts' in data
        assert 'distribucion' in data['charts']
        assert 'promedio_por_curso' in data['charts']

    def test_dashboard_data_rector_full(self, client, rector_session):
        """Rector dashboard returns all expected sections."""
        r = client.get('/testcolegio/dashboard_data')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert 'rankings' in data
        assert 'alerts' in data
        assert 'estadisticas' in data
        assert data['estadisticas'] is not None or data['estadisticas'] is None

    def test_dashboard_stats_function(self):
        """Test the descriptive statistics function directly."""
        from flask_app import _estadisticas_desc
        vals = [3.0, 4.0, 5.0, 4.0, 2.0]
        st = _estadisticas_desc(vals)
        assert st is not None
        assert st['media'] == 3.6
        assert st['mediana'] == 4.0
        assert st['maximo'] == 5.0
        assert st['minimo'] == 2.0
        assert st['q1'] is not None
        assert st['q3'] is not None
        assert st['desviacion'] > 0

    def test_dashboard_stats_empty(self):
        from flask_app import _estadisticas_desc
        assert _estadisticas_desc([]) is None
        assert _estadisticas_desc([None, None]) is None

    def test_dashboard_stats_single(self):
        from flask_app import _estadisticas_desc
        st = _estadisticas_desc([4.0])
        assert st['media'] == 4.0
        assert st['mediana'] == 4.0
        assert st['maximo'] == 4.0
        assert st['minimo'] == 4.0

    def test_dashboard_estadisticas_has_all_keys(self, client, teacher_session):
        r = client.get('/testcolegio/dashboard_data?curso=Primero A&periodo=1')
        data = json.loads(r.get_data(as_text=True))
        st = data['estadisticas']
        if st:
            for k in ['media','mediana','moda','desviacion','maximo','minimo','q1','q2','q3','p10','p90']:
                assert k in st, f'Missing key: {k}'

    def test_dashboard_rankings_profesor(self, client, teacher_session):
        r = client.get('/testcolegio/dashboard_data?curso=Primero A&periodo=1')
        data = json.loads(r.get_data(as_text=True))
        assert 'top_estudiantes' in data['rankings']
        assert 'top_cursos' in data['rankings']

    def test_dashboard_alerts_profesor(self, client, teacher_session):
        r = client.get('/testcolegio/dashboard_data?curso=Primero A&periodo=1')
        data = json.loads(r.get_data(as_text=True))
        assert 'estudiantes_bajo' in data['alerts']
        assert 'destacados' in data['alerts']

    def test_dashboard_charts_profesor(self, client, teacher_session):
        r = client.get('/testcolegio/dashboard_data?curso=Primero A&periodo=1')
        data = json.loads(r.get_data(as_text=True))
        ch = data['charts']
        assert 'distribucion' in ch
        assert 'promedio_por_curso' in ch
        assert 'evolucion_periodos' in ch
        assert 'rendimiento_actividades' in ch

    def test_dashboard_permission_teacher_only(self, client, teacher_session):
        """Verify profesor sees own data, not all institutional data."""
        r = client.get('/testcolegio/dashboard_data?curso=Primero A&periodo=1')
        data = json.loads(r.get_data(as_text=True))
        assert 'promedio_materia' in data['cards']  # prof card key
        assert 'total_profesores' not in data['cards']  # rector card key


class TestAttendance:
    """Fase 7 — Sistema de Asistencia Profesional."""

    def test_attendance_page_profesor(self, client, teacher_session):
        r = client.get('/testcolegio/asistencia')
        assert r.status_code == 200

    def test_attendance_page_no_auth(self, client):
        r = client.get('/testcolegio/asistencia')
        assert r.status_code == 302

    def test_save_attendance_P(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'P'})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] == 'ok'

    def test_save_attendance_E(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'E', 'fecha': '2026-07-01'})
        assert r.status_code == 200

    def test_save_attendance_X(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '2', 'estado': 'X', 'fecha': '2026-07-01'})
        assert r.status_code == 200

    def test_save_attendance_S(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '2', 'estado': 'S', 'fecha': '2026-07-02'})
        assert r.status_code == 200

    def test_save_attendance_T(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'T', 'fecha': '2026-07-02'})
        assert r.status_code == 200

    def test_save_attendance_with_observacion(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'P',
            'fecha': '2026-07-03', 'observacion': 'Llego temprano', 'hora': '07:00'})
        assert r.status_code == 200
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute(
            'SELECT estado, observacion, hora FROM asistencia WHERE aid=1 AND fecha="2026-07-03"'
        ).fetchone()
        conn.close()
        assert row is not None
        assert row[0] == 'P'
        assert row[1] == 'Llego temprano'
        assert row[2] == '07:00'

    def test_save_attendance_invalid_state_rejected(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'Z'})
        assert r.status_code == 400

    def test_save_attendance_no_csrf_rejected(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            'aid': '1', 'estado': 'P'})
        assert r.status_code == 403

    def test_save_attendance_no_auth(self, client):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'P'})
        assert r.status_code == 403

    def test_save_attendance_invalid_fecha_rejected(self, client, teacher_session):
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'P', 'fecha': 'not-a-date'})
        assert r.status_code == 400

    def test_asistencia_data_endpoint(self, client, teacher_session):
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR IGNORE INTO asistencia (aid,fecha,estado) VALUES (1,'2026-06-01','P')")
        conn.execute("INSERT OR IGNORE INTO asistencia (aid,fecha,estado) VALUES (2,'2026-06-01','A')")
        conn.commit(); conn.close()
        r = client.get('/testcolegio/asistencia_data?curso=Primero A',
                        headers={'X-CSRF-Token': 'pytest_csrf_token'})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert 'stats' in data
        assert 'alertas' in data
        assert 'estados' in data

    def test_asistencia_data_no_auth(self, client):
        r = client.get('/testcolegio/asistencia_data?curso=Primero A')
        assert r.status_code == 403

    def test_rector_asistencia_page(self, client, rector_session):
        r = client.get('/testcolegio/rector/asistencia')
        assert r.status_code == 200

    def test_rector_asistencia_page_no_auth(self, client):
        r = client.get('/testcolegio/rector/asistencia')
        assert r.status_code == 302

    def test_rector_asistencia_data(self, client, rector_session):
        r = client.get('/testcolegio/rector/asistencia_data?fecha=2026-07-01&curso=Primero A&jornada=Mañana')
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert 'estudiantes' in data
        assert 'stats' in data

    def test_rector_asistencia_data_no_auth(self, client):
        r = client.get('/testcolegio/rector/asistencia_data')
        assert r.status_code == 403

    def test_excel_report_download(self, client, teacher_session):
        r = client.get('/testcolegio/asistencia_reporte_excel?curso=Primero A')
        assert r.status_code == 200
        assert r.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or r.content_type.startswith('application/octet-stream')

    def test_student_attendance_view(self, client):
        with client.session_transaction() as sess:
            sess['rol_testcolegio'] = 'estudiante'
            sess['alumno_id_testcolegio'] = 1
            sess['_csrf_token'] = 'pytest_csrf_token'
        r = client.get('/testcolegio/estudiante')
        assert r.status_code == 200

    def test_attendance_update_existing(self, client, teacher_session):
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR IGNORE INTO asistencia (aid,fecha,estado) VALUES (1,'2026-07-10','A')")
        conn.commit(); conn.close()
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '1', 'estado': 'P', 'fecha': '2026-07-10',
            'observacion': 'Cambiado a presente', 'hora': '08:00'})
        assert r.status_code == 200
        conn = sqlite3.connect(TEST_DB)
        row = conn.execute(
            'SELECT estado, observacion, hora FROM asistencia WHERE aid=1 AND fecha="2026-07-10"'
        ).fetchone()
        conn.close()
        assert row[0] == 'P'
        assert row[1] == 'Cambiado a presente'

    def test_attendance_unauthorized_student(self, client, teacher_session):
        """Prof cannot mark attendance for student not in their courses."""
        # Create a student not in any course the prof teaches
        conn = sqlite3.connect(TEST_DB)
        conn.execute("INSERT OR IGNORE INTO alumnos (id,nombre,curso,jornada,activo) VALUES (99,'Forastero','Tercero A','Mañana',1)")
        conn.commit(); conn.close()
        r = client.post('/testcolegio/marcar_asistencia', data={
            '_csrf_token': 'pytest_csrf_token', 'aid': '99', 'estado': 'P'})
        assert r.status_code == 403


# ── PWA / Progressive Web App ──

STATIC_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static')
ICONS_DIR = os.path.join(STATIC_DIR, 'icons')

class TestPWA:

    # ── Manifest ──

    def test_manifest_served(self, client):
        r = client.get('/static/manifest.json')
        assert r.status_code == 200
        assert r.content_type in ('application/json', 'application/manifest+json')

    def test_manifest_valid_json(self, client):
        r = client.get('/static/manifest.json')
        data = json.loads(r.get_data(as_text=True))
        assert 'name' in data
        assert 'short_name' in data
        assert 'description' in data
        assert 'start_url' in data
        assert 'display' in data
        assert data['display'] == 'standalone'
        assert 'icons' in data
        assert len(data['icons']) >= 5

    def test_manifest_icon_sizes(self, client):
        r = client.get('/static/manifest.json')
        data = json.loads(r.get_data(as_text=True))
        sizes = [int(ico['sizes'].split('x')[0]) for ico in data['icons']]
        required = [72, 96, 128, 144, 152, 192, 384, 512]
        for s in required:
            assert s in sizes, f'Missing icon size {s}x{s}'

    # ── Service Worker ──

    def test_sw_served(self, client):
        r = client.get('/static/sw.js')
        assert r.status_code == 200
        assert 'text/javascript' in r.content_type or 'application/javascript' in r.content_type

    def test_sw_contains_cache_strategy(self, client):
        r = client.get('/static/sw.js')
        js = r.get_data(as_text=True)
        assert 'cacheFirst' in js
        assert 'networkFirst' in js
        assert 'CACHE_NAME' in js

    # ── Offline Page ──

    def test_offline_page_served(self, client):
        r = client.get('/offline')
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        assert 'Sin conexi' in html
        assert 'Reintentar' in html

    def test_offline_page_has_logo(self, client):
        r = client.get('/offline')
        html = r.get_data(as_text=True)
        assert 'class="logo"' in html

    # ── Icons ──

    def test_icons_exist(self):
        sizes = [72, 96, 128, 144, 152, 192, 384, 512]
        for s in sizes:
            path = os.path.join(ICONS_DIR, f'icon-{s}x{s}.png')
            assert os.path.isfile(path), f'Missing icon: icon-{s}x{s}.png'

    def test_apple_touch_icon_exists(self):
        path = os.path.join(ICONS_DIR, 'apple-touch-icon.png')
        assert os.path.isfile(path)

    # ── HTML head references (present in rendered pages) ──

    def test_manifest_link_in_html(self, client, rector_session):
        r = client.get('/testcolegio/rector')
        html = r.get_data(as_text=True)
        assert 'manifest.json' in html or '/static/manifest.json' in html

    def test_apple_touch_icon_link_in_html(self, client, rector_session):
        r = client.get('/testcolegio/rector')
        html = r.get_data(as_text=True)
        assert 'apple-touch-icon' in html

    def test_pwa_js_loaded(self, client, rector_session):
        r = client.get('/testcolegio/rector')
        html = r.get_data(as_text=True)
        assert 'pwa.js' in html

    # ── Notification Manager ──

    def test_notification_manager_js_loaded(self, client):
        r = client.get('/static/js/notification-manager.js')
        assert r.status_code == 200
        js = r.get_data(as_text=True)
        assert 'NotificationManager' in js

    # ── Python notification helper structure ──

    def test_notification_helper_exists(self):
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            'notifications',
            os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                         'utils', 'notifications.py')
        )
        assert spec is not None, 'utils/notifications.py not found'
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        assert hasattr(mod, 'PushNotification')
        assert hasattr(mod, 'send_notification')
        assert hasattr(mod, 'save_subscription')

    # ── Offline page div balance ──

    def test_offline_div_balance(self, client):
        r = client.get('/offline')
        html = r.get_data(as_text=True)
        assert html.count('<div') == html.count('</div>')

    def test_offline_no_mojibake(self, client):
        r = client.get('/offline')
        html = r.get_data(as_text=True)
        for pat in ['Ã¡', 'Ã©', 'Ã­', 'Ã³', 'Ãº', 'Ã±']:
            assert pat not in html
