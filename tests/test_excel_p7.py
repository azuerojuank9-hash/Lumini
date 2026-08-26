"""P7 — Excel como herramienta real (notas, estudiantes, asistencia, reportes).

Cubre:
  - Servicio excel_service (extensión, notas 0-5, fechas, estados asistencia).
  - Importar Notas: página con pasos, preview (vacío/columnas/notas/estudiante/
    duplicados), confirmar con revalidación server-side, CSRF y auditoría.
  - Migrar-excel: acepta .xlsx (fix) y rechaza .csv.
  - Reportes rector: allow-list (sin SQLi), CSRF, exportación .xlsx.
  - Excel institucional (rector + directora): importar estudiantes con dedupe,
    exportar estudiantes/cursos, y permisos (profesor bloqueado).
  - Importar asistencia: plantilla compatible, preview/confirmar, CSRF, auditoría.
"""

import io
import json
import os
import sqlite3
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ['FLASK_ENV'] = 'development'
os.environ['ENV'] = 'development'

import pytest

from flask_app import app
from test_app import TEST_DB, seed_test_db

seed_test_db()

SLUG = 'testcolegio'
CSRF = 'p7_excel_csrf'
MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@pytest.fixture
def client():
    app.config['TESTING'] = True
    with app.test_client() as c:
        yield c


@pytest.fixture
def csrf(client):
    with client.session_transaction() as sess:
        sess['_csrf_token'] = CSRF
    return CSRF


def db():
    conn = sqlite3.connect(TEST_DB)
    conn.row_factory = sqlite3.Row
    return conn


@pytest.fixture
def teacher(client, csrf):
    with client.session_transaction() as sess:
        sess[f'profesor_id_{SLUG}'] = 1
        sess[f'rol_{SLUG}'] = 'profesor'
        sess[f'jornada_{SLUG}'] = 'Mañana'
        sess[f'materia_{SLUG}'] = 'Matemáticas'
        sess['_csrf_token'] = CSRF


@pytest.fixture
def rector(client, csrf):
    with client.session_transaction() as sess:
        sess[f'rector_id_{SLUG}'] = 99
        sess['_csrf_token'] = CSRF


@pytest.fixture
def directora(client, csrf):
    conn = db()
    did = conn.execute("SELECT id FROM directoras WHERE usuario='directora'").fetchone()
    conn.close()
    assert did, 'Falta directora de test'
    with client.session_transaction() as sess:
        sess[f'directora_id_{SLUG}'] = did['id']
        sess['_csrf_token'] = CSRF


def make_xlsx(rows, sheet='Datos'):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def primer_alumno_primero_a():
    conn = db()
    a = conn.execute(
        "SELECT id, nombre FROM alumnos WHERE curso='Primero A' AND jornada='Mañana' AND activo=1 ORDER BY id LIMIT 1"
    ).fetchone()
    conn.close()
    return a


def primera_actividad():
    conn = db()
    a = conn.execute(
        "SELECT id, nombre FROM actividades WHERE profesor_id=1 AND materia='Matemáticas' "
        "AND curso='Primero A' AND COALESCE(periodo,1)=1 ORDER BY orden LIMIT 1"
    ).fetchone()
    conn.close()
    return a


# ── Servicio Excel ────────────────────────────────────────────────────────

class TestExcelService:
    def test_extension_excel_valida(self):
        from app.services.excel_service import extension_excel_valida
        assert extension_excel_valida('notas.xlsx')
        assert extension_excel_valida('NOTAS.XLSX')
        assert not extension_excel_valida('notas.csv')
        assert not extension_excel_valida('notas.xls')
        assert not extension_excel_valida('')

    def test_parsear_nota(self):
        from app.services.excel_service import parsear_nota
        assert parsear_nota(4.5) == (4.5, None)
        assert parsear_nota('3,75') == (3.75, None)
        assert parsear_nota('') == (None, None)
        assert parsear_nota(None) == (None, None)
        val, err = parsear_nota(6.0)
        assert err and '0.0' in err and '5.0' in err
        val, err = parsear_nota(-1)
        assert err
        val, err = parsear_nota('abc')
        assert err == 'valor no numérico'

    def test_parsear_nota_respeta_escala(self):
        from app.services.excel_service import parsear_nota
        val, err = parsear_nota(8.5, escala_min=1.0, escala_max=10.0)
        assert val == 8.5 and err is None
        val, err = parsear_nota(6.0, escala_min=1.0, escala_max=10.0)
        assert val == 6.0 and err is None
        val, err = parsear_nota(11.0, escala_min=1.0, escala_max=10.0)
        assert err and '10.0' in err
        val, err = parsear_nota(0.5, escala_min=1.0, escala_max=10.0)
        assert err and '1.0' in err

    def test_parsear_fecha(self):
        from app.services.excel_service import parsear_fecha
        assert parsear_fecha('2026-08-10') == ('2026-08-10', None)
        assert parsear_fecha('') == (None, None)
        assert parsear_fecha('10/08/2026')[1] is not None

    def test_parsear_estado_asistencia(self):
        from app.services.excel_service import parsear_estado_asistencia
        assert parsear_estado_asistencia('P') == ('P', None)
        assert parsear_estado_asistencia('presente') == ('P', None)
        assert parsear_estado_asistencia('T') == ('T', None)
        assert parsear_estado_asistencia('') == (None, None)
        assert parsear_estado_asistencia('Z')[1] is not None

    def test_leer_workbook_vacio(self):
        from app.services.excel_service import leer_workbook
        with pytest.raises(ValueError) as ei:
            leer_workbook(b'no es un xlsx')
        assert 'vacío' in str(ei.value) or 'formato' in str(ei.value).lower() or 'archivo' in str(ei.value)

    def test_leer_workbook_excede_columnas(self):
        from app.services.excel_service import leer_workbook
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for i in range(130):
            ws.cell(row=1, column=i + 1, value=f'C{i}')
        bio = io.BytesIO()
        wb.save(bio)
        with pytest.raises(ValueError) as ei:
            leer_workbook(bio.getvalue(), max_columnas=120)
        assert 'columnas' in str(ei.value)


# ── Importar Notas (P7) ───────────────────────────────────────────────────

class TestImportarNotasP7:
    def test_importar_notas_page_tiene_pasos(self, client, teacher, csrf):
        r = client.get(f'/{SLUG}/importar_notas?curso=Primero A&periodo=1')
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        assert 'Previsualizar' in html
        assert 'Confirmar importación' in html
        assert 'csrf_token' in html

    def test_preview_archivo_vacio(self, client, teacher, csrf):
        bio = make_xlsx([['N°', 'Estudiante', 'AID', 'Promedio']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'vacio.xlsx')})
        assert r.status_code == 400
        data = json.loads(r.get_data(as_text=True))
        assert 'mensaje' in data

    def test_preview_rechaza_csv(self, client, teacher, csrf):
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (io.BytesIO(b'x,y,z'), 'notas.csv')})
        assert r.status_code == 400
        data = json.loads(r.get_data(as_text=True))
        assert 'xlsx' in data['mensaje']

    def test_preview_columnas_invalidas(self, client, teacher, csrf):
        bio = make_xlsx([['Col1', 'Col2', 'Col3'], [1, 2, 3]])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'mal.xlsx')})
        assert r.status_code == 400
        data = json.loads(r.get_data(as_text=True))
        assert 'N°' in data['mensaje'] or 'N\u00b0' in data['mensaje']

    def test_preview_nota_fuera_de_rango(self, client, teacher, csrf):
        """Escala 1-10 del testcolegio: 11.0 queda fuera; 6.0 es válido."""
        alumno = primer_alumno_primero_a()
        actividad = primera_actividad()
        assert alumno and actividad
        bio = make_xlsx([['N°', 'Estudiante', 'AID', actividad['nombre'], 'Promedio'],
                         [1, alumno['nombre'], alumno['id'], 11.0, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'rango.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is False
        assert data['filas'][0]['ok'] is False
        assert any('rango' in e.lower() for e in data['filas'][0]['errors'])

    def test_preview_nota_dentro_escala_10(self, client, teacher, csrf):
        """Una nota de 6.0 (escala 1-10) no debe marcarse como error."""
        alumno = primer_alumno_primero_a()
        actividad = primera_actividad()
        assert alumno and actividad
        bio = make_xlsx([['N°', 'Estudiante', 'AID', actividad['nombre'], 'Promedio'],
                         [1, alumno['nombre'], alumno['id'], 6.0, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'escala10.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is True
        assert data['filas'][0]['ok'] is True

    def test_preview_reutiliza_formato_exportacion(self, client, teacher, csrf):
        """El Excel exportado por LUMINI se puede volver a importar (roundtrip)."""
        r = client.get(f'/{SLUG}/exportar_notas?curso=Primero A&periodo=1')
        assert r.status_code == 200
        assert r.content_type == MIME_XLSX
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (io.BytesIO(r.data), 'exportado.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is True, data.get('filas', [{}])[0].get('errors')
        assert data['total'] >= 28
        assert data['validos'] == data['total']

    def test_importar_28_estudiantes_y_confirmar(self, client, teacher, csrf):
        """Escenario de la prueba de evaluación: 28 estudiantes, varias
        actividades y notas distintas dentro de la escala 1-10."""
        conn = db()
        alumnos = conn.execute(
            "SELECT id, nombre FROM alumnos WHERE curso='Primero A' AND jornada='Mañana' "
            "AND activo=1 ORDER BY id LIMIT 28").fetchall()
        actividades = conn.execute(
            "SELECT id, nombre FROM actividades WHERE profesor_id=1 AND materia='Matemáticas' "
            "AND curso='Primero A' AND jornada='Mañana' AND COALESCE(periodo,1)=1 "
            "AND nombre IN ('Tarea 1','Examen 1') ORDER BY orden LIMIT 2").fetchall()
        conn.close()
        assert len(alumnos) == 28
        assert len(actividades) == 2
        filas = [['N°', 'Estudiante', 'AID', actividades[0]['nombre'], actividades[1]['nombre'],
                  'Evaluación', 'Autoevaluación', 'Promedio']]
        for i, al in enumerate(alumnos, 1):
            filas.append([i, al['nombre'], al['id'], round(4 + i % 5, 1), round(6 + i % 3, 1),
                          round(5.5 + i % 3, 1), round(6.5 + i % 2, 1), ''])
        bio = make_xlsx(filas)
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, '28.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is True, data.get('filas', [{}])[0].get('errors')
        assert data['total'] == 28 and data['errores'] == 0
        r = client.post(f'/{SLUG}/importar_notas/confirmar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'data': json.dumps(data)})
        assert r.status_code == 200
        res = json.loads(r.get_data(as_text=True))
        assert res['status'] == 'ok'
        conn = db()
        saved = conn.execute(
            'SELECT COUNT(*) c FROM notas WHERE aid IN (%s) AND actividad_id IN (%s)'
            % (','.join('?' * 28), ','.join('?' * 2)),
            tuple(a['id'] for a in alumnos) + (actividades[0]['id'], actividades[1]['id'])).fetchone()['c']
        conn.close()
        assert saved == 28 * 2

    def test_preview_estudiante_inexistente(self, client, teacher, csrf):
        bio = make_xlsx([['N°', 'Estudiante', 'AID', 'Promedio'],
                         [1, 'Estudiante Que No Existe', 999999, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'inexistente.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is False
        assert data['filas'][0]['ok'] is False
        assert any('estudiante' in e.lower() for e in data['filas'][0]['errors'])

    def test_preview_estudiante_duplicado(self, client, teacher, csrf):
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N°', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], alumno['id'], ''],
                         [2, alumno['nombre'], alumno['id'], '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'dup.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is False
        assert data['filas'][1]['ok'] is False
        assert any('duplicado' in e.lower() for e in data['filas'][1]['errors'])

    def test_confirmar_sin_csrf(self, client, teacher, csrf):
        r = client.post(f'/{SLUG}/importar_notas/confirmar', data={'curso': 'Primero A', 'periodo': '1', 'data': '{}'})
        assert r.status_code == 403

    def test_confirmar_rechaza_all_ok_falso(self, client, teacher, csrf):
        data = {'all_ok': False, 'filas': [], 'nuevas_actividades': []}
        r = client.post(f'/{SLUG}/importar_notas/confirmar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'data': json.dumps(data)})
        assert r.status_code == 400

    def test_confirmar_rechaza_estudiante_ajeno(self, client, teacher, csrf):
        """Un estudiante de otro curso no debe poder guardarse vía confirmar
        aunque el payload diga all_ok (revalidación server-side)."""
        conn = db()
        ajeno = conn.execute(
            "SELECT id, nombre FROM alumnos WHERE curso='Segundo A' AND jornada='Mañana' AND activo=1 LIMIT 1"
        ).fetchone()
        conn.close()
        assert ajeno
        data = {
            'all_ok': True,
            'curso': 'Primero A',
            'filas': [{'fila': 2, 'aid': ajeno['id'], 'alumno': dict(ajeno),
                       'ok': True, 'errors': [], 'changes': {}}],
            'nuevas_actividades': [],
        }
        r = client.post(f'/{SLUG}/importar_notas/confirmar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'data': json.dumps(data)})
        assert r.status_code == 400

    def test_confirmar_guarda_y_audita(self, client, teacher, csrf):
        alumno = primer_alumno_primero_a()
        actividad = primera_actividad()
        assert alumno and actividad
        bio = make_xlsx([['N°', 'Estudiante', 'AID', actividad['nombre'], 'Promedio'],
                         [1, alumno['nombre'], alumno['id'], 4.5, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'ok.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is True
        r = client.post(f'/{SLUG}/importar_notas/confirmar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'data': json.dumps(data)})
        assert r.status_code == 200
        res = json.loads(r.get_data(as_text=True))
        assert res['status'] == 'ok'
        conn = db()
        row = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?',
                           (alumno['id'], actividad['id'])).fetchone()
        assert row and abs(row['val'] - 4.5) < 0.001
        audit = conn.execute(
            "SELECT COUNT(*) as c FROM auditoria_notas WHERE aid=? AND motivo='Importacion masiva Excel'",
            (alumno['id'],)).fetchone()['c']
        conn.close()
        assert audit >= 1


# ── Importar Notas: robustez AID/Nombre ─────────────────────────────────────

class TestImportarNotasRobustez:
    """Tests for AID parsing edge cases and name normalization."""

    def test_aid_integer(self, client, teacher, csrf):
        """AID as integer should work."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], alumno['id'], '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'int.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True
        assert data['filas'][0]['aid'] == alumno['id']

    def test_aid_as_float_in_excel(self, client, teacher, csrf):
        """AID as float (15.0) should be parsed correctly."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], float(alumno['id']), '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'float.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True
        assert data['filas'][0]['aid'] == alumno['id']

    def test_aid_as_string_integer(self, client, teacher, csrf):
        """AID as string '15' should be parsed correctly."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], str(alumno['id']), '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'str_int.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True
        assert data['filas'][0]['aid'] == alumno['id']

    def test_aid_as_string_float(self, client, teacher, csrf):
        """AID as string '15.0' (Excel text format) should be parsed correctly."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], f'{alumno["id"]}.0', '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'str_float.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True
        assert data['filas'][0]['aid'] == alumno['id']

    def test_aid_as_string_with_comma(self, client, teacher, csrf):
        """AID as string '15,0' (comma decimal) should be parsed correctly."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], f'{alumno["id"]},0', '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'str_comma.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True
        assert data['filas'][0]['aid'] == alumno['id']

    def test_aid_as_string_with_spaces(self, client, teacher, csrf):
        """AID as string ' 15 ' (extra spaces) should be parsed correctly."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], f' {alumno["id"]} ', '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'str_spaces.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True
        assert data['filas'][0]['aid'] == alumno['id']

    def test_name_lowercase_matches(self, client, teacher, csrf):
        """Lowercase name should match DB name via normalization."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'].lower(), 999999, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'lower.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True

    def test_name_uppercase_matches(self, client, teacher, csrf):
        """Uppercase name should match DB name via normalization."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'].upper(), 999999, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'upper.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True

    def test_name_with_extra_spaces(self, client, teacher, csrf):
        """Name with leading/trailing spaces should match."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, f'  {alumno["nombre"]}  ', 999999, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'spaces.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True

    def test_wrong_aid_correct_name_matches(self, client, teacher, csrf):
        """Wrong AID but correct name should find student by name."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], 999999, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'wrongaid.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is True
        assert data['filas'][0]['aid'] == alumno['id']

    def test_student_from_other_course_rejected(self, client, teacher, csrf):
        """Student from a different course must be rejected."""
        conn = db()
        ajeno = conn.execute(
            "SELECT id, nombre FROM alumnos WHERE curso='Segundo A' "
            "AND jornada='Ma\u00f1ana' AND activo=1 LIMIT 1"
        ).fetchone()
        conn.close()
        assert ajeno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, ajeno['nombre'], ajeno['id'], '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'ajeno.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is False
        assert any('no encontrado' in e.lower() for e in data['filas'][0]['errors'])

    def test_real_duplicate_student_detected(self, client, teacher, csrf):
        """Same student appearing twice should be flagged as duplicate."""
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, alumno['nombre'], alumno['id'], ''],
                         [2, alumno['nombre'], alumno['id'], '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'dup.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][1]['ok'] is False
        assert any('duplicado' in e.lower() for e in data['filas'][1]['errors'])

    def test_28_students_valid(self, client, teacher, csrf):
        """28 students with correct AIDs should all be valid."""
        conn = db()
        alumnos = conn.execute(
            "SELECT id, nombre FROM alumnos WHERE curso='Primero A' "
            "AND jornada='Ma\u00f1ana' AND activo=1 ORDER BY id LIMIT 28"
        ).fetchall()
        conn.close()
        assert len(alumnos) == 28
        filas = [['N\u00b0', 'Estudiante', 'AID', 'Promedio']]
        for i, al in enumerate(alumnos, 1):
            filas.append([i, al['nombre'], al['id'], round(4 + i % 5, 1)])
        bio = make_xlsx(filas)
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, '28.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is True
        assert data['total'] == 28
        assert data['errores'] == 0

    def test_plantilla_roundtrip(self, client, teacher, csrf):
        """Plantilla downloaded from LUMINI can be re-imported."""
        r = client.get(f'/{SLUG}/plantilla_notas?curso=Primero A&periodo=1')
        assert r.status_code == 200
        r2 = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (io.BytesIO(r.data), 'plantilla.xlsx')})
        assert r2.status_code == 200
        data = json.loads(r2.get_data(as_text=True))
        assert data['all_ok'] is True
        assert data['total'] >= 28
        assert data['validos'] == data['total']

    def test_no_duplicate_error_when_student_not_found(self, client, teacher, csrf):
        """A student not found should NOT also get 'duplicado' error."""
        bio = make_xlsx([['N\u00b0', 'Estudiante', 'AID', 'Promedio'],
                         [1, 'Fantasma Uno', 999998, ''],
                         [2, 'Fantasma Dos', 999998, '']])
        r = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'ghost.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['filas'][0]['ok'] is False
        assert data['filas'][1]['ok'] is False
        assert any('no encontrado' in e.lower() for e in data['filas'][0]['errors'])
        assert not any('duplicado' in e.lower() for e in data['filas'][0]['errors'])
        assert not any('duplicado' in e.lower() for e in data['filas'][1]['errors'])

    def test_export_roundtrip(self, client, teacher, csrf):
        """Exported notas Excel can be re-imported."""
        r = client.get(f'/{SLUG}/exportar_notas?curso=Primero A&periodo=1')
        assert r.status_code == 200
        r2 = client.post(f'/{SLUG}/importar_notas/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (io.BytesIO(r.data), 'export.xlsx')})
        assert r2.status_code == 200
        data = json.loads(r2.get_data(as_text=True))
        assert data['all_ok'] is True
        assert data['validos'] == data['total']


# ── Migrar-Excel (obsoleto pero funcional) ────────────────────────────────

class TestMigrarExcel:
    def test_analizar_acepta_xlsx(self, client, teacher, csrf):
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['#', 'Estudiante', 'Trabajo 1'],
                         [1, alumno['nombre'], 4.0]])
        r = client.post(f'/{SLUG}/migrar-excel/analizar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (bio, 'migra.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['status'] in ('ok', 'warning')
        assert data['total'] >= 1

    def test_analizar_rechaza_csv(self, client, teacher, csrf):
        r = client.post(f'/{SLUG}/migrar-excel/analizar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'periodo': '1',
            'archivo': (io.BytesIO(b'a,b,c'), 'migra.csv')})
        assert r.status_code == 400

    def test_analizar_sin_csrf(self, client, teacher, csrf):
        bio = make_xlsx([['#', 'Estudiante'], [1, 'X']])
        r = client.post(f'/{SLUG}/migrar-excel/analizar', data={
            'curso': 'Primero A', 'periodo': '1', 'archivo': (bio, 'x.xlsx')})
        assert r.status_code == 403

    def test_previsualizar_sin_csrf(self, client, teacher, csrf):
        r = client.post(f'/{SLUG}/migrar/previsualizar',
                        json={'contenido': 'a', 'tipo': 'estudiantes'})
        assert r.status_code == 403

    def test_ejecutar_sin_csrf(self, client, teacher, csrf):
        r = client.post(f'/{SLUG}/migrar/ejecutar',
                        json={'contenido': 'a', 'tipo': 'estudiantes'})
        assert r.status_code == 403


# ── Reportes del rector (seguridad) ───────────────────────────────────────

class TestReportesSeguridad:
    def test_tablas_allowlist(self, client, rector, csrf):
        r = client.get(f'/{SLUG}/reportes/tablas')
        assert r.status_code == 200
        tablas = r.get_json()['tablas']
        assert 'alumnos' in tablas
        assert 'notas' in tablas
        assert 'asistencia' in tablas
        assert 'sqlite_master' not in tablas

    def test_ejecutar_tabla_no_permitida(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/ejecutar',
                        json={'tabla': 'sqlite_master', 'campos': ['name']},
                        headers={'X-CSRF-Token': CSRF})
        assert r.status_code == 400

    def test_ejecutar_columna_no_permitida(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/ejecutar',
                        json={'tabla': 'alumnos', 'campos': ['id', 'nombre; DROP']},
                        headers={'X-CSRF-Token': CSRF})
        assert r.status_code == 400

    def test_ejecutar_sqli_en_tabla(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/ejecutar',
                        json={'tabla': 'alumnos; DROP TABLE alumnos', 'campos': ['id']},
                        headers={'X-CSRF-Token': CSRF})
        assert r.status_code == 400

    def test_ejecutar_sin_csrf(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/ejecutar',
                        json={'tabla': 'alumnos', 'campos': ['id']})
        assert r.status_code in (400, 403)

    def test_ejecutar_valido(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/ejecutar',
                        json={'tabla': 'alumnos', 'campos': ['id', 'nombre']},
                        headers={'X-CSRF-Token': CSRF})
        assert r.status_code == 200
        body = r.get_json()
        assert body['columnas'] == ['id', 'nombre']
        assert body['total'] >= 1

    def test_exportar_excel_rector(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/exportar_excel',
                        json={'tabla': 'alumnos', 'campos': ['id', 'nombre', 'curso']},
                        headers={'X-CSRF-Token': CSRF})
        assert r.status_code == 200
        assert r.content_type == MIME_XLSX
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.data))
        assert wb.active.max_row >= 2

    def test_exportar_excel_tabla_invalida(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/exportar_excel',
                        json={'tabla': 'no_existe', 'campos': ['id']},
                        headers={'X-CSRF-Token': CSRF})
        assert r.status_code == 400

    def test_exportar_excel_sin_csrf(self, client, rector, csrf):
        r = client.post(f'/{SLUG}/reportes/exportar_excel',
                        json={'tabla': 'alumnos', 'campos': ['id']})
        assert r.status_code == 403


# ── Excel institucional (permisos + estudiantes) ──────────────────────────

class TestExcelInstitucional:
    def test_profesor_no_accede(self, client, teacher, csrf):
        r = client.get(f'/{SLUG}/institucional/excel')
        assert r.status_code == 302

    def test_profesor_no_exporta(self, client, teacher, csrf):
        r = client.post(f'/{SLUG}/reportes/exportar_excel',
                        json={'tabla': 'alumnos', 'campos': ['id']},
                        headers={'X-CSRF-Token': CSRF})
        assert r.status_code == 403

    def test_profesor_no_importa_estudiantes(self, client, teacher, csrf):
        r = client.get(f'/{SLUG}/institucional/importar_estudiantes')
        assert r.status_code == 302

    def test_directora_accede(self, client, directora, csrf):
        r = client.get(f'/{SLUG}/institucional/excel')
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        assert 'Exportar estudiantes' in html
        assert 'Importar estudiantes' in html

    def test_directora_exporta_estudiantes(self, client, directora, csrf):
        r = client.get(f'/{SLUG}/institucional/exportar_estudiantes')
        assert r.status_code == 200
        assert r.content_type == MIME_XLSX
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.data))
        headers = [c.value for c in next(wb.active.iter_rows(max_row=1))]
        assert 'Nombre' in headers

    def test_directora_exporta_cursos(self, client, directora, csrf):
        r = client.get(f'/{SLUG}/institucional/exportar_cursos')
        assert r.status_code == 200
        assert r.content_type == MIME_XLSX

    def test_importar_estudiantes_page(self, client, directora, csrf):
        r = client.get(f'/{SLUG}/institucional/importar_estudiantes')
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        assert 'Importar estudiantes' in html
        assert 'csrf_token' in html

    def test_importar_estudiantes_sin_csrf(self, client, directora, csrf):
        bio = make_xlsx([['Nombre', 'Documento'], ['X']])
        r = client.post(f'/{SLUG}/institucional/importar_estudiantes/preview', data={
            'curso': 'Primero A', 'jornada': 'Mañana', 'archivo': (bio, 'e.xlsx')})
        assert r.status_code == 403

    def test_importar_estudiantes_preview_nuevo_existente(self, client, directora, csrf):
        conn = db()
        existente = conn.execute(
            "SELECT nombre FROM alumnos WHERE curso='Primero A' AND jornada='Mañana' LIMIT 1"
        ).fetchone()
        conn.close()
        assert existente
        nombre_nuevo = f'P7 Nuevo {datetime.now().strftime("%H%M%S")}'
        bio = make_xlsx([['Nombre', 'Documento'],
                         [existente['nombre'], ''],
                         [nombre_nuevo, '12345']])
        r = client.post(f'/{SLUG}/institucional/importar_estudiantes/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'jornada': 'Mañana',
            'archivo': (bio, 'est.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        estados = {f['estado'] for f in data['filas']}
        assert 'existe' in estados
        assert 'nuevo' in estados
        assert data['nuevos'] >= 1

    def test_importar_estudiantes_preview_duplicado(self, client, directora, csrf):
        nombre_nuevo = f'P7 Dup {datetime.now().strftime("%H%M%S")}'
        bio = make_xlsx([['Nombre'],
                         [nombre_nuevo],
                         [nombre_nuevo]])
        r = client.post(f'/{SLUG}/institucional/importar_estudiantes/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'jornada': 'Mañana',
            'archivo': (bio, 'dup.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['errores'] >= 1
        assert any('duplicado' in (f['errores'][0] or '') for f in data['filas'] if f['errores'])

    def test_importar_estudiantes_confirmar(self, client, directora, csrf):
        nombre_nuevo = f'P7 Guardar {datetime.now().strftime("%H%M%S")}'
        bio = make_xlsx([['Nombre', 'Documento'], [nombre_nuevo, '999']])
        r = client.post(f'/{SLUG}/institucional/importar_estudiantes/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'jornada': 'Mañana',
            'archivo': (bio, 's.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['nuevos'] == 1
        r = client.post(f'/{SLUG}/institucional/importar_estudiantes/confirmar', data={
            '_csrf_token': CSRF, 'data': json.dumps(data)})
        assert r.status_code == 200
        res = json.loads(r.get_data(as_text=True))
        assert res['status'] == 'ok' and res['insertados'] == 1
        conn = db()
        row = conn.execute('SELECT id FROM alumnos WHERE nombre=? AND curso=? AND jornada=?',
                           (nombre_nuevo, 'Primero A', 'Mañana')).fetchone()
        audit = conn.execute("SELECT COUNT(*) as c FROM audit_log WHERE accion='importar_estudiantes'").fetchone()['c']
        conn.close()
        assert row is not None
        assert audit >= 1


# ── Importar asistencia ───────────────────────────────────────────────────

class TestImportarAsistencia:
    FECHA = '2026-08-10'
    FECHA2 = '2026-08-11'

    def test_page(self, client, teacher, csrf):
        r = client.get(f'/{SLUG}/importar_asistencia?curso=Primero A')
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        assert 'Importar asistencia' in html
        assert 'csrf_token' in html

    def test_plantilla_compatible(self, client, teacher, csrf):
        """El Excel actual de asistencia sirve como plantilla para el importador."""
        r = client.get(f'/{SLUG}/asistencia_reporte_excel?curso=Primero A')
        assert r.status_code == 200
        assert r.content_type == MIME_XLSX
        r = client.post(f'/{SLUG}/importar_asistencia/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A',
            'archivo': (io.BytesIO(r.data), 'plantilla_asistencia.xlsx')})
        assert r.status_code == 200

    def test_preview_valido(self, client, teacher, csrf):
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['#', 'Estudiante', self.FECHA, self.FECHA2],
                         [1, alumno['nombre'], 'P', 'T']])
        r = client.post(f'/{SLUG}/importar_asistencia/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A',
            'archivo': (bio, 'asis.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is True
        assert data['fechas'] == [self.FECHA, self.FECHA2]
        assert data['filas'][0]['cambios'][self.FECHA]['estado'] == 'P'

    def test_preview_estado_invalido(self, client, teacher, csrf):
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['#', 'Estudiante', self.FECHA],
                         [1, alumno['nombre'], 'Z']])
        r = client.post(f'/{SLUG}/importar_asistencia/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A',
            'archivo': (bio, 'bad.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is False
        assert any('estado no válido' in e for e in data['filas'][0]['errores'])

    def test_preview_estudiante_inexistente(self, client, teacher, csrf):
        bio = make_xlsx([['#', 'Estudiante', self.FECHA],
                         ['', 'Nadie De Este Curso', 'P']])
        r = client.post(f'/{SLUG}/importar_asistencia/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A',
            'archivo': (bio, 'no.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is False
        assert any('estudiante' in e for e in data['filas'][0]['errores'])

    def test_preview_sin_csrf(self, client, teacher, csrf):
        bio = make_xlsx([['#', 'Estudiante', self.FECHA]])
        r = client.post(f'/{SLUG}/importar_asistencia/preview', data={
            'curso': 'Primero A', 'archivo': (bio, 'x.xlsx')})
        assert r.status_code == 403

    def test_preview_archivo_vacio(self, client, teacher, csrf):
        bio = make_xlsx([['#', 'Estudiante']])
        r = client.post(f'/{SLUG}/importar_asistencia/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'archivo': (bio, 'v.xlsx')})
        assert r.status_code == 400

    def test_confirmar_guarda_y_audita(self, client, teacher, csrf):
        alumno = primer_alumno_primero_a()
        assert alumno
        bio = make_xlsx([['#', 'Estudiante', self.FECHA],
                         [1, alumno['nombre'], 'P']])
        r = client.post(f'/{SLUG}/importar_asistencia/preview', data={
            '_csrf_token': CSRF, 'curso': 'Primero A',
            'archivo': (bio, 'ok.xlsx')})
        assert r.status_code == 200
        data = json.loads(r.get_data(as_text=True))
        assert data['all_ok'] is True
        r = client.post(f'/{SLUG}/importar_asistencia/confirmar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'data': json.dumps(data)})
        assert r.status_code == 200
        res = json.loads(r.get_data(as_text=True))
        assert res['status'] == 'ok'
        assert res['updated'] >= 1
        conn = db()
        row = conn.execute(
            'SELECT estado FROM asistencia WHERE aid=? AND fecha=?', (alumno['id'], self.FECHA)).fetchone()
        audit = conn.execute(
            "SELECT COUNT(*) as c FROM audit_log WHERE accion='asistencia_importada'").fetchone()['c']
        conn.close()
        assert row and row['estado'] == 'P'
        assert audit >= 1

    def test_confirmar_sin_csrf(self, client, teacher, csrf):
        r = client.post(f'/{SLUG}/importar_asistencia/confirmar',
                        data={'curso': 'Primero A', 'data': '{}'})
        assert r.status_code == 403

    def test_confirmar_rechaza_all_ok_falso(self, client, teacher, csrf):
        data = {'all_ok': False, 'fechas': [self.FECHA], 'filas': []}
        r = client.post(f'/{SLUG}/importar_asistencia/confirmar', data={
            '_csrf_token': CSRF, 'curso': 'Primero A', 'data': json.dumps(data)})
        assert r.status_code == 400
