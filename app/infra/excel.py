from app.infra.database import conectar
from app.infra.grades import _promedio_ponderado


def _excel_armar_wb(slug, prof, materia, jornada, curso_sel, periodo, actividades, alumnos):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = 'Notas'
    header_fill = PatternFill('solid', fgColor='6D28D9')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    headers = ['N°', 'Estudiante', 'AID'] + [a['nombre'] for a in actividades] + ['Evaluación', 'Autoevaluación', 'Promedio']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    conn = conectar(slug)
    try:
        aid_list = [a['id'] for a in alumnos]
        if aid_list:
            placeholders = ','.join('?' * len(aid_list))
            notas_all = conn.execute(
                f'''SELECT n.aid, n.actividad_id, n.val FROM notas n
                    JOIN actividades ac ON ac.id=n.actividad_id
                    WHERE n.aid IN ({placeholders}) AND ac.materia=? AND ac.jornada=? AND ac.curso=?
                    AND COALESCE(ac.periodo,1)=? AND ac.profesor_id=?''',
                (*aid_list, materia, jornada, curso_sel, periodo, prof['id'])).fetchall()
            notas_by_aid = {}
            for r in notas_all:
                notas_by_aid.setdefault(r['aid'], {})[r['actividad_id']] = r['val']
            evals_all = conn.execute(
                f'''SELECT aid, evaluacion, autoevaluacion FROM evaluaciones
                    WHERE aid IN ({placeholders}) AND profesor_id=? AND materia=? AND jornada=?
                    AND COALESCE(periodo,1)=?''',
                (*aid_list, prof['id'], materia, jornada, periodo)).fetchall()
            evals_by_aid = {r['aid']: {'ev': r['evaluacion'], 'auto': r['autoevaluacion']} for r in evals_all}
        else:
            notas_by_aid, evals_by_aid = {}, {}
        for i, a in enumerate(alumnos, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).alignment = center
            ws.cell(row=row, column=2, value=a['nombre'])
            ws.cell(row=row, column=3, value=a['id']).alignment = center
            for j, act in enumerate(actividades, 4):
                val = None
                if a['id'] in notas_by_aid:
                    val = notas_by_aid[a['id']].get(act['id'])
                if val is not None:
                    ws.cell(row=row, column=j, value=float(val)).alignment = center
            ev = evals_by_aid.get(a['id'], {})
            ev_val = ev.get('ev')
            auto_val = ev.get('auto')
            ecol = 4 + len(actividades)
            if ev_val is not None:
                ws.cell(row=row, column=ecol, value=float(ev_val)).alignment = center
            if auto_val is not None:
                ws.cell(row=row, column=ecol + 1, value=float(auto_val)).alignment = center
            notas_dict = notas_by_aid.get(a['id'], {})
            prom = _promedio_ponderado([notas_dict.get(act['id']) for act in actividades], ev_val, auto_val)
            ws.cell(row=row, column=ecol + 2, value=round(prom, 2) if prom is not None else '').alignment = center
    finally:
        conn.close()
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    for j in range(len(actividades)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 14
    return wb
