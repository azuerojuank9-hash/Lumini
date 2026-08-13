import logging
from datetime import datetime, timedelta
from io import BytesIO

from flask import Blueprint, Response, jsonify, redirect, render_template, request, url_for

from app.services.attendance_service import (
    COLORES_ASISTENCIA,
    ESTADOS_ASISTENCIA,
    build_asistencia_calendario,
    compute_asistencia_alertas,
    compute_asistencia_stats,
)
from app.utils.security import validar_csrf

logger = logging.getLogger(__name__)

attendance_bp = Blueprint('attendance', __name__)


def _fa():
    import flask_app
    return flask_app


@attendance_bp.route('/<slug>/asistencia', methods=['GET'])
def asistencia(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('teacher.seleccionar_jornada', slug=slug))
    colegio = f.get_colegio(slug)
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else None)
    fecha_sel = request.args.get('fecha', datetime.today().strftime('%Y-%m-%d'))
    try:
        fecha_dt = datetime.strptime(fecha_sel, '%Y-%m-%d') if fecha_sel else datetime.today()
    except ValueError:
        fecha_sel = datetime.today().strftime('%Y-%m-%d')
        fecha_dt = datetime.today()
    fecha_sel_dia_anterior = (fecha_dt - timedelta(days=1)).strftime('%Y-%m-%d')
    fecha_sel_dia_siguiente = (fecha_dt + timedelta(days=1)).strftime('%Y-%m-%d')
    hoy_fecha = datetime.today().strftime('%Y-%m-%d')
    hoy_hora = datetime.today().strftime('%H:%M')
    if not curso_sel:
        return render_template('asistencia.html', profesor=prof, slug=slug, colegio=colegio,
                               materia=materia, jornada=jornada, mis_cursos=mis_cursos,
                               curso_sel=None, estudiantes=[], fecha_sel=fecha_sel,
                               fecha_sel_dia_anterior=fecha_sel_dia_anterior,
                               fecha_sel_dia_siguiente=fecha_sel_dia_siguiente,
                               hoy_fecha=hoy_fecha, hoy_hora=hoy_hora,
                               estados_asistencia=ESTADOS_ASISTENCIA,
                               colores_asistencia=COLORES_ASISTENCIA,
                               materias_jornadas=f.get_materias_profesor(slug, prof['id']))
    conn = f.conectar(slug)
    try:
        from app.repositories.attendance_repository import get_asistencia_for_date, get_students_by_curso
        alumnos = get_students_by_curso(conn, curso_sel, jornada)
        asis_rows = []
        if alumnos:
            aid_tuple = tuple(a['id'] for a in alumnos)
            asis_rows = get_asistencia_for_date(conn, fecha_sel, list(aid_tuple))
        asis_map = {r['aid']: {'estado': r['estado'], 'observacion': r['observacion'] or '', 'hora': r['hora'] or ''} for r in asis_rows}
        datos = []
        for a in alumnos:
            info = asis_map.get(a['id'], {})
            datos.append({
                'id': a['id'], 'nombre': a['nombre'],
                'num_curso': a['num_curso'],
                'asistencia': info.get('estado', ''),
                'observacion': info.get('observacion', ''),
                'hora': info.get('hora', ''),
            })
        stats = compute_asistencia_stats(conn, curso=curso_sel, jornada=jornada)
    finally:
        conn.close()
    return render_template('asistencia.html', profesor=prof, slug=slug, colegio=colegio,
                           materia=materia, jornada=jornada, mis_cursos=mis_cursos,
                           curso_sel=curso_sel, estudiantes=datos, fecha_sel=fecha_sel,
                           fecha_sel_dia_anterior=fecha_sel_dia_anterior,
                           fecha_sel_dia_siguiente=fecha_sel_dia_siguiente,
                           hoy_fecha=hoy_fecha, hoy_hora=hoy_hora,
                           stats=stats,
                           estados_asistencia=ESTADOS_ASISTENCIA,
                           colores_asistencia=COLORES_ASISTENCIA,
                           materias_jornadas=f.get_materias_profesor(slug, prof['id']))


@attendance_bp.route('/<slug>/marcar_asistencia', methods=['POST'])
def marcar_asistencia(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return ('', 403)
    if not validar_csrf():
        return ('Error CSRF', 403)
    aid = request.form.get('aid', type=int)
    estado = request.form.get('estado')
    fecha = request.form.get('fecha', '')
    observacion = request.form.get('observacion', '').strip()
    hora = request.form.get('hora', '')
    if aid is None or (not estado and not observacion):
        return ('', 400)
    if estado and estado not in ESTADOS_ASISTENCIA:
        return ('', 400)
    if fecha:
        try:
            datetime.strptime(fecha, '%Y-%m-%d')
        except ValueError:
            return ('', 400)
    jornada, materia = f.get_sesion_jornada_materia(slug)
    conn = f.conectar(slug)
    cursos_prof = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not cursos_prof:
        conn.close()
        return ('', 403)
    from app.repositories.attendance_repository import upsert_asistencia, verify_student_in_cursos
    if not verify_student_in_cursos(conn, aid, cursos_prof, jornada):
        conn.close()
        return ('', 403)
    upsert_asistencia(conn, aid, fecha if fecha else None, estado, observacion, hora, 'profesor', prof['id'])
    conn.commit()
    f.audit_log(slug, prof['id'], 'asistencia_editada', 'asistencia', aid,
                None, {'estado': estado, 'observacion': observacion, 'hora': hora})
    conn.close()
    return jsonify({'status': 'ok'})


@attendance_bp.route('/<slug>/asistencia_data')
def asistencia_data(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'error': 'Sin jornada/materia'}), 400
    conn = f.conectar(slug)
    try:
        curso = request.args.get('curso', '')
        if not curso:
            conn.close()
            return jsonify({'error': 'Curso requerido'}), 400
        stats = compute_asistencia_stats(conn, curso=curso, jornada=jornada)
        alertas = compute_asistencia_alertas(conn, slug, curso, jornada)
        calendario = build_asistencia_calendario(conn, curso, jornada)
    finally:
        conn.close()
    return jsonify({
        'stats': stats,
        'alertas': alertas,
        'calendario': {k: dict(v) for k, v in calendario.items()},
        'estados': dict(ESTADOS_ASISTENCIA),
        'colores': COLORES_ASISTENCIA,
    })


@attendance_bp.route('/<slug>/asistencia_reporte_excel')
def asistencia_reporte_excel(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return ('', 403)
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return ('', 400)
    conn = f.conectar(slug)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        wb = Workbook()
        ws = wb.active
        ws.title = 'Asistencia'
        hd_font = Font(bold=True, color='FFFFFF', size=11)
        hd_fill = PatternFill('solid', fgColor='1E293B')
        thin = Side(style='thin', color='334155')
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        curso = request.args.get('curso', '')
        if not curso:
            conn.close()
            return ('Curso requerido', 400)
        from app.repositories.attendance_repository import get_asistencia_full, get_students_by_curso
        alumnos = get_students_by_curso(conn, curso, jornada)
        if not alumnos:
            conn.close()
            return ('Sin estudiantes', 404)
        aids = [a['id'] for a in alumnos]
        asis_rows = get_asistencia_full(conn, aids)
        fechas = sorted(set(r['fecha'] for r in asis_rows))
        header = ['#', 'Estudiante'] + fechas
        ws.append(header)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hd_font
            cell.fill = hd_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        asis_map = {}
        for r in asis_rows:
            asis_map.setdefault(r['aid'], {})[r['fecha']] = {'estado': r['estado'], 'obs': r['observacion'] or ''}
        for i, a in enumerate(alumnos, start=2):
            ws.append([a['num_curso'], a['nombre']] + [asis_map.get(a['id'], {}).get(f, {}).get('estado', '') for f in fechas])
            for c in range(1, len(header) + 1):
                ws.cell(row=i, column=c).border = border
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for ci in range(3, len(header) + 1):
            ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = 7
    finally:
        conn.close()
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=asistencia_{slug}_{curso}.xlsx'})


# ── Importación de asistencia desde Excel ──

def _perfil_profesor(f, slug):
    """Valida profesor activo con jornada/materia en sesión."""
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return None, None, None
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return None, None, None
    return prof, jornada, materia


@attendance_bp.route('/<slug>/importar_asistencia', methods=['GET'])
def importar_asistencia(slug):
    f = _fa()
    prof, jornada, materia = _perfil_profesor(f, slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    colegio = f.get_colegio(slug)
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    return render_template('importar_asistencia.html', slug=slug, colegio=colegio, profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel,
                           materia=materia, jornada=jornada,
                           estados_asistencia=ESTADOS_ASISTENCIA)


@attendance_bp.route('/<slug>/importar_asistencia/preview', methods=['POST'])
def importar_asistencia_preview(slug):
    from app.services.excel_service import extension_excel_valida, leer_workbook, parsear_estado_asistencia, parsear_fecha
    from app.repositories.attendance_repository import get_students_by_curso
    f = _fa()
    prof, jornada, materia = _perfil_profesor(f, slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    curso_sel = request.form.get('curso', '')
    if not curso_sel:
        return jsonify({'status': 'error', 'mensaje': 'Selecciona un curso.'}), 400
    if 'archivo' not in request.files:
        return jsonify({'status': 'error', 'mensaje': 'No se envió ningún archivo.'}), 400
    archivo = request.files['archivo']
    if not archivo.filename or not extension_excel_valida(archivo.filename):
        return jsonify({'status': 'error', 'mensaje': 'El archivo debe ser .xlsx.'}), 400
    try:
        headers, filas = leer_workbook(archivo.read())
    except ValueError as e:
        return jsonify({'status': 'error', 'mensaje': str(e)}), 400
    if len(headers) < 3:
        return jsonify({'status': 'error',
                        'mensaje': 'Formato inválido. Se esperan columnas #, Estudiante y al menos una fecha.'}), 400
    fechas_headers = headers[2:]
    fechas = []
    errores_fecha = []
    for h in fechas_headers:
        fecha, err = parsear_fecha(h)
        if err:
            errores_fecha.append(f'Columna "{h}": {err}.')
        else:
            fechas.append(fecha)
    if not fechas:
        return jsonify({'status': 'error',
                        'mensaje': 'No se encontraron columnas de fecha válidas (AAAA-MM-DD).',
                        'errores': errores_fecha[:5]}), 400
    conn = f.conectar(slug)
    try:
        alumnos = get_students_by_curso(conn, curso_sel, jornada)
        if not alumnos:
            return jsonify({'status': 'error', 'mensaje': 'El curso no tiene estudiantes.'}), 400
        alumnos_by_num = {}
        for a in alumnos:
            if a['num_curso'] is not None:
                alumnos_by_num[str(a['num_curso']).strip()] = a
        alumnos_by_nombre = {a['nombre'].strip().lower(): a for a in alumnos}
        preview_rows = []
        all_ok = True
        for nro, vals in filas:
            nombre = str(vals[1]).strip() if len(vals) > 1 and vals[1] is not None else ''
            num = str(vals[0]).strip() if vals and vals[0] is not None else ''
            errores = []
            alumno = None
            if nombre:
                alumno = alumnos_by_nombre.get(nombre.lower())
            if not alumno and num:
                alumno = alumnos_by_num.get(num)
            if not alumno:
                errores.append('estudiante no encontrado en el curso')
                all_ok = False
            cambios = {}
            for fecha in fechas:
                col_idx = 2 + fechas.index(fecha)
                v = vals[col_idx] if col_idx < len(vals) else ''
                estado, err = parsear_estado_asistencia(v)
                if err:
                    errores.append(f'{fecha}: {err}')
                    all_ok = False
                    cambios[fecha] = {'estado': None}
                elif estado is not None:
                    cambios[fecha] = {'estado': estado}
            preview_rows.append({'fila': nro, 'aid': alumno['id'] if alumno else None,
                                 'nombre': alumno['nombre'] if alumno else (nombre or f'Fila {nro}'),
                                 'ok': len(errores) == 0, 'errores': errores, 'cambios': cambios})
        validos = sum(1 for r in preview_rows if r['ok'])
        return jsonify({'status': 'ok' if all_ok else 'warning', 'curso': curso_sel,
                        'fechas': fechas, 'filas': preview_rows, 'total': len(preview_rows),
                        'validos': validos, 'errores': len(preview_rows) - validos,
                        'all_ok': all_ok})
    finally:
        conn.close()


@attendance_bp.route('/<slug>/importar_asistencia/confirmar', methods=['POST'])
def importar_asistencia_confirmar(slug):
    import json as _json
    from app.services.excel_service import parsear_estado_asistencia, parsear_fecha
    from app.repositories.attendance_repository import upsert_asistencia
    f = _fa()
    prof, jornada, materia = _perfil_profesor(f, slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    curso_sel = request.form.get('curso', '')
    data_json = request.form.get('data', '')
    if not data_json:
        return jsonify({'status': 'error', 'mensaje': 'No hay datos para guardar.'}), 400
    try:
        data = _json.loads(data_json)
    except (_json.JSONDecodeError, TypeError):
        return jsonify({'status': 'error', 'mensaje': 'Datos inválidos.'}), 400
    if not data.get('all_ok'):
        return jsonify({'status': 'error', 'mensaje': 'Hay errores que deben corregirse primero.'}), 400
    fechas = data.get('fechas', [])
    filas = data.get('filas', [])
    if not fechas or not filas:
        return jsonify({'status': 'error', 'mensaje': 'No hay datos para importar.'}), 400
    conn = f.conectar(slug)
    try:
        for fecha in fechas:
            if parsear_fecha(fecha)[1]:
                conn.close()
                return jsonify({'status': 'error', 'mensaje': 'Fecha no válida en los datos.'}), 400
        updated = 0
        for fila in filas:
            if not fila.get('ok') or not fila.get('aid'):
                continue
            aid = fila['aid']
            for fecha, ch in (fila.get('cambios') or {}).items():
                estado, err = parsear_estado_asistencia(ch.get('estado'))
                if err or estado is None:
                    continue
                upsert_asistencia(conn, aid, fecha, estado, usuario_tipo='profesor', usuario_id=prof['id'])
                conn.commit()
                f.audit_log(slug, prof['id'], 'asistencia_importada', 'asistencia', aid,
                            valor_nuevo={'fecha': fecha, 'estado': estado})
                updated += 1
        conn.commit()
    except Exception as e:
        conn.close()
        logger.exception('Error confirmando importación de asistencia: %s', e)
        return jsonify({'status': 'error', 'mensaje': 'Error al guardar. Intenta de nuevo.'}), 500
    conn.close()
    return jsonify({'status': 'ok',
                    'mensaje': f'Importación completada. {updated} registros de asistencia actualizados.',
                    'updated': updated})
