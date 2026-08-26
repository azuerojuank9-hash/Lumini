"""Teacher routes — activities, grades, history, Excel import/export."""

import json
import logging
import sqlite3
from datetime import datetime
from io import BytesIO

from flask import Blueprint, Response, jsonify, redirect, render_template, request, session, url_for

from app.infra.attendance import _asistencia_stats
from app.repositories.notification_repository import get_notificaciones_no_leidas_count
from app.services.excel_service import (
    extension_excel_valida,
    leer_workbook,
    revalidar_importacion_notas,
)
from app.utils.security import extension_permitida, validar_csrf

logger = logging.getLogger(__name__)

teacher_bp = Blueprint('teacher', __name__)


def _fa():
    import flask_app
    return flask_app


def _colegio(slug):
    f = _fa()
    f.require_colegio(slug)
    return f.get_colegio(slug)


def _prof(slug):
    return _fa().get_profesor(slug)


def _session(slug):
    return _fa().get_sesion_jornada_materia(slug)


# ── Helper: build Excel workbook ──────────────────────────────────────
def _excel_armar_wb(slug, prof, materia, jornada, curso_sel, periodo, actividades, alumnos):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from app.repositories.grade_repository import get_all_evaluaciones_for_curso, get_all_notas_for_curso
    from app.services.grade_service import promedio_ponderado

    wb = Workbook()
    ws = wb.active
    ws.title = 'Notas'
    header_fill = PatternFill('solid', fgColor='6D28D9')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    headers = ['N\u00b0', 'Estudiante', 'AID'] + [a['nombre'] for a in actividades] + ['Evaluaci\u00f3n', 'Autoevaluaci\u00f3n', 'Promedio']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    aid_list = [a['id'] for a in alumnos]
    notas_by_aid = {}
    evals_by_aid = {}
    if aid_list:
        notas_all = get_all_notas_for_curso(slug, aid_list, materia, jornada, curso_sel, periodo, prof['id'])
        for r in notas_all:
            notas_by_aid.setdefault(r['aid'], {})[r['actividad_id']] = r['val']
        evals_all = get_all_evaluaciones_for_curso(slug, aid_list, prof['id'], materia, jornada, periodo)
        evals_by_aid = {r['aid']: {'ev': r['evaluacion'], 'auto': r['autoevaluacion']} for r in evals_all}

    for i, a in enumerate(alumnos, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=a['nombre'])
        ws.cell(row=row, column=3, value=a['id']).alignment = center
        for j, act in enumerate(actividades, 4):
            val = None
            if a['id'] in notas_by_aid:
                val = notas_by_aid[a['id']].get(act['id'])
            if val is not None:
                ws.cell(row=row, column=j, value=float(val)).alignment = center
        ev = evals_by_aid.get(a['id'], {})
        ev_val = ev.get('ev')
        auto_val = ev.get('auto')
        ecol = 4 + len(actividades)
        if ev_val is not None:
            ws.cell(row=row, column=ecol, value=float(ev_val)).alignment = center
        if auto_val is not None:
            ws.cell(row=row, column=ecol + 1, value=float(auto_val)).alignment = center
        notas_dict = notas_by_aid.get(a['id'], {})
        prom = promedio_ponderado([notas_dict.get(act['id']) for act in actividades], ev_val, auto_val)
        ws.cell(row=row, column=ecol + 2, value=round(prom, 2) if prom is not None else '').alignment = center

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    for j in range(len(actividades)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 14
    return wb


# ── Route: seleccionar jornada/materia ───────────────────────────────
@teacher_bp.route('/<slug>/seleccionar', methods=['GET', 'POST'])
def seleccionar_jornada(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    colegio = f.get_colegio(slug)
    materias_jornadas = f.get_materias_profesor(slug, prof['id'])

    if not materias_jornadas:
        return render_template('seleccionar_jornada.html',
                               slug=slug, colegio=colegio, profesor=prof,
                               materias_jornadas=[],
                               error='No tienes materias asignadas. Contacta al administrador.')

    if request.method == 'POST':
        if not validar_csrf():
            return ('Error CSRF', 403)
        materia = request.form.get('materia', '').strip()
        jornada = request.form.get('jornada', '').strip()
        if materia and jornada:
            session[f'materia_{slug}'] = materia
            session[f'jornada_{slug}'] = jornada
            return redirect(url_for('teacher.home', slug=slug))

    if len(materias_jornadas) == 1:
        session[f'materia_{slug}'] = materias_jornadas[0]['materia']
        session[f'jornada_{slug}'] = materias_jornadas[0]['jornada']
        return redirect(url_for('teacher.home', slug=slug))

    return render_template('seleccionar_jornada.html',
                           slug=slug, colegio=colegio, profesor=prof,
                           materias_jornadas=materias_jornadas)


# ── Route: home (main gradebook) ─────────────────────────────────────
@teacher_bp.route('/<slug>/')
@teacher_bp.route('/<slug>')
def home(slug):
    from app.repositories.grade_repository import get_actividades_by_curso, get_alumnos_by_curso
    from app.services.grade_service import get_notas_mapped, promedio_ponderado, promedio_simple

    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('teacher.seleccionar_jornada', slug=slug))
    colegio = f.get_colegio(slug)
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else None)
    periodo_sel = request.args.get('periodo', 1, type=int)

    from app.models.schema import conectar
    conn = conectar(slug)
    try:
        alumnos = actividades = agenda = []
        if curso_sel and curso_sel in mis_cursos:
            alumnos = get_alumnos_by_curso(slug, curso_sel, jornada)
            if alumnos:
                case_parts = ['WHEN ? THEN ?' for _ in alumnos]
                id_list = [a['id'] for a in alumnos]
                params = []
                for i, a in enumerate(alumnos, 1):
                    params.extend([a['id'], i])
                params.extend(id_list)
                conn.execute(
                    f'UPDATE alumnos SET num_curso = CASE id {" ".join(case_parts)} END WHERE id IN ({",".join("?" * len(alumnos))})',
                    params)
                conn.commit()
                alumnos = get_alumnos_by_curso(slug, curso_sel, jornada)
            actividades = get_actividades_by_curso(slug, prof['id'], materia, jornada, curso_sel, periodo_sel)
            agenda = conn.execute(
                'SELECT * FROM compromisos WHERE materia=? AND curso=? AND jornada=? ORDER BY fecha',
                (materia, curso_sel, jornada)).fetchall()

        MESES = {'01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr', '05': 'May', '06': 'Jun',
                 '07': 'Jul', '08': 'Ago', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'}
        datos = []
        aid_list = [a['id'] for a in alumnos]
        notas_by_aid = {}
        evals_by_aid = {}
        if aid_list:
            from app.services.grade_service import get_notas_mapped
            notas_by_aid, evals_by_aid = get_notas_mapped(slug, aid_list, materia, jornada, curso_sel, periodo_sel, prof['id'])

        from app.models.schema import conectar as conectar2
        conn2 = conectar2(slug)
        asis_all = {}
        asis_ultimo = {}
        obs_all = {}
        if aid_list:
            placeholders = ','.join('?' * len(aid_list))
            rows_asistencia = conn2.execute(
                f'SELECT aid, fecha, estado FROM asistencia WHERE aid IN ({placeholders}) ORDER BY aid, fecha',
                aid_list).fetchall()
            for r in rows_asistencia:
                asis_all.setdefault(r['aid'], []).append(r)
            rows_ultimo = conn2.execute(
                f'SELECT aid, estado FROM asistencia WHERE aid IN ({placeholders}) AND fecha=date("now")',
                aid_list).fetchall()
            asis_ultimo = {r['aid']: r['estado'] for r in rows_ultimo}
            rows_obs = conn2.execute(
                f'SELECT id, aid, materia, texto, fecha FROM observaciones WHERE aid IN ({placeholders}) AND materia=? ORDER BY aid, fecha DESC',
                (*aid_list, materia)).fetchall()
            for r in rows_obs:
                obs_all.setdefault(r['aid'], []).append(r)
        conn2.close()

        for a in alumnos:
            notas_raw = notas_by_aid.get(a['id'], [])
            notas_map = {nr['actividad_id']: {'val': nr['val'], 'id': nr['id']} for nr in notas_raw}
            ev = evals_by_aid.get(a['id'])
            eval_v = ev['evaluacion'] if ev and ev['evaluacion'] is not None else None
            auto_v = ev['autoevaluacion'] if ev and ev['autoevaluacion'] is not None else None
            vals = [nr['val'] for nr in notas_raw]
            prom = promedio_simple(vals)
            nota_final = promedio_ponderado(vals, eval_v, auto_v)
            historial_raw = asis_all.get(a['id'], [])
            hist_meses = {}
            for h in historial_raw:
                if h['fecha']:
                    p2 = h['fecha'].split('-')
                    if len(p2) >= 2:
                        label = f"{MESES.get(p2[1], p2[1])} {p2[0]}"
                        hist_meses.setdefault(label, []).append({'fecha': h['fecha'], 'estado': h['estado']})
            ult_estado = asis_ultimo.get(a['id'])
            obs = obs_all.get(a['id'], [])
            datos.append({
                'id': a['id'], 'num_curso': a['num_curso'],
                'nombre': a['nombre'], 'curso': a['curso'],
                'promedio': prom, 'nota_final': nota_final, 'notas_map': notas_map,
                'evaluacion': eval_v if eval_v is not None else '',
                'autoevaluacion': auto_v if auto_v is not None else '',
                'asistencia': ult_estado or '-',
                'historial_meses': hist_meses,
                'observaciones': [dict(o) for o in obs],
            })

        promedios = [d['promedio'] for d in datos if d['promedio'] is not None]
        prom_gral = round(sum(promedios) / len(promedios), 2) if promedios else None
        mejor = max(datos, key=lambda x: x['promedio'] or 0, default={'nombre': 'N/A', 'promedio': None})

        DIAS = ['Lunes', 'Martes', 'Mi\u00e9rcoles', 'Jueves', 'Viernes', 'S\u00e1bado', 'Domingo']
        hoy_idx = datetime.today().weekday()
        hoy_nombre = DIAS[hoy_idx] if hoy_idx < 7 else ''
        hoy_fecha = datetime.today().strftime('%Y-%m-%d')
        total_alumnos = conn.execute(
            f'SELECT COUNT(*) as c FROM alumnos WHERE curso IN ({",".join("?" * len(mis_cursos))}) AND jornada=? AND activo=1',
            (*mis_cursos, jornada)
        ).fetchone()['c'] if mis_cursos else 0
        horario_hoy = conn.execute(
            'SELECT * FROM horarios_curso WHERE materia=? AND jornada=? AND dia=? ORDER BY franja',
            (materia, jornada, hoy_nombre)
        ).fetchall() if curso_sel else []
        asis_hoy = conn.execute(
            "SELECT COUNT(DISTINCT aid) as total FROM asistencia WHERE fecha=?",
            (hoy_fecha,)
        ).fetchone()
        asistencia_hoy = asis_hoy['total'] if asis_hoy else 0
        notas_pend = 0
        if curso_sel and actividades:
            act_ids = [a['id'] for a in actividades]
            placeholders = ','.join('?' * len(act_ids))
            rows_present = conn.execute(
                f'SELECT actividad_id, COUNT(DISTINCT aid) as cnt FROM notas n JOIN alumnos al ON al.id=n.aid AND al.activo=1 WHERE n.actividad_id IN ({placeholders}) GROUP BY n.actividad_id',
                act_ids).fetchall()
            present_sum = sum(r['cnt'] for r in rows_present)
            total_alumnos_curso = conn.execute(
                'SELECT COUNT(*) as c FROM alumnos WHERE curso=? AND jornada=? AND activo=1',
                (curso_sel, jornada)
            ).fetchone()['c']
            notas_pend = total_alumnos_curso * len(act_ids) - present_sum
        alertas = []
        if curso_sel:
            rows = conn.execute(
                "SELECT a.nombre, a.id, COUNT(*) as faltas FROM asistencia asis JOIN alumnos a ON a.id=asis.aid WHERE asis.estado='A' AND a.curso=? AND a.jornada=? AND a.activo=1 GROUP BY asis.aid HAVING faltas > 1 ORDER BY faltas DESC LIMIT 5",
                (curso_sel, jornada)).fetchall()
            for r in rows:
                alertas.append({'nombre': r['nombre'], 'faltas': r['faltas']})
            for e in datos:
                if e['promedio'] is not None and e['promedio'] < 3.0:
                    alertas.append({'nombre': e['nombre'], 'promedio': e['promedio']})
            alertas = alertas[:5]
        pendientes = f.comunicaciones_pendientes(slug, 'profesor', prof['id'])
        num_periodos = int(colegio['num_periodos']) if colegio and colegio['num_periodos'] else 4
        pc = f.periodo_cerrado(slug, periodo_sel) if curso_sel and materia else False
        error_msg = request.args.get('error', '')
        if error_msg == 'periodo_cerrado':
            error_msg = 'El per\u00edodo est\u00e1 cerrado. No se pueden crear actividades.'
        solicitudes_pend = conn.execute(
            'SELECT COUNT(*) as c FROM solicitudes_modificacion WHERE profesor_id=? AND estado=? AND slug=?',
            (prof['id'], 'pendiente', slug)).fetchone()['c'] if curso_sel else 0
        abrir_nueva_actividad = request.args.get('nueva_actividad') == '1'
        abrir_analitica = request.args.get('analitica') == '1'
    finally:
        conn.close()

    return render_template('index.html',
                           profesor=prof, mis_cursos=mis_cursos, curso_sel=curso_sel,
                           estudiantes=datos, actividades=actividades, compromisos=agenda,
                           prom_general=prom_gral, mejor=mejor, slug=slug, colegio=colegio,
                           num_periodos=num_periodos, periodo_sel=periodo_sel,
                           materia=materia, jornada=jornada,
                           materias_jornadas=f.get_materias_profesor(slug, prof['id']),
                           hoy_nombre=hoy_nombre, hoy_fecha=hoy_fecha,
                           total_alumnos=total_alumnos, horario_hoy=horario_hoy,
                           asistencia_hoy=asistencia_hoy, notas_pend=notas_pend,
                           alertas=alertas,
                           periodo_cerrado=pc,
                           error_msg=error_msg,
                           solicitudes_pendientes_mod=solicitudes_pend,
                           comunicaciones_pendientes=pendientes,
                           abrir_nueva_actividad=abrir_nueva_actividad,
                           abrir_analitica=abrir_analitica)


# ── ACTIVIDADES ──────────────────────────────────────────────────────

@teacher_bp.route('/<slug>/nueva_actividad', methods=['POST'])
def nueva_actividad(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf():
        return ('Error CSRF', 403)
    jornada, materia = f.get_sesion_jornada_materia(slug)
    nombre = request.form.get('nombre', '').strip()
    curso_sel = request.form.get('curso_sel', '')
    periodo = request.form.get('periodo_sel', 1, type=int)
    if nombre and curso_sel and materia and jornada:
        if f.periodo_cerrado(slug, periodo):
            return redirect(url_for('teacher.home', slug=slug, curso=curso_sel, periodo=periodo, error='periodo_cerrado'))
        from app.repositories.grade_repository import create_actividad, get_max_orden_actividad
        max_ord = get_max_orden_actividad(slug, prof['id'], materia, jornada, curso_sel, periodo)
        create_actividad(slug, prof['id'], materia, jornada, curso_sel, nombre, max_ord + 1, periodo)
    return redirect(url_for('teacher.home', slug=slug, curso=curso_sel, periodo=periodo))


@teacher_bp.route('/<slug>/borrar_actividad/<int:act_id>', methods=['POST'])
def borrar_actividad(slug, act_id):
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    from app.models.schema import conectar
    conn = conectar(slug)
    act = conn.execute('SELECT profesor_id, curso, materia, COALESCE(periodo,1) as p FROM actividades WHERE id=?', (act_id,)).fetchone()
    if not act or act['profesor_id'] != prof['id']:
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Actividad no encontrada'}), 404
    if f.periodo_cerrado(slug, act['p']):
        conn.close()
        return jsonify({'status': 'error', 'codigo': 'PERIODO_CERRADO', 'mensaje': 'El per\u00edodo est\u00e1 cerrado.'}), 403
    notas_borradas = conn.execute('SELECT aid, val FROM notas WHERE actividad_id=?', (act_id,)).fetchall()
    conn.execute('DELETE FROM entregas WHERE actividad_id=?', (act_id,))
    conn.execute('DELETE FROM solicitudes_modificacion WHERE actividad_id=?', (act_id,))
    conn.execute('DELETE FROM notas WHERE actividad_id=?', (act_id,))
    conn.execute('DELETE FROM actividades WHERE id=?', (act_id,))
    conn.commit()
    curso = act['curso']
    jornada_ctx, materia_ctx = f.get_sesion_jornada_materia(slug)
    for n in notas_borradas:
        f.auditar_nota(slug, prof['id'], 'profesor', 'eliminacion', 'notas', n['aid'],
                       act['curso'], materia_ctx or act['materia'], act['p'],
                       campo='nota', actividad_id=act_id,
                       valor_anterior=n['val'], valor_nuevo=None)
    f.auditar_nota(slug, prof['id'], 'profesor', 'eliminacion', 'actividades', None,
                   act['curso'], materia_ctx or act['materia'], act['p'],
                   actividad_id=act_id, valor_anterior=act_id, valor_nuevo=None,
                   motivo='Actividad eliminada')
    conn.close()
    return jsonify({'status': 'ok', 'actividad_id': act_id, 'redirect': url_for('teacher.home', slug=slug, curso=curso)})


@teacher_bp.route('/<slug>/actividades/crear', methods=['POST'])
def actividades_crear(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status': 'error', 'mensaje': 'Sesion no valida'}), 400
    data = request.get_json(silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    curso_sel = (data.get('curso') or '').strip()
    periodo = data.get('periodo', 1)
    tipo = data.get('tipo', 'taller')
    peso = data.get('peso')
    fecha_limite = data.get('fecha_limite') or ''
    hora_limite = data.get('hora_limite') or ''
    descripcion = (data.get('descripcion') or '').strip()
    observaciones = (data.get('observaciones') or '').strip()
    estado_act = data.get('estado', 'borrador')
    competencia = (data.get('competencia') or '').strip()
    entrega_digital = 1 if data.get('entrega_digital') else 0
    if not nombre or not curso_sel:
        return jsonify({'status': 'error', 'mensaje': 'Nombre y curso requeridos'}), 400
    if f.periodo_cerrado(slug, periodo):
        return jsonify({'status': 'error', 'codigo': 'PERIODO_CERRADO', 'mensaje': 'Periodo cerrado'}), 403
    from app.models.schema import conectar
    from app.repositories.grade_repository import create_actividad, get_max_orden_actividad
    conn = conectar(slug)
    try:
        max_ord = get_max_orden_actividad(slug, prof['id'], materia, jornada, curso_sel, periodo)
        act_id = create_actividad(slug, prof['id'], materia, jornada, curso_sel, nombre, max_ord + 1, periodo,
                                  tipo, peso, fecha_limite or None, hora_limite or None,
                                  descripcion, observaciones, estado_act, competencia, entrega_digital)
        f.audit_log(slug, prof['id'], 'actividad_creada', 'actividades',
                    registro_id=act_id,
                    valor_nuevo={'nombre': nombre, 'tipo': tipo, 'curso': curso_sel, 'materia': materia})
        f.auditar_nota(slug, prof['id'], 'profesor', 'creacion', 'actividades', None,
                       curso_sel, materia, periodo, actividad_id=act_id,
                       valor_nuevo=nombre, motivo='Actividad creada via API')
        from app.services.activity_service import create_calendar_event
        create_calendar_event(slug, conn, nombre, tipo, descripcion, fecha_limite, hora_limite,
                              curso_sel, prof['id'])
        return jsonify({
            'status': 'ok',
            'actividad': {
                'id': act_id, 'nombre': nombre, 'tipo': tipo, 'orden': max_ord + 1,
                'peso': peso, 'fecha_limite': fecha_limite, 'hora_limite': hora_limite,
                'descripcion': descripcion, 'observaciones': observaciones,
                'estado_act': estado_act, 'competencia': competencia,
                'entrega_digital': entrega_digital, 'periodo': periodo
            }
        })
    except Exception as e:
        logger.error(f'Error creando actividad: {e}')
        return jsonify({'status': 'error', 'mensaje': 'Error al guardar'}), 500
    finally:
        conn.close()


@teacher_bp.route('/<slug>/actividades/<int:act_id>', methods=['PUT'])
def actividades_editar(slug, act_id):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    from app.models.schema import conectar
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Actividad no encontrada'}), 404
    changes = []
    for field in ('nombre', 'tipo', 'peso', 'fecha_limite', 'hora_limite', 'descripcion', 'observaciones', 'estado_act', 'competencia', 'entrega_digital'):
        if field in data:
            old_val = act[field]
            new_val = data[field]
            if str(old_val) != str(new_val):
                changes.append(field + '=?')
                conn.execute(f'UPDATE actividades SET {field}=? WHERE id=?', (new_val, act_id))
                f.audit_log(slug, prof['id'], 'actividad_' + field, 'actividades', registro_id=act_id,
                            valor_anterior={field: old_val}, valor_nuevo={field: new_val})
    if not changes:
        conn.close()
        act_dict = dict(act)
        act_dict.pop('profesor_id', None)
        return jsonify({'status': 'ok', 'actividad': act_dict})
    conn.commit()
    updated = conn.execute('SELECT * FROM actividades WHERE id=?', (act_id,)).fetchone()
    conn.close()
    act_dict = dict(updated)
    act_dict.pop('profesor_id', None)
    return jsonify({'status': 'ok', 'actividad': act_dict})


@teacher_bp.route('/<slug>/actividades/<int:act_id>/estado', methods=['POST'])
def actividades_cambiar_estado(slug, act_id):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    if request.is_json:
        estado = (request.get_json(silent=True) or {}).get('estado', '').strip()
    else:
        estado = (request.form.get('estado') or '').strip()
    if estado not in ('borrador', 'publicada', 'cerrada', 'archivada'):
        return jsonify({'status': 'error', 'mensaje': 'Estado invalido'}), 400
    from app.models.schema import conectar
    conn = conectar(slug)
    act = conn.execute('SELECT id, estado_act FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Actividad no encontrada'}), 404
    if estado == 'archivada' and act['estado_act'] not in ('cerrada', 'publicada'):
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Solo actividades cerradas o publicadas pueden archivarse'}), 400
    conn.execute('UPDATE actividades SET estado_act=? WHERE id=?', (estado, act_id))
    conn.commit()
    f.audit_log(slug, prof['id'], 'actividad_estado', 'actividades', registro_id=act_id,
                valor_anterior={'estado': act['estado_act']}, valor_nuevo={'estado': estado})
    conn.close()
    return jsonify({'status': 'ok', 'estado': estado})


@teacher_bp.route('/<slug>/actividades/<int:act_id>/detalle', methods=['GET'])
def actividades_detalle(slug, act_id):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    from app.models.schema import conectar
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Actividad no encontrada'}), 404
    act_dict = dict(act)
    act_dict.pop('profesor_id', None)
    conn.close()
    return jsonify({'status': 'ok', 'actividad': act_dict})


@teacher_bp.route('/<slug>/actividades/<int:act_id>/duplicar', methods=['POST'])
def actividades_duplicar(slug, act_id):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    from app.models.schema import conectar
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Actividad no encontrada'}), 404
    from app.services.activity_service import duplicar_actividad_logic
    new_id, nuevo_nombre = duplicar_actividad_logic(slug, act, prof['id'], conn)
    f.audit_log(slug, prof['id'], 'actividad_duplicada', 'actividades', registro_id=new_id,
                valor_nuevo={'desde': act_id, 'nombre': nuevo_nombre})
    nueva = dict(conn.execute('SELECT * FROM actividades WHERE id=?', (new_id,)).fetchone())
    conn.close()
    nueva.pop('profesor_id', None)
    return jsonify({'status': 'ok', 'actividad': nueva})


@teacher_bp.route('/<slug>/actividades/<int:act_id>/historial', methods=['GET'])
def actividades_historial(slug, act_id):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    from app.models.schema import conectar
    conn = conectar(slug)
    act = conn.execute('SELECT id FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Actividad no encontrada'}), 404
    rows = conn.execute(
        '''SELECT usuario_id, accion, valor_anterior, valor_nuevo, creado, ip
           FROM audit_log WHERE tabla='actividades' AND registro_id=?
           ORDER BY id DESC LIMIT 100''', (act_id,)).fetchall()
    conn.close()
    return jsonify({'status': 'ok', 'historial': [dict(r) for r in rows]})


@teacher_bp.route('/<slug>/actividades/<int:act_id>/estadisticas', methods=['GET'])
def actividades_estadisticas(slug, act_id):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    from app.models.schema import conectar
    conn = conectar(slug)
    act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (act_id, prof['id'])).fetchone()
    if not act:
        conn.close()
        return jsonify({'status': 'error', 'mensaje': 'Actividad no encontrada'}), 404
    notas = conn.execute(
        'SELECT n.val, a.nombre, a.id as aid FROM notas n JOIN alumnos a ON a.id=n.aid WHERE n.actividad_id=?',
        (act_id,)).fetchall()
    vals = [float(r['val']) for r in notas if r['val'] is not None]
    total = len(vals)
    prom = round(sum(vals) / total, 2) if total else None
    mx = max(vals) if total else None
    mn = min(vals) if total else None
    aprobados = sum(1 for v in vals if v >= 3.0)
    reprobados = sum(1 for v in vals if v < 3.0 and v > 0)
    distribucion = {'1': 0, '2': 0, '3': 0, '4': 0, '5': 0}
    for v in vals:
        k = str(int(v))
        if k in distribucion:
            distribucion[k] += 1
    conn.close()
    return jsonify({'status': 'ok', 'estadisticas': {
        'total_notas': total, 'promedio': prom, 'max': mx, 'min': mn,
        'aprobados': aprobados, 'reprobados': reprobados, 'distribucion': distribucion
    }})


@teacher_bp.route('/<slug>/reordenar_actividades', methods=['POST'])
def reordenar_actividades(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    ordenes = data.get('orden', [])
    if not ordenes:
        return jsonify({'status': 'error', 'mensaje': 'Sin datos'}), 400
    from app.models.schema import conectar
    conn = conectar(slug)
    try:
        for item in ordenes:
            act_id = item.get('id')
            orden = item.get('orden', 0)
            conn.execute('UPDATE actividades SET orden=? WHERE id=? AND profesor_id=?',
                         (orden, act_id, prof['id']))
        conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f'Error reordenando: {e}')
        return jsonify({'status': 'error', 'mensaje': 'Error al reordenar'}), 500
    finally:
        conn.close()


# ── NOTAS ────────────────────────────────────────────────────────────

@teacher_bp.route('/<slug>/notas/batch', methods=['POST'])
def notas_batch(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    notas = data.get('notas', [])
    if not notas:
        return jsonify({'status': 'error', 'mensaje': 'Sin datos'}), 400
    from app.models.schema import conectar
    conn = conectar(slug)
    errors = []
    saved = []
    try:
        cerrados = {r[0] for r in conn.execute(
            "SELECT periodo FROM periodos_estado WHERE estado='cerrado'").fetchall()}
        cfg = f.config_get(slug)
        try:
            escala_min = float(cfg.get('escala_min', 0.0))
            escala_max = float(cfg.get('escala_max', 5.0))
        except (TypeError, ValueError):
            escala_min, escala_max = 0.0, 5.0

        valid_items = []
        for item in notas:
            aid = item.get('aid')
            actividad_id = item.get('actividad_id')
            val = item.get('val')
            if aid is None or actividad_id is None:
                errors.append({'aid': aid, 'actividad_id': actividad_id, 'error': 'Datos invalidos'})
                continue
            if val is not None and (not isinstance(val, (int, float)) or isinstance(val, bool) or val != val):
                errors.append({'aid': aid, 'actividad_id': actividad_id, 'val': val, 'error': 'Valor invalido'})
                continue
            if val is not None and (val < escala_min or val > escala_max):
                errors.append({'aid': aid, 'actividad_id': actividad_id, 'val': val, 'error': 'Valor fuera de rango'})
                continue
            valid_items.append(item)

        act_ids = list({it['actividad_id'] for it in valid_items})
        if act_ids:
            ph = ','.join('?' for _ in act_ids)
            act_rows = conn.execute(
                f'SELECT id, profesor_id, curso, materia, jornada, COALESCE(periodo,1) as p FROM actividades WHERE id IN ({ph})',
                act_ids).fetchall()
        else:
            act_rows = []
        acts_map = {r[0]: r for r in act_rows}

        aid_act_pairs = [(it['aid'], it['actividad_id']) for it in valid_items]
        old_map = {}
        if aid_act_pairs:
            batch_size = 900
            for i in range(0, len(aid_act_pairs), batch_size):
                chunk = aid_act_pairs[i:i+batch_size]
                conditions = ' OR '.join('(' + ' AND '.join(f'{col}=?' for col in ('aid','actividad_id')) + ')' for _ in chunk)
                flat = []
                for a, ac in chunk:
                    flat.extend([a, ac])
                old_rows = conn.execute(
                    f'SELECT aid, actividad_id, val FROM notas WHERE {conditions}',
                    flat).fetchall()
                for r in old_rows:
                    old_map[(r[0], r[1])] = r[2]

        audit_rows = []
        for item in valid_items:
            aid = item['aid']
            actividad_id = item['actividad_id']
            val = item.get('val')
            act = acts_map.get(actividad_id)
            if not act or act[1] != prof['id']:
                errors.append({'aid': aid, 'actividad_id': actividad_id, 'error': 'No autorizado'})
                continue
            if act[5] in cerrados:
                errors.append({'aid': aid, 'actividad_id': actividad_id, 'error': 'Periodo cerrado'})
                continue
            old_val = old_map.get((aid, actividad_id))
            if val is None:
                if old_val is not None:
                    conn.execute('DELETE FROM notas WHERE aid=? AND actividad_id=?', (aid, actividad_id))
                    old_map[(aid, actividad_id)] = None
            elif old_val is not None:
                conn.execute('UPDATE notas SET val=? WHERE aid=? AND actividad_id=?', (val, aid, actividad_id))
                old_map[(aid, actividad_id)] = val
            else:
                conn.execute('INSERT INTO notas (aid, actividad_id, val) VALUES (?,?,?)', (aid, actividad_id, val))
                old_map[(aid, actividad_id)] = val
            from flask import request as flask_request
            audit_rows.append((
                prof['id'], 'profesor', flask_request.remote_addr,
                act[2], act[3], act[5],
                'modificacion', 'notas', None, aid, actividad_id, 'nota',
                json.dumps(old_val) if old_val is not None else None,
                json.dumps(val) if val is not None else None, None
            ))
            saved.append({'aid': aid, 'actividad_id': actividad_id, 'val_anterior': old_val, 'val_nuevo': val,
                          'curso': act[2], 'materia': act[3], 'jornada': act[4], 'periodo': act[5]})

        if audit_rows:
            conn.executemany(
                '''INSERT INTO auditoria_notas
                   (usuario_id, rol, ip, curso, materia, periodo, tipo_accion, tabla, registro_id, aid, actividad_id, campo, valor_anterior, valor_nuevo, motivo)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                audit_rows)

        conn.commit()

        recalcs = {}
        for s in saved:
            key = (s['aid'], s['curso'], s['materia'], s['jornada'], s['periodo'])
            if key in recalcs:
                continue
            prom, nf = f.calcular_stats_y_nota_final(conn, slug, s['aid'], s['curso'], s['materia'], s['jornada'], s['periodo'], prof['id'])
            recalcs[key] = {'promedio': prom, 'nota_final': nf}

        calculos = {}
        for (aid, *_), v in recalcs.items():
            calculos[aid] = v

        return jsonify({'status': 'ok', 'saved': len(saved), 'errors': errors,
                        'snapshot': saved[-5:] if saved else [], 'calculos': calculos})
    except Exception as e:
        logger.error(f'Error batch notas: {e}')
        return jsonify({'status': 'error', 'mensaje': 'Error al guardar'}), 500
    finally:
        conn.close()


@teacher_bp.route('/<slug>/notas/deshacer', methods=['POST'])
def notas_deshacer(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'status': 'error', 'mensaje': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    aid = data.get('aid')
    actividad_id = data.get('actividad_id')
    val_anterior = data.get('val')
    if None in (aid, actividad_id):
        return jsonify({'status': 'error', 'mensaje': 'Datos invalidos'}), 400
    from app.models.schema import conectar
    conn = conectar(slug)
    try:
        act = conn.execute(
            'SELECT profesor_id, curso, materia, COALESCE(periodo,1) as p FROM actividades WHERE id=?',
            (actividad_id,)).fetchone()
        if not act or act['profesor_id'] != prof['id']:
            conn.close()
            return jsonify({'status': 'error', 'mensaje': 'No autorizado'}), 403
        if val_anterior is not None:
            conn.execute('''INSERT INTO notas (aid, actividad_id, val) VALUES (?,?,?)
                            ON CONFLICT(aid, actividad_id) DO UPDATE SET val=excluded.val''',
                         (aid, actividad_id, val_anterior))
        else:
            conn.execute('DELETE FROM notas WHERE aid=? AND actividad_id=?', (aid, actividad_id))
        conn.commit()
        f.auditar_nota(slug, prof['id'], 'profesor', 'deshacer', 'notas', aid,
                       act['curso'], act['materia'], act['p'],
                       campo='nota', actividad_id=actividad_id,
                       valor_nuevo=val_anterior, motivo='Deshacer')
        conn.close()
        return jsonify({'status': 'ok', 'val': val_anterior})
    except Exception as e:
        logger.error('deshacer_nota: %s', e)
        conn.close()
        return jsonify({'status': 'error', 'mensaje': str(e)}), 500


@teacher_bp.route('/<slug>/guardar_nota', methods=['POST'])
def guardar_nota(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    aid          = request.form.get('aid', type=int)
    actividad_id = request.form.get('actividad_id', type=int)
    val          = request.form.get('val', type=float)
    if None in (aid, actividad_id, val): return ('', 400)
    conn = f.conectar(slug)
    act = conn.execute(
        'SELECT a.id, a.profesor_id, a.curso, a.materia, COALESCE(a.periodo,1) as p FROM actividades a WHERE a.id=?',
        (actividad_id,)).fetchone()
    if not act:
        conn.close()
        return ('', 404)
    if act['profesor_id'] != prof['id']:
        conn.close()
        return ('', 403)
    alumno = conn.execute('SELECT id FROM alumnos WHERE id=? AND curso=? AND activo=1',
                          (aid, act['curso'])).fetchone()
    if not alumno:
        conn.close()
        return ('', 403)
    if f.periodo_cerrado(slug, act['p']):
        conn.close()
        return jsonify({'status':'error','codigo':'PERIODO_CERRADO','mensaje':'El per\u00edodo est\u00e1 cerrado.'}), 403
    old = conn.execute(
        'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
        (aid, actividad_id)).fetchone()
    old_val = old['val'] if old else None
    conn.execute(
        '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
           ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
        (aid, actividad_id, val))
    f.audit_log(slug, prof['id'], 'nota_editada', 'notas', registro_id=None,
              valor_anterior={'aid': aid, 'actividad_id': actividad_id, 'val': old_val},
              valor_nuevo={'aid': aid, 'actividad_id': actividad_id, 'val': val},
              conn=conn)
    tipo_nota = 'creacion' if old_val is None else 'modificacion'
    jornada, materia = f.get_sesion_jornada_materia(slug)
    f.auditar_nota(slug, prof['id'], 'profesor', tipo_nota, 'notas', aid,
                 act['curso'], materia, act['p'],
                 campo='nota', actividad_id=actividad_id,
                 valor_anterior=old_val, valor_nuevo=val, conn=conn)
    prom_est, nf = f.calcular_stats_y_nota_final(conn, slug, aid, act['curso'], materia, jornada, act['p'], prof['id'])
    curso_stats = f.calcular_stats_curso(conn, slug, act['curso'], materia, jornada, act['p'], prof['id'])
    conn.commit()
    conn.close()
    logger.info('guardar_nota: aid=%d actividad_id=%d val=%s prom_est=%s nf=%s', aid, actividad_id, val, prom_est, nf)
    return jsonify({'status':'ok','promedio':prom_est,'nota_final':nf,'promedio_curso':curso_stats['promedio_curso'],'notas_pendientes':curso_stats['notas_pendientes']})


@teacher_bp.route('/<slug>/historial_notas/<int:aid>')
def historial_notas(slug, aid):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return ('', 403)
    conn = f.conectar(slug)
    rows = conn.execute(
        '''SELECT a.id, a.tipo_accion, a.tabla, a.campo, a.valor_anterior, a.valor_nuevo,
                  a.creado, a.materia, a.periodo, a.motivo, a.aid,
                  COALESCE(ac.nombre, '') as actividad_nombre
           FROM auditoria_notas a
           LEFT JOIN actividades ac ON ac.id = a.actividad_id
           WHERE a.aid = ?
           ORDER BY a.creado DESC
           LIMIT 200''',
        (aid,)).fetchall()
    conn.close()
    return jsonify([{
        'id': r['id'], 'tipo_accion': r['tipo_accion'], 'tabla': r['tabla'],
        'campo': r['campo'], 'valor_anterior': r['valor_anterior'],
        'valor_nuevo': r['valor_nuevo'], 'creado': r['creado'],
        'materia': r['materia'], 'periodo': r['periodo'], 'motivo': r['motivo'],
        'actividad_nombre': r['actividad_nombre'], 'aid': r['aid'],
    } for r in rows])


@teacher_bp.route('/<slug>/historial_curso')
def historial_curso(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return ('', 403)
    curso = request.args.get('curso', '')
    periodo = request.args.get('periodo', 1, type=int)
    if not curso: return jsonify([])
    conn = f.conectar(slug)
    rows = conn.execute(
        '''SELECT a.id, a.tipo_accion, a.tabla, a.campo, a.valor_anterior, a.valor_nuevo,
                  a.creado, a.materia, a.periodo, a.motivo, a.aid,
                  COALESCE(ac.nombre, '') as actividad_nombre
           FROM auditoria_notas a
           LEFT JOIN actividades ac ON ac.id = a.actividad_id
           WHERE a.curso = ? AND a.materia = ? AND a.periodo = ? AND a.usuario_id = ?
           ORDER BY a.creado DESC
           LIMIT 500''',
        (curso, request.args.get('materia', ''), periodo, prof['id'])).fetchall()
    conn.close()
    return jsonify([{
        'id': r['id'], 'tipo_accion': r['tipo_accion'], 'tabla': r['tabla'],
        'campo': r['campo'], 'valor_anterior': r['valor_anterior'],
        'valor_nuevo': r['valor_nuevo'], 'creado': r['creado'],
        'materia': r['materia'], 'periodo': r['periodo'], 'motivo': r['motivo'],
        'actividad_nombre': r['actividad_nombre'], 'aid': r['aid'],
    } for r in rows])


@teacher_bp.route('/<slug>/guardar_evaluacion', methods=['POST'])
def guardar_evaluacion(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return ('', 403)
    if not validar_csrf(): return ('Error CSRF', 403)
    jornada, materia = f.get_sesion_jornada_materia(slug)
    aid     = request.form.get('aid', type=int)
    ev_raw  = request.form.get('evaluacion')
    au_raw  = request.form.get('autoevaluacion')
    ev      = float(ev_raw) if ev_raw not in (None, '') else None
    au      = float(au_raw) if au_raw not in (None, '') else None
    ev_set  = 'evaluacion' in request.form
    au_set  = 'autoevaluacion' in request.form
    periodo = request.form.get('periodo', 1, type=int)
    curso   = request.form.get('curso', '')
    if aid is None: return ('', 400)
    conn = f.conectar(slug)
    if f.periodo_cerrado(slug, periodo):
        conn.close()
        return jsonify({'status':'error','codigo':'PERIODO_CERRADO','mensaje':'El per\u00edodo est\u00e1 cerrado.'}), 403
    if not curso:
        cursos_prof = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
        curso = cursos_prof[0] if cursos_prof else ''
    existing = conn.execute(
        '''SELECT evaluacion, autoevaluacion FROM evaluaciones
           WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?''',
        (aid, prof['id'], materia, jornada, periodo)
    ).fetchone()
    old_eval = existing['evaluacion'] if existing else None
    old_auto = existing['autoevaluacion'] if existing else None
    ev_final = ev if ev_set else old_eval
    au_final = au if au_set else old_auto
    try:
        conn.execute(
            '''INSERT INTO evaluaciones
               (aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
               VALUES (?,?,?,?,?,?,?)
               ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
               DO UPDATE SET evaluacion=excluded.evaluacion, autoevaluacion=excluded.autoevaluacion''',
            (aid, prof['id'], materia, jornada, ev_final, au_final, periodo))
        conn.commit()
    except sqlite3.OperationalError as e:
        conn.rollback()
        if 'ON CONFLICT clause does not match' in str(e):
            logger.warning(f'[{slug}] ON CONFLICT fall\u00f3 en guardar_evaluacion, reparando...')
            f._recrear_si_unique_incorrecto(conn, slug, 'evaluaciones',
                '(aid,profesor_id,materia,jornada,periodo)',
                '''CREATE TABLE evaluaciones (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    aid INTEGER NOT NULL, profesor_id INTEGER NOT NULL,
                    materia TEXT NOT NULL, jornada TEXT NOT NULL DEFAULT "Ma\u00f1ana",
                    evaluacion REAL, autoevaluacion REAL, periodo INTEGER DEFAULT 1,
                    UNIQUE(aid,profesor_id,materia,jornada,periodo))''',
                '''(id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
                   SELECT id,aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,
                          COALESCE(periodo,1) FROM evaluaciones_old''')
            conn.execute(
                '''INSERT INTO evaluaciones
                   (aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                   DO UPDATE SET evaluacion=excluded.evaluacion, autoevaluacion=excluded.autoevaluacion''',
                (aid, prof['id'], materia, jornada, ev_final, au_final, periodo))
            conn.commit()
        else:
            raise
    f.audit_log(slug, prof['id'], 'evaluacion_editada', 'evaluaciones', registro_id=None,
              valor_anterior={'aid': aid, 'evaluacion': old_eval, 'autoevaluacion': old_auto},
              valor_nuevo={'aid': aid, 'evaluacion': ev_final, 'autoevaluacion': au_final})
    if ev_set and ev_final != old_eval:
        tipo_ev = 'modificacion' if old_eval is not None else 'creacion'
        f.auditar_nota(slug, prof['id'], 'profesor', tipo_ev, 'evaluaciones', aid,
                     curso, materia, periodo, campo='evaluacion',
                     valor_anterior=old_eval, valor_nuevo=ev_final)
    if au_set and au_final != old_auto:
        tipo_au = 'modificacion' if old_auto is not None else 'creacion'
        f.auditar_nota(slug, prof['id'], 'profesor', tipo_au, 'evaluaciones', aid,
                     curso, materia, periodo, campo='autoevaluacion',
                     valor_anterior=old_auto, valor_nuevo=au_final)
    prom_est, nf = f.calcular_stats_y_nota_final(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
    curso_stats = f.calcular_stats_curso(conn, slug, curso, materia, jornada, periodo, prof['id'])
    conn.close()
    return jsonify({'status':'ok','promedio':prom_est,'nota_final':nf,'promedio_curso':curso_stats['promedio_curso'],'notas_pendientes':curso_stats['notas_pendientes']})


@teacher_bp.route('/<slug>/guardar_nota_batch', methods=['POST'])
def guardar_nota_batch(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status':'error'}), 403
    if not validar_csrf():
        return jsonify({'status':'error'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error'}), 400
    curso = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    raw = request.form.get('batch', '')
    if not raw:
        return jsonify({'status':'error'}), 400
    try:
        items = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'status':'error'}), 400
    conn = f.conectar(slug)
    resultados = {}
    try:
        cerrados = {r[0] for r in conn.execute(
            "SELECT periodo FROM periodos_estado WHERE estado='cerrado'").fetchall()}
        for item in items:
            aid = item.get('aid')
            actividad_id = item.get('actividad_id')
            val = item.get('val')
            if None in (aid, actividad_id, val):
                continue
            act = conn.execute(
                'SELECT a.id, a.profesor_id, a.curso, a.materia, a.jornada, COALESCE(a.periodo,1) as p FROM actividades a WHERE a.id=?',
                (actividad_id,)).fetchone()
            if not act or act['profesor_id'] != prof['id']: continue
            if act['p'] in cerrados:
                continue
            old = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (aid, actividad_id)).fetchone()
            old_val = old['val'] if old else None
            conn.execute(
                '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
                   ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
                (aid, actividad_id, val))
            f.audit_log(slug, prof['id'], 'nota_editada', 'notas', registro_id=None,
                      valor_anterior={'aid': aid, 'actividad_id': actividad_id, 'val': old_val},
                      valor_nuevo={'aid': aid, 'actividad_id': actividad_id, 'val': val},
                      conn=conn)
            tipo_nota = 'creacion' if old_val is None else 'modificacion'
            f.auditar_nota(slug, prof['id'], 'profesor', tipo_nota, 'notas', aid,
                         act['curso'], materia, act['p'],
                         campo='nota', actividad_id=actividad_id,
                         valor_anterior=old_val, valor_nuevo=val, conn=conn)
        conn.commit()
        aids = set(item.get('aid') for item in items if item.get('aid'))
        for aid in aids:
            prom_est, nf = f.calcular_stats_y_nota_final(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
            resultados[aid] = {'promedio': prom_est, 'nota_final': nf}
    except Exception as e:
        conn.close()
        logger.error(f'Error en guardar_nota_batch: {e}')
        return jsonify({'status':'error'}), 500
    curso_stats = f.calcular_stats_curso(conn, slug, curso, materia, jornada, periodo, prof['id'])
    conn.close()
    return jsonify({'status':'ok', 'resultados': resultados, 'stats_curso': curso_stats})


@teacher_bp.route('/<slug>/guardar_evaluacion_batch', methods=['POST'])
def guardar_evaluacion_batch(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'status':'error'}), 403
    if not validar_csrf():
        return jsonify({'status':'error'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error'}), 400
    curso = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    raw = request.form.get('batch', '')
    if not raw:
        return jsonify({'status':'error'}), 400
    try:
        items = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'status':'error'}), 400
    conn = f.conectar(slug)
    resultados = {}
    try:
        for item in items:
            aid = item.get('aid')
            if not aid: continue
            ev = item.get('evaluacion', type=float)
            au = item.get('autoevaluacion', type=float)
            existing = conn.execute(
                '''SELECT evaluacion, autoevaluacion FROM evaluaciones
                   WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?''',
                (aid, prof['id'], materia, jornada, periodo)).fetchone()
            old_eval = existing['evaluacion'] if existing else None
            old_auto = existing['autoevaluacion'] if existing else None
            ev_final = ev if ev is not None else old_eval
            au_final = au if au is not None else old_auto
            conn.execute(
                '''INSERT INTO evaluaciones
                   (aid,profesor_id,materia,jornada,evaluacion,autoevaluacion,periodo)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                   DO UPDATE SET evaluacion=excluded.evaluacion, autoevaluacion=excluded.autoevaluacion''',
                (aid, prof['id'], materia, jornada, ev_final, au_final, periodo))
            if ev is not None:
                tipo_ev = 'creacion' if old_eval is None else 'modificacion'
                f.auditar_nota(slug, prof['id'], 'profesor', tipo_ev, 'evaluaciones', aid,
                             curso, materia, periodo, campo='evaluacion',
                             valor_anterior=old_eval, valor_nuevo=ev_final)
            if au is not None:
                tipo_au = 'creacion' if old_auto is None else 'modificacion'
                f.auditar_nota(slug, prof['id'], 'profesor', tipo_au, 'evaluaciones', aid,
                             curso, materia, periodo, campo='autoevaluacion',
                             valor_anterior=old_auto, valor_nuevo=au_final)
        conn.commit()
        aids = set(item.get('aid') for item in items if item.get('aid'))
        for aid in aids:
            prom_est, nf = f.calcular_stats_y_nota_final(conn, slug, aid, curso, materia, jornada, periodo, prof['id'])
            resultados[aid] = {'promedio': prom_est, 'nota_final': nf}
    except Exception as e:
        conn.close()
        logger.error(f'Error en guardar_evaluacion_batch: {e}')
        return jsonify({'status':'error'}), 500
    curso_stats = f.calcular_stats_curso(conn, slug, curso, materia, jornada, periodo, prof['id'])
    conn.close()
    return jsonify({'status':'ok', 'resultados': resultados, 'stats_curso': curso_stats})


@teacher_bp.route('/<slug>/solicitar_modificacion', methods=['POST'])
def solicitar_modificacion(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    aid             = request.form.get('aid', type=int)
    actividad_id    = request.form.get('actividad_id', type=int)
    tipo            = request.form.get('tipo', '').strip()
    valor_solicitado = request.form.get('valor_solicitado', type=float)
    motivo          = request.form.get('motivo', '').strip()
    periodo         = request.form.get('periodo', 1, type=int)
    if None in (aid, valor_solicitado) or not motivo or not tipo:
        return jsonify({'status':'error','mensaje':'Datos incompletos'}), 400
    if valor_solicitado < 0 or valor_solicitado > 5:
        return jsonify({'status':'error','mensaje':'Nota debe estar entre 0 y 5'}), 400
    if tipo not in ('actividad', 'evaluacion', 'autoevaluacion'):
        return jsonify({'status':'error','mensaje':'Tipo inv\u00e1lido'}), 400
    if not f.periodo_cerrado(slug, periodo):
        return jsonify({'status':'error','mensaje':'El per\u00edodo no est\u00e1 cerrado'}), 400
    conn = f.conectar(slug)
    jornada_ctx, materia_ctx = f.get_sesion_jornada_materia(slug)
    materia = materia_ctx or ''
    jornada = jornada_ctx or ''
    curso = ''
    valor_actual = None
    if tipo == 'actividad':
        if actividad_id is None:
            conn.close()
            return jsonify({'status':'error','mensaje':'actividad_id requerido para tipo actividad'}), 400
        act = conn.execute(
            'SELECT id, profesor_id, curso, materia, COALESCE(periodo,1) as p FROM actividades WHERE id=?',
            (actividad_id,)).fetchone()
        if not act:
            conn.close()
            return jsonify({'status':'error','mensaje':'Actividad no encontrada'}), 404
        if act['profesor_id'] != prof['id']:
            conn.close()
            return jsonify({'status':'error','mensaje':'No eres el propietario de esta actividad'}), 403
        materia = act['materia']
        curso = act['curso']
        nota_db = conn.execute(
            'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
            (aid, actividad_id)).fetchone()
        valor_actual = nota_db['val'] if nota_db else None
    elif tipo == 'evaluacion':
        curso = request.form.get('curso', '')
        alumno = conn.execute(
            'SELECT evaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND COALESCE(periodo,1)=?',
            (aid, prof['id'], materia, periodo)).fetchone()
        valor_actual = alumno['evaluacion'] if alumno else None
    elif tipo == 'autoevaluacion':
        curso = request.form.get('curso', '')
        alumno = conn.execute(
            'SELECT autoevaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND COALESCE(periodo,1)=?',
            (aid, prof['id'], materia, periodo)).fetchone()
        valor_actual = alumno['autoevaluacion'] if alumno else None
    if not curso:
        curso = request.form.get('curso', '')
    conn.execute(
        '''INSERT INTO solicitudes_modificacion
           (slug, aid, profesor_id, materia, curso, jornada, periodo, tipo, actividad_id,
            valor_actual, valor_solicitado, motivo, estado)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'pendiente')''',
        (slug, aid, prof['id'], materia, curso, jornada, periodo, tipo, actividad_id,
         str(valor_actual) if valor_actual is not None else None, str(valor_solicitado), motivo))
    sid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.commit()
    f.auditar_nota(slug, prof['id'], 'profesor', 'solicitud_creada', 'solicitudes_modificacion', aid,
                 curso, materia, periodo, campo=tipo, actividad_id=actividad_id,
                 valor_anterior=valor_actual, valor_nuevo=valor_solicitado,
                 motivo='Solicitud #%d: %s' % (sid, motivo))
    rectores = conn.execute('SELECT id FROM rectores WHERE activo=1').fetchall()
    for r in rectores:
        f.crear_notificacion(slug, 'rector', r['id'],
            'Nueva solicitud de modificaci\u00f3n de %s' % prof['nombre'],
            'El profesor %s solicita cambiar %s de %s a %s. Motivo: %s' % (
                prof['nombre'], tipo, valor_actual or 'sin nota', valor_solicitado, motivo),
            link=url_for('rector.rector_solicitudes', slug=slug))
    conn.close()
    return jsonify({'status':'ok','mensaje':'Solicitud enviada correctamente.','id':sid})


@teacher_bp.route('/<slug>/plantilla_notas')
def plantilla_notas(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('teacher.seleccionar_jornada', slug=slug))
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = f.conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    alumnos = []
    if curso_sel:
        conn = f.conectar(slug)
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
            (curso_sel, jornada)).fetchall()
        conn.close()
    wb = f._excel_armar_wb(slug, prof, materia, jornada, curso_sel, periodo, actividades, alumnos)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'plantilla_lumini_{slug}_{curso_sel}_{periodo}.xlsx'
    return Response(bio.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})


@teacher_bp.route('/<slug>/exportar_notas')
def exportar_notas(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('teacher.seleccionar_jornada', slug=slug))
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = f.conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    alumnos = []
    if curso_sel:
        conn = f.conectar(slug)
        alumnos = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre COLLATE NOCASE',
            (curso_sel, jornada)).fetchall()
        conn.close()
    wb = f._excel_armar_wb(slug, prof, materia, jornada, curso_sel, periodo, actividades, alumnos)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'notas_{slug}_{curso_sel}_{periodo}.xlsx'
    return Response(bio.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})


@teacher_bp.route('/<slug>/importar_notas', methods=['GET'])
def importar_notas(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('teacher.seleccionar_jornada', slug=slug))
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = f.conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    return render_template('importar_notas.html', slug=slug, colegio=f.get_colegio(slug), profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel, periodo=periodo,
                           materia=materia, jornada=jornada, actividades=actividades)


@teacher_bp.route('/<slug>/importar_notas/preview', methods=['POST'])
def importar_notas_preview(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    cfg = f.config_get(slug)
    try:
        escala_min = float(cfg.get('escala_min', 0.0))
        escala_max = float(cfg.get('escala_max', 5.0))
    except (TypeError, ValueError):
        escala_min, escala_max = 0.0, 5.0
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    if 'archivo' not in request.files:
        return jsonify({'status':'error','mensaje':'No se envio ningun archivo'}), 400
    archivo = request.files['archivo']
    if not archivo.filename or not archivo.filename.lower().endswith('.xlsx'):
        return jsonify({'status':'error','mensaje':'El archivo debe ser .xlsx'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f'Error al leer archivo de notas: {e}')
        # Mostrar la causa concreta del error en lugar de un mensaje genérico
        mensaje_error = 'Error al leer el archivo.'
        try:
            from openpyxl.utils.exceptions import InvalidFileException
            from zipfile import BadZipFile
            if isinstance(e, (BadZipFile, InvalidFileException)):
                mensaje_error = 'El archivo no es un Excel .xlsx válido o está corrupto.'
            elif hasattr(e, 'msg') and e.msg:
                mensaje_error = f'Error al leer el archivo: {e.msg}'
        except ImportError:
            if str(e):
                mensaje_error = f'Error al leer el archivo: {e}'
        return jsonify({'status':'error','mensaje':mensaje_error}), 400
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    import unicodedata as _ud
    def _norm(v):
        return str(v).strip().lstrip('\ufeff') if v is not None else ''
    def _norm_name(v):
        s = _norm(v).lower()
        s = ''.join(c for c in _ud.normalize('NFKD', s) if not _ud.combining(c))
        return s
    if not header_row or _norm(header_row[0]) != 'N\u00b0':
        return jsonify({'status':'error','mensaje':'Formato de archivo invalido. La primera columna debe ser N\u00b0'}), 400
    rows_data = list(ws.iter_rows(min_row=2, values_only=False))
    if not rows_data:
        return jsonify({'status':'error','mensaje':'El archivo no contiene datos'}), 400
    conn = f.conectar(slug)
    try:
        actividades_existentes = conn.execute(
            '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
               AND COALESCE(periodo,1)=? ORDER BY orden''',
            (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
        existing_act_names = {a['nombre']: a for a in actividades_existentes}
        alumnos_curso = conn.execute(
            'SELECT * FROM alumnos WHERE curso=? AND jornada=? AND activo=1',
            (curso_sel, jornada)).fetchall()
        alumno_by_name_norm = {}
        for _al in alumnos_curso:
            _k = _norm_name(_al['nombre'])
            if _k in alumno_by_name_norm:
                alumno_by_name_norm[_k] = None
            else:
                alumno_by_name_norm[_k] = _al
        act_cols = []
        eval_col = auto_col = None
        for col_idx, h in enumerate(header_row):
            if col_idx <= 2:
                continue
            h_str = _norm(h)
            if h_str == 'Evaluaci\u00f3n':
                eval_col = col_idx
            elif h_str == 'Autoevaluaci\u00f3n':
                auto_col = col_idx
            elif h_str == 'Promedio':
                continue
            elif h_str:
                act_cols.append((col_idx, h_str))
        new_activities = []
        col_map = {}
        max_orden = max([a['orden'] for a in actividades_existentes], default=0)
        for col_idx, act_name in act_cols:
            if act_name in existing_act_names:
                col_map[col_idx] = {'tipo': 'actividad', 'nombre': act_name, 'actividad_id': existing_act_names[act_name]['id']}
            else:
                max_orden += 1
                col_map[col_idx] = {'tipo': 'actividad', 'nombre': act_name, 'actividad_id': None, 'orden': max_orden}
                new_activities.append({'nombre': act_name, 'orden': max_orden})
        all_ok = True
        preview_rows = []
        aid_set = set()
        for row_cells in rows_data:
            cells = [c.value for c in row_cells]
            if not any(c is not None for c in cells):
                continue
            raw_nombre = str(cells[1]).strip() if cells[1] is not None else ''
            raw_aid = cells[2]
            row_errors = []
            aid = None
            if raw_aid is not None:
                try:
                    aid = int(float(str(raw_aid).replace(',', '.').strip()))
                except (ValueError, TypeError):
                    pass
            alumno = None
            if aid:
                al = conn.execute('SELECT * FROM alumnos WHERE id=? AND jornada=? AND activo=1', (aid, jornada)).fetchone()
                if al and al['curso'] == curso_sel:
                    alumno = al
            if not alumno and raw_nombre:
                al = alumno_by_name_norm.get(_norm_name(raw_nombre))
                if al is None and _norm_name(raw_nombre) in alumno_by_name_norm:
                    row_errors.append('Nombre duplicado en el curso, incluye el AID correcto')
                    all_ok = False
                elif al:
                    alumno = al
                    aid = al['id']
            if not alumno and not any('duplicado' in e for e in row_errors):
                row_errors.append('Estudiante no encontrado en este curso')
                all_ok = False
            if alumno and aid:
                if aid in aid_set:
                    row_errors.append('Estudiante duplicado en el archivo')
                    all_ok = False
                aid_set.add(aid)
            changes = {}
            for col_idx, cinfo in col_map.items():
                raw_val = cells[col_idx] if col_idx < len(cells) else None
                val = None
                if raw_val is not None:
                    try:
                        val = float(str(raw_val).replace(',', '.'))
                        if val < escala_min or val > escala_max:
                            row_errors.append(f'{cinfo["nombre"]}: nota fuera de rango ({escala_min}-{escala_max})')
                            all_ok = False
                            continue
                    except (ValueError, TypeError):
                        row_errors.append(f'{cinfo["nombre"]}: valor invalido')
                        all_ok = False
                        continue
                changes[f'act_{col_idx}'] = {'tipo': 'actividad', 'actividad_id': cinfo.get('actividad_id'),
                                               'valor': val, 'nombre_col': cinfo['nombre']}
            if alumno:
                if eval_col is not None and eval_col < len(cells):
                    raw_val = cells[eval_col]
                    if raw_val is not None:
                        try:
                            val = float(str(raw_val).replace(',', '.'))
                            if val < escala_min or val > escala_max:
                                row_errors.append('Evaluaci\u00f3n fuera de rango')
                                all_ok = False
                            else:
                                changes['eval'] = {'tipo': 'evaluacion', 'valor': val}
                        except (ValueError, TypeError):
                            row_errors.append('Evaluaci\u00f3n invalida')
                            all_ok = False
                if auto_col is not None and auto_col < len(cells):
                    raw_val = cells[auto_col]
                    if raw_val is not None:
                        try:
                            val = float(str(raw_val).replace(',', '.'))
                            if val < escala_min or val > escala_max:
                                row_errors.append('Autoevaluaci\u00f3n fuera de rango')
                                all_ok = False
                            else:
                                changes['auto'] = {'tipo': 'autoevaluacion', 'valor': val}
                        except (ValueError, TypeError):
                            row_errors.append('Autoevaluaci\u00f3n invalida')
                            all_ok = False
            preview_rows.append({
                'fila': row_cells[0].row,
                'aid': aid,
                'nombre': raw_nombre,
                'alumno': dict(alumno) if alumno else None,
                'errors': row_errors,
                'changes': changes,
                'ok': len(row_errors) == 0,
            })
    finally:
        conn.close()
    return jsonify({
        'status': 'ok' if all_ok else 'error',
        'total': len(preview_rows),
        'validos': sum(1 for r in preview_rows if r['ok']),
        'errores': sum(1 for r in preview_rows if not r['ok']),
        'filas': preview_rows,
        'nuevas_actividades': new_activities,
        'curso': curso_sel,
        'periodo': periodo,
        'all_ok': all_ok,
    })


@teacher_bp.route('/<slug>/importar_notas/confirmar', methods=['POST'])
def importar_notas_confirmar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    data_json = request.form.get('data', '')
    if not data_json:
        return jsonify({'status':'error','mensaje':'No hay datos para guardar'}), 400
    try:
        data = json.loads(data_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'status':'error','mensaje':'Datos invalidos'}), 400
    if not data.get('all_ok'):
        return jsonify({'status':'error','mensaje':'Hay errores que deben corregirse primero'}), 400
    cfg = f.config_get(slug)
    try:
        escala_min = float(cfg.get('escala_min', 0.0))
        escala_max = float(cfg.get('escala_max', 5.0))
    except (TypeError, ValueError):
        escala_min, escala_max = 0.0, 5.0
    conn = f.conectar(slug)
    try:
        ok_rev, errores_rev = revalidar_importacion_notas(
            data, conn, prof, materia, jornada, curso_sel, periodo,
            escala_min=escala_min, escala_max=escala_max)
    except Exception as e:
        conn.close()
        logger.error(f'Error revalidando importacion: {e}')
        return jsonify({'status':'error','mensaje':'Error al validar los datos. Intenta de nuevo.'}), 500
    if not ok_rev:
        conn.close()
        return jsonify({'status':'error','mensaje':'Hay errores que deben corregirse primero',
                        'errores': errores_rev[:20]}), 400
    try:
        new_act_names = {}
        for na in data.get('nuevas_actividades', []):
            conn.execute(
                'INSERT INTO actividades (nombre, profesor_id, materia, jornada, curso, orden, periodo) VALUES (?,?,?,?,?,?,?)',
                (na['nombre'], prof['id'], materia, jornada, curso_sel, na['orden'], periodo))
            act_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            new_act_names[na['nombre']] = act_id
        updated = 0
        for fila in data['filas']:
            if not fila.get('ok') or not fila.get('alumno'):
                continue
            aid = fila['aid']
            for key, ch in fila.get('changes', {}).items():
                if ch['tipo'] == 'actividad':
                    act_id = ch.get('actividad_id')
                    if act_id is None:
                        act_id = new_act_names.get(ch.get('nombre_col', ''))
                    if act_id is None:
                        continue
                    existing = conn.execute(
                        'SELECT val FROM notas WHERE aid=? AND actividad_id=?', (aid, act_id)).fetchone()
                    old_val = existing['val'] if existing else None
                    if old_val != ch['valor']:
                        if ch['valor'] is not None:
                            conn.execute(
                                '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
                                   ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
                                (aid, act_id, ch['valor']))
                        elif existing:
                            conn.execute('DELETE FROM notas WHERE aid=? AND actividad_id=?', (aid, act_id))
                        f.auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'notas', aid,
                                     curso_sel, materia, periodo,
                                     campo='nota', actividad_id=act_id,
                                     valor_anterior=old_val, valor_nuevo=ch['valor'],
                                     motivo='Importacion masiva Excel')
                        updated += 1
                elif ch['tipo'] == 'evaluacion':
                    existing = conn.execute(
                        'SELECT evaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['evaluacion'] if existing else None
                    if old_val != ch['valor']:
                        if ch['valor'] is not None:
                            conn.execute(
                                '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,evaluacion,periodo)
                                   VALUES (?,?,?,?,?,?)
                                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                                   DO UPDATE SET evaluacion=excluded.evaluacion''',
                                (aid, prof['id'], materia, jornada, ch['valor'], periodo))
                        elif existing:
                            conn.execute(
                                'UPDATE evaluaciones SET evaluacion=NULL WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                                (aid, prof['id'], materia, jornada, periodo))
                        f.auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='evaluacion',
                                     valor_anterior=old_val, valor_nuevo=ch['valor'],
                                     motivo='Importacion masiva Excel')
                        updated += 1
                elif ch['tipo'] == 'autoevaluacion':
                    existing = conn.execute(
                        'SELECT autoevaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['autoevaluacion'] if existing else None
                    if old_val != ch['valor']:
                        if ch['valor'] is not None:
                            conn.execute(
                                '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,autoevaluacion,periodo)
                                   VALUES (?,?,?,?,?,?)
                                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                                   DO UPDATE SET autoevaluacion=excluded.autoevaluacion''',
                                (aid, prof['id'], materia, jornada, ch['valor'], periodo))
                        elif existing:
                            conn.execute(
                                'UPDATE evaluaciones SET autoevaluacion=NULL WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                                (aid, prof['id'], materia, jornada, periodo))
                        f.auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='autoevaluacion',
                                     valor_anterior=old_val, valor_nuevo=ch['valor'],
                                     motivo='Importacion masiva Excel')
                        updated += 1
        conn.commit()
    except Exception as e:
        conn.close()
        logger.error(f'Error al guardar: {e}')
        return jsonify({'status':'error','mensaje':'Error al guardar. Intenta de nuevo.'}), 500
    conn.close()
    return jsonify({'status':'ok', 'mensaje': f'Importacion completada. {updated} valores actualizados.', 'updated': updated})


@teacher_bp.route('/<slug>/migrar-excel', methods=['GET'])
def migrar_excel(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return redirect(url_for('teacher.seleccionar_jornada', slug=slug))
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else '')
    periodo = request.args.get('periodo', 1, type=int)
    conn = f.conectar(slug)
    actividades = conn.execute(
        '''SELECT * FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=? ORDER BY orden''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    conn.close()
    return render_template('migrar_excel.html', slug=slug, colegio=f.get_colegio(slug), profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel, periodo=periodo,
                           materia=materia, jornada=jornada, actividades=actividades)


@teacher_bp.route('/<slug>/migrar-excel/analizar', methods=['POST'])
def migrar_excel_analizar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    if 'archivo' not in request.files:
        return jsonify({'status':'error','mensaje':'No se envio ningun archivo'}), 400
    archivo = request.files['archivo']
    if not archivo.filename or not extension_excel_valida(archivo.filename):
        return jsonify({'status':'error','mensaje':'Formato no valido. Usa .xlsx'}), 400
    import os
    import tempfile
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        archivo.save(tmp.name)
        tmp.close()
        import openpyxl
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) > 2001:
            return jsonify({'status':'error','mensaje':'El archivo tiene demasiadas filas (max 2000).'}), 400
        if len(rows) < 2:
            return jsonify({'status':'error','mensaje':'El archivo debe tener al menos una fila de encabezados y una fila de datos'}), 400
        encabezados = [str(c).strip() if c is not None else '' for c in rows[0]]
        conn = f.conectar(slug)
        alumnos = conn.execute(
            'SELECT id, nombre FROM alumnos WHERE curso=? AND activo=1 ORDER BY nombre',
            (curso_sel,)).fetchall()
        actividades = conn.execute(
            '''SELECT id, nombre, orden FROM actividades WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
               AND COALESCE(periodo,1)=? ORDER BY orden''',
            (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
        conn.close()
        acts_map = {a['nombre'].strip().lower(): a for a in actividades}
        col_detectadas = []
        col_pendientes = []
        for i, h in enumerate(encabezados):
            hl = h.strip().lower()
            if not hl:
                col_pendientes.append({'indice': i, 'nombre': h, 'tipo': 'skip'})
                continue
            act_match = acts_map.get(hl)
            if act_match:
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'actividad_id': act_match['id'], 'actividad_nombre': act_match['nombre']})
            elif hl in ('nombre', 'nombre del estudiante', 'estudiante', 'alumno', 'alumno(a)'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'nombre'})
            elif hl in ('documento', 'id', 'identificacion', 'codigo', 'c\u00f3digo', 'cedula', 'c\u00e9dula'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'documento'})
            elif hl in ('evaluacion', 'evaluaci\u00f3n', 'eva', 'nota evaluacion', 'nota evaluaci\u00f3n', 'eval'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'evaluacion'})
            elif hl in ('autoevaluacion', 'autoevaluaci\u00f3n', 'auto-evaluacion', 'auto', 'auto-evaluaci\u00f3n'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'autoevaluacion'})
            elif hl in ('proyecto', 'nota proyecto'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'proyecto'})
            elif hl in ('recuperacion', 'recuperaci\u00f3n', 'recu', 'nota recuperacion', 'nota recuperaci\u00f3n'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'recuperacion'})
            elif hl in ('nota final', 'nota definitiva', 'definitiva', 'promedio', 'final'):
                col_detectadas.append({'indice': i, 'nombre': h, 'tipo': 'auto', 'campo': 'nota_final'})
            else:
                col_pendientes.append({'indice': i, 'nombre': h, 'tipo': 'manual'})
        filas = []
        for idx, row in enumerate(rows[1:], start=2):
            vals = [str(v).strip() if v is not None else '' for v in row]
            nombre_al = ''
            documento_al = ''
            errores = []
            for col in col_detectadas:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                if col.get('campo') == 'nombre' and v:
                    nombre_al = v
                elif col.get('campo') == 'documento' and v:
                    documento_al = v
            for col in col_pendientes:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                if v:
                    errores.append(f'Columna "{col["nombre"]}" sin asignar: "{v}"')
            alumno_match = None
            if documento_al:
                for a in alumnos:
                    if str(a['documento']).strip() == documento_al:
                        alumno_match = a
                        break
            if not alumno_match and nombre_al:
                na = nombre_al.strip().lower()
                for a in alumnos:
                    if a['nombre'].strip().lower() == na:
                        alumno_match = a
                        break
            if not alumno_match and nombre_al:
                partes = nombre_al.strip().lower().split()
                for a in alumnos:
                    an = a['nombre'].strip().lower()
                    if all(p in an for p in partes):
                        alumno_match = a
                        break
            ok = alumno_match is not None
            if not ok and not errores:
                errores.append('Estudiante no encontrado en el curso')
            valores = {}
            for col in col_detectadas:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                if col.get('actividad_id') and v:
                    try:
                        vn = v.replace(',', '.')
                        valores[col['indice']] = round(float(vn), 2)
                    except ValueError:
                        valores[col['indice']] = v
                elif col.get('campo') in ('evaluacion','autoevaluacion','proyecto','recuperacion','nota_final') and v:
                    try:
                        vn = v.replace(',', '.')
                        valores[col['indice']] = round(float(vn), 2)
                    except ValueError:
                        valores[col['indice']] = v
                else:
                    valores[col['indice']] = v
            for col in col_pendientes:
                v = vals[col['indice']] if col['indice'] < len(vals) else ''
                valores[col['indice']] = v if v else None
            filas.append({
                'fila': idx, 'alumno': alumno_match['id'] if alumno_match else None,
                'nombre': alumno_match['nombre'] if alumno_match else (nombre_al or f'Fila {idx}'),
                'ok': ok, 'aid': alumno_match['id'] if alumno_match else None,
                'errores': errores, 'valores': valores
            })
        nuevas_actividades = []
        for col in col_pendientes:
            hl = col['nombre'].strip().lower()
            if hl and hl not in acts_map:
                next_orden = (max((a['orden'] for a in actividades), default=0) + 1 + len(nuevas_actividades))
                nuevas_actividades.append({'nombre': col['nombre'].strip(), 'orden': next_orden})
                col['tipo'] = 'auto'
                col['actividad_nombre'] = col['nombre'].strip()
        validos = sum(1 for f in filas if f['ok'])
        errores_count = sum(1 for f in filas if not f['ok'])
        status = 'ok' if errores_count == 0 else 'warning'
        return jsonify({
            'status': status,
            'columnas': col_detectadas + col_pendientes,
            'columnas_pendientes': col_pendientes,
            'filas': filas,
            'total': len(filas),
            'validos': validos,
            'errores': errores_count,
            'nuevas_actividades': nuevas_actividades,
            'actividades_existentes': [{'id': a['id'], 'nombre': a['nombre']} for a in actividades]
        })
    except Exception as e:
        logger.exception(f'Error analizando archivo: {e}')
        return jsonify({'status':'error','mensaje':'Error al procesar el archivo: '+str(e)}), 500
    finally:
        os.unlink(tmp.name)


@teacher_bp.route('/<slug>/migrar-excel/confirmar', methods=['POST'])
def migrar_excel_confirmar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'status':'error','mensaje':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'status':'error','mensaje':'Error CSRF'}), 403
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'status':'error','mensaje':'Sesion no valida'}), 400
    curso_sel = request.form.get('curso', '')
    periodo = request.form.get('periodo', 1, type=int)
    data_json = request.form.get('data', '')
    if not data_json:
        return jsonify({'status':'error','mensaje':'No hay datos para guardar'}), 400
    try:
        data = json.loads(data_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'status':'error','mensaje':'Datos invalidos'}), 400
    conn = f.conectar(slug)
    try:
        new_act_ids = {}
        for na in data.get('nuevas_actividades', []):
            conn.execute(
                'INSERT INTO actividades (nombre, profesor_id, materia, jornada, curso, orden, periodo) VALUES (?,?,?,?,?,?,?)',
                (na['nombre'], prof['id'], materia, jornada, curso_sel, na['orden'], periodo))
            act_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            new_act_ids[na['nombre']] = act_id
        data.get('column_mapping', {})
        updated = 0
        for fila in data.get('filas', []):
            if not fila.get('ok') or not fila.get('aid'):
                continue
            aid = fila['aid']
            for col in data.get('columnas', []):
                str(col['indice'])
                val = fila.get('valores', {}).get(col['indice'])
                if val is None or val == '':
                    continue
                try:
                    val_num = round(float(str(val).replace(',', '.')), 2)
                except (ValueError, TypeError):
                    continue
                if col.get('actividad_id'):
                    act_id = col['actividad_id']
                    existing = conn.execute(
                        'SELECT val FROM notas WHERE aid=? AND actividad_id=?', (aid, act_id)).fetchone()
                    old_val = existing['val'] if existing else None
                    if old_val != val_num:
                        conn.execute(
                            '''INSERT INTO notas (aid,actividad_id,val) VALUES (?,?,?)
                               ON CONFLICT(aid,actividad_id) DO UPDATE SET val=excluded.val''',
                            (aid, act_id, val_num))
                        f.auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'notas', aid,
                                     curso_sel, materia, periodo,
                                     campo='nota', actividad_id=act_id,
                                     valor_anterior=old_val, valor_nuevo=val_num,
                                     motivo='Migracion desde Excel')
                        updated += 1
                elif col.get('campo') in ('evaluacion',):
                    existing = conn.execute(
                        'SELECT evaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['evaluacion'] if existing else None
                    if old_val != val_num:
                        conn.execute(
                            '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,evaluacion,periodo)
                               VALUES (?,?,?,?,?,?)
                               ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                               DO UPDATE SET evaluacion=excluded.evaluacion''',
                            (aid, prof['id'], materia, jornada, val_num, periodo))
                        f.auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='evaluacion',
                                     valor_anterior=old_val, valor_nuevo=val_num,
                                     motivo='Migracion desde Excel')
                        updated += 1
                elif col.get('campo') in ('autoevaluacion',):
                    existing = conn.execute(
                        'SELECT autoevaluacion FROM evaluaciones WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                        (aid, prof['id'], materia, jornada, periodo)).fetchone()
                    old_val = existing['autoevaluacion'] if existing else None
                    if old_val != val_num:
                        if old_val is not None:
                            conn.execute(
                                '''INSERT INTO evaluaciones (aid,profesor_id,materia,jornada,autoevaluacion,periodo)
                                   VALUES (?,?,?,?,?,?)
                                   ON CONFLICT(aid,profesor_id,materia,jornada,periodo)
                                   DO UPDATE SET autoevaluacion=excluded.autoevaluacion''',
                                (aid, prof['id'], materia, jornada, val_num, periodo))
                        else:
                            conn.execute(
                                'UPDATE evaluaciones SET autoevaluacion=? WHERE aid=? AND profesor_id=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                                (val_num, aid, prof['id'], materia, jornada, periodo))
                        f.auditar_nota(slug, prof['id'], 'profesor', 'modificacion', 'evaluaciones', aid,
                                     curso_sel, materia, periodo, campo='autoevaluacion',
                                     valor_anterior=old_val, valor_nuevo=val_num,
                                     motivo='Migracion desde Excel')
                        updated += 1
        conn.commit()
    except Exception as e:
        conn.close()
        logger.error(f'Error al confirmar migracion: {e}')
        return jsonify({'status':'error','mensaje':'Error al guardar. Intenta de nuevo.'}), 500
    conn.close()
    return jsonify({'status':'ok', 'mensaje': f'Migracion completada. {updated} valores guardados.', 'updated': updated})


@teacher_bp.route('/<slug>/observaciones_json', methods=['POST'])
def observaciones_json(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'observaciones':[]})
    if not validar_csrf(): return jsonify({'observaciones':[]}), 403
    data = request.get_json(silent=True) or {}
    aid = data.get('aid')
    if not aid: return jsonify({'observaciones':[]})
    conn = f.conectar(slug)
    rows = conn.execute(
        'SELECT id, materia, texto, fecha FROM observaciones WHERE aid=? ORDER BY fecha DESC LIMIT 20',
        (aid,)).fetchall()
    conn.close()
    return jsonify({'observaciones':[dict(r) for r in rows]})


@teacher_bp.route('/<slug>/recalcular/<int:aid>')
def recalcular(slug, aid):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso_sel = request.args.get('curso') or ''
        periodo = request.args.get('periodo', 1, type=int)
        if not curso_sel:
            al = conn.execute('SELECT curso FROM alumnos WHERE id=?', (aid,)).fetchone()
            curso_sel = al['curso'] if al else ''
        prom, nf = f.calcular_stats_y_nota_final(conn, slug, aid, curso_sel, materia, jornada, periodo, prof['id'])
        return jsonify({'promedio':prom,'nota_final':nf})
    except Exception as e:
        logger.error('stats_estudiante: %s', e)
        return jsonify({'error':str(e)}), 500
    finally:
        conn.close()


@teacher_bp.route('/<slug>/curso/analitica')
def curso_analitica(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso:
            conn.close(); return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute(
            'SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
        act_ids = [r['id'] for r in conn.execute(
            'SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()]
        notas_por_act = {}
        for act_id in act_ids:
            vals = [float(r['val']) for r in conn.execute(
                'SELECT val FROM notas WHERE actividad_id=?', (act_id,)).fetchall() if r['val'] is not None]
            notas_por_act[act_id] = vals
        todos_vals = [v for vals in notas_por_act.values() for v in vals]
        promedios = []
        estudiantes_data = []
        for al in alumnos:
            vals_al = []
            for act_id in act_ids:
                r = conn.execute(
                    'SELECT val FROM notas WHERE aid=? AND actividad_id=?',
                    (al['id'], act_id)).fetchone()
                if r and r['val'] is not None:
                    vals_al.append(float(r['val']))
            prom = round(sum(vals_al)/len(vals_al),2) if vals_al else None
            if prom is not None:
                promedios.append(prom)
            estudiantes_data.append({'id':al['id'],'nombre':al['nombre'],'promedio':prom,'notas_count':len(vals_al)})
        max_val = max(todos_vals) if todos_vals else None
        min_val = min(todos_vals) if todos_vals else None
        prom_curso = round(sum(promedios)/len(promedios),2) if promedios else None
        aprobando = sum(1 for p in promedios if p is not None and p >= 3.0)
        perdiendo = sum(1 for p in promedios if p is not None and p < 3.0)
        sin_notas = sum(1 for e in estudiantes_data if e['promedio'] is None)
        riesgo_alto = sum(1 for p in promedios if p is not None and p < 2.8)
        riesgo_medio = sum(1 for p in promedios if p is not None and 2.8 <= p <= 3.5)
        acts_info = []
        for act_id in act_ids:
            vals = notas_por_act.get(act_id, [])
            a = conn.execute('SELECT nombre, tipo, peso, competencia FROM actividades WHERE id=?', (act_id,)).fetchone()
            if a:
                acts_info.append({
                    'id':act_id,'nombre':a['nombre'],'tipo':a['tipo'],'peso':a['peso'],
                    'promedio':round(sum(vals)/len(vals),2) if vals else None,
                    'cantidad':len(vals),
                    'aprobados':sum(1 for v in vals if v>=3.0),
                    'reprobados':sum(1 for v in vals if v<3.0 and v>0),
                    'total_estudiantes':len(alumnos)
                })
        dist = {'1':0,'2':0,'3':0,'4':0,'5':0}
        for v in todos_vals:
            k = str(int(v))
            if k in dist: dist[k] += 1
        return jsonify({
            'promedio_curso':prom_curso,'max':max_val,'min':min_val,
            'aprobando':aprobando,'perdiendo':perdiendo,'sin_notas':sin_notas,
            'riesgo_alto':riesgo_alto,'riesgo_medio':riesgo_medio,
            'total_estudiantes':len(alumnos),'actividades':acts_info,
            'distribucion':dist,'total_notas':len(todos_vals)
        })
    finally:
        conn.close()


@teacher_bp.route('/<slug>/curso/ranking')
def curso_ranking(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute(
            'SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
        act_ids = [r['id'] for r in conn.execute(
            'SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()]
        data = []
        for al in alumnos:
            vals = []
            for act_id in act_ids:
                r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                if r and r['val'] is not None: vals.append(float(r['val']))
            prom = round(sum(vals)/len(vals),2) if vals else None
            ev = conn.execute(
                'SELECT evaluacion, autoevaluacion FROM evaluaciones WHERE aid=? AND materia=? AND jornada=? AND COALESCE(periodo,1)=?',
                (al['id'], materia, jornada, periodo)).fetchone()
            nf = None
            if prom is not None:
                e = ev['evaluacion'] if ev else None
                a = ev['autoevaluacion'] if ev else None
                nf = round(prom*0.65 + (e or 0)*0.25 + (a or 0)*0.10, 2)
            data.append({'id':al['id'],'nombre':al['nombre'],'promedio':prom,'nota_final':nf,'notas_count':len(vals)})
        data.sort(key=lambda x: (x['promedio'] or 0), reverse=True)
        for i, d in enumerate(data):
            d['posicion'] = i + 1
        return jsonify({'ranking':data[:30],'total':len(data)})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/estudiante/<int:aid>/tendencia')
def estudiante_tendencia(slug, aid):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        cfg = f.config_get(slug)
        escala_max = float(cfg.get('escala_max', 5.0)) if cfg else 5.0
        rows = conn.execute(
            '''SELECT n.val, n.actividad_id, ac.nombre as act_nombre, ac.orden, ac.periodo
               FROM notas n JOIN actividades ac ON ac.id=n.actividad_id
               WHERE n.aid=? AND ac.materia=? AND ac.jornada=? AND ac.profesor_id=?
               ORDER BY ac.periodo, ac.orden''',
            (aid, materia, jornada, prof['id'])).fetchall()
        puntos = [{'orden':r['orden'],'valor':float(r['val']),'nombre':r['act_nombre'],
                    'periodo':r['periodo'],'actividad_id':r['actividad_id']} for r in rows if r['val'] is not None]
        acum = []; running = []
        for p in puntos:
            acum.append(p['valor'])
            running.append(round(sum(acum)/len(acum),2))
        for i, p in enumerate(puntos):
            p['promedio_acumulado'] = running[i] if i < len(running) else p['valor']
        if len(puntos) >= 2:
            vals = [p['valor'] for p in puntos]
            n = len(vals)
            x_avg = (n - 1) / 2
            y_avg = sum(vals) / n
            num = sum((i - x_avg) * (v - y_avg) for i, v in enumerate(vals))
            den = sum((i - x_avg) ** 2 for i in range(n))
            slope = num / den if den != 0 else 0
            pred = y_avg + slope * (n + 2 - x_avg)
            pred = max(0, min(escala_max, round(pred, 2)))
            confianza = min(95, max(30, int(100 - abs(slope) * 20)))
        else:
            pred = puntos[-1]['valor'] if puntos else None
            confianza = 30 if pred else 0
        all_vals = []
        for p in puntos:
            r = conn.execute('SELECT val FROM notas WHERE actividad_id=?', (p['actividad_id'],)).fetchall()
            all_vals.extend([float(x['val']) for x in r if x['val'] is not None])
        prom_curso = round(sum(all_vals)/len(all_vals),2) if all_vals else None
        prom_est = puntos[-1]['promedio_acumulado'] if puntos else None
        diff = None
        if prom_est is not None and prom_curso is not None and prom_curso > 0:
            diff = round((prom_est - prom_curso) / prom_curso * 100, 1)
        return jsonify({
            'puntos':puntos,
            'prediccion':pred,
            'confianza':confianza,
            'promedio_estudiante':prom_est,
            'promedio_curso':prom_curso,
            'diferencia_porcentual':diff,
            'escala_max':escala_max
        })
    finally:
        conn.close()


@teacher_bp.route('/<slug>/observaciones/sugerir', methods=['POST'])
def observaciones_sugerir(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    aid = data.get('aid')
    cambio = data.get('cambio', '')
    if not aid: return jsonify({'sugerencia':''})
    conn = f.conectar(slug)
    try:
        al = conn.execute('SELECT nombre FROM alumnos WHERE id=?', (aid,)).fetchone()
        nombre = al['nombre'] if al else 'El estudiante'
        if cambio == 'bajo':
            sugerencia = f'{nombre} disminuy\u00f3 su rendimiento acad\u00e9mico.'
        elif cambio == 'subio':
            sugerencia = f'{nombre} presenta una mejora constante en su rendimiento.'
        else:
            sugerencia = ''
        return jsonify({'sugerencia':sugerencia})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/alertas')
def curso_alertas(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute(
            'SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre',
            (curso, jornada)).fetchall()
        act_ids = [r['id'] for r in conn.execute(
            'SELECT id, nombre, fecha_limite FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()]
        alertas = []
        for al in alumnos:
            vals = []
            for act_id in act_ids:
                r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                if r and r['val'] is not None: vals.append(float(r['val']))
            prom = round(sum(vals)/len(vals),2) if vals else None
            if prom is not None and prom < 2.8:
                ausencias = conn.execute('SELECT COUNT(*) as c FROM asistencia WHERE aid=? AND estado=?', (al['id'], 'A')).fetchone()
                aus = ausencias['c'] if ausencias else 0
                sin_entregar = sum(1 for act_id in act_ids if not conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone())
                alertas.append({'tipo':'riesgo','alumno_id':al['id'],'nombre':al['nombre'],'promedio':prom,'ausencias':aus,'sin_entregar':sin_entregar})
            elif prom is None:
                alertas.append({'tipo':'sin_notas','alumno_id':al['id'],'nombre':al['nombre']})
        for act_id in act_ids:
            a = conn.execute('SELECT nombre, fecha_limite FROM actividades WHERE id=?', (act_id,)).fetchone()
            if a and a['fecha_limite']:
                try:
                    from datetime import datetime, timedelta
                    fl = datetime.strptime(a['fecha_limite'][:10], '%Y-%m-%d') if isinstance(a['fecha_limite'], str) else a['fecha_limite']
                    if fl < datetime.now() + timedelta(days=3) and fl >= datetime.now():
                        alertas.append({'tipo':'proximo_vencer','actividad_id':act_id,'nombre':a['nombre'],'fecha':a['fecha_limite'][:10]})
                except (ValueError, TypeError):
                    pass
        for al in alumnos:
            aus = conn.execute('SELECT COUNT(*) as c FROM asistencia WHERE aid=? AND estado=?', (al['id'], 'A')).fetchone()
            if aus and aus['c'] >= 4:
                if not any(a.get('alumno_id')==al['id'] for a in alertas):
                    alertas.append({'tipo':'muchas_ausencias','alumno_id':al['id'],'nombre':al['nombre'],'ausencias':aus['c']})
        return jsonify({'alertas':alertas})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/institucional/dashboard')
def institucional_dashboard(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso', (materia, jornada, prof['id'])).fetchall()
        data = {'cursos':[],'promedio_institucional':0,'total_estudiantes':0,'activos':0}
        prom_sum=0; prom_count=0; total_al=0
        for c in cursos:
            curso_name = c['curso']
            alumnos = conn.execute('SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso_name, jornada)).fetchall()
            total_al += len(alumnos)
            act_ids = [r['id'] for r in conn.execute('SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?', (materia, jornada, curso_name, prof['id'])).fetchall()]
            promedios = []
            for al in alumnos:
                vals = []
                for act_id in act_ids:
                    r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                    if r and r['val'] is not None: vals.append(float(r['val']))
                if vals: promedios.append(round(sum(vals)/len(vals),2))
            prom_curso = round(sum(promedios)/len(promedios),2) if promedios else None
            if prom_curso is not None: prom_sum += prom_curso; prom_count += 1
            data['cursos'].append({'curso':curso_name,'promedio':prom_curso,'estudiantes':len(alumnos),'aprobados':sum(1 for p in promedios if p>=3) if promedios else 0,'perdiendo':sum(1 for p in promedios if p<3) if promedios else 0})
        data['promedio_institucional'] = round(prom_sum/prom_count,2) if prom_count else None
        data['total_estudiantes'] = total_al
        data['destacados'] = []
        return jsonify(data)
    finally:
        conn.close()


@teacher_bp.route('/<slug>/actividades/list')
def actividades_list(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        acts = conn.execute(
            'SELECT id, nombre, tipo, peso, fecha_limite, estado_act FROM actividades '
            'WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        result = []
        alunos_count = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso, jornada)).fetchone()
        total_al = alunos_count['c'] if alunos_count else 0
        for a in acts:
            notas_count = conn.execute('SELECT COUNT(*) as c FROM notas WHERE actividad_id=? AND val IS NOT NULL', (a['id'],)).fetchone()
            graded = notas_count['c'] if notas_count else 0
            pending = total_al - graded
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id=? AND val IS NOT NULL', (a['id'],)).fetchall()]
            prom = round(sum(vals)/len(vals),2) if vals else None
            result.append({
                'id':a['id'],'nombre':a['nombre'],'tipo':a['tipo'],'peso':a['peso'],
                'fecha_limite':a['fecha_limite'],'estado_act':a['estado_act'],
                'promedio':prom,'pendientes':pending,'total_estudiantes':total_al
            })
        return jsonify({'actividades':result})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/actividades/masiva', methods=['POST'])
def actividades_masiva(slug):
    f = _fa()
    f.require_colegio(slug)
    if not f.validar_csrf():
        return jsonify({'error': 'CSRF inválido'}), 403
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    data = request.get_json(silent=True) or {}
    accion = data.get('accion')
    ids = data.get('ids', [])
    if not ids or not accion: return jsonify({'error':'Datos incompletos'}), 400
    conn = f.conectar(slug)
    try:
        placeholders = ','.join('?' for _ in ids)
        params = ids
        # Bloquear acciones sobre actividades de períodos cerrados.
        periodos = [r['periodo'] for r in conn.execute(
            f'SELECT DISTINCT periodo FROM actividades WHERE id IN ({placeholders}) AND profesor_id=?',
            params + [prof['id']]).fetchall() if r['periodo'] is not None]
        for p in periodos:
            if f.periodo_cerrado(slug, p):
                return jsonify({'status':'error','codigo':'PERIODO_CERRADO',
                                'mensaje':'El per\u00edodo est\u00e1 cerrado.'}), 403
        if accion == 'eliminar':
            for aid in ids:
                conn.execute('DELETE FROM entregas WHERE actividad_id=?', (aid,))
                conn.execute('DELETE FROM solicitudes_modificacion WHERE actividad_id=?', (aid,))
                conn.execute('DELETE FROM notas WHERE actividad_id=?', (aid,))
            conn.execute(f'DELETE FROM actividades WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'publicar':
            conn.execute(f'UPDATE actividades SET estado_act=\'publicada\' WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'cerrar':
            conn.execute(f'UPDATE actividades SET estado_act=\'cerrada\' WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'archivar':
            conn.execute(f'UPDATE actividades SET estado_act=\'archivada\' WHERE id IN ({placeholders}) AND profesor_id=?', params + [prof['id']])
        elif accion == 'duplicar':
            for aid in ids:
                act = conn.execute('SELECT * FROM actividades WHERE id=? AND profesor_id=?', (aid, prof['id'])).fetchone()
                if act:
                    old_notas = conn.execute('SELECT aid, val FROM notas WHERE actividad_id=?', (aid,)).fetchall()
                    orden = act['orden'] if act['orden'] is not None else 0
                    conn.execute(
                        'INSERT INTO actividades (nombre, tipo, peso, fecha_limite, estado_act, materia, jornada, curso, periodo, profesor_id, orden) '
                        'VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                        (act['nombre']+' (copia)', act['tipo'], act['peso'], act['fecha_limite'], 'borrador',
                         act['materia'], act['jornada'], act['curso'], act['periodo'], prof['id'], orden))
                    new_id = conn.execute('SELECT last_insert_rowid() as lid').fetchone()['lid']
                    for n in old_notas:
                        conn.execute('INSERT INTO notas (actividad_id, aid, val) VALUES (?,?,?)', (new_id, n['aid'], n['val']))
        elif accion == 'cambiar_peso':
            peso = data.get('peso')
            if peso is not None:
                conn.execute(f'UPDATE actividades SET peso=? WHERE id IN ({placeholders}) AND profesor_id=?', [peso] + params + [prof['id']])
        elif accion == 'cambiar_tipo':
            tipo = data.get('tipo')
            if tipo:
                conn.execute(f'UPDATE actividades SET tipo=? WHERE id IN ({placeholders}) AND profesor_id=?', [tipo] + params + [prof['id']])
        elif accion == 'cambiar_fecha':
            fecha = data.get('fecha')
            if fecha:
                conn.execute(f'UPDATE actividades SET fecha_limite=? WHERE id IN ({placeholders}) AND profesor_id=?', [fecha] + params + [prof['id']])
        conn.commit()
        try:
            f.audit_log(slug, prof['id'], 'actividad_'+accion, 'actividades', registro_id=None,
                      valor_anterior={'ids': ids, 'extra': {k: v for k, v in data.items() if k not in ('ids', 'accion')}},
                      valor_nuevo={'accion': accion, 'ids': ids})
        except Exception as e:
            logger.warning('[%s] audit masiva: %s', slug, e)
        return jsonify({'status':'ok'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/validar', methods=['POST'])
def curso_validar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    if not validar_csrf(): return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = data.get('curso', '')
        periodo = data.get('periodo', 1)
        notas_data = data.get('notas', {})
        warnings = []
        empty_count = 0
        for aid, acts in notas_data.items():
            for act_id, val in acts.items():
                if val is None or val == '':
                    empty_count += 1
        if empty_count > 0:
            warnings.append({'tipo':'warning','mensaje':f'{empty_count} nota(s) vac\u00eda(s) que se ignorar\u00e1n'})
        for aid, acts in notas_data.items():
            vals = [v for v in acts.values() if v is not None and v != '']
            if len(vals) > 1 and len(set(str(v) for v in vals)) == 1 and len(vals) > 2:
                warnings.append({'tipo':'info','mensaje':f'Estudiante #{aid}: todas las notas son iguales ({vals[0]})'})
                break
        acts = conn.execute(
            'SELECT id, nombre FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? AND (peso IS NULL OR peso=0)',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        for a in acts:
            warnings.append({'tipo':'warning','mensaje':f'Actividad "{a["nombre"]}" sin peso asignado'})
        total_peso = conn.execute(
            'SELECT COALESCE(SUM(peso),0) as s FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? AND peso IS NOT NULL',
            (materia, jornada, curso, periodo, prof['id'])).fetchone()
        if total_peso and total_peso['s'] > 100:
            warnings.append({'tipo':'warning','mensaje':f'El peso total de actividades es {total_peso["s"]}% (m\u00e1ximo 100%)'})
        return jsonify({'warnings':warnings})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/sugerencias')
def curso_sugerencias(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        alumnos = conn.execute('SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre', (curso, jornada)).fetchall()
        acts = conn.execute(
            'SELECT id, nombre, tipo FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        act_ids = [a['id'] for a in acts]
        sugerencias = []
        for al in alumnos:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL', act_ids+[al['id']]).fetchall()]
            prom = round(sum(vals)/len(vals),2) if vals else None
            if prom is not None and prom < 3.0:
                sugerencias.append({'tipo':'recuperacion','alumno_id':al['id'],'nombre':al['nombre'],'promedio':prom,'accion':'Necesita recuperaci\u00f3n o refuerzo'})
        for al in alumnos:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL ORDER BY actividad_id', act_ids+[al['id']]).fetchall()]
            if len(vals) >= 4:
                half = len(vals)//2
                first_half = sum(vals[:half])/half
                second_half = sum(vals[half:])/half
                diff = second_half - first_half
                if diff > 0.5:
                    sugerencias.append({'tipo':'mejora','alumno_id':al['id'],'nombre':al['nombre'],'diferencia':round(diff,2),'accion':'Ha mejorado su rendimiento, reconocerlo'})
                elif diff < -0.5:
                    sugerencias.append({'tipo':'empeoro','alumno_id':al['id'],'nombre':al['nombre'],'diferencia':round(diff,2),'accion':'Ha disminuido su rendimiento, revisar causa'})
        for a in acts:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id=? AND val IS NOT NULL', (a['id'],)).fetchall()]
            if vals:
                prom_act = sum(vals)/len(vals)
                if prom_act < 2.5:
                    sugerencias.append({'tipo':'dificil','actividad_id':a['id'],'nombre':a['nombre'],'promedio':round(prom_act,2),'accion':'Actividad demasiado dif\u00edcil, considerar ajuste'})
                elif prom_act > 4.5:
                    sugerencias.append({'tipo':'facil','actividad_id':a['id'],'nombre':a['nombre'],'promedio':round(prom_act,2),'accion':'Actividad demasiado f\u00e1cil, subir nivel'})
        for al in alumnos:
            vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL ORDER BY actividad_id', act_ids+[al['id']]).fetchall()]
            for i in range(1, len(vals)):
                if vals[i-1] - vals[i] >= 1.0:
                    sugerencias.append({'tipo':'observacion','alumno_id':al['id'],'nombre':al['nombre'],'accion':f'Repentina ca\u00edda de {vals[i-1]} a {vals[i]}, generar observaci\u00f3n'})
                    break
        return jsonify({'sugerencias':sugerencias})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/comparar')
def curso_comparar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso', (materia, jornada, prof['id'])).fetchall()
        result = []
        for c in cursos:
            curso_name = c['curso']
            alumnos = conn.execute('SELECT id, nombre FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre', (curso_name, jornada)).fetchall()
            acts = conn.execute('SELECT id, nombre FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=? ORDER BY orden', (materia, jornada, curso_name, prof['id'])).fetchall()
            act_ids = [a['id'] for a in acts]
            promedios = []
            total_vals = 0
            for al in alumnos:
                vals = [float(r['val']) for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL', act_ids+[al['id']]).fetchall()]
                if vals:
                    promedios.append(sum(vals)/len(vals))
                    total_vals += len(vals)
            prom_curso = round(sum(promedios)/len(promedios),2) if promedios else 0
            approved = sum(1 for p in promedios if p >= 3)
            at_risk = sum(1 for p in promedios if p < 2.8)
            result.append({
                'curso':curso_name,
                'estudiantes':len(alumnos),
                'actividades':len(acts),
                'promedio':prom_curso,
                'aprobados':approved,
                'en_riesgo':at_risk,
                'total_notas':total_vals
            })
        return jsonify({'cursos':result})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/timeline')
def curso_timeline(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        periodo = request.args.get('periodo', 1, type=int)
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        events = []
        logs = conn.execute(
            "SELECT creado, accion, tabla, registro_id FROM audit_log WHERE tabla='actividades' AND "
            "registro_id IN (SELECT id FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=?) "
            "ORDER BY creado DESC LIMIT 50",
            (materia, jornada, curso, periodo, prof['id'])).fetchall()
        for l in logs:
            act = conn.execute('SELECT nombre FROM actividades WHERE id=?', (l['registro_id'],)).fetchone()
            name = act['nombre'] if act else f'#{l["registro_id"]}'
            events.append({'tipo':'actividad_creada','titulo':f'Actividad "{name}" {l["accion"]}','fecha':l['creado'] or ''})
        notas_raw = conn.execute(
            'SELECT DISTINCT an.creado, an.actividad_id, ac.nombre as act_nombre FROM auditoria_notas an '
            'JOIN actividades ac ON ac.id=an.actividad_id '
            'WHERE ac.materia=? AND ac.jornada=? AND ac.curso=? AND ac.profesor_id=? AND an.creado IS NOT NULL '
            'ORDER BY an.creado DESC LIMIT 30',
            (materia, jornada, curso, prof['id'])).fetchall()
        for nr in notas_raw:
            events.append({'tipo':'notas_registradas','actividad_id':nr['actividad_id'],'titulo':f'Notas en "{nr["act_nombre"]}"','fecha':nr['creado'] or ''})
        events.sort(key=lambda e: e.get('fecha','') or '', reverse=True)
        return jsonify({'eventos':events[:30]})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/institucional/centro-control')
def institucional_centro_control(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        cursos = conn.execute(
            'SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso',
            (materia, jornada, prof['id'])).fetchall()
        total_estudiantes = 0
        total_actividades = 0
        cursos_data = []
        prom_sum = 0.0
        prom_count = 0
        for c in cursos:
            curso_name = c['curso']
            alumnos = conn.execute('SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso_name, jornada)).fetchall()
            total_estudiantes += len(alumnos)
            acts = conn.execute('SELECT id, nombre, estado_act FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?', (materia, jornada, curso_name, prof['id'])).fetchall()
            total_actividades += len(acts)
            act_ids = [a['id'] for a in acts]
            promedios = []
            riesgo_count = 0
            sin_notas_count = 0
            for al in alumnos:
                vals = []
                for act_id in act_ids:
                    r = conn.execute('SELECT val FROM notas WHERE aid=? AND actividad_id=?', (al['id'], act_id)).fetchone()
                    if r and r['val'] is not None: vals.append(float(r['val']))
                if vals:
                    p = sum(vals)/len(vals)
                    promedios.append(p)
                    if p < 2.8: riesgo_count += 1
                else:
                    sin_notas_count += 1
            prom_curso = round(sum(promedios)/len(promedios),2) if promedios else 0
            if promedios: prom_sum += prom_curso; prom_count += 1
            aprobados = sum(1 for p in promedios if p >= 3)
            inasistencia = conn.execute('SELECT COUNT(*) as c FROM asistencia a JOIN alumnos al ON al.id=a.aid WHERE al.curso=? AND al.jornada=? AND a.estado=?', (curso_name, jornada, 'A')).fetchone()
            cursos_data.append({
                'curso':curso_name,'estudiantes':len(alumnos),'actividades':len(acts),
                'promedio':prom_curso,'aprobados':aprobados,'riesgo':riesgo_count,
                'sin_notas':sin_notas_count,'inasistencia':inasistencia['c'] if inasistencia else 0
            })
        try:
            pend_comms = conn.execute("SELECT COUNT(*) as c FROM comunicaciones WHERE estado='pendiente'").fetchone()
        except Exception:
            logger.debug('dashboard_stats: tabla comunicaciones no encontrada')
            pend_comms = None
        try:
            hoy_eventos = conn.execute("SELECT COUNT(*) as c FROM eventos_calendario WHERE fecha_inicio=date('now')").fetchone()
        except Exception:
            logger.debug('dashboard_stats: tabla eventos_calendario no encontrada')
            hoy_eventos = None
        conn.execute("SELECT creado, accion FROM audit_log WHERE usuario_id=? ORDER BY creado DESC LIMIT 5", (prof['id'],)).fetchall()
        return jsonify({
            'promedio_institucional':round(prom_sum/prom_count,2) if prom_count else 0,
            'total_estudiantes':total_estudiantes,
            'total_actividades':total_actividades,
            'cursos':cursos_data,
            'comunicados_pendientes':pend_comms['c'] if pend_comms else 0,
            'eventos_hoy':hoy_eventos['c'] if hoy_eventos else 0,
            'actividades_sin_publicar':sum(1 for c in cursos_data for a in conn.execute('SELECT COUNT(*) as c FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=? AND estado_act=?', (materia, jornada, c['curso'], prof['id'], 'borrador')).fetchall() if a['c']>0),
            'alertas':{'total':sum(c['riesgo']+c['sin_notas'] for c in cursos_data),'riesgo':sum(c['riesgo'] for c in cursos_data),'sin_notas':sum(c['sin_notas'] for c in cursos_data)}
        })
    finally:
        conn.close()


@teacher_bp.route('/<slug>/smart-hub')
def smart_hub(slug):
    f = _fa()
    f.require_colegio(slug)
    user_type = session.get('user_type', 'profesor')
    user_id = session.get('user_id') or session.get('profesor_id')
    conn = f.conectar(slug)
    try:
        result = {'rol':user_type}
        if user_type == 'profesor':
            prof = f.get_profesor(slug)
            if not prof: return jsonify({'error':'No autorizado'}), 403
            jornada, materia = f.get_sesion_jornada_materia(slug)
            cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? ORDER BY curso', (materia, jornada, prof['id'])).fetchall()
            cursos_data = []
            total_riesgo = 0
            total_sin_notas = 0
            act_sin_publicar = 0
            for c in cursos:
                cn = c['curso']
                acts = conn.execute('SELECT id, estado_act FROM actividades WHERE materia=? AND jornada=? AND curso=? AND profesor_id=?', (materia, jornada, cn, prof['id'])).fetchall()
                act_ids = [a['id'] for a in acts]
                for a in acts:
                    if a['estado_act'] == 'borrador': act_sin_publicar += 1
                alumnos = conn.execute('SELECT id FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (cn, jornada)).fetchall()
                riesgo = 0
                sin_notas = 0
                for al in alumnos:
                    vals = [r['val'] for r in conn.execute('SELECT val FROM notas WHERE actividad_id IN ('+','.join('?'*len(act_ids))+') AND aid=? AND val IS NOT NULL', act_ids+[al['id']]).fetchall()]
                    if vals:
                        p = sum(vals)/len(vals)
                        if p < 2.8: riesgo += 1
                    else:
                        sin_notas += 1
                total_riesgo += riesgo
                total_sin_notas += sin_notas
                cursos_data.append({'curso':cn,'riesgo':riesgo,'sin_notas':sin_notas})
            result['cursos'] = cursos_data
            result['total_riesgo'] = total_riesgo
            result['total_sin_notas'] = total_sin_notas
            result['actividades_sin_publicar'] = act_sin_publicar
        elif user_type == 'rector' or user_type == 'directora':
            total_al = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE activo=1').fetchone()
            result['total_estudiantes'] = total_al['c'] if total_al else 0
            criticos = conn.execute("""
                SELECT a.curso, ROUND(AVG(CASE WHEN n.val IS NOT NULL THEN n.val ELSE 0 END),2) as prom,
                       COUNT(DISTINCT a.id) as estudiantes
                FROM alumnos a LEFT JOIN notas n ON n.aid=a.id
                WHERE a.activo=1
                GROUP BY a.curso ORDER BY prom ASC LIMIT 5
            """).fetchall()
            result['cursos_criticos'] = [{'curso':r['curso'],'promedio':r['prom'],'estudiantes':r['estudiantes']} for r in criticos]
            result['total_alertas'] = 0
        elif user_type == 'padre' or user_type == 'estudiante':
            aid = request.args.get('aid', type=int) or user_id
            al = conn.execute('SELECT id, nombre, curso FROM alumnos WHERE id=?', (aid,)).fetchone()
            if al:
                result['alumno'] = {'id':al['id'],'nombre':al['nombre'],'curso':al['curso']}
        return jsonify(result)
    finally:
        conn.close()


@teacher_bp.route('/<slug>/notas/pagina')
def notas_pagina(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    per_page = min(per_page, 100)
    offset = (page - 1) * per_page
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        curso = request.args.get('curso', '')
        if not curso: return jsonify({'error':'Curso requerido'}), 400
        count = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE curso=? AND jornada=? AND activo=1', (curso, jornada)).fetchone()
        total = count['c'] if count else 0
        alumnos = conn.execute(
            'SELECT id, nombre, num_curso FROM alumnos WHERE curso=? AND jornada=? AND activo=1 ORDER BY nombre LIMIT ? OFFSET ?',
            (curso, jornada, per_page, offset)).fetchall()
        acts = conn.execute(
            'SELECT id, nombre, tipo, peso, estado_act, fecha_limite FROM actividades WHERE materia=? AND jornada=? AND curso=? AND COALESCE(periodo,1)=? AND profesor_id=? ORDER BY orden',
            (materia, jornada, curso, request.args.get('periodo',1,type=int), prof['id'])).fetchall()
        act_ids = [a['id'] for a in acts]
        notas_map = {}
        rows = []
        if alumnos and act_ids:
            placeholders = ','.join('?'*len(act_ids))
            aids = [al['id'] for al in alumnos]
            rows = conn.execute(
                f'SELECT aid, actividad_id, val FROM notas WHERE aid IN ({",".join("?"*len(aids))}) AND actividad_id IN ({placeholders})',
                aids + act_ids).fetchall()
            for r in rows:
                key = (r['aid'], r['actividad_id'])
                notas_map[key] = r['val']
        return jsonify({
            'page':page,'per_page':per_page,'total':total,'total_pages':max(1,-(-total//per_page)),
            'actividades':[{'id':a['id'],'nombre':a['nombre'],'tipo':a['tipo'],'peso':a['peso'],'estado_act':a['estado_act'],'fecha_limite':a['fecha_limite']} for a in acts],
            'alumnos':[{'id':a['id'],'nombre':a['nombre'],'num_curso':a['num_curso']} for a in alumnos],
            'notas':{f"{r['aid']}_{r['actividad_id']}":{'aid':r['aid'],'actividad_id':r['actividad_id'],'val':r['val']} for r in rows}
        })
    finally:
        conn.close()


@teacher_bp.route('/<slug>/analitica/comparar')
def analitica_comparar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof: return jsonify({'error':'No autorizado'}), 403
    tipo = request.args.get('tipo', 'periodos')
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        cfg = f.config_get(slug)
        escala_max = float(cfg.get('escala_max', 5.0)) if cfg else 5.0
        result = {'tipo':tipo,'datos':[],'escala_max':escala_max}
        if tipo == 'periodos':
            for p in [1,2,3,4]:
                acts = conn.execute(
                    'SELECT id FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? AND COALESCE(periodo,1)=?',
                    (materia, jornada, prof['id'], p)).fetchall()
                act_ids = [a['id'] for a in acts]
                promedios = []
                if act_ids:
                    rows = conn.execute(
                        'SELECT n.val FROM notas n WHERE n.actividad_id IN ('+','.join('?'*len(act_ids))+') AND n.val IS NOT NULL',
                        act_ids).fetchall()
                    vals = [float(r['val']) for r in rows]
                    if vals: promedios.append(sum(vals)/len(vals))
                result['datos'].append({'periodo':p,'promedio':round(sum(promedios)/len(promedios),2) if promedios else 0,'actividades':len(acts)})
        elif tipo == 'docentes':
            profs = conn.execute('SELECT DISTINCT p.id, p.nombre FROM profesores p JOIN actividades a ON a.profesor_id=p.id WHERE a.materia=? AND a.jornada=?', (materia, jornada)).fetchall()
            for p in profs:
                vals = [float(r['val']) for r in conn.execute(
                    'SELECT n.val FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE a.profesor_id=? AND a.materia=? AND a.jornada=? AND n.val IS NOT NULL',
                    (p['id'], materia, jornada)).fetchall()]
                prom = round(sum(vals)/len(vals),2) if vals else 0
                result['datos'].append({'nombre':p['nombre'],'promedio':prom,'notas':len(vals)})
        return jsonify(result)
    finally:
        conn.close()


# ── Dashboard ─────────────────────────────────────────────────────────────
@teacher_bp.route('/<slug>/dashboard')
def dashboard(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    rector = f.get_rector(slug)
    if not prof and not rector:
        return redirect(url_for('auth.login', slug=slug))
    colegio = f.get_colegio(slug)
    num_periodos = colegio['num_periodos'] if colegio and 'num_periodos' in colegio.keys() else 4
    conn = f.conectar(slug)
    materias_list = []
    hoy_schedule = []
    comunicados_recientes = []
    notif_count = 0
    asistencia_stats = None
    if prof:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia or '', jornada or '')
        instance = 'profesor'
        nombre = prof['nombre']
        dias_semana = ['Lun', 'Mar', 'Mi\u00e9', 'Jue', 'Vie', 'S\u00e1b', 'Dom']
        hoy_idx = datetime.now().weekday()
        hoy_dia = dias_semana[hoy_idx]
        for curso in mis_cursos[:3]:
            filas = conn.execute(
                'SELECT dia, franja, num, materia FROM horarios_curso WHERE curso=? AND jornada=? AND dia=? ORDER BY franja',
                (curso, jornada, hoy_dia)).fetchall()
            for r in filas:
                hoy_schedule.append({'curso': curso, 'dia': r['dia'], 'franja': r['franja'], 'num': r['num'], 'materia': r['materia']})
        fechas = conn.execute(
            'SELECT c.id, c.titulo, c.fecha_publicacion, r.nombre as creado_por FROM comunicaciones c JOIN rectores r ON r.id=c.rector_id WHERE c.activo=1 AND c.estado=\'publicado\' ORDER BY c.fecha_publicacion DESC LIMIT 5').fetchall()
        comunicados_recientes = [dict(r) for r in fechas]
        notif_count = get_notificaciones_no_leidas_count(conn, 'profesor', prof['id'])
        if mis_cursos:
            asistencia_stats = _asistencia_stats(conn, curso=mis_cursos[0], jornada=jornada)
    elif rector:
        jornada = ''
        materia = ''
        mis_cursos = [r['curso'] for r in conn.execute(
            'SELECT DISTINCT curso FROM alumnos WHERE activo=1 ORDER BY curso').fetchall()]
        materias_list = [r['materia'] for r in conn.execute(
            'SELECT DISTINCT materia FROM asignaciones_materia ORDER BY materia').fetchall()]
        instance = 'rector'
        nombre = rector['nombre']
        notif_count = get_notificaciones_no_leidas_count(conn, 'rector', rector['id'])
    conn.close()
    colegio_dash = f.get_colegio(slug)
    dias_es = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
    meses_es = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre']
    ahora = datetime.now()
    hoy_str = f"{dias_es[ahora.weekday()]}, {ahora.day} de {meses_es[ahora.month-1]} de {ahora.year}"
    return render_template('dashboard.html', slug=slug, colegio=colegio_dash, instance=instance, nombre=nombre,
                           num_periodos=num_periodos, mis_cursos=mis_cursos,
                           materias_list=materias_list if rector else [materia] if materia else [],
                           jornada=jornada, materia=materia,
                           hoy_schedule=hoy_schedule, comunicados_recientes=comunicados_recientes,
                           notif_count=notif_count, asistencia_stats=asistencia_stats,
                           hoy_str=hoy_str)


@teacher_bp.route('/<slug>/dashboard_data')
def dashboard_data(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    rector = f.get_rector(slug)
    if not prof and not rector:
        return jsonify({'error': 'no_auth'}), 401
    curso = request.args.get('curso') or None
    materia = request.args.get('materia') or None
    jornada_sel = request.args.get('jornada') or None
    periodo = request.args.get('periodo', type=int) or None
    conn = f.conectar(slug)
    try:
        if prof:
            sess_jornada, sess_materia = f.get_sesion_jornada_materia(slug)
            m = materia or sess_materia or ''
            j = jornada_sel or sess_jornada or ''
            data = f._dashboard_profesor_data(conn, slug, prof, curso, m, j, periodo)
        else:
            data = f._dashboard_rector_data(conn, slug, rector)
    finally:
        conn.close()
    return jsonify(data)


@teacher_bp.route('/<slug>/nuevo_trabajo', methods=['POST'])
def nuevo_trabajo(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf():
        return ('Error CSRF', 403)
    jornada, materia = f.get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso_sel', '')
    conn = f.conectar(slug)
    from app.repositories.student_repository import create_compromiso
    create_compromiso(conn, request.form.get('titulo'), request.form.get('fecha'), materia, curso_sel, jornada)
    conn.commit()
    conn.close()
    return redirect(url_for('teacher.home', slug=slug, curso=curso_sel))


@teacher_bp.route('/<slug>/borrar_trabajo/<int:id_t>', methods=['POST'])
def borrar_trabajo(slug, id_t):
    f = _fa()
    if not validar_csrf():
        return redirect(url_for('teacher.home', slug=slug))
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    conn = f.conectar(slug)
    from app.repositories.student_repository import delete_compromiso, get_compromiso_curso
    c = get_compromiso_curso(conn, id_t)
    curso = c['curso'] if c else ''
    delete_compromiso(conn, id_t, materia)
    conn.commit()
    conn.close()
    return redirect(url_for('teacher.home', slug=slug, curso=curso))


@teacher_bp.route('/<slug>/registrar', methods=['POST'])
def registrar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    if not validar_csrf():
        return ('Error CSRF', 403)
    jornada, materia = f.get_sesion_jornada_materia(slug)
    nom = request.form.get('nombre', '').strip()
    cur = request.form.get('curso', '').strip()
    curso_sel = request.form.get('curso_sel', cur)
    if nom and cur and jornada:
        conn = f.conectar(slug)
        with conn:
            from app.services.student_service import register_student
            register_student(conn, nom, cur, jornada)
            f.audit_log(slug, prof['id'], 'crear', 'alumnos')
        conn.close()
    return redirect(url_for('teacher.home', slug=slug, curso=curso_sel))


@teacher_bp.route('/<slug>/archivar_alumno/<int:id>', methods=['POST'])
def archivar_alumno(slug, id):
    f = _fa()
    if not validar_csrf():
        return redirect(url_for('teacher.home', slug=slug))
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso', '')
    conn = f.conectar(slug)
    from app.repositories.student_repository import get_student_curso
    alumno = get_student_curso(conn, id)
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not alumno or alumno['curso'] not in mis_cursos:
        conn.close()
        return ('No autorizado', 403)
    from app.services.student_service import archive_student_action
    archive_student_action(conn, id)
    f.audit_log(slug, prof['id'], 'archivar', 'alumnos', id)
    conn.commit()
    conn.close()
    return redirect(url_for('teacher.home', slug=slug, curso=curso_sel))


@teacher_bp.route('/<slug>/reactivar_alumno/<int:id>', methods=['POST'])
def reactivar_alumno(slug, id):
    f = _fa()
    if not validar_csrf():
        return redirect(url_for('teacher.archivados', slug=slug))
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso', '')
    conn = f.conectar(slug)
    from app.repositories.student_repository import get_student_curso
    alumno = get_student_curso(conn, id)
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not alumno or alumno['curso'] not in mis_cursos:
        conn.close()
        return ('No autorizado', 403)
    from app.services.student_service import reactivate_student_action
    reactivate_student_action(conn, id)
    f.audit_log(slug, prof['id'], 'reactivar', 'alumnos', id)
    conn.commit()
    conn.close()
    return redirect(url_for('teacher.archivados', slug=slug, curso=curso_sel))


@teacher_bp.route('/<slug>/eliminar_alumno/<int:id>', methods=['POST'])
def eliminar_alumno(slug, id):
    f = _fa()
    if not validar_csrf():
        return redirect(url_for('teacher.archivados', slug=slug))
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    curso_sel = request.form.get('curso', '')
    conn = f.conectar(slug)
    from app.repositories.student_repository import get_student_curso
    alumno = get_student_curso(conn, id)
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    if not alumno or alumno['curso'] not in mis_cursos:
        conn.close()
        return ('No autorizado', 403)
    from app.services.student_service import delete_student_action
    delete_student_action(conn, id)
    f.audit_log(slug, prof['id'], 'eliminar', 'alumnos', id)
    conn.commit()
    conn.close()
    return redirect(url_for('teacher.archivados', slug=slug, curso=curso_sel))


@teacher_bp.route('/<slug>/archivados')
def archivados(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return redirect(url_for('auth.login', slug=slug))
    jornada, materia = f.get_sesion_jornada_materia(slug)
    colegio = f.get_colegio(slug)
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada)
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else None)
    conn = f.conectar(slug)
    from app.services.student_service import build_archivados_context
    ctx = build_archivados_context(conn, jornada, mis_cursos, curso_sel)
    conn.close()
    return render_template('archivados.html',
                           slug=slug, colegio=colegio, profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel,
                           alumnos_archivados=ctx['alumnos_arch'],
                           profesores_archivados=ctx['profs_arch'],
                           profesores_activos=ctx['profesores_activos'])


@teacher_bp.route('/<slug>/archivar_profesor/<int:id>', methods=['POST'])
def archivar_profesor(slug, id):
    f = _fa()
    if not validar_csrf():
        return jsonify({'ok': False, 'mensaje': 'Error CSRF'}), 403
    f.require_colegio(slug)
    rector = f.get_rector(slug)
    if not rector:
        return jsonify({'ok': False, 'mensaje': 'Solo el rector puede archivar profesores'}), 403
    conn = f.conectar(slug)
    from app.repositories.student_repository import archive_profesor
    archive_profesor(conn, id)
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@teacher_bp.route('/<slug>/archivar_profesor_con_reasignacion', methods=['POST'])
def archivar_profesor_con_reasignacion(slug):
    f = _fa()
    f.require_colegio(slug)
    rector = f.get_rector(slug)
    if not rector:
        return jsonify({'ok': False, 'mensaje': 'Solo el rector puede archivar profesores'})
    if not validar_csrf():
        return jsonify({'ok': False, 'mensaje': 'Error CSRF'})
    profesor_id = request.form.get('profesor_id', type=int)
    prof_destino_id = request.form.get('prof_destino_id', type=int)
    cursos_reasignar = request.form.getlist('cursos_reasignar')
    if not profesor_id:
        return jsonify({'ok': False, 'mensaje': 'Datos incompletos.'})
    conn = f.conectar(slug)
    try:
        from app.repositories.student_repository import (
            archive_profesor,
            copy_asignacion_curso,
            ensure_asignacion_materia,
            reasignar_actividades,
            reasignar_evaluaciones,
            remove_asignacion_curso,
        )
        if prof_destino_id and cursos_reasignar:
            for item in cursos_reasignar:
                partes = item.split('|')
                if len(partes) != 3:
                    continue
                curso, mat, jor = partes
                reasignar_actividades(conn, profesor_id, prof_destino_id, curso, mat, jor)
                reasignar_evaluaciones(conn, profesor_id, prof_destino_id, curso, jor)
                copy_asignacion_curso(conn, prof_destino_id, mat, jor, curso)
                remove_asignacion_curso(conn, profesor_id, mat, jor, curso)
                ensure_asignacion_materia(conn, prof_destino_id, mat, jor)
        archive_profesor(conn, profesor_id)
        conn.commit()
        return jsonify({'ok': True, 'mensaje': 'Profesor archivado correctamente.'})
    except Exception as e:
        conn.rollback()
        logger.error(f'Error al archivar profesor {profesor_id} en {slug}: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al archivar. Intenta de nuevo.'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/reactivar_profesor/<int:id>', methods=['POST'])
def reactivar_profesor(slug, id):
    f = _fa()
    if not validar_csrf():
        return jsonify({'ok': False, 'mensaje': 'Error CSRF'}), 403
    f.require_colegio(slug)
    rector = f.get_rector(slug)
    if not rector:
        return jsonify({'ok': False, 'mensaje': 'Solo el rector puede reactivar profesores'}), 403
    conn = f.conectar(slug)
    from app.repositories.student_repository import reactivate_profesor
    reactivate_profesor(conn, id)
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@teacher_bp.route('/<slug>/eliminar_profesor/<int:id>', methods=['POST'])
def eliminar_profesor(slug, id):
    f = _fa()
    if not validar_csrf():
        return jsonify({'ok': False, 'mensaje': 'Error CSRF'}), 403
    f.require_colegio(slug)
    rector = f.get_rector(slug)
    if not rector:
        return jsonify({'ok': False, 'mensaje': 'Solo el rector puede eliminar profesores'}), 403
    conn = f.conectar(slug)
    from app.repositories.student_repository import delete_profesor
    delete_profesor(conn, id)
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@teacher_bp.route('/<slug>/comunicados')
def comunicados_list(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        comms = conn.execute(
            'SELECT id, titulo, contenido, destinatario_tipo, destinatario_valor, fecha_programada, fecha_creacion, estado FROM comunicaciones ORDER BY fecha_creacion DESC LIMIT 50',
            ()).fetchall()
        result = []
        for c in comms:
            result.append({
                'id': c['id'], 'titulo': c['titulo'], 'contenido': c['contenido'],
                'destinatario_tipo': c['destinatario_tipo'], 'destinatario_valor': c['destinatario_valor'],
                'fecha_programada': c['fecha_programada'], 'creado': c['fecha_creacion'],
                'estado': c['estado']
            })
        return jsonify({'comunicados': result})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/comunicados/crear', methods=['POST'])
def comunicados_crear(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'CSRF inválido'}), 403
    data = request.get_json(silent=True) or {}
    titulo = data.get('titulo', '')
    contenido = data.get('contenido', '')
    destinatario_tipo = data.get('destinatario_tipo', 'todos')
    destinatario_valor = data.get('destinatario_valor', '')
    programada = data.get('programada')
    if not titulo:
        return jsonify({'error': 'Titulo requerido'}), 400
    conn = f.conectar(slug)
    try:
        conn.execute(
            "INSERT INTO comunicaciones (rector_id, titulo, contenido, destinatario_tipo, destinatario_valor, fecha_programada, estado, fecha_creacion) VALUES (?,?,?,?,?,?,'publicado',datetime('now'))",
            (prof['id'], titulo, contenido, destinatario_tipo, destinatario_valor, programada))
        conn.commit()
        return jsonify({'status': 'ok'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/comunicados/<int:cid>/leer', methods=['POST'])
def comunicados_leer(slug, cid):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    conn = f.conectar(slug)
    try:
        conn.execute("UPDATE comunicaciones SET estado='leido' WHERE id=?", (cid,))
        conn.commit()
        return jsonify({'status': 'ok'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/calendario')
def calendario_list(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        jornada, materia = f.get_sesion_jornada_materia(slug)
        events = []
        acts = conn.execute(
            'SELECT id, nombre, fecha_limite, tipo, curso FROM actividades WHERE materia=? AND jornada=? AND profesor_id=? AND fecha_limite IS NOT NULL ORDER BY fecha_limite',
            (materia, jornada, prof['id'])).fetchall()
        for a in acts:
            events.append({'id': a['id'], 'titulo': a['nombre'], 'fecha': a['fecha_limite'][:10], 'tipo': 'evaluacion', 'curso': a['curso']})
        evs = conn.execute('SELECT id, titulo, fecha_inicio, tipo FROM eventos_calendario ORDER BY fecha_inicio', ()).fetchall()
        for e in evs:
            events.append({'id': e['id'], 'titulo': e['titulo'], 'fecha': e['fecha_inicio'], 'tipo': e['tipo'] or 'evento', 'curso': ''})
        events.sort(key=lambda e: e.get('fecha', ''))
        return jsonify({'eventos': events})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/estudiante/<int:aid>/expediente')
def estudiante_expediente(slug, aid):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        al = conn.execute('SELECT * FROM alumnos WHERE id=?', (aid,)).fetchone()
        if not al:
            return jsonify({'error': 'No encontrado'}), 404
        notas = conn.execute(
            'SELECT ac.nombre as actividad, ac.tipo, n.val, ac.fecha_limite FROM notas n JOIN actividades ac ON ac.id=n.actividad_id WHERE n.aid=? ORDER BY ac.fecha_limite',
            (aid,)).fetchall()
        asistencias = conn.execute('SELECT fecha, estado FROM asistencia WHERE aid=? ORDER BY fecha DESC LIMIT 30', (aid,)).fetchall()
        observaciones = conn.execute('SELECT texto, fecha FROM observaciones WHERE aid=? ORDER BY fecha DESC LIMIT 20', (aid,)).fetchall()
        sanciones = conn.execute("SELECT texto, fecha, tipo FROM observador_registros WHERE aid=? AND tipo IN ('llamado','compromiso') ORDER BY fecha DESC LIMIT 10", (aid,)).fetchall()
        reconocimientos = conn.execute("SELECT texto, fecha, tipo FROM observador_registros WHERE aid=? AND tipo='positivo' ORDER BY fecha DESC LIMIT 10", (aid,)).fetchall()
        return jsonify({
            'alumno': {'id': al['id'], 'nombre': al['nombre'], 'curso': al['curso'], 'email_acudiente': al['email_acudiente']},
            'notas': [{'actividad': n['actividad'], 'tipo': n['tipo'], 'val': n['val'], 'fecha': n['fecha_limite']} for n in notas],
            'asistencias': [{'fecha': a['fecha'], 'presente': a['estado'] == 'P'} for a in asistencias],
            'observaciones': [{'texto': o['texto'], 'fecha': o['fecha']} for o in observaciones],
            'sanciones': [{'texto': s['texto'], 'fecha': s['fecha'], 'tipo': s['tipo']} for s in sanciones] if sanciones else [],
            'reconocimientos': [{'texto': r['texto'], 'fecha': r['fecha'], 'tipo': r['tipo']} for r in reconocimientos] if reconocimientos else []
        })
    finally:
        conn.close()


@teacher_bp.route('/<slug>/auditoria')
def auditoria_list(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 30, type=int)
    offset = (page - 1) * per_page
    conn = f.conectar(slug)
    try:
        count = conn.execute('SELECT COUNT(*) as c FROM audit_log WHERE usuario_id=?', (prof['id'],)).fetchone()
        total = count['c'] if count else 0
        logs = conn.execute(
            'SELECT id, accion, tabla, registro_id, valor_anterior, valor_nuevo, ip, creado FROM audit_log WHERE usuario_id=? ORDER BY creado DESC LIMIT ? OFFSET ?',
            (prof['id'], per_page, offset)).fetchall()
        return jsonify({
            'total': total, 'page': page, 'per_page': per_page,
            'items': [{'id': l['id'], 'accion': l['accion'], 'tabla': l['tabla'], 'registro_id': l['registro_id'],
                       'valor_anterior': l['valor_anterior'], 'valor_nuevo': l['valor_nuevo'], 'ip': l['ip'], 'creado': l['creado']} for l in logs]
        })
    finally:
        conn.close()


@teacher_bp.route('/<slug>/config', methods=['GET', 'POST'])
def school_config(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if request.method == 'POST' and not f.validar_csrf():
        return jsonify({'error': 'CSRF inválido'}), 403
    conn = f.conectar(slug)
    try:
        if request.method == 'GET':
            conf = conn.execute('SELECT * FROM config_institucion WHERE slug=?', (slug,)).fetchone()
            if conf:
                conf = dict(conf)
            return jsonify({
                'nombre': conf.get('nombre_institucion', '') if conf else '',
                'lema': conf.get('lema', '') if conf else '',
                'logo': conf.get('logo', '') if conf else '',
                'primary_color': conf.get('primary_color', '#3b82f6') if conf else '#3b82f6',
                'secondary_color': conf.get('secondary_color', '#6366f1') if conf else '#6366f1'
            })
        else:
            data = request.get_json(silent=True) or {}
            if conf := conn.execute('SELECT id FROM config_institucion WHERE slug=?', (slug,)).fetchone():
                conn.execute('UPDATE config_institucion SET nombre_institucion=?, lema=?, logo=?, primary_color=?, secondary_color=? WHERE id=?',
                             (data.get('nombre', ''), data.get('lema', ''), data.get('logo', ''), data.get('primary_color', '#3b82f6'), data.get('secondary_color', '#6366f1'), conf['id']))
            else:
                conn.execute('INSERT INTO config_institucion (slug, nombre_institucion, lema, logo, primary_color, secondary_color) VALUES (?,?,?,?,?,?)',
                             (slug, data.get('nombre', ''), data.get('lema', ''), data.get('logo', ''), data.get('primary_color', '#3b82f6'), data.get('secondary_color', '#6366f1')))
            conn.commit()
            return jsonify({'status': 'ok'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/plantillas')
def plantillas_list(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    conn = f.conectar(slug)
    try:
        from app.services.template_service import list_templates
        return jsonify({'plantillas': list_templates(conn, prof['id'])})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/plantillas/crear', methods=['POST'])
def plantillas_crear(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    nombre = data.get('nombre', '')
    tipo = data.get('tipo', 'tarea')
    peso = data.get('peso', 10)
    descripcion = data.get('descripcion', '')
    conn = f.conectar(slug)
    try:
        from app.services.template_service import create
        ok, err = create(conn, prof['id'], nombre, tipo, peso, descripcion)
        if not ok:
            return jsonify({'error': err}), 400
        return jsonify({'status': 'ok'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/plantillas/aplicar', methods=['POST'])
def plantillas_aplicar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    tmpl_id = data.get('plantilla_id')
    curso = data.get('curso', '')
    materia = data.get('materia', '')
    if not tmpl_id or not curso:
        return jsonify({'error': 'Datos incompletos'}), 400
    conn = f.conectar(slug)
    try:
        from app.services.template_service import apply_template
        ses_jornada, ses_materia = f.get_sesion_jornada_materia(slug)
        jornada = data.get('jornada') or ses_jornada or 'ma\u00f1ana'
        if not materia:
            materia = ses_materia
        periodo = data.get('periodo', 1)
        ok, err = apply_template(conn, prof['id'], tmpl_id, curso, materia, jornada, periodo)
        if not ok:
            return jsonify({'error': err}), 404
        return jsonify({'status': 'ok'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/plantillas/eliminar/<int:tid>', methods=['POST'])
def plantillas_eliminar(slug, tid):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    conn = f.conectar(slug)
    try:
        from app.services.template_service import delete
        delete(conn, tid, prof['id'])
        return jsonify({'status': 'ok'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/planificacion/copiar', methods=['POST'])
def planificacion_copiar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    origen_curso = data.get('origen_curso')
    destino_cursos = data.get('destino_cursos', [])
    if not origen_curso or not destino_cursos:
        return jsonify({'error': 'Datos incompletos'}), 400
    conn = f.conectar(slug)
    try:
        from app.services.planning_service import copy_planning
        jornada, materia = f.get_sesion_jornada_materia(slug)
        count = copy_planning(conn, prof['id'], materia, jornada, origen_curso, destino_cursos)
        return jsonify({'status': 'ok', 'copiadas': count})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/migrar/previsualizar', methods=['POST'])
def migrar_previsualizar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    contenido = data.get('contenido', '')
    tipo = data.get('tipo', 'estudiantes')
    if not contenido:
        return jsonify({'error': 'Contenido requerido'}), 400
    try:
        from app.services.migration_service import preview_migration
        headers, rows, sugg, total = preview_migration(contenido, tipo)
        return jsonify({'headers': headers, 'filas': rows, 'sugerencias': sugg, 'total': total})
    except Exception as e:
        logger.error('preview_migracion: %s', e)
        return jsonify({'error': str(e)}), 400


@teacher_bp.route('/<slug>/migrar/ejecutar', methods=['POST'])
def migrar_ejecutar(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    contenido = data.get('contenido', '')
    tipo = data.get('tipo', 'estudiantes')
    mapeo = data.get('mapeo', {})
    if not contenido:
        return jsonify({'error': 'Contenido requerido'}), 400
    conn = f.conectar(slug)
    try:
        from app.services.migration_service import execute_migration
        jornada, materia = f.get_sesion_jornada_materia(slug)
        count = execute_migration(conn, contenido, tipo, mapeo, prof['id'], materia, jornada)
        return jsonify({'status': 'ok', 'importados': count})
    except Exception as e:
        logger.error('ejecutar_migracion: %s', e)
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']


@teacher_bp.route('/<slug>/horarios', methods=['GET', 'POST'])
def horarios(slug):
    f = _fa()
    f.require_colegio(slug)
    if not session.get(f'rol_{slug}'):
        return redirect(url_for('auth.login', slug=slug))
    prof = f.get_profesor(slug)
    colegio = f.get_colegio(slug)
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if prof and (not jornada or not materia):
        return redirect(url_for('teacher.seleccionar_jornada', slug=slug))
    mis_cursos = f.get_cursos_profesor(slug, prof['id'], materia, jornada) if prof else []
    curso_sel = request.args.get('curso', mis_cursos[0] if mis_cursos else None)

    if request.method == 'POST':
        if not validar_csrf():
            return redirect(url_for('teacher.horarios', slug=slug))
        c = f.conectar(slug)
        try:
            dia = request.form.get('dia', '')
            franja = request.form.get('franja', '')
            num = request.form.get('num', '').strip()
            mat = request.form.get('materia', '').strip()
            profesor = request.form.get('profesor', '').strip()
            curso_p = request.form.get('curso', curso_sel)
            if mat or profesor:
                c.execute(
                    '''INSERT INTO horarios_curso (curso,jornada,dia,franja,num,materia,profesor)
                       VALUES (?,?,?,?,?,?,?)
                       ON CONFLICT(curso,jornada,dia,franja) DO UPDATE SET
                           num=excluded.num, materia=excluded.materia, profesor=excluded.profesor''',
                    (curso_p, jornada, dia, franja, num, mat, profesor))
            else:
                c.execute(
                    'DELETE FROM horarios_curso WHERE curso=? AND jornada=? AND dia=? AND franja=?',
                    (curso_p, jornada, dia, franja))
            c.commit()
        finally:
            c.close()
        return ('', 204)
    c = f.conectar(slug)
    filas = []
    if curso_sel:
        filas = c.execute(
            'SELECT dia, franja, num, materia, profesor FROM horarios_curso WHERE curso=? AND jornada=?',
            (curso_sel, jornada)).fetchall()
    c.close()
    horario_map = {
        f"{r['dia']}_{r['franja']}": {'num': r['num'], 'materia': r['materia'], 'profesor': r['profesor']}
        for r in filas}
    return render_template('horarios.html', slug=slug, colegio=colegio, profesor=prof,
                           mis_cursos=mis_cursos, curso_sel=curso_sel, dias=DIAS_SEMANA,
                           horario_map=horario_map, jornada=jornada)


@teacher_bp.route('/<slug>/home')
def home_dashboard(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    rector = f.get_rector(slug)
    pid = session.get(f'padre_id_{slug}')
    aid = session.get(f'alumno_id_{slug}')
    conn = f.conectar(slug)
    try:
        if rector:
            total_est = conn.execute('SELECT COUNT(*) as c FROM alumnos WHERE activo=1').fetchone()['c']
            total_prof = conn.execute('SELECT COUNT(*) as c FROM profesores WHERE activo=1').fetchone()['c']
            riesgo = conn.execute('SELECT COUNT(DISTINCT n.aid) as c FROM notas n JOIN actividades a ON a.id=n.actividad_id JOIN alumnos al ON al.id=n.aid WHERE al.activo=1 GROUP BY n.aid HAVING AVG(n.val) < 3.0').fetchall()
            pendientes = conn.execute('SELECT COUNT(*) as c FROM solicitudes_modificacion WHERE estado="pendiente"').fetchone()['c']
            cal_events = conn.execute("SELECT COUNT(*) as c FROM eventos_calendario WHERE DATE(fecha)>=DATE('now') AND DATE(fecha)<=DATE('now','+7 days')").fetchone()['c']
            return jsonify({'rol': 'rector', 'total_estudiantes': total_est, 'total_profesores': total_prof, 'riesgo': len(riesgo), 'pendientes': pendientes, 'eventos': cal_events})
        elif prof:
            cursos = conn.execute('SELECT DISTINCT curso FROM actividades WHERE profesor_id=?', (prof['id'],)).fetchall()
            acts_pend = conn.execute("SELECT COUNT(*) as c FROM actividades WHERE profesor_id=? AND estado_act='borrador'", (prof['id'],)).fetchone()['c']
            riesgo = conn.execute('SELECT COUNT(DISTINCT n.aid) as c FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE a.profesor_id=? GROUP BY n.aid HAVING AVG(n.val) < 3.0', (prof['id'],)).fetchall()
            return jsonify({'rol': 'docente', 'nombre': prof['nombre'], 'cursos': len(cursos), 'actividades_pendientes': acts_pend, 'riesgo': len(riesgo)})
        elif pid:
            hijos = conn.execute('SELECT a.id, a.nombre, a.curso FROM alumno_padre ap JOIN alumnos a ON a.id=ap.alumno_id WHERE ap.padre_id=?', (pid,)).fetchall()
            hijos_data = []
            for h in hijos:
                prom = conn.execute('SELECT COALESCE(AVG(n.val),0) as p FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.aid=?', (h['id'],)).fetchone()
                hijos_data.append({'id': h['id'], 'nombre': h['nombre'], 'curso': h['curso'], 'promedio': round(prom['p'], 2)})
            return jsonify({'rol': 'padre', 'hijos': hijos_data})
        elif aid:
            al = conn.execute('SELECT * FROM alumnos WHERE id=?', (aid,)).fetchone()
            return jsonify({'rol': 'estudiante', 'nombre': al['nombre'], 'curso': al['curso']})
        else:
            return jsonify({'rol': 'anonimo'})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/ai/ask', methods=['POST'])
def ai_ask(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    rector = f.get_rector(slug)
    if not prof and not rector:
        return jsonify({'error': 'No autorizado'}), 403
    if not validar_csrf():
        return jsonify({'error': 'Error CSRF'}), 403
    data = request.get_json(silent=True) or {}
    pregunta = data.get('pregunta', '').lower().strip()
    if not pregunta:
        return jsonify({'error': 'Pregunta requerida'}), 400
    conn = f.conectar(slug)
    try:
        respuesta = {'pregunta': pregunta, 'respuesta': '', 'datos': []}
        if any(p in pregunta for p in ['riesgo', 'perder', 'reprobar', 'bajo rendimiento', 'recuperacion']):
            curso = None
            for c in ['grado ', 'curso ']:
                idx = pregunta.find(c)
                if idx >= 0:
                    curso = pregunta[idx + len(c):].split()[0] if pregunta[idx + len(c):].split() else None
                    break
            q = 'SELECT a.id, a.nombre, a.curso, ROUND(AVG(n.val),2) as prom FROM alumnos a JOIN notas n ON n.aid=a.id WHERE a.activo=1'
            params = []
            if curso:
                q += ' AND a.curso=?'
                params.append(curso)
            q += ' GROUP BY a.id HAVING prom < 3.0 ORDER BY prom ASC LIMIT 20'
            rows = conn.execute(q, params).fetchall()
            respuesta['respuesta'] = f'Se encontraron {len(rows)} estudiantes con riesgo académico (prom<3.0){" en " + curso if curso else ""}.'
            respuesta['datos'] = [{'nombre': r['nombre'], 'curso': r['curso'], 'promedio': r['prom']} for r in rows]
        elif any(p in pregunta for p in ['materia baja', 'promedio bajo', 'peor materia', 'materia más baja', 'nota baja']):
            rows = conn.execute('SELECT a.materia, ROUND(AVG(n.val),2) as prom FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY a.materia ORDER BY prom ASC LIMIT 10').fetchall()
            respuesta['respuesta'] = f"Materias con menor promedio: {rows[0]['materia']} ({rows[0]['prom']})" if rows else 'No hay datos suficientes.'
            respuesta['datos'] = [{'materia': r['materia'], 'promedio': r['prom']} for r in rows]
        elif any(p in pregunta for p in ['docente sin', 'profesor sin', 'sin registrar', 'días sin', 'no ha registrado']):
            from datetime import datetime, timedelta
            days = 10
            for w in pregunta.split():
                if w.isdigit():
                    days = int(w)
                    break
            cutoff = (datetime.now() - timedelta(days=days)).isoformat()
            rows = conn.execute('''SELECT p.id, p.nombre, MAX(n.created_at) as ultima_nota
                FROM profesores p JOIN actividades a ON a.profesor_id=p.id
                LEFT JOIN notas n ON n.actividad_id=a.id
                GROUP BY p.id HAVING ultima_nota IS NULL OR ultima_nota < ?''', (cutoff,)).fetchall()
            respuesta['respuesta'] = f'{len(rows)} docentes no han registrado notas en los últimos {days} días.'
            respuesta['datos'] = [{'nombre': r['nombre'], 'ultima_nota': r['ultima_nota'] or 'Nunca'} for r in rows]
        elif any(p in pregunta for p in ['inasistencia', 'ausencia', 'falta', 'asistencia baja', 'no asiste']):
            rows = conn.execute('SELECT a.nombre, a.curso, COUNT(*) as faltas FROM asistencia_v2 av JOIN alumnos a ON a.id=av.alumno_id WHERE av.estado IN ("X","E") AND a.activo=1 GROUP BY av.alumno_id HAVING faltas > 3 ORDER BY faltas DESC LIMIT 20').fetchall()
            respuesta['respuesta'] = f'{len(rows)} estudiantes tienen más de 3 inasistencias.'
            respuesta['datos'] = [{'nombre': r['nombre'], 'curso': r['curso'], 'faltas': r['faltas']} for r in rows]
        elif any(p in pregunta for p in ['mejorado', 'mejora', 'progreso', 'subió', 'aumentó']):
            rows = conn.execute('''SELECT a.nombre, a.curso, ROUND(AVG(CASE WHEN n.created_at < date("now","-15 days") THEN n.val END),2) as antes,
                ROUND(AVG(CASE WHEN n.created_at >= date("now","-15 days") THEN n.val END),2) as despues
                FROM notas n JOIN actividades ac ON ac.id=n.actividad_id JOIN alumnos a ON a.id=n.aid
                WHERE n.val IS NOT NULL GROUP BY n.aid HAVING despues > antes AND antes > 0
                ORDER BY (despues-antes) DESC LIMIT 15''').fetchall()
            respuesta['respuesta'] = f'{len(rows)} estudiantes han mejorado su rendimiento recientemente.'
            respuesta['datos'] = [{'nombre': r['nombre'], 'curso': r['curso'], 'antes': r['antes'], 'despues': r['despues']} for r in rows]
        elif any(p in pregunta for p in ['curso crítico', 'curso necesita', 'intervención', 'peor curso', 'curso bajo']):
            rows = conn.execute('SELECT a.curso, ROUND(AVG(n.val),2) as prom, COUNT(DISTINCT n.aid) as estudiantes FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY a.curso ORDER BY prom ASC LIMIT 10').fetchall()
            respuesta['respuesta'] = f"Curso con menor rendimiento: {rows[0]['curso']} ({rows[0]['prom']})" if rows else 'No hay datos.'
            respuesta['datos'] = [{'curso': r['curso'], 'promedio': r['prom'], 'estudiantes': r['estudiantes']} for r in rows]
        elif any(p in pregunta for p in ['resume', 'resumen', 'rendimiento de', 'grado ']):
            rows = conn.execute('SELECT a.curso, ROUND(AVG(n.val),2) as prom, COUNT(DISTINCT a.id) as acts FROM notas n JOIN actividades a ON a.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY a.curso ORDER BY a.curso').fetchall()
            resp = '<div style="display:flex;flex-direction:column;gap:4px;">'
            for r in rows:
                color = 'var(--success)' if r['prom'] >= 3.5 else 'var(--warning)' if r['prom'] >= 3.0 else 'var(--danger)'
                resp += f'<div style="display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid var(--border);"><span>{r["curso"]}</span><span style="font-weight:700;color:{color};">{r["prom"]}</span></div>'
            resp += '</div>'
            respuesta['respuesta'] = f'Rendimiento por curso ({len(rows)} cursos):'
            respuesta['html'] = resp
            respuesta['datos'] = [dict(r) for r in rows]
        elif any(p in pregunta for p in ['actividad perdiendo', 'actividad más', 'actividad difícil', 'tarea difícil', 'examen difícil']):
            rows = conn.execute('SELECT ac.nombre, ac.tipo, ac.curso, ROUND(AVG(n.val),2) as prom, COUNT(*) as notas FROM notas n JOIN actividades ac ON ac.id=n.actividad_id WHERE n.val IS NOT NULL GROUP BY ac.id HAVING prom < 3.0 ORDER BY prom ASC LIMIT 15').fetchall()
            respuesta['respuesta'] = f'{len(rows)} actividades tienen promedio menor a 3.0.'
            respuesta['datos'] = [{'nombre': r['nombre'], 'tipo': r['tipo'], 'curso': r['curso'], 'promedio': r['prom']} for r in rows]
        elif any(p in pregunta for p in ['bajaron', 'disminuyó', 'respecto', 'anterior', 'comparado']):
            rows = conn.execute('''
                SELECT a.materia,
                    ROUND(AVG(CASE WHEN COALESCE(a.periodo,1) <= 2 THEN n.val END),2) as p1,
                    ROUND(AVG(CASE WHEN COALESCE(a.periodo,1) > 2 THEN n.val END),2) as p2
                FROM notas n JOIN actividades a ON a.id=n.actividad_id
                WHERE n.val IS NOT NULL GROUP BY a.materia HAVING p2 < p1
                ORDER BY (p1-p2) DESC LIMIT 10''').fetchall()
            respuesta['respuesta'] = f'{len(rows)} materias bajaron su promedio respecto al período anterior.' if rows else 'No se detectaron bajas significativas.'
            respuesta['datos'] = [{'materia': r['materia'], 'periodo_anterior': r['p1'], 'periodo_actual': r['p2'], 'diferencia': round(r['p1'] - r['p2'], 2)} for r in rows]
        else:
            respuesta['respuesta'] = 'No entendí la pregunta. Intenta preguntar sobre: estudiantes en riesgo, materias con promedio bajo, docentes sin notas, inasistencias, o rendimiento por curso.'
        return jsonify(respuesta)
    except Exception as e:
        logger.error('asistente_analisis: %s', e)
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@teacher_bp.route('/<slug>/get_notas')
def get_notas(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'notas': []})
    aid = request.args.get('aid', type=int)
    if not aid:
        return jsonify({'notas': []})
    conn = f.conectar(slug)
    try:
        rows = conn.execute(
            '''SELECT a.nombre as actividad, n.val as valor
               FROM notas n
               JOIN actividades a ON a.id = n.actividad_id
               WHERE n.aid = ?
               ORDER BY a.fecha_limite DESC, a.nombre''',
            (aid,)).fetchall()
        return jsonify({'notas': [dict(r) for r in rows]})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/get_asistencia')
def get_asistencia(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'asistencias': []})
    aid = request.args.get('aid', type=int)
    if not aid:
        return jsonify({'asistencias': []})
    conn = f.conectar(slug)
    try:
        rows = conn.execute(
            '''SELECT fecha, CASE WHEN estado='P' THEN 1 ELSE 0 END as presente
               FROM asistencia WHERE aid = ?
               ORDER BY fecha DESC LIMIT 30''',
            (aid,)).fetchall()
        return jsonify({'asistencias': [dict(r) for r in rows]})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/get_observaciones')
def get_observaciones(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'observaciones': []})
    aid = request.args.get('aid', type=int)
    if not aid:
        return jsonify({'observaciones': []})
    conn = f.conectar(slug)
    try:
        rows = conn.execute(
            'SELECT id, materia, texto, fecha FROM observaciones WHERE aid=? ORDER BY fecha DESC LIMIT 20',
            (aid,)).fetchall()
        return jsonify({'observaciones': [dict(r) for r in rows]})
    finally:
        conn.close()


@teacher_bp.route('/<slug>/cursos')
def get_cursos_json(slug):
    f = _fa()
    f.require_colegio(slug)
    prof = f.get_profesor(slug)
    if not prof:
        return jsonify({'cursos': []})
    jornada, materia = f.get_sesion_jornada_materia(slug)
    if not jornada or not materia:
        return jsonify({'cursos': []})
    conn = f.conectar(slug)
    try:
        rows = conn.execute(
            'SELECT DISTINCT curso FROM asignaciones_curso WHERE profesor_id=? AND materia=? AND jornada=? ORDER BY curso',
            (prof['id'], materia, jornada)).fetchall()
        return jsonify({'cursos': [{'nombre': r['curso']} for r in rows]})
    finally:
        conn.close()
