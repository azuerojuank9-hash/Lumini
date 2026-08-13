"""Servicio centralizado de importación/exportación Excel (P7).

Agrupa la lectura segura de .xlsx, la validación de datos y la generación
de libros exportables para que importar/exportar notas, estudiantes y
asistencia compartan una única lógica.
"""
import io

EXTENSIONES_EXCEL = {'.xlsx'}
EXCEL_MAX_FILAS = 2000
EXCEL_MAX_COLUMNAS = 120

ESTADOS_ASISTENCIA = {
    'P': 'Presente', 'A': 'Ausente', 'T': 'Tardanza',
    'E': 'Excusa', 'X': 'Permiso', 'S': 'Salida anticipada',
}
_MAPA_ESTADOS = {nombre.lower(): k for k, nombre in ESTADOS_ASISTENCIA.items()}
_MAPA_ESTADOS.update({k.lower(): k for k in ESTADOS_ASISTENCIA})


def extension_excel_valida(filename):
    """Valida la extensión del archivo (no confía solo en esta: se usa junto
    con el límite de tamaño y la revalidación server-side)."""
    ext = ('.' + filename.rsplit('.', 1)[-1]).lower() if '.' in filename else ''
    return ext in EXTENSIONES_EXCEL


def leer_workbook(archivo_bytes, max_filas=EXCEL_MAX_FILAS, max_columnas=EXCEL_MAX_COLUMNAS):
    """Lee un .xlsx de forma segura (read_only). Devuelve (encabezados, filas)
    donde filas es [(nro_fila, [valores...]), ...]. Levanta ValueError con
    mensajes amigables al usuario."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(archivo_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError('El archivo no es un Excel .xlsx válido. Verifica el formato.')
    try:
        ws = wb.active
        if ws.max_column and ws.max_column > max_columnas:
            raise ValueError(
                f'El archivo tiene demasiadas columnas ({ws.max_column}). '
                f'Máximo permitido: {max_columnas}.')
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            raise ValueError('El archivo está vacío.')
        headers = [str(c).strip() if c is not None else '' for c in first]
        if not any(headers):
            raise ValueError('El archivo no tiene encabezados.')
        filas = []
        for i, row in enumerate(rows_iter, start=2):
            if i - 1 > max_filas:
                raise ValueError(
                    f'El archivo tiene más de {max_filas} filas de datos. '
                    f'Redúcelo e inténtalo de nuevo.')
            vals = [c if c is not None else '' for c in row]
            if not any(str(v).strip() for v in vals):
                continue
            filas.append((i, vals))
        if not filas:
            raise ValueError('El archivo no contiene datos.')
        return headers, filas
    finally:
        wb.close()


def parsear_nota(valor, nombre_columna='', escala_min=0.0, escala_max=5.0):
    """Convierte un valor a nota válida dentro de la escala del colegio.
    Devuelve (float|None, error|None)."""
    if valor is None or str(valor).strip() == '':
        return None, None
    try:
        num = float(str(valor).replace(',', '.'))
    except (ValueError, TypeError):
        return None, 'valor no numérico'
    if num < escala_min or num > escala_max:
        return None, f'la nota debe estar entre {escala_min} y {escala_max}'
    return round(num, 2), None


def parsear_fecha(valor):
    """Convierte a fecha ISO YYYY-MM-DD. Devuelve (fecha|None, error|None)."""
    if valor is None or str(valor).strip() == '':
        return None, None
    from datetime import datetime
    try:
        dt = datetime.strptime(str(valor).strip()[:10], '%Y-%m-%d')
        return dt.date().isoformat(), None
    except ValueError:
        return None, 'fecha no válida (usa AAAA-MM-DD)'


def parsear_estado_asistencia(valor):
    """Convierte un estado de asistencia (P/A/T/E/X/S o su nombre en español)
    a la clave canónica. Devuelve (estado|None, error|None)."""
    if valor is None or str(valor).strip() == '':
        return None, None
    v = str(valor).strip()
    if v.upper() in ESTADOS_ASISTENCIA:
        return v.upper(), None
    key = _MAPA_ESTADOS.get(v.lower())
    if key:
        return key, None
    return None, 'estado no válido (P/A/T/E/X/S)'


def revalidar_importacion_notas(data, conn, prof, materia, jornada, curso_sel, periodo,
                                escala_min=0.0, escala_max=5.0):
    """Revalida en el servidor el payload de importar_notas/confirmar para no
    confiar en los datos enviados por el cliente. Devuelve (ok, errores)."""
    errores = []
    if not isinstance(data, dict) or not data.get('all_ok'):
        return False, ['Datos de importación inválidos.']
    actividades = conn.execute(
        '''SELECT id, nombre FROM actividades
           WHERE profesor_id=? AND materia=? AND jornada=? AND curso=?
           AND COALESCE(periodo,1)=?''',
        (prof['id'], materia, jornada, curso_sel, periodo)).fetchall()
    acts_map = {a['nombre']: a['id'] for a in actividades}
    ids_actividades = set(acts_map.values())
    nombres_nuevas = {
        na.get('nombre') for na in data.get('nuevas_actividades', [])
        if na.get('nombre')
    }
    filas = data.get('filas', [])
    if not filas:
        return False, ['No hay filas para importar.']
    for fila in filas:
        if not fila.get('ok'):
            continue
        nfila = fila.get('fila', '?')
        aid = fila.get('aid')
        if not isinstance(aid, int):
            errores.append(f'Fila {nfila}: estudiante no válido.')
            continue
        alumno = conn.execute(
            'SELECT id FROM alumnos WHERE id=? AND curso=? AND jornada=? AND activo=1',
            (aid, curso_sel, jornada)).fetchone()
        if not alumno:
            errores.append(f'Fila {nfila}: estudiante no encontrado en el curso.')
            continue
        for ch in (fila.get('changes') or {}).values():
            valor = ch.get('valor')
            if ch.get('tipo') == 'actividad':
                act_id = ch.get('actividad_id')
                nombre_col = ch.get('nombre_col', '')
                if act_id is None:
                    if nombre_col not in nombres_nuevas:
                        errores.append(
                            f'Fila {nfila}: actividad "{nombre_col}" no reconocida.')
                        continue
                elif act_id not in ids_actividades:
                    errores.append(
                        f'Fila {nfila}: la actividad no pertenece a esta materia/curso.')
                    continue
                num, err = parsear_nota(valor, nombre_col, escala_min=escala_min, escala_max=escala_max)
                if err:
                    errores.append(f'Fila {nfila}: {nombre_col}: {err}.')
            elif ch.get('tipo') == 'evaluacion':
                num, err = parsear_nota(valor, escala_min=escala_min, escala_max=escala_max)
                if err:
                    errores.append(f'Fila {nfila}: Evaluación: {err}.')
            elif ch.get('tipo') == 'autoevaluacion':
                num, err = parsear_nota(valor, escala_min=escala_min, escala_max=escala_max)
                if err:
                    errores.append(f'Fila {nfila}: Autoevaluación: {err}.')
    return len(errores) == 0, errores


def wb_desde_filas(headers, filas, titulo_hoja='Datos', fill='6D28D9'):
    """Construye un Workbook openpyxl desde encabezados y filas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_hoja
    header_fill = PatternFill('solid', fgColor=fill)
    header_font = Font(bold=True, size=11, color='FFFFFF')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    for i, row in enumerate(filas, start=2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v).border = border
    for idx in range(len(headers)):
        letra = get_column_letter(idx + 1)
        ws.column_dimensions[letra].width = min(40, max(12, len(str(headers[idx])) + 4))
    return wb


def xlsx_bytes(wb):
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
